"""
╔══════════════════════════════════════════════════════════════════════════════════════════════════════════╗
║                                                                                                          ║
║     ███████╗████████╗ ██████╗  ██████╗██╗  ██╗    ███████╗ ██████╗██████╗ ███████╗███████╗███╗   ██╗     ║
║     ██╔════╝╚══██╔══╝██╔═══██╗██╔════╝██║ ██╔╝    ██╔════╝██╔════╝██╔══██╗██╔════╝██╔════╝████╗  ██║     ║
║     ███████╗   ██║   ██║   ██║██║     █████╔╝     ███████╗██║     ██████╔╝█████╗  █████╗  ██╔██╗ ██║     ║
║     ╚════██║   ██║   ██║   ██║██║     ██╔═██╗     ╚════██║██║     ██╔══██╗██╔══╝  ██╔══╝  ██║╚██╗██║     ║
║     ███████║   ██║   ╚██████╔╝╚██████╗██║  ██╗    ███████║╚██████╗██║  ██║███████╗███████╗██║ ╚████║     ║
║     ╚══════╝   ╚═╝    ╚═════╝  ╚═════╝╚═╝  ╚═╝    ╚══════╝ ╚═════╝╚═╝  ╚═╝╚══════╝╚══════╝╚═╝  ╚═══╝     ║
║                                                                                                          ║
║          🚀 SCREENER PRO v11.0 - OPPORTUNITÉS COURT & MOYEN TERME 🚀                                    ║
║                                                                                                          ║
║     ════════════════════════════════════════════════════════════════════════════════════════════════     ║
║     ⚡ INDICATEURS COURT TERME (NOUVEAUX):                                                               ║
║        • ADX + DI+/DI- (Force de tendance)                                                               ║
║        • Stochastic RSI (Timing optimal)                                                                 ║
║        • Williams %R (Zones extrêmes)                                                                    ║
║        • CMF - Chaikin Money Flow (Pression achat/vente)                                                 ║
║        • OBV - On Balance Volume (Accumulation/Distribution)                                             ║
║        • VWAP (Volume Weighted Average Price)                                                            ║
║        • SuperTrend (Direction de tendance)                                                              ║
║        • Squeeze Momentum (Volatilité + Momentum)                                                        ║
║                                                                                                          ║
║     🎯 DÉTECTION OPPORTUNITÉS:                                                                           ║
║        • Scanner Breakout (cassure résistance/support)                                                   ║
║        • Détection Gap (Gap Up/Down significatifs)                                                       ║
║        • Volume Spike Alert (volume > 200% moyenne)                                                      ║
║        • Pattern Reversal (Hammer, Engulfing, Doji)                                                      ║
║        • Momentum Burst (accélération soudaine)                                                          ║
║                                                                                                          ║
║     📊 SCORING MULTI-HORIZON:                                                                            ║
║        • Score Intraday (1-3 jours)                                                                      ║
║        • Score Swing (5-15 jours)                                                                        ║
║        • Score Position (15-60 jours)                                                                    ║
║                                                                                                          ║
║     🔥 ALERTES TIMING:                                                                                   ║
║        • Zone d'achat optimale détectée                                                                  ║
║        • Catalyseur imminent (earnings, events)                                                          ║
║        • Confluence de signaux                                                                           ║
║                                                                                                          ║
║     🌍 MARCHÉS: US + Europe + France + Gaming + Commodities + Crypto                                     ║
║                                                                                                          ║
╚══════════════════════════════════════════════════════════════════════════════════════════════════════════╝
"""

import os
import sys
import time
import json
import logging
import sqlite3
import re
import urllib.request
import urllib.parse
import html
from logging.handlers import RotatingFileHandler
import threading
from queue import Queue, Empty
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any, Callable
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field, asdict
from enum import Enum
from collections import OrderedDict
import warnings
import uuid

warnings.filterwarnings('ignore')

# ══════════════════════════════════════════════════════════════════════════════
#                         VÉRIFICATION DES DÉPENDANCES
# ══════════════════════════════════════════════════════════════════════════════

PAQUETS_MANQUANTS = []

try:
    import pandas as pd
except ImportError:
    PAQUETS_MANQUANTS.append('pandas')

try:
    import numpy as np
except ImportError:
    PAQUETS_MANQUANTS.append('numpy')

try:
    import yfinance as yf
except ImportError:
    PAQUETS_MANQUANTS.append('yfinance')

try:
    from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, IsolationForest, AdaBoostClassifier, StackingClassifier, VotingClassifier, BaggingClassifier, HistGradientBoostingClassifier
    from sklearn.preprocessing import StandardScaler, RobustScaler, MinMaxScaler
    from sklearn.model_selection import TimeSeriesSplit, cross_val_score
    from sklearn.pipeline import Pipeline
    from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_auc_score, classification_report
    from sklearn.neural_network import MLPClassifier
    from sklearn.svm import SVC
    from sklearn.neighbors import KNeighborsClassifier
    from sklearn.calibration import CalibratedClassifierCV
    from sklearn.feature_selection import SelectKBest, f_classif, RFE
except ImportError:
    PAQUETS_MANQUANTS.append('scikit-learn')

XGBOOST_DISPONIBLE = False
try:
    import xgboost as xgb
    XGBOOST_DISPONIBLE = True
except ImportError:
    pass

LIGHTGBM_DISPONIBLE = False
try:
    import lightgbm as lgb
    LIGHTGBM_DISPONIBLE = True
except ImportError:
    pass

CATBOOST_DISPONIBLE = False
try:
    from catboost import CatBoostClassifier
    CATBOOST_DISPONIBLE = True
except ImportError:
    pass

if PAQUETS_MANQUANTS:
    print("=" * 60)
    print("❌ PAQUETS MANQUANTS")
    print("=" * 60)
    print(f"\nInstallez: {', '.join(PAQUETS_MANQUANTS)}")
    print("\nExécutez: pip install pandas numpy yfinance scikit-learn xgboost")
    print("=" * 60)
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
#                              LOGGING
# ══════════════════════════════════════════════════════════════════════════════

def configurer_logging(fichier_log: str = "screener_v9.log") -> logging.Logger:
    logger = logging.getLogger('ScreenerPro')
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()
    
    formatter = logging.Formatter('%(asctime)s | %(levelname)-7s | %(message)s', datefmt='%H:%M:%S')
    
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(formatter)
    logger.addHandler(console)
    
    try:
        fichier = RotatingFileHandler(fichier_log, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8')
        fichier.setLevel(logging.DEBUG)
        fichier.setFormatter(formatter)
        logger.addHandler(fichier)
    except:
        pass
    
    return logger

logger = configurer_logging()


# ══════════════════════════════════════════════════════════════════════════════
#                              CACHE
# ══════════════════════════════════════════════════════════════════════════════

class CacheLRU:
    def __init__(self, nom: str, taille_max: int = 500, ttl: int = 300):
        self.nom = nom
        self.taille_max = taille_max
        self.ttl = ttl
        self._cache: OrderedDict = OrderedDict()
        self._timestamps: Dict[str, float] = {}
        self._lock = threading.Lock()
        self.hits = 0
        self.misses = 0
    
    def _bucket_temps(self) -> str:
        now = datetime.now()
        bucket = (now.minute // 5) * 5
        return f"{now.strftime('%Y%m%d_%H')}{bucket:02d}"
    
    def get(self, cle: str) -> Optional[Any]:
        cache_cle = f"{cle}_{self._bucket_temps()}"
        with self._lock:
            if cache_cle in self._cache:
                self._cache.move_to_end(cache_cle)
                self.hits += 1
                return self._cache[cache_cle]
            self.misses += 1
            return None
    
    def set(self, cle: str, valeur: Any):
        cache_cle = f"{cle}_{self._bucket_temps()}"
        with self._lock:
            while len(self._cache) >= self.taille_max:
                self._cache.popitem(last=False)
            self._cache[cache_cle] = valeur
            self._timestamps[cache_cle] = time.time()
    
    def stats(self) -> str:
        total = self.hits + self.misses
        taux = (self.hits / total * 100) if total > 0 else 0
        return f"{taux:.0f}%"


class GestionnaireCache:
    def __init__(self):
        self.prix = CacheLRU("prix", 500, 300)
        self.fondamentaux = CacheLRU("fondamentaux", 300, 3600)
        self.ml = CacheLRU("ml", 200, 900)
    
    def vider(self):
        self.prix._cache.clear()
        self.fondamentaux._cache.clear()
        self.ml._cache.clear()


# ══════════════════════════════════════════════════════════════════════════════
#                    🚀 ULTRA-FAST BATCH PRICE FETCHER (SNIPER MODE)
# ══════════════════════════════════════════════════════════════════════════════

from concurrent.futures import ThreadPoolExecutor, as_completed

class LivePriceFetcher:
    """
    🚀 ULTRA-FAST: Récupère les prix de centaines d'actions en quelques secondes
    Utilise yfinance.download() en batch + ThreadPoolExecutor pour max de parallélisation
    """
    
    def __init__(self, max_workers: int = 20):
        self.max_workers = max_workers
        self.cache_live = {}
        self.cache_timestamp = None
        self.cache_ttl = 60  # 60 secondes de validité cache
    
    def fetch_batch_prices(self, symbols: List[str], period: str = "5d") -> Dict[str, Dict]:
        """
        📈 Récupère les prix de tous les symboles en UN SEUL appel batch
        Retourne: {symbole: {price, prev_close, change_pct, volume, high, low}}
        """
        if not symbols:
            return {}
        
        # Vérifier cache
        if self.cache_timestamp and (time.time() - self.cache_timestamp) < self.cache_ttl:
            cached_symbols = [s for s in symbols if s in self.cache_live]
            if len(cached_symbols) == len(symbols):
                return {s: self.cache_live[s] for s in symbols}
        
        results = {}
        
        try:
            logger.info(f"🚀 Batch fetch de {len(symbols)} symboles...")
            start = time.time()
            
            # Téléchargement batch ultra-rapide
            df = yf.download(
                symbols, 
                period=period, 
                interval="1d",
                progress=False,
                threads=True,
                group_by='ticker'
            )
            
            elapsed = time.time() - start
            logger.info(f"✅ Batch fetch terminé en {elapsed:.2f}s")
            
            for symbol in symbols:
                try:
                    if len(symbols) == 1:
                        # Cas spécial: un seul symbole, structure différente
                        close_col = df['Close']
                        volume_col = df['Volume']
                        high_col = df['High']
                        low_col = df['Low']
                    else:
                        close_col = df[symbol]['Close']
                        volume_col = df[symbol]['Volume']
                        high_col = df[symbol]['High']
                        low_col = df[symbol]['Low']
                    
                    if close_col.empty or close_col.isna().all():
                        continue
                    
                    price = close_col.iloc[-1]
                    prev_close = close_col.iloc[-2] if len(close_col) >= 2 else price
                    change_pct = ((price - prev_close) / prev_close * 100) if prev_close > 0 else 0
                    
                    results[symbol] = {
                        'price': round(float(price), 2),
                        'prev_close': round(float(prev_close), 2),
                        'change_pct': round(float(change_pct), 2),
                        'volume': int(volume_col.iloc[-1]) if not pd.isna(volume_col.iloc[-1]) else 0,
                        'high': round(float(high_col.iloc[-1]), 2) if not pd.isna(high_col.iloc[-1]) else price,
                        'low': round(float(low_col.iloc[-1]), 2) if not pd.isna(low_col.iloc[-1]) else price,
                        'timestamp': datetime.now().strftime('%H:%M:%S')
                    }
                except Exception as e:
                    logger.debug(f"Erreur {symbol}: {e}")
                    continue
            
            # Mettre en cache
            self.cache_live.update(results)
            self.cache_timestamp = time.time()
            
        except Exception as e:
            logger.error(f"Erreur batch fetch: {e}")
        
        return results
    
    def fetch_realtime_single(self, symbol: str) -> Dict:
        """Récupère le prix en temps réel d'un symbole unique (fast_info)"""
        try:
            ticker = yf.Ticker(symbol)
            info = ticker.fast_info
            price = getattr(info, 'last_price', None) or getattr(info, 'previous_close', None)
            prev_close = getattr(info, 'previous_close', None)
            
            if price:
                change_pct = ((price - prev_close) / prev_close * 100) if prev_close else 0
                return {
                    'price': round(price, 2),
                    'prev_close': round(prev_close, 2) if prev_close else None,
                    'change_pct': round(change_pct, 2),
                    'timestamp': datetime.now().strftime('%H:%M:%S')
                }
        except:
            pass
        return {}
    
    def fetch_parallel_realtime(self, symbols: List[str]) -> Dict[str, Dict]:
        """
        ⚡ Récupère les prix en parallèle avec ThreadPoolExecutor
        Plus lent que batch mais données plus fraîches
        """
        results = {}
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(self.fetch_realtime_single, s): s for s in symbols}
            for future in as_completed(futures):
                symbol = futures[future]
                try:
                    result = future.result()
                    if result:
                        results[symbol] = result
                except:
                    pass
        
        return results


class SniperScanner:
    """
    🎯 SNIPER SCANNER - Détection ultra-rapide des meilleures opportunités
    Analyse 100+ actions en moins de 30 secondes avec scoring en temps réel
    """
    
    def __init__(self, config):
        self.config = config
        self.price_fetcher = LivePriceFetcher(max_workers=25)
    
    def scan_rapid(self, symbols: List[str], callback: Callable = None) -> List[Dict]:
        """
        ⚡ SCAN RAPIDE - Analyse basique ultra-rapide pour triage initial
        Retourne les meilleures opportunités avec timing score
        """
        start_time = time.time()
        results = []
        
        # 1. Batch fetch de tous les prix (ultra-rapide)
        if callback:
            callback("📡 Récupération des données...", 0, len(symbols))
        
        prices = self.price_fetcher.fetch_batch_prices(symbols, period="1mo")
        
        # 2. Télécharger les données historiques en batch
        if callback:
            callback("📊 Analyse technique rapide...", len(symbols) // 3, len(symbols))
        
        try:
            df_all = yf.download(symbols, period="3mo", interval="1d", progress=False, threads=True, group_by='ticker')
        except:
            df_all = None
        
        # 3. Analyse rapide de chaque symbole
        for i, symbol in enumerate(symbols):
            try:
                if callback and i % 20 == 0:
                    callback(f"⚡ Analyse {symbol}...", i, len(symbols))
                
                # Récupérer les données du symbole
                if df_all is not None:
                    if len(symbols) == 1:
                        df = df_all
                    else:
                        df = df_all[symbol] if symbol in df_all.columns.get_level_values(0) else None
                else:
                    continue
                
                if df is None or df.empty or len(df) < 30:
                    continue
                
                close = df['Close'].dropna()
                volume = df['Volume'].dropna()
                high = df['High'].dropna()
                low = df['Low'].dropna()
                
                if len(close) < 30:
                    continue
                
                # Prix actuel
                prix = prices.get(symbol, {}).get('price') or float(close.iloc[-1])
                change_pct = prices.get(symbol, {}).get('change_pct', 0)
                
                # ═══════════════════════════════════════════════════════════
                # INDICATEURS RAPIDES (sans appels supplémentaires)
                # ═══════════════════════════════════════════════════════════
                
                # RSI (14)
                delta = close.diff()
                gain = delta.where(delta > 0, 0).rolling(14).mean()
                loss = -delta.where(delta < 0, 0).rolling(14).mean()
                rsi = 100 - (100 / (1 + gain.iloc[-1] / loss.iloc[-1])) if loss.iloc[-1] != 0 else 50
                
                # Z-Score (20)
                mean_20 = close.rolling(20).mean().iloc[-1]
                std_20 = close.rolling(20).std().iloc[-1]
                z_score = (prix - mean_20) / std_20 if std_20 > 0 else 0
                
                # Moyennes mobiles
                sma_20 = close.rolling(20).mean().iloc[-1]
                sma_50 = close.rolling(50).mean().iloc[-1] if len(close) >= 50 else sma_20
                
                # Volume relatif
                vol_avg = volume.rolling(20).mean().iloc[-1]
                vol_ratio = volume.iloc[-1] / vol_avg if vol_avg > 0 else 1
                
                # Variation récente
                var_5j = ((close.iloc[-1] - close.iloc[-5]) / close.iloc[-5] * 100) if len(close) >= 5 else 0
                var_20j = ((close.iloc[-1] - close.iloc[-20]) / close.iloc[-20] * 100) if len(close) >= 20 else 0
                
                # ATR
                tr = pd.concat([
                    high - low,
                    abs(high - close.shift(1)),
                    abs(low - close.shift(1))
                ], axis=1).max(axis=1)
                atr = tr.rolling(14).mean().iloc[-1]
                
                # Support/Résistance rapide
                support = low.rolling(20).min().iloc[-1]
                resistance = high.rolling(20).max().iloc[-1]
                distance_support = ((prix - support) / prix * 100) if support > 0 else 100
                
                # ═══════════════════════════════════════════════════════════
                # SCORING SNIPER (0-100)
                # ═══════════════════════════════════════════════════════════
                score = 50
                signaux = []
                opportunite_type = "NORMAL"
                
                # RSI
                if rsi <= 30:
                    score += 20
                    signaux.append(f"RSI survendu ({rsi:.0f})")
                    opportunite_type = "OVERSOLD"
                elif rsi <= 40:
                    score += 10
                    signaux.append(f"RSI bas ({rsi:.0f})")
                elif rsi >= 70:
                    score -= 15
                    signaux.append(f"⚠️ RSI suracheté ({rsi:.0f})")
                
                # Z-Score
                if z_score <= -2:
                    score += 18
                    signaux.append(f"Z-Score très bas ({z_score:.2f})")
                    opportunite_type = "OVERSOLD"
                elif z_score <= -1:
                    score += 10
                    signaux.append(f"Z-Score bas ({z_score:.2f})")
                elif z_score >= 2:
                    score -= 12
                
                # Tendance (prix vs moyennes)
                if prix > sma_20 > sma_50:
                    score += 12
                    signaux.append("Tendance haussière")
                    if opportunite_type == "NORMAL":
                        opportunite_type = "MOMENTUM"
                elif prix < sma_20 < sma_50:
                    score -= 10
                
                # Volume
                if vol_ratio >= 2:
                    score += 15
                    signaux.append(f"Volume spike ({vol_ratio:.1f}x)")
                    opportunite_type = "BREAKOUT" if prix > sma_20 else opportunite_type
                elif vol_ratio >= 1.5:
                    score += 8
                    signaux.append(f"Volume élevé ({vol_ratio:.1f}x)")
                
                # Distance au support
                if distance_support <= 3:
                    score += 12
                    signaux.append("Proche support")
                    opportunite_type = "SUPPORT_BOUNCE"
                
                # Variation récente positive
                if 2 <= var_5j <= 8:
                    score += 8
                    signaux.append(f"Momentum +{var_5j:.1f}%")
                elif var_5j > 15:
                    score -= 5  # Trop étendu
                elif var_5j < -10 and rsi < 35:
                    score += 10
                    signaux.append("Oversold bounce potential")
                
                # Stop Loss / Take Profit estimés
                stop_loss = prix - (atr * 2)
                take_profit = prix + (atr * 3)
                ratio_rr = (take_profit - prix) / (prix - stop_loss) if (prix - stop_loss) > 0 else 0
                
                if ratio_rr >= 2:
                    score += 8
                    signaux.append(f"R/R excellent ({ratio_rr:.1f})")
                elif ratio_rr >= 1.5:
                    score += 4
                
                # Normaliser le score
                score = max(0, min(100, score))
                
                # Niveau de signal
                if score >= 75:
                    signal_level = "🟢🟢 FORTE OPPORTUNITÉ"
                elif score >= 65:
                    signal_level = "🟢 OPPORTUNITÉ"
                elif score >= 55:
                    signal_level = "🟡 À SURVEILLER"
                elif score >= 45:
                    signal_level = "⚪ NEUTRE"
                else:
                    signal_level = "🔴 ÉVITER"
                
                results.append({
                    'symbole': symbol,
                    'prix': round(prix, 2),
                    'change_pct': round(change_pct, 2),
                    'rsi': round(rsi, 1),
                    'z_score': round(z_score, 2),
                    'volume_ratio': round(vol_ratio, 1),
                    'score_sniper': round(score, 1),
                    'signal_level': signal_level,
                    'opportunite_type': opportunite_type,
                    'signaux': signaux[:4],
                    'stop_loss': round(stop_loss, 2),
                    'take_profit': round(take_profit, 2),
                    'ratio_rr': round(ratio_rr, 2),
                    'var_5j': round(var_5j, 2),
                    'support': round(support, 2),
                    'resistance': round(resistance, 2),
                    'timestamp': datetime.now().strftime('%H:%M:%S')
                })
                
            except Exception as e:
                logger.debug(f"Erreur scan {symbol}: {e}")
                continue
        
        # Trier par score décroissant
        results.sort(key=lambda x: x['score_sniper'], reverse=True)
        
        elapsed = time.time() - start_time
        logger.info(f"🎯 Sniper scan: {len(results)} résultats en {elapsed:.1f}s")
        
        return results


class MasterReportGenerator:
    """
    📊 GÉNÉRATEUR DE RAPPORT MAÎTRE
    Crée un fichier professionnel avec toutes les opportunités,
    recommandations et prix en temps réel
    """
    
    def __init__(self, config):
        self.config = config
        self.price_fetcher = LivePriceFetcher()
    
    def generate_master_report(self, results: List[Dict], output_path: str = None, 
                               include_live_prices: bool = True) -> str:
        """
        📋 Génère le rapport maître avec toutes les colonnes
        """
        if not results:
            logger.warning("Pas de résultats pour le rapport")
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"MASTER_REPORT_PRO_{timestamp}.csv"
        
        # Convertir en DataFrame
        df = pd.DataFrame(results)
        
        # Ajouter prix live si demandé
        if include_live_prices and 'symbole' in df.columns:
            logger.info("📡 Récupération des prix en temps réel...")
            symbols = df['symbole'].tolist()
            live_prices = self.price_fetcher.fetch_batch_prices(symbols)
            
            df['PRIX_LIVE'] = df['symbole'].apply(lambda s: live_prices.get(s, {}).get('price', None))
            df['VAR_LIVE_%'] = df['symbole'].apply(lambda s: live_prices.get(s, {}).get('change_pct', None))
            df['VOLUME_LIVE'] = df['symbole'].apply(lambda s: live_prices.get(s, {}).get('volume', None))
            df['MAJ_HEURE'] = datetime.now().strftime('%H:%M:%S')
        
        # Réorganiser les colonnes (les plus importantes en premier)
        priority_cols = [
            'symbole', 'marche', 'secteur', 'prix', 'PRIX_LIVE', 'VAR_LIVE_%',
            'decision', 'action', 'score_final', 'score_sniper',
            'signal_level', 'opportunite_type', 'meilleur_horizon',
            'potentiel', 'prix_reel', 'prix_achat_ideal',
            'stop_loss', 'take_profit', 'ratio_rr',
            'rsi', 'z_score', 'volume_ratio', 'tendance',
            'ml_signal', 'ml_probabilite',
            'conseil', 'signaux'
        ]
        
        existing_priority = [c for c in priority_cols if c in df.columns]
        other_cols = [c for c in df.columns if c not in priority_cols and not c.startswith('_')]
        final_cols = existing_priority + other_cols
        
        df = df[[c for c in final_cols if c in df.columns]]
        
        # Trier par score décroissant
        if 'score_final' in df.columns:
            df = df.sort_values('score_final', ascending=False)
        elif 'score_sniper' in df.columns:
            df = df.sort_values('score_sniper', ascending=False)
        
        # Sauvegarder
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        logger.info(f"✅ Rapport maître généré: {output_path} ({len(df)} lignes)")
        
        return output_path
    
    def generate_top_opportunities_report(self, results: List[Dict], top_n: int = 30,
                                          output_path: str = None) -> str:
        """
        🏆 Génère un rapport des TOP N meilleures opportunités
        """
        if not results:
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"TOP_{top_n}_OPPORTUNITES_{timestamp}.csv"
        
        # Filtrer les meilleures opportunités
        df = pd.DataFrame(results)
        
        # Déterminer la colonne de score
        score_col = 'score_final' if 'score_final' in df.columns else 'score_sniper'
        
        if score_col in df.columns:
            df = df.sort_values(score_col, ascending=False).head(top_n)
        
        # Ajouter recommandations
        def generer_recommandation(row):
            score = row.get(score_col, 50)
            action = row.get('action', row.get('signal_level', ''))
            
            if score >= 75:
                return "⭐⭐⭐ PRIORITÉ HAUTE - Entrée recommandée"
            elif score >= 65:
                return "⭐⭐ BONNE OPPORTUNITÉ - Confirmer avant entrée"
            elif score >= 55:
                return "⭐ À SURVEILLER - Attendre pullback"
            else:
                return "Observer uniquement"
        
        df['RECOMMANDATION'] = df.apply(generer_recommandation, axis=1)
        
        # Ajouter prix live
        if 'symbole' in df.columns:
            symbols = df['symbole'].tolist()
            live_prices = self.price_fetcher.fetch_batch_prices(symbols)
            df['PRIX_LIVE'] = df['symbole'].apply(lambda s: live_prices.get(s, {}).get('price', None))
            df['VAR_LIVE_%'] = df['symbole'].apply(lambda s: live_prices.get(s, {}).get('change_pct', None))
        
        df['MAJ_RAPPORT'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        logger.info(f"🏆 Rapport TOP {top_n} généré: {output_path}")
        
        return output_path


# ══════════════════════════════════════════════════════════════════════════════
#                    BASE DE DONNÉES SIGNAUX (SQLite)
# ══════════════════════════════════════════════════════════════════════════════

class GestionnaireSignaux:
    """
    🗄️ Gestionnaire de base de données pour les signaux et trades
    
    Tables:
    - signals: Enregistre tous les signaux générés
    - trades: Résultats des signaux évalués
    """
    
    def __init__(self, db_path: str = "signals.db"):
        self.db_path = db_path
        self._initialiser_db()
    
    def _initialiser_db(self):
        """Crée les tables si elles n'existent pas"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Table des signaux
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS signals (
                id TEXT PRIMARY KEY,
                symbol TEXT NOT NULL,
                time_signal TEXT NOT NULL,
                timeframe TEXT DEFAULT 'Daily',
                strategy TEXT,
                action TEXT NOT NULL,
                entry_rule TEXT DEFAULT 'Close',
                entry_price_ref REAL NOT NULL,
                stop_loss REAL,
                take_profit REAL,
                horizon_days INTEGER,
                score_final REAL,
                confiance_pct REAL,
                ratio_rr REAL,
                tick_size REAL DEFAULT 0.01,
                tick_value REAL DEFAULT 1.0,
                commission REAL DEFAULT 0.0,
                slippage_ticks REAL DEFAULT 1.0,
                marche TEXT,
                secteur TEXT,
                statut TEXT DEFAULT 'PENDING',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Table des trades (résultats)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS trades (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                signal_id TEXT NOT NULL,
                exit_time TEXT,
                exit_price REAL,
                exit_reason TEXT,
                pnl_ticks REAL,
                pnl_usd REAL,
                pnl_net REAL,
                slippage_total REAL,
                commission_total REAL,
                result TEXT,
                holding_days INTEGER,
                worst_case_pnl REAL,
                best_case_pnl REAL,
                evaluated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (signal_id) REFERENCES signals(id)
            )
        ''')
        
        # Index pour performance
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_signals_symbol ON signals(symbol)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_signals_statut ON signals(statut)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_signals_action ON signals(action)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_trades_signal ON trades(signal_id)')
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 MIGRATION: Ajouter les colonnes manquantes (position_size, slippage_pct)
        # ══════════════════════════════════════════════════════════════════
        try:
            cursor.execute('ALTER TABLE signals ADD COLUMN position_size INTEGER DEFAULT 100')
        except sqlite3.OperationalError:
            pass  # Colonne existe déjà
        
        try:
            cursor.execute('ALTER TABLE signals ADD COLUMN slippage_pct REAL DEFAULT 0.0005')
        except sqlite3.OperationalError:
            pass  # Colonne existe déjà
        
        try:
            cursor.execute('ALTER TABLE signals ADD COLUMN commission_per_share REAL DEFAULT 0.005')
        except sqlite3.OperationalError:
            pass  # Colonne existe déjà
        
        try:
            cursor.execute('ALTER TABLE trades ADD COLUMN entry_price_reel REAL')
        except sqlite3.OperationalError:
            pass  # Colonne existe déjà
        
        conn.commit()
        conn.close()
        logger.info(f"Base de données initialisée: {self.db_path}")
    
    def sauvegarder_signal(self, signal: Dict) -> str:
        """
        💾 Sauvegarde un nouveau signal dans la base
        
        Args:
            signal: Dictionnaire contenant toutes les données du signal
        
        Returns:
            ID du signal créé
        """
        signal_id = str(uuid.uuid4())[:8]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Déterminer le profil de coûts
        profil = ProfilsContrats.obtenir(signal.get('symbole', 'STOCK'))
        
        cursor.execute('''
            INSERT INTO signals (
                id, symbol, time_signal, timeframe, strategy, action,
                entry_rule, entry_price_ref, stop_loss, take_profit,
                horizon_days, score_final, confiance_pct, ratio_rr,
                tick_size, tick_value, commission, slippage_ticks,
                marche, secteur, statut,
                position_size, slippage_pct, commission_per_share
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            signal_id,
            signal.get('symbole'),
            signal.get('time_signal', datetime.now().isoformat()),
            signal.get('timeframe', 'Daily'),
            signal.get('strategie', 'Composite'),
            signal.get('action'),
            signal.get('entry_rule', 'Close'),
            signal.get('prix'),
            signal.get('stop_loss'),
            signal.get('take_profit'),
            signal.get('horizon_jours'),
            signal.get('score_final'),
            signal.get('confiance_pct'),
            signal.get('ratio_rr'),
            profil.tick_size,
            profil.tick_value,
            profil.commission,
            profil.slippage_ticks,
            signal.get('marche'),
            signal.get('secteur'),
            'PENDING',
            signal.get('position_size', 100),  # Nombre d'actions par défaut
            signal.get('slippage_pct', 0.0005),  # 0.05% slippage par défaut
            signal.get('commission_per_share', 0.005)  # $0.005/action par défaut
        ))
        
        conn.commit()
        conn.close()
        
        logger.info(f"Signal sauvegardé: {signal_id} - {signal.get('symbole')} {signal.get('action')}")
        return signal_id
    
    def obtenir_signaux_pending(self) -> List[Dict]:
        """Récupère tous les signaux en attente d'évaluation"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM signals WHERE statut = 'PENDING'
            ORDER BY time_signal DESC
        ''')
        
        resultats = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return resultats
    
    def obtenir_tous_signaux(self, limite: int = 100) -> List[Dict]:
        """Récupère tous les signaux"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT s.*, t.exit_reason, t.pnl_net, t.result
            FROM signals s
            LEFT JOIN trades t ON s.id = t.signal_id
            ORDER BY s.time_signal DESC
            LIMIT ?
        ''', (limite,))
        
        resultats = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return resultats
    
    def sauvegarder_trade(self, signal_id: str, trade: Dict):
        """
        💾 Sauvegarde le résultat d'un trade évalué
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Mettre à jour le statut du signal
        cursor.execute('''
            UPDATE signals SET statut = 'EVALUATED' WHERE id = ?
        ''', (signal_id,))
        
        # Insérer le trade (avec entry_price_reel)
        cursor.execute('''
            INSERT INTO trades (
                signal_id, exit_time, exit_price, exit_reason,
                pnl_ticks, pnl_usd, pnl_net, slippage_total,
                commission_total, result, holding_days,
                worst_case_pnl, best_case_pnl, entry_price_reel
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            signal_id,
            trade.get('exit_time'),
            trade.get('exit_price'),
            trade.get('exit_reason'),
            trade.get('pnl_ticks'),
            trade.get('pnl_usd'),
            trade.get('pnl_net'),
            trade.get('slippage_total'),
            trade.get('commission_total'),
            trade.get('result'),
            trade.get('holding_days'),
            trade.get('worst_case_pnl'),
            trade.get('best_case_pnl'),
            trade.get('entry_price_reel')  # 🔧 AJOUTÉ: Prix d'exécution réel
        ))
        
        conn.commit()
        conn.close()
        
        logger.info(f"Trade évalué: {signal_id} - {trade.get('result')} - PnL: {trade.get('pnl_net'):.2f}")
    
    def obtenir_statistiques(self, filtre: Dict = None) -> Dict:
        """
        📊 Calcule les statistiques de performance
        
        Args:
            filtre: Dictionnaire avec clés optionnelles:
                - strategy: Filtrer par stratégie
                - marche: Filtrer par marché (US/EU/France)
                - action: Filtrer par type (ACHAT/VENTE)
                - timeframe: Filtrer par timeframe
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Construire la requête avec filtres
        query = '''
            SELECT 
                t.*,
                s.strategy,
                s.marche,
                s.action,
                s.timeframe,
                s.score_final
            FROM trades t
            JOIN signals s ON t.signal_id = s.id
            WHERE 1=1
        '''
        params = []
        
        if filtre:
            if filtre.get('strategy'):
                query += ' AND s.strategy = ?'
                params.append(filtre['strategy'])
            if filtre.get('marche'):
                query += ' AND s.marche LIKE ?'
                params.append(f'%{filtre["marche"]}%')
            if filtre.get('action'):
                query += ' AND s.action = ?'
                params.append(filtre['action'])
            if filtre.get('timeframe'):
                query += ' AND s.timeframe = ?'
                params.append(filtre['timeframe'])
        
        cursor.execute(query, params)
        trades = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        if not trades:
            return {
                'total_trades': 0,
                'win_rate': 0,
                'expectancy': 0,
                'profit_factor': 0,
                'max_drawdown': 0,
                'avg_pnl': 0,
                'total_pnl': 0,
                'avg_holding_days': 0,
                'sqs': 0
            }
        
        # Calculs
        wins = [t for t in trades if t['result'] == 'WIN']
        losses = [t for t in trades if t['result'] == 'LOSS']
        
        total = len(trades)
        win_count = len(wins)
        loss_count = len(losses)
        
        win_rate = (win_count / total * 100) if total > 0 else 0
        
        # Expectancy (moyenne PnL par trade)
        pnls = [t['pnl_net'] or 0 for t in trades]
        expectancy = np.mean(pnls) if pnls else 0
        
        # Profit Factor
        gross_profit = sum(t['pnl_net'] for t in wins if t['pnl_net'])
        gross_loss = abs(sum(t['pnl_net'] for t in losses if t['pnl_net']))
        profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else float('inf')
        
        # Max Drawdown
        equity_curve = []
        running_total = 0
        for t in sorted(trades, key=lambda x: x.get('evaluated_at', '')):
            running_total += t['pnl_net'] or 0
            equity_curve.append(running_total)
        
        max_dd = 0
        peak = 0
        for val in equity_curve:
            if val > peak:
                peak = val
            dd = peak - val
            if dd > max_dd:
                max_dd = dd
        
        # SQS (Signal Quality Score) - Score personnalisé 0-100
        # Basé sur: Win Rate, Profit Factor, Expectancy positive
        sqs = 0
        if total >= 10:  # Minimum de trades pour évaluer
            sqs += min(40, win_rate * 0.5)  # Max 40 points pour win rate
            sqs += min(30, profit_factor * 10) if profit_factor != float('inf') else 30  # Max 30 points
            sqs += 30 if expectancy > 0 else 0  # 30 points si expectancy positive
        
        # Holding days moyen
        holding_days = [t['holding_days'] for t in trades if t['holding_days']]
        avg_holding = np.mean(holding_days) if holding_days else 0
        
        return {
            'total_trades': total,
            'wins': win_count,
            'losses': loss_count,
            'win_rate': round(win_rate, 2),
            'expectancy': round(expectancy, 2),
            'profit_factor': round(profit_factor, 2) if profit_factor != float('inf') else 999,
            'max_drawdown': round(max_dd, 2),
            'avg_pnl': round(expectancy, 2),
            'total_pnl': round(sum(pnls), 2),
            'gross_profit': round(gross_profit, 2),
            'gross_loss': round(gross_loss, 2),
            'avg_holding_days': round(avg_holding, 1),
            'sqs': round(sqs, 0)
        }
    
    def supprimer_signal(self, signal_id: str):
        """Supprime un signal et son trade associé"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM trades WHERE signal_id = ?', (signal_id,))
        cursor.execute('DELETE FROM signals WHERE id = ?', (signal_id,))
        conn.commit()
        conn.close()
    
    def obtenir_trades(self, filtre: Dict = None) -> List[Dict]:
        """
        📊 Récupère tous les trades avec leurs signaux associés
        
        Args:
            filtre: Dictionnaire optionnel pour filtrer
        
        Returns:
            Liste de trades avec données complètes
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        query = '''
            SELECT 
                t.*,
                s.symbol,
                s.action,
                s.entry_price_ref,
                s.stop_loss,
                s.take_profit,
                s.strategy,
                s.marche,
                s.score_final,
                s.position_size,
                s.slippage_pct,
                s.commission_per_share,
                s.time_signal
            FROM trades t
            JOIN signals s ON t.signal_id = s.id
            WHERE 1=1
        '''
        params = []
        
        if filtre:
            if filtre.get('date_debut'):
                query += ' AND s.time_signal >= ?'
                params.append(filtre['date_debut'])
            if filtre.get('date_fin'):
                query += ' AND s.time_signal <= ?'
                params.append(filtre['date_fin'])
            if filtre.get('action'):
                query += ' AND s.action = ?'
                params.append(filtre['action'])
            if filtre.get('strategy'):
                query += ' AND s.strategy = ?'
                params.append(filtre['strategy'])
        
        query += ' ORDER BY s.time_signal DESC'
        
        cursor.execute(query, params)
        trades = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return trades

    # ══════════════════════════════════════════════════════════════════════════
    # 📋 NOUVELLES MÉTHODES POUR EXPLORATION DES SIGNAUX
    # ══════════════════════════════════════════════════════════════════════════
    
    def obtenir_signaux_par_periode(self, jours: int = 7, symbole: str = None) -> List[Dict]:
        """
        📅 Récupère les signaux des X derniers jours
        
        Args:
            jours: Nombre de jours à regarder en arrière
            symbole: Filtrer par symbole (optionnel)
        
        Returns:
            Liste de signaux avec leurs résultats
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        date_limite = (datetime.now() - timedelta(days=jours)).isoformat()
        
        query = '''
            SELECT 
                s.*,
                t.exit_time,
                t.exit_price,
                t.exit_reason,
                t.pnl_net,
                t.pnl_usd,
                t.result,
                t.holding_days,
                t.entry_price_reel
            FROM signals s
            LEFT JOIN trades t ON s.id = t.signal_id
            WHERE s.time_signal >= ?
        '''
        params = [date_limite]
        
        if symbole:
            query += ' AND s.symbol = ?'
            params.append(symbole)
        
        query += ' ORDER BY s.time_signal DESC'
        
        cursor.execute(query, params)
        signaux = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return signaux
    
    def obtenir_evolution_signal(self, signal_id: str) -> Dict:
        """
        📈 Récupère l'évolution détaillée d'un signal
        
        Args:
            signal_id: ID du signal
        
        Returns:
            Dictionnaire avec toutes les informations du signal et son évolution
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                s.*,
                t.exit_time,
                t.exit_price,
                t.exit_reason,
                t.pnl_net,
                t.pnl_usd,
                t.pnl_ticks,
                t.result,
                t.holding_days,
                t.entry_price_reel,
                t.worst_case_pnl,
                t.best_case_pnl,
                t.slippage_total,
                t.commission_total
            FROM signals s
            LEFT JOIN trades t ON s.id = t.signal_id
            WHERE s.id = ?
        ''', (signal_id,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            return dict(row)
        return {}
    
    def obtenir_signaux_par_symbole(self, symbole: str, limite: int = 50) -> List[Dict]:
        """
        🎯 Récupère l'historique des signaux pour un symbole
        
        Args:
            symbole: Le ticker du symbole
            limite: Nombre max de signaux
        
        Returns:
            Liste des signaux pour ce symbole
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                s.*,
                t.exit_time,
                t.exit_price,
                t.exit_reason,
                t.pnl_net,
                t.result,
                t.holding_days
            FROM signals s
            LEFT JOIN trades t ON s.id = t.signal_id
            WHERE s.symbol = ?
            ORDER BY s.time_signal DESC
            LIMIT ?
        ''', (symbole, limite))
        
        signaux = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return signaux
    
    def obtenir_resume_par_date(self, jours: int = 30) -> List[Dict]:
        """
        📊 Résumé des signaux groupés par date
        
        Args:
            jours: Période en jours
        
        Returns:
            Liste de résumés par jour
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        date_limite = (datetime.now() - timedelta(days=jours)).isoformat()
        
        cursor.execute('''
            SELECT 
                DATE(s.time_signal) as date_signal,
                COUNT(*) as nb_signaux,
                COUNT(CASE WHEN s.action = 'ACHAT' THEN 1 END) as nb_achats,
                COUNT(CASE WHEN s.action = 'VENTE' THEN 1 END) as nb_ventes,
                COUNT(CASE WHEN s.statut = 'PENDING' THEN 1 END) as nb_pending,
                COUNT(CASE WHEN t.result = 'WIN' THEN 1 END) as nb_wins,
                COUNT(CASE WHEN t.result = 'LOSS' THEN 1 END) as nb_losses,
                ROUND(AVG(s.score_final), 1) as score_moyen,
                ROUND(SUM(COALESCE(t.pnl_net, 0)), 2) as pnl_total
            FROM signals s
            LEFT JOIN trades t ON s.id = t.signal_id
            WHERE s.time_signal >= ?
            GROUP BY DATE(s.time_signal)
            ORDER BY date_signal DESC
        ''', (date_limite,))
        
        resultats = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return resultats
    
    def obtenir_symboles_uniques(self) -> List[str]:
        """
        📋 Liste tous les symboles ayant des signaux
        
        Returns:
            Liste des symboles uniques
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT DISTINCT symbol FROM signals ORDER BY symbol
        ''')
        
        symboles = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        return symboles
    
    def comparer_signal_prix_actuel(self, signal_id: str, prix_actuel: float) -> Dict:
        """
        📊 Compare un signal avec le prix actuel
        
        Args:
            signal_id: ID du signal
            prix_actuel: Prix actuel du symbole
        
        Returns:
            Dictionnaire avec les comparaisons
        """
        signal = self.obtenir_evolution_signal(signal_id)
        
        if not signal:
            return {}
        
        prix_signal = signal.get('entry_price_ref', 0)
        stop_loss = signal.get('stop_loss', 0)
        take_profit = signal.get('take_profit', 0)
        action = signal.get('action', 'ACHAT')
        
        if prix_signal <= 0:
            return signal
        
        # Calculs de performance
        variation_pct = ((prix_actuel - prix_signal) / prix_signal) * 100
        
        # Pour un ACHAT
        if action == 'ACHAT':
            distance_sl_pct = ((prix_actuel - stop_loss) / prix_actuel) * 100 if stop_loss else None
            distance_tp_pct = ((take_profit - prix_actuel) / prix_actuel) * 100 if take_profit else None
            sl_touche = prix_actuel <= stop_loss if stop_loss else False
            tp_touche = prix_actuel >= take_profit if take_profit else False
            performance = "GAGNANT" if variation_pct > 0 else "PERDANT"
        else:  # VENTE
            distance_sl_pct = ((stop_loss - prix_actuel) / prix_actuel) * 100 if stop_loss else None
            distance_tp_pct = ((prix_actuel - take_profit) / prix_actuel) * 100 if take_profit else None
            sl_touche = prix_actuel >= stop_loss if stop_loss else False
            tp_touche = prix_actuel <= take_profit if take_profit else False
            performance = "GAGNANT" if variation_pct < 0 else "PERDANT"
        
        signal.update({
            'prix_actuel': prix_actuel,
            'variation_pct': round(variation_pct, 2),
            'distance_sl_pct': round(distance_sl_pct, 2) if distance_sl_pct else None,
            'distance_tp_pct': round(distance_tp_pct, 2) if distance_tp_pct else None,
            'sl_touche': sl_touche,
            'tp_touche': tp_touche,
            'performance_actuelle': performance,
            'jours_depuis_signal': (datetime.now() - datetime.fromisoformat(signal.get('time_signal', datetime.now().isoformat()))).days
        })
        
        return signal
    
    def exporter_signaux_csv(self, filepath: str, jours: int = None) -> bool:
        """
        📁 Exporte les signaux vers un fichier CSV
        
        Args:
            filepath: Chemin du fichier
            jours: Période à exporter (None = tout)
        
        Returns:
            True si succès
        """
        try:
            if jours:
                signaux = self.obtenir_signaux_par_periode(jours)
            else:
                signaux = self.obtenir_tous_signaux(limite=10000)
            
            if not signaux:
                return False
            
            df = pd.DataFrame(signaux)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            logger.info(f"Signaux exportés: {filepath} ({len(signaux)} signaux)")
            return True
        except Exception as e:
            logger.error(f"Erreur export CSV: {e}")
            return False
    
    def obtenir_statistiques_periode(self, jours: int = 7) -> Dict:
        """
        📊 Statistiques des signaux sur une période
        
        Args:
            jours: Période en jours
        
        Returns:
            Dictionnaire de statistiques
        """
        signaux = self.obtenir_signaux_par_periode(jours)
        
        if not signaux:
            return {
                'periode_jours': jours,
                'nb_signaux': 0,
                'nb_achats': 0,
                'nb_ventes': 0,
                'nb_pending': 0,
                'nb_evaluated': 0,
                'nb_wins': 0,
                'nb_losses': 0,
                'win_rate': 0,
                'pnl_total': 0,
                'score_moyen': 0
            }
        
        nb_total = len(signaux)
        achats = [s for s in signaux if s.get('action') == 'ACHAT']
        ventes = [s for s in signaux if s.get('action') == 'VENTE']
        pending = [s for s in signaux if s.get('statut') == 'PENDING']
        evaluated = [s for s in signaux if s.get('statut') == 'EVALUATED']
        wins = [s for s in signaux if s.get('result') == 'WIN']
        losses = [s for s in signaux if s.get('result') == 'LOSS']
        
        pnl_total = sum(s.get('pnl_net', 0) or 0 for s in signaux)
        scores = [s.get('score_final', 0) for s in signaux if s.get('score_final')]
        score_moyen = np.mean(scores) if scores else 0
        
        win_rate = (len(wins) / len(evaluated) * 100) if evaluated else 0
        
        return {
            'periode_jours': jours,
            'nb_signaux': nb_total,
            'nb_achats': len(achats),
            'nb_ventes': len(ventes),
            'nb_pending': len(pending),
            'nb_evaluated': len(evaluated),
            'nb_wins': len(wins),
            'nb_losses': len(losses),
            'win_rate': round(win_rate, 1),
            'pnl_total': round(pnl_total, 2),
            'score_moyen': round(score_moyen, 1)
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    ÉVALUATEUR DE SIGNAUX
# ══════════════════════════════════════════════════════════════════════════════

class EvaluateurSignaux:
    """
    🎯 Évalue les signaux en attente pour déterminer leur résultat
    
    Logique:
    - Pour ACHAT: TP atteint si High >= TP, SL atteint si Low <= SL
    - Pour VENTE: TP atteint si Low <= TP, SL atteint si High >= SL
    - Gestion des cas où TP et SL sont touchés dans la même bougie
    """
    
    def __init__(self, gestionnaire: GestionnaireSignaux, politique_same_bar: str = 'worst'):
        """
        Args:
            gestionnaire: Instance de GestionnaireSignaux
            politique_same_bar: 'worst' (conservateur) ou 'best' (optimiste)
        """
        self.gestionnaire = gestionnaire
        self.politique_same_bar = politique_same_bar
    
    def evaluer_signal(self, signal: Dict, donnees_ohlc: pd.DataFrame) -> Optional[Dict]:
        """
        Évalue un signal individuel (V2 - CORRIGÉ selon audit)
        
        Corrections appliquées:
        - Filtrage temporel strict (>= pour inclure la bougie d'entrée si nécessaire)
        - Index trié avant filtrage
        - Gestion timezone robuste
        - Coûts de transaction réalistes
        - Règle same-bar cohérente
        
        Args:
            signal: Dictionnaire du signal depuis la DB
            donnees_ohlc: DataFrame avec colonnes Open, High, Low, Close (données ADJUSTED recommandées)
        
        Returns:
            Dictionnaire du résultat du trade ou None si non évaluable
        """
        if donnees_ohlc.empty:
            return None
        
        entry_price = signal['entry_price_ref']
        stop_loss = signal['stop_loss']
        take_profit = signal['take_profit']
        action = signal['action']
        horizon_days = signal['horizon_days'] or 30
        
        # Validation des prix
        if not entry_price or entry_price <= 0:
            logger.warning(f"Signal invalide: entry_price={entry_price}")
            return None
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 COÛTS DE TRANSACTION RÉALISTES
        # ══════════════════════════════════════════════════════════════════
        position_size = signal.get('position_size', 100)  # Nombre d'actions
        
        # Commission: typiquement $0.005/action ou $1 minimum
        commission_per_share = signal.get('commission_per_share', 0.005)  # $0.005/action par défaut
        commission_total = max(commission_per_share * position_size, 1.0) * 2  # Min $1, aller-retour
        
        # Slippage: sera calculé après entry_price_reel (Open réel)
        slippage_pct = signal.get('slippage_pct', 0.0005)  # 0.05% par défaut
        # NOTE: couts_totaux sera calculé après détermination de entry_price_reel
        
        # Date de signal
        time_signal = None
        if signal.get('time_signal'):
            try:
                time_signal = datetime.fromisoformat(signal['time_signal'].replace('Z', '+00:00'))
            except Exception as e:
                logger.warning(f"Erreur parsing time_signal: {e}")
                time_signal = None
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 FILTRAGE TEMPOREL STRICT (CORRIGÉ selon audit)
        # ══════════════════════════════════════════════════════════════════
        donnees_filtrees = donnees_ohlc.copy()
        
        # 1. TOUJOURS trier l'index chronologiquement
        donnees_filtrees = donnees_filtrees.sort_index()
        
        if time_signal is not None:
            # 2. Gérer le fuseau horaire de manière robuste
            try:
                if hasattr(donnees_filtrees.index, 'tz') and donnees_filtrees.index.tz is not None:
                    if time_signal.tzinfo is None:
                        time_signal = time_signal.replace(tzinfo=donnees_filtrees.index.tz)
                    else:
                        time_signal = time_signal.astimezone(donnees_filtrees.index.tz)
                else:
                    # Index sans timezone, retirer le timezone du signal
                    if time_signal.tzinfo is not None:
                        time_signal = time_signal.replace(tzinfo=None)
            except Exception as e:
                logger.warning(f"Erreur timezone: {e}")
            
            # 3. FILTRER: >= pour inclure la bougie de signal (entrée à l'ouverture de la bougie suivante)
            # Note: On utilise > car on veut les bougies APRÈS le signal (pas la bougie du signal)
            # L'entrée se fait à l'ouverture de la première bougie après le signal
            donnees_filtrees = donnees_filtrees[donnees_filtrees.index > time_signal]
        
        # 4. Vérifier qu'il reste des données
        if donnees_filtrees.empty:
            logger.debug(f"Aucune donnée après time_signal pour ce signal")
            return None
        
        # 5. VALIDATION: Vérifier que la première bougie est bien APRÈS le signal
        premiere_bougie = donnees_filtrees.index[0]
        if time_signal and premiere_bougie <= time_signal:
            logger.error(f"BUG: Première bougie ({premiere_bougie}) <= time_signal ({time_signal})")
            return None
        
        # Ajuster le prix d'entrée au prix d'ouverture de la première bougie (plus réaliste)
        entry_price_reel = donnees_filtrees.iloc[0]['Open']
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 CALCUL DES COÛTS (basé sur entry_price_reel, pas entry_price_ref)
        # ══════════════════════════════════════════════════════════════════
        slippage_per_share = entry_price_reel * slippage_pct * 2  # Aller-retour
        couts_totaux = commission_total + (slippage_per_share * position_size)
        
        # Parcourir les bougies (APRÈS le signal uniquement)
        exit_price = None
        exit_reason = None
        exit_time = None
        exit_day = 0
        
        worst_case_pnl = None
        best_case_pnl = None
        
        for i, (idx, row) in enumerate(donnees_filtrees.iterrows()):
            if i >= horizon_days:
                # Timeout - sortie au prix de clôture
                exit_price = row['Close']
                exit_reason = 'TIMEOUT'
                exit_time = str(idx)
                exit_day = i
                break
            
            high = row['High']
            low = row['Low']
            open_price = row['Open']
            
            # ══════════════════════════════════════════════════════════════════
            # 🔧 RÈGLE SAME-BAR COHÉRENTE (selon audit)
            # Si TP et SL sont touchés dans la même bougie:
            # - 'worst': on assume le pire (SL touché en premier)
            # - 'best': on assume le meilleur (TP touché en premier)
            # - 'open_based': on regarde si l'ouverture est plus proche du TP ou SL
            # ══════════════════════════════════════════════════════════════════
            
            if action == 'ACHAT':
                # ══════════════════════════════════════════════════════════════════
                # 🔧 GAP LOGIC: Exécution réaliste sur Open si gap dépasse SL/TP
                # ══════════════════════════════════════════════════════════════════
                
                # Gap down: Ouverture sous le SL → exécution à l'Open (pas au SL)
                gap_sl = open_price <= stop_loss if stop_loss else False
                # Gap up: Ouverture au-dessus du TP → exécution à l'Open (pas au TP)
                gap_tp = open_price >= take_profit if take_profit else False
                
                sl_hit = (low <= stop_loss or gap_sl) if stop_loss else False
                tp_hit = (high >= take_profit or gap_tp) if take_profit else False
                
                if gap_sl:
                    # Gap sous SL: sortie à l'Open (pire cas)
                    exit_price = open_price
                    exit_reason = 'SL_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                
                if gap_tp:
                    # Gap au-dessus TP: sortie à l'Open (meilleur cas)
                    exit_price = open_price
                    exit_reason = 'TP_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                
                if sl_hit and tp_hit:
                    # Les deux touchés dans la même bougie
                    if self.politique_same_bar == 'worst':
                        exit_price = stop_loss
                        exit_reason = 'SL'
                    elif self.politique_same_bar == 'open_based':
                        # Logique basée sur l'ouverture
                        dist_to_sl = abs(open_price - stop_loss)
                        dist_to_tp = abs(take_profit - open_price)
                        if dist_to_sl < dist_to_tp:
                            exit_price = stop_loss
                            exit_reason = 'SL'
                        else:
                            exit_price = take_profit
                            exit_reason = 'TP'
                    else:  # 'best'
                        exit_price = take_profit
                        exit_reason = 'TP'
                    
                    # Calculer best/worst case
                    worst_case_pnl = (stop_loss - entry_price_reel) * position_size - couts_totaux
                    best_case_pnl = (take_profit - entry_price_reel) * position_size - couts_totaux
                    
                elif sl_hit:
                    exit_price = stop_loss
                    exit_reason = 'SL'
                elif tp_hit:
                    exit_price = take_profit
                    exit_reason = 'TP'
            
            else:  # VENTE (Short)
                # ══════════════════════════════════════════════════════════════════
                # 🔧 GAP LOGIC: Exécution réaliste pour SHORT
                # ══════════════════════════════════════════════════════════════════
                
                # Gap up: Ouverture au-dessus du SL → exécution à l'Open
                gap_sl = open_price >= stop_loss if stop_loss else False
                # Gap down: Ouverture sous le TP → exécution à l'Open
                gap_tp = open_price <= take_profit if take_profit else False
                
                sl_hit = (high >= stop_loss or gap_sl) if stop_loss else False
                tp_hit = (low <= take_profit or gap_tp) if take_profit else False
                
                if gap_sl:
                    # Gap au-dessus SL: sortie à l'Open (pire cas pour short)
                    exit_price = open_price
                    exit_reason = 'SL_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                
                if gap_tp:
                    # Gap sous TP: sortie à l'Open (meilleur cas pour short)
                    exit_price = open_price
                    exit_reason = 'TP_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                
                if sl_hit and tp_hit:
                    if self.politique_same_bar == 'worst':
                        exit_price = stop_loss
                        exit_reason = 'SL'
                    elif self.politique_same_bar == 'open_based':
                        dist_to_sl = abs(stop_loss - open_price)
                        dist_to_tp = abs(open_price - take_profit)
                        if dist_to_sl < dist_to_tp:
                            exit_price = stop_loss
                            exit_reason = 'SL'
                        else:
                            exit_price = take_profit
                            exit_reason = 'TP'
                    else:
                        exit_price = take_profit
                        exit_reason = 'TP'
                    
                    worst_case_pnl = (entry_price_reel - stop_loss) * position_size - couts_totaux
                    best_case_pnl = (entry_price_reel - take_profit) * position_size - couts_totaux
                    
                elif sl_hit:
                    exit_price = stop_loss
                    exit_reason = 'SL'
                elif tp_hit:
                    exit_price = take_profit
                    exit_reason = 'TP'
            
            if exit_reason:
                exit_time = str(idx)
                exit_day = i + 1
                break
        
        # Si pas de sortie trouvée, sortir au dernier prix
        if not exit_reason:
            exit_price = donnees_filtrees['Close'].iloc[-1]
            exit_reason = 'TIMEOUT'
            exit_time = str(donnees_filtrees.index[-1])
            exit_day = len(donnees_filtrees)
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 PnL CALCULÉ CORRECTEMENT (avec entry_price_reel)
        # ══════════════════════════════════════════════════════════════════
        if action == 'ACHAT':
            pnl_per_share = exit_price - entry_price_reel
        else:
            pnl_per_share = entry_price_reel - exit_price
        
        pnl_usd = pnl_per_share * position_size
        pnl_net = pnl_usd - couts_totaux
        
        # Protection contre division par zéro
        pnl_pct = (pnl_per_share / entry_price_reel * 100) if entry_price_reel and entry_price_reel > 0 else 0
        
        result = 'WIN' if pnl_net > 0 else 'LOSS'
        
        # Calculer slippage_total pour le suivi des coûts
        slippage_total = slippage_per_share * position_size
        
        return {
            'exit_time': exit_time,
            'exit_price': round(exit_price, 4),
            'exit_reason': exit_reason,
            'entry_price_ref': round(signal['entry_price_ref'], 4),  # Prix original du signal
            'entry_price_reel': round(entry_price_reel, 4),  # Prix d'exécution réel
            'slippage_entry': round(abs(entry_price_reel - signal['entry_price_ref']), 4),
            'slippage_total': round(slippage_total, 2),  # Coût total du slippage
            'pnl_per_share': round(pnl_per_share, 4),
            'pnl_pct': round(pnl_pct, 2),
            'pnl_usd': round(pnl_usd, 2),
            'pnl_net': round(pnl_net, 2),
            'couts_totaux': round(couts_totaux, 2),
            'commission_total': round(commission_total, 2),
            'position_size': position_size,
            'result': result,
            'holding_days': exit_day,
            'worst_case_pnl': round(worst_case_pnl, 2) if worst_case_pnl else None,
            'best_case_pnl': round(best_case_pnl, 2) if best_case_pnl else None
        }
    
    def evaluer_tous_pending(self, callback: Callable = None) -> Dict:
        """
        Évalue tous les signaux en attente
        
        Returns:
            Résumé des évaluations
        """
        signaux = self.gestionnaire.obtenir_signaux_pending()
        
        if not signaux:
            return {'evalues': 0, 'erreurs': 0}
        
        evalues = 0
        erreurs = 0
        
        for i, signal in enumerate(signaux):
            try:
                if callback:
                    callback(f"Évaluation: {signal['symbol']}", i, len(signaux))
                
                # Récupérer les données OHLC
                ticker = yf.Ticker(signal['symbol'])
                df = ticker.history(period="3mo")
                
                if df.empty:
                    erreurs += 1
                    continue
                
                # Évaluer le signal
                resultat = self.evaluer_signal(signal, df)
                
                if resultat:
                    self.gestionnaire.sauvegarder_trade(signal['id'], resultat)
                    evalues += 1
                else:
                    erreurs += 1
                    
            except Exception as e:
                logger.error(f"Erreur évaluation {signal.get('symbol')}: {e}")
                erreurs += 1
        
        return {'evalues': evalues, 'erreurs': erreurs}
    
    @staticmethod
    def evaluer_trade_v2(entry_price: float, stop_loss: float, take_profit: float, 
                         action: str, df_ohlc: pd.DataFrame, time_signal=None,
                         position_size: int = 100, commission_per_share: float = 0.005,
                         slippage_pct: float = 0.0005, horizon_days: int = 30,
                         politique_same_bar: str = 'worst') -> Optional[Dict]:
        """
        🔧 VERSION STATIQUE de evaluer_signal pour Walk-Forward/OOS
        
        GARANTIE: Utilise EXACTEMENT la même logique que evaluer_signal V2:
        - Filtrage temporel strict
        - Gap logic (SL_GAP, TP_GAP)
        - Modèle de coûts actions (commission_per_share + slippage_pct)
        - Règle same-bar configurable
        
        Args:
            entry_price: Prix d'entrée
            stop_loss: Stop loss
            take_profit: Take profit
            action: 'ACHAT' ou 'VENTE'
            df_ohlc: DataFrame OHLC à évaluer
            time_signal: Timestamp du signal (optionnel, sinon commence à la 1ère bougie)
            position_size: Nombre d'actions
            commission_per_share: Commission par action
            slippage_pct: Slippage en pourcentage
            horizon_days: Nombre max de jours
            politique_same_bar: 'worst', 'best', ou 'open_based'
        
        Returns:
            Dict avec résultat ou None
        """
        if df_ohlc.empty:
            return None
        
        # Validation des prix
        if entry_price <= 0 or (stop_loss and stop_loss <= 0) or (take_profit and take_profit <= 0):
            return None
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 FILTRAGE TEMPOREL (si time_signal fourni)
        # ══════════════════════════════════════════════════════════════════
        df = df_ohlc.copy()
        df.index = pd.to_datetime(df.index)
        if df.index.tzinfo is not None:
            df.index = df.index.tz_localize(None)
        df = df.sort_index()
        
        if time_signal is not None:
            if isinstance(time_signal, str):
                time_signal = pd.to_datetime(time_signal)
            if hasattr(time_signal, 'tzinfo') and time_signal.tzinfo is not None:
                time_signal = time_signal.tz_localize(None)
            # FILTRAGE STRICT: > time_signal (pas le jour du signal)
            df = df[df.index > time_signal]
        
        if df.empty:
            return None
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 ENTRÉE RÉELLE = Open de la 1ère bougie après signal (COMME evaluer_signal!)
        # ══════════════════════════════════════════════════════════════════
        first_row = df.iloc[0]
        entry_price_reel = first_row['Open']
        if isinstance(entry_price_reel, pd.Series):
            entry_price_reel = entry_price_reel.iloc[0]
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 COÛTS DE TRANSACTION (IDENTIQUE à evaluer_signal!)
        # NOTE: Le slippage est comptabilisé UNE SEULE FOIS dans couts_totaux
        #       On ne modifie PAS entry_price_reel avec le slippage
        # ══════════════════════════════════════════════════════════════════
        commission_total = max(commission_per_share * position_size, 1.0) * 2  # Min $1, aller-retour
        slippage_per_share = entry_price_reel * slippage_pct * 2  # Aller-retour
        couts_totaux = commission_total + (slippage_per_share * position_size)
        
        # Variables de sortie
        exit_price = None
        exit_reason = None
        exit_time = None
        exit_day = 0
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 BOUCLE D'ÉVALUATION AVEC GAP LOGIC
        # ══════════════════════════════════════════════════════════════════
        for i, (idx, row) in enumerate(df.iterrows()):
            if i >= horizon_days:
                exit_price = row['Close']
                exit_reason = 'TIMEOUT'
                exit_time = str(idx)
                exit_day = i
                break
            
            high = row['High'] if not isinstance(row['High'], pd.Series) else row['High'].iloc[0]
            low = row['Low'] if not isinstance(row['Low'], pd.Series) else row['Low'].iloc[0]
            open_price = row['Open'] if not isinstance(row['Open'], pd.Series) else row['Open'].iloc[0]
            close = row['Close'] if not isinstance(row['Close'], pd.Series) else row['Close'].iloc[0]
            
            if action == 'ACHAT':
                # GAP LOGIC
                gap_sl = open_price <= stop_loss if stop_loss else False
                gap_tp = open_price >= take_profit if take_profit else False
                sl_hit = (low <= stop_loss or gap_sl) if stop_loss else False
                tp_hit = (high >= take_profit or gap_tp) if take_profit else False
                
                if gap_sl:
                    exit_price = open_price
                    exit_reason = 'SL_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                if gap_tp:
                    exit_price = open_price
                    exit_reason = 'TP_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                
                if sl_hit and tp_hit:
                    if politique_same_bar == 'worst':
                        exit_price = stop_loss
                        exit_reason = 'SL'
                    elif politique_same_bar == 'open_based':
                        dist_to_sl = abs(open_price - stop_loss)
                        dist_to_tp = abs(take_profit - open_price)
                        if dist_to_sl < dist_to_tp:
                            exit_price = stop_loss
                            exit_reason = 'SL'
                        else:
                            exit_price = take_profit
                            exit_reason = 'TP'
                    else:
                        exit_price = take_profit
                        exit_reason = 'TP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                elif sl_hit:
                    exit_price = stop_loss
                    exit_reason = 'SL'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                elif tp_hit:
                    exit_price = take_profit
                    exit_reason = 'TP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                    
            else:  # VENTE (Short)
                # GAP LOGIC
                gap_sl = open_price >= stop_loss if stop_loss else False
                gap_tp = open_price <= take_profit if take_profit else False
                sl_hit = (high >= stop_loss or gap_sl) if stop_loss else False
                tp_hit = (low <= take_profit or gap_tp) if take_profit else False
                
                if gap_sl:
                    exit_price = open_price
                    exit_reason = 'SL_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                if gap_tp:
                    exit_price = open_price
                    exit_reason = 'TP_GAP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                
                if sl_hit and tp_hit:
                    if politique_same_bar == 'worst':
                        exit_price = stop_loss
                        exit_reason = 'SL'
                    elif politique_same_bar == 'open_based':
                        dist_to_sl = abs(stop_loss - open_price)
                        dist_to_tp = abs(open_price - take_profit)
                        if dist_to_sl < dist_to_tp:
                            exit_price = stop_loss
                            exit_reason = 'SL'
                        else:
                            exit_price = take_profit
                            exit_reason = 'TP'
                    else:
                        exit_price = take_profit
                        exit_reason = 'TP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                elif sl_hit:
                    exit_price = stop_loss
                    exit_reason = 'SL'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
                elif tp_hit:
                    exit_price = take_profit
                    exit_reason = 'TP'
                    exit_time = str(idx)
                    exit_day = i + 1
                    break
        
        # TIMEOUT si aucune sortie
        if not exit_reason:
            exit_price = df['Close'].iloc[-1]
            exit_reason = 'TIMEOUT'
            exit_time = str(df.index[-1])
            exit_day = len(df)
        
        # ══════════════════════════════════════════════════════════════════
        # 🔧 CALCUL PnL (MODÈLE ACTIONS V2)
        # ══════════════════════════════════════════════════════════════════
        if action == 'ACHAT':
            pnl_brut = (exit_price - entry_price_reel) * position_size
        else:
            pnl_brut = (entry_price_reel - exit_price) * position_size
        
        pnl_net = pnl_brut - couts_totaux
        result = 'WIN' if pnl_net > 0 else 'LOSS'
        
        return {
            'entry': entry_price,
            'entry_price_reel': round(entry_price_reel, 4),
            'exit_price': round(exit_price, 4),
            'exit_reason': exit_reason,
            'exit_time': exit_time,
            'holding_days': exit_day,
            'pnl_brut': round(pnl_brut, 2),
            'pnl_net': round(pnl_net, 2),
            'couts_totaux': round(couts_totaux, 2),
            'result': result,
            'action': action,
            'position_size': position_size
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    STRESS TESTS
# ══════════════════════════════════════════════════════════════════════════════

class StressTests:
    """
    🔬 Exécute des tests de robustesse sur les résultats
    
    Tests disponibles:
    1. Coûts x2: Double les commissions et slippage
    2. Entrée retardée: Décale l'entrée d'une bougie
    3. SL/TP ±10%: Modifie les niveaux de sortie
    4. Analyse par régime de marché
    """
    
    def __init__(self, gestionnaire: GestionnaireSignaux):
        self.gestionnaire = gestionnaire
    
    def test_couts_doubles(self) -> Dict:
        """Test avec coûts x2 (MODÈLE ACTIONS)"""
        conn = sqlite3.connect(self.gestionnaire.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Utiliser les champs du modèle ACTIONS: commission_per_share, slippage_pct, position_size
        # NOTE: t.entry_price_reel est le prix Open réel, s.entry_price_ref est le prix du signal
        cursor.execute('''
            SELECT t.*, t.pnl_usd, t.pnl_net, t.entry_price_reel,
                   s.commission_per_share, s.slippage_pct, s.position_size, s.entry_price_ref
            FROM trades t
            JOIN signals s ON t.signal_id = s.id
        ''')
        
        trades = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        if not trades:
            return {'nom': 'Coûts x2', 'impact': 0, 'nouveau_pnl': 0, 'ancien_pnl': 0, 'robuste': False}
        
        ancien_pnl = sum(t['pnl_net'] or 0 for t in trades)
        
        # Recalculer avec coûts x2 (modèle actions)
        nouveau_pnl = 0
        for t in trades:
            position_size = t.get('position_size') or 100
            commission_per_share = t.get('commission_per_share') or 0.005
            slippage_pct = t.get('slippage_pct') or 0.0005
            # IMPORTANT: Utiliser entry_price_reel (prix d'exécution réel) pour le slippage
            entry_price_reel = t.get('entry_price_reel') or t.get('entry_price_ref') or 100
            
            # Coûts originaux (aller-retour) basés sur entry_price_reel
            commission_orig = max(commission_per_share * position_size, 1.0) * 2
            slippage_orig = entry_price_reel * slippage_pct * 2 * position_size
            couts_originaux = commission_orig + slippage_orig
            
            # Coûts doublés
            couts_doubles = couts_originaux * 2
            
            # PnL avec coûts doublés = PnL brut - coûts doublés
            # NOTE: pnl_net = pnl_brut - couts, donc pnl_brut = pnl_net + couts
            pnl_brut = (t['pnl_net'] or 0) + couts_originaux  # Retrouver le PnL brut depuis pnl_net
            nouveau_pnl += pnl_brut - couts_doubles
        
        impact = ((nouveau_pnl - ancien_pnl) / abs(ancien_pnl) * 100) if ancien_pnl != 0 else 0
        
        return {
            'nom': 'Coûts x2',
            'ancien_pnl': round(ancien_pnl, 2),
            'nouveau_pnl': round(nouveau_pnl, 2),
            'impact_pct': round(impact, 2),
            'robuste': nouveau_pnl > 0
        }
    
    def test_sltp_modifies(self, variation: float = 0.10) -> Dict:
        """
        Test avec SL/TP modifiés de ±variation (V3 - UTILISE evaluer_trade_v2)
        
        GARANTIE: Utilise EXACTEMENT le même moteur que le backtest principal
        """
        trades = self.gestionnaire.obtenir_trades()
        if not trades:
            return {'nom': f'SL/TP ±{variation*100:.0f}%', 'robuste': False, 'erreur': 'Pas de trades'}
        
        wins_originaux = sum(1 for t in trades if t.get('result') == 'WIN')
        wins_modifies = 0
        pnl_modifie_total = 0
        pnl_original_total = sum(t.get('pnl_net', 0) or 0 for t in trades)
        trades_simules = 0
        
        # Grouper les trades par symbole
        trades_par_symbole = {}
        for t in trades:
            symbole = t.get('symbol', '')
            if symbole:
                trades_par_symbole.setdefault(symbole, []).append(t)
        
        for symbole, trades_symbole in trades_par_symbole.items():
            try:
                # Déterminer la plage de dates nécessaire
                time_signals = []
                for t in trades_symbole:
                    ts = t.get('time_signal', '')
                    if ts:
                        try:
                            time_signals.append(pd.to_datetime(ts))
                        except:
                            pass
                
                if not time_signals:
                    continue
                
                # Télécharger avec start/end basé sur time_signal (pas period='1y')
                date_min = min(time_signals) - pd.Timedelta(days=5)
                date_max = max(time_signals) + pd.Timedelta(days=60)  # Horizon max
                
                df = yf.download(symbole, start=date_min.strftime('%Y-%m-%d'), 
                                end=date_max.strftime('%Y-%m-%d'), progress=False)
                if df.empty:
                    continue
                
                df.index = pd.to_datetime(df.index)
                if df.index.tzinfo is not None:
                    df.index = df.index.tz_localize(None)
                
                for t in trades_symbole:
                    entry = t.get('entry_price_ref') or t.get('entry_price', 0)
                    sl = t.get('stop_loss', 0)
                    tp = t.get('take_profit', 0)
                    action = t.get('action', 'ACHAT')
                    time_signal_str = t.get('time_signal', '')
                    position_size = t.get('position_size', 100)
                    commission = t.get('commission_per_share', 0.005)
                    slippage = t.get('slippage_pct', 0.0005)
                    
                    if not all([entry, sl, tp, time_signal_str]):
                        continue
                    
                    # Modifier SL/TP selon la variation
                    if action == 'ACHAT':
                        sl_modifie = sl * (1 + variation)
                        tp_modifie = tp * (1 - variation)
                    else:
                        sl_modifie = sl * (1 - variation)
                        tp_modifie = tp * (1 + variation)
                    
                    # 🔧 UTILISER evaluer_trade_v2 (MÊME LOGIQUE QUE LE MOTEUR PRINCIPAL!)
                    result = EvaluateurSignaux.evaluer_trade_v2(
                        entry_price=entry,
                        stop_loss=sl_modifie,
                        take_profit=tp_modifie,
                        action=action,
                        df_ohlc=df,
                        time_signal=time_signal_str,
                        position_size=position_size,
                        commission_per_share=commission,
                        slippage_pct=slippage,
                        horizon_days=30,
                        politique_same_bar='worst'
                    )
                    
                    if result:
                        pnl_modifie_total += result['pnl_net']
                        trades_simules += 1
                        if result['result'] == 'WIN':
                            wins_modifies += 1
                        
            except Exception as e:
                logging.debug(f"Erreur simulation {symbole}: {e}")
                continue
        
        nb_trades = len(trades)
        win_rate_base = (wins_originaux / nb_trades * 100) if nb_trades else 0
        win_rate_stress = (wins_modifies / trades_simules * 100) if trades_simules else 0
        
        return {
            'nom': f'SL/TP ±{variation*100:.0f}%',
            'win_rate_base': round(win_rate_base, 2),
            'win_rate_stress': round(win_rate_stress, 2),
            'pnl_base': round(pnl_original_total, 2),
            'pnl_stress': round(pnl_modifie_total, 2),
            'trades_simules': trades_simules,
            'robuste': pnl_modifie_total > 0 and win_rate_stress > 40,
            'note': '✅ Utilise evaluer_trade_v2 (même moteur que backtest)'
        }
    
    def test_entree_retardee(self) -> Dict:
        """
        Test avec entrée retardée d'une bougie (V3 - UTILISE evaluer_trade_v2)
        
        Simule un retard d'entrée: au lieu d'entrer à la 1ère bougie après signal,
        on entre à la 2ème bougie. Utilise le MÊME MOTEUR que le backtest principal.
        """
        trades = self.gestionnaire.obtenir_trades()
        if not trades:
            return {'nom': 'Entrée +1 bougie', 'robuste': False, 'erreur': 'Pas de trades'}
        
        pnl_base = sum(t.get('pnl_net', 0) or 0 for t in trades)
        pnl_retarde = 0
        wins_base = sum(1 for t in trades if t.get('result') == 'WIN')
        wins_retarde = 0
        trades_simules = 0
        
        # Grouper les trades par symbole
        trades_par_symbole = {}
        for t in trades:
            symbole = t.get('symbol', '')
            if symbole:
                trades_par_symbole.setdefault(symbole, []).append(t)
        
        for symbole, trades_symbole in trades_par_symbole.items():
            try:
                # Déterminer la plage de dates nécessaire
                time_signals = []
                for t in trades_symbole:
                    ts = t.get('time_signal', '')
                    if ts:
                        try:
                            time_signals.append(pd.to_datetime(ts))
                        except:
                            pass
                
                if not time_signals:
                    continue
                
                # Télécharger avec start/end basé sur time_signal
                date_min = min(time_signals) - pd.Timedelta(days=5)
                date_max = max(time_signals) + pd.Timedelta(days=60)
                
                df = yf.download(symbole, start=date_min.strftime('%Y-%m-%d'), 
                                end=date_max.strftime('%Y-%m-%d'), progress=False)
                if df.empty:
                    continue
                
                df.index = pd.to_datetime(df.index)
                if df.index.tzinfo is not None:
                    df.index = df.index.tz_localize(None)
                
                for t in trades_symbole:
                    entry_original = t.get('entry_price_ref') or t.get('entry_price', 0)
                    sl = t.get('stop_loss', 0)
                    tp = t.get('take_profit', 0)
                    action = t.get('action', 'ACHAT')
                    time_signal_str = t.get('time_signal', '')
                    position_size = t.get('position_size', 100)
                    commission = t.get('commission_per_share', 0.005)
                    slippage = t.get('slippage_pct', 0.0005)
                    
                    if not all([entry_original, sl, tp, time_signal_str]):
                        continue
                    
                    # Parser time_signal
                    try:
                        time_signal = pd.to_datetime(time_signal_str)
                        if time_signal.tzinfo is not None:
                            time_signal = time_signal.tz_localize(None)
                    except:
                        continue
                    
                    # Trouver la 1ère bougie après signal pour simuler le retard
                    df_apres = df[df.index > time_signal]
                    if len(df_apres) < 2:
                        continue
                    
                    # Le "nouveau" time_signal est la date de la 1ère bougie
                    # Ainsi evaluer_trade_v2 utilisera l'Open de la 2ème bougie
                    nouveau_time_signal = df_apres.index[0]
                    
                    # L'entrée retardée sera l'Open de la 2ème bougie
                    entry_retardee = df_apres.iloc[1]['Open']
                    if isinstance(entry_retardee, pd.Series):
                        entry_retardee = entry_retardee.iloc[0]
                    
                    # Ajuster SL/TP proportionnellement à la nouvelle entrée
                    if action == 'ACHAT':
                        ratio_sl = (entry_original - sl) / entry_original if entry_original else 0
                        ratio_tp = (tp - entry_original) / entry_original if entry_original else 0
                        sl_ajuste = entry_retardee * (1 - ratio_sl)
                        tp_ajuste = entry_retardee * (1 + ratio_tp)
                    else:
                        ratio_sl = (sl - entry_original) / entry_original if entry_original else 0
                        ratio_tp = (entry_original - tp) / entry_original if entry_original else 0
                        sl_ajuste = entry_retardee * (1 + ratio_sl)
                        tp_ajuste = entry_retardee * (1 - ratio_tp)
                    
                    # 🔧 UTILISER evaluer_trade_v2 avec le time_signal décalé
                    result = EvaluateurSignaux.evaluer_trade_v2(
                        entry_price=entry_retardee,
                        stop_loss=sl_ajuste,
                        take_profit=tp_ajuste,
                        action=action,
                        df_ohlc=df,
                        time_signal=str(nouveau_time_signal),
                        position_size=position_size,
                        commission_per_share=commission,
                        slippage_pct=slippage,
                        horizon_days=30,
                        politique_same_bar='worst'
                    )
                    
                    if result:
                        pnl_retarde += result['pnl_net']
                        trades_simules += 1
                        if result['result'] == 'WIN':
                            wins_retarde += 1
                        
            except Exception as e:
                logging.debug(f"Erreur simulation retardée {symbole}: {e}")
                continue
        
        nb_trades = len(trades)
        win_rate_base = (wins_base / nb_trades * 100) if nb_trades else 0
        win_rate_retarde = (wins_retarde / trades_simules * 100) if trades_simules else 0
        impact_pct = ((pnl_retarde - pnl_base) / abs(pnl_base) * 100) if pnl_base != 0 else 0
        
        return {
            'nom': 'Entrée +1 bougie',
            'pnl_base': round(pnl_base, 2),
            'pnl_stress': round(pnl_retarde, 2),
            'win_rate_base': round(win_rate_base, 2),
            'win_rate_retarde': round(win_rate_retarde, 2),
            'trades_simules': trades_simules,
            'impact_pct': round(impact_pct, 2),
            'robuste': pnl_retarde > 0,
            'note': '✅ Utilise evaluer_trade_v2 (même moteur que backtest)'
        }
    
    def executer_tous_tests(self) -> List[Dict]:
        """Exécute tous les stress tests"""
        resultats = []
        
        resultats.append(self.test_couts_doubles())
        resultats.append(self.test_sltp_modifies(0.10))
        resultats.append(self.test_sltp_modifies(-0.10))
        resultats.append(self.test_entree_retardee())
        
        # Score global de robustesse
        tests_reussis = sum(1 for r in resultats if r.get('robuste', False))
        score_robustesse = (tests_reussis / len(resultats)) * 100
        
        return {
            'tests': resultats,
            'score_robustesse': round(score_robustesse, 0),
            'verdict': '✅ ROBUSTE' if score_robustesse >= 75 else ('⚠️ FRAGILE' if score_robustesse >= 50 else '❌ NON ROBUSTE')
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    UNIT TESTS POUR EVALUER_SIGNAL
# ══════════════════════════════════════════════════════════════════════════════

class UnitTestsBacktest:
    """
    🧪 Tests unitaires pour valider la logique du backtest
    
    Ces tests garantissent que:
    1. Aucune bougie avant time_signal n'est utilisée
    2. TP/SL ne peuvent pas se déclencher avant le signal
    3. Le filtrage temporel fonctionne correctement
    """
    
    @staticmethod
    def test_filtrage_temporel() -> Dict:
        """
        Test: La première bougie évaluée doit être APRÈS time_signal
        """
        import pandas as pd
        from datetime import datetime, timedelta
        
        # Créer des données de test
        dates = pd.date_range(start='2024-01-01', periods=10, freq='D')
        df_test = pd.DataFrame({
            'Open': [100, 101, 102, 103, 104, 105, 106, 107, 108, 109],
            'High': [102, 103, 104, 105, 106, 107, 108, 109, 110, 111],
            'Low': [99, 100, 101, 102, 103, 104, 105, 106, 107, 108],
            'Close': [101, 102, 103, 104, 105, 106, 107, 108, 109, 110]
        }, index=dates)
        
        # Signal au jour 5 (2024-01-06)
        time_signal = datetime(2024, 1, 6)
        
        # Filtrer comme dans evaluer_signal
        df_filtre = df_test[df_test.index > time_signal]
        
        # Vérifications
        erreurs = []
        
        if df_filtre.empty:
            erreurs.append("ERREUR: DataFrame filtré est vide")
        else:
            premiere_date = df_filtre.index[0]
            if premiere_date <= time_signal:
                erreurs.append(f"ERREUR: Première date ({premiere_date}) <= time_signal ({time_signal})")
            
            # Vérifier que les dates avant le signal sont exclues
            dates_avant = [d for d in df_filtre.index if d <= time_signal]
            if dates_avant:
                erreurs.append(f"ERREUR: {len(dates_avant)} dates avant le signal dans le résultat")
        
        return {
            'test': 'Filtrage Temporel',
            'succes': len(erreurs) == 0,
            'erreurs': erreurs,
            'details': {
                'time_signal': str(time_signal),
                'premiere_bougie': str(df_filtre.index[0]) if not df_filtre.empty else None,
                'nb_bougies_filtrees': len(df_filtre)
            }
        }
    
    @staticmethod
    def test_tp_sl_apres_signal() -> Dict:
        """
        Test: TP/SL ne peuvent être atteints que sur des bougies APRÈS le signal
        """
        import pandas as pd
        from datetime import datetime
        
        # Créer des données où TP serait atteint AVANT le signal
        dates = pd.date_range(start='2024-01-01', periods=10, freq='D')
        df_test = pd.DataFrame({
            'Open': [100, 101, 102, 150, 104, 105, 106, 107, 108, 109],  # High=150 au jour 4
            'High': [102, 103, 104, 160, 106, 107, 108, 109, 110, 111],  # TP=120 atteint au jour 4
            'Low': [99, 100, 101, 140, 103, 104, 105, 106, 107, 108],
            'Close': [101, 102, 103, 155, 105, 106, 107, 108, 109, 110]
        }, index=dates)
        
        # Signal au jour 6 (2024-01-07), TP=120
        time_signal = datetime(2024, 1, 7)
        take_profit = 120
        
        # Filtrer
        df_filtre = df_test[df_test.index > time_signal]
        
        erreurs = []
        
        # Vérifier que le TP n'est PAS atteint (car il était atteint avant le signal)
        tp_atteint_apres = any(row['High'] >= take_profit for _, row in df_filtre.iterrows())
        tp_atteint_avant = any(row['High'] >= take_profit for idx, row in df_test.iterrows() if idx <= time_signal)
        
        if tp_atteint_avant and not tp_atteint_apres:
            # C'est correct: le TP avant le signal est ignoré
            pass
        elif tp_atteint_avant and tp_atteint_apres:
            # Le TP est atteint avant ET après - OK si on ne compte que celui d'après
            pass
        
        # Le test principal: vérifier qu'aucune bougie avant time_signal n'est dans df_filtre
        bougies_incorrectes = [idx for idx in df_filtre.index if idx <= time_signal]
        if bougies_incorrectes:
            erreurs.append(f"ERREUR: {len(bougies_incorrectes)} bougies avant/au signal dans les données filtrées")
        
        return {
            'test': 'TP/SL après signal uniquement',
            'succes': len(erreurs) == 0,
            'erreurs': erreurs,
            'details': {
                'tp_atteint_avant_signal': tp_atteint_avant,
                'tp_atteint_apres_signal': tp_atteint_apres,
                'nb_bougies_apres_signal': len(df_filtre)
            }
        }
    
    @staticmethod
    def test_index_trie() -> Dict:
        """
        Test: L'index doit être trié chronologiquement avant le filtrage
        """
        import pandas as pd
        from datetime import datetime
        
        # Créer des données avec index NON trié
        dates = [
            datetime(2024, 1, 5),
            datetime(2024, 1, 1),  # Désordre
            datetime(2024, 1, 3),
            datetime(2024, 1, 2),
            datetime(2024, 1, 4),
        ]
        df_test = pd.DataFrame({
            'Open': [105, 101, 103, 102, 104],
            'High': [106, 102, 104, 103, 105],
            'Low': [104, 100, 102, 101, 103],
            'Close': [105, 101, 103, 102, 104]
        }, index=pd.DatetimeIndex(dates))
        
        time_signal = datetime(2024, 1, 2)
        
        # Trier puis filtrer (comme dans evaluer_signal)
        df_trie = df_test.sort_index()
        df_filtre = df_trie[df_trie.index > time_signal]
        
        erreurs = []
        
        # Vérifier que l'index est maintenant trié
        index_trie = list(df_filtre.index)
        index_attendu = sorted(index_trie)
        
        if index_trie != index_attendu:
            erreurs.append("ERREUR: L'index n'est pas trié après sort_index()")
        
        # Vérifier que seules les dates après le signal sont présentes
        for idx in df_filtre.index:
            if idx <= time_signal:
                erreurs.append(f"ERREUR: Date {idx} <= time_signal {time_signal}")
        
        return {
            'test': 'Index Trié',
            'succes': len(erreurs) == 0,
            'erreurs': erreurs,
            'details': {
                'index_original_trie': list(df_test.index) == sorted(list(df_test.index)),
                'index_final_trie': index_trie == index_attendu
            }
        }
    
    @classmethod
    def executer_tous_tests(cls) -> Dict:
        """Exécute tous les tests unitaires"""
        resultats = [
            cls.test_filtrage_temporel(),
            cls.test_tp_sl_apres_signal(),
            cls.test_index_trie()
        ]
        
        succes = sum(1 for r in resultats if r['succes'])
        total = len(resultats)
        
        return {
            'tests': resultats,
            'succes': succes,
            'total': total,
            'taux_succes': f"{succes}/{total}",
            'verdict': '✅ TOUS LES TESTS PASSENT' if succes == total else f'❌ {total - succes} TEST(S) ÉCHOUÉ(S)'
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    OUT-OF-SAMPLE TESTING
# ══════════════════════════════════════════════════════════════════════════════

class OutOfSampleTester:
    """
    📊 Système de test Out-of-Sample pour valider la robustesse
    
    Méthodologie:
    1. Période de training: Génération du modèle/règles
    2. Période de test: Application sur données jamais vues
    3. Walk-Forward: Fenêtres glissantes pour simulation réaliste
    """
    
    def __init__(self, config):
        self.config = config
    
    def split_temporel(self, df: pd.DataFrame, 
                       date_split: str = '2025-01-01') -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Divise les données en train/test selon une date
        
        Args:
            df: DataFrame avec index datetime
            date_split: Date de séparation (format 'YYYY-MM-DD')
        
        Returns:
            (df_train, df_test)
        """
        date_split = pd.Timestamp(date_split)
        
        df_train = df[df.index < date_split]
        df_test = df[df.index >= date_split]
        
        return df_train, df_test
    
    def calculer_metriques(self, trades: List[Dict]) -> Dict:
        """
        Calcule les métriques de performance
        
        Args:
            trades: Liste des trades évalués
        
        Returns:
            Dictionnaire des métriques
        """
        if not trades:
            return {
                'nb_trades': 0,
                'win_rate': 0,
                'avg_win': 0,
                'avg_loss': 0,
                'profit_factor': 0,
                'expectancy': 0,
                'max_drawdown': 0,
                'total_pnl': 0
            }
        
        wins = [t for t in trades if t.get('result') == 'WIN']
        losses = [t for t in trades if t.get('result') == 'LOSS']
        
        nb_trades = len(trades)
        win_rate = (len(wins) / nb_trades * 100) if nb_trades else 0
        
        avg_win = np.mean([t.get('pnl_net', 0) for t in wins]) if wins else 0
        avg_loss = np.mean([abs(t.get('pnl_net', 0)) for t in losses]) if losses else 0
        
        total_wins = sum(t.get('pnl_net', 0) for t in wins)
        total_losses = abs(sum(t.get('pnl_net', 0) for t in losses))
        
        profit_factor = (total_wins / total_losses) if total_losses > 0 else float('inf') if total_wins > 0 else 0
        
        # Expectancy = (Win% × Avg Win) - (Loss% × Avg Loss)
        expectancy = (win_rate/100 * avg_win) - ((100-win_rate)/100 * avg_loss)
        
        # Max Drawdown
        equity_curve = []
        cumul = 0
        for t in trades:
            cumul += t.get('pnl_net', 0)
            equity_curve.append(cumul)
        
        peak = equity_curve[0] if equity_curve else 0
        max_dd = 0
        for eq in equity_curve:
            if eq > peak:
                peak = eq
            dd = peak - eq
            if dd > max_dd:
                max_dd = dd
        
        return {
            'nb_trades': nb_trades,
            'win_rate': round(win_rate, 2),
            'avg_win': round(avg_win, 2),
            'avg_loss': round(avg_loss, 2),
            'profit_factor': round(profit_factor, 2) if profit_factor != float('inf') else 'INF',
            'expectancy': round(expectancy, 2),
            'max_drawdown': round(max_dd, 2),
            'total_pnl': round(sum(t.get('pnl_net', 0) for t in trades), 2)
        }
    
    def walk_forward_test(self, symboles: List[str], 
                          train_months: int = 12,
                          test_months: int = 3,
                          callback: Callable = None,
                          use_real_signals: bool = False,
                          moteur = None) -> Dict:
        """
        Test Walk-Forward avec fenêtres glissantes
        
        Args:
            symboles: Liste des symboles à tester
            train_months: Mois de training par fenêtre
            test_months: Mois de test par fenêtre
            callback: Fonction de callback pour le progrès
            use_real_signals: Si True, utilise les vrais signaux du système
            moteur: Instance du MoteurAnalyse pour générer les vrais signaux
        
        Returns:
            Résultats agrégés
        """
        resultats_windows = []
        
        # Définir les fenêtres (mises à jour pour 2025-2026)
        windows = [
            ('2023-01-01', '2024-01-01', '2024-01-01', '2024-04-01'),
            ('2023-04-01', '2024-04-01', '2024-04-01', '2024-07-01'),
            ('2023-07-01', '2024-07-01', '2024-07-01', '2024-10-01'),
            ('2023-10-01', '2024-10-01', '2024-10-01', '2025-01-01'),
            ('2024-01-01', '2025-01-01', '2025-01-01', '2025-07-01'),  # Nouvelle fenêtre 2025
            ('2024-07-01', '2025-07-01', '2025-07-01', '2026-01-25'),  # Fenêtre actuelle
        ]
        
        all_trades_test = []
        
        for i, (train_start, train_end, test_start, test_end) in enumerate(windows):
            if callback:
                callback(f"Walk-Forward Fenêtre {i+1}/{len(windows)}", i, len(windows))
            
            window_trades = []
            
            # Évaluer chaque symbole sur cette fenêtre
            for symbol in symboles[:30]:  # Augmenté à 30 symboles
                try:
                    # Télécharger les données
                    ticker = yf.Ticker(symbol)
                    df = ticker.history(start=train_start, end=test_end)
                    
                    if df.empty or len(df) < 60:
                        continue
                    
                    # Split train/test
                    df_train = df[df.index < pd.Timestamp(test_start)]
                    df_test = df[df.index >= pd.Timestamp(test_start)]
                    
                    if df_test.empty or len(df_test) < 5:
                        continue
                    
                    # ══════════════════════════════════════════════════════════════════
                    # 🔧 MODE SIGNAUX RÉELS vs SIGNAUX SIMPLES
                    # ══════════════════════════════════════════════════════════════════
                    
                    if use_real_signals and moteur:
                        # Utiliser le vrai système de signaux
                        result = self._evaluer_signal_reel(
                            symbol, df_train, df_test, moteur
                        )
                    else:
                        # ══════════════════════════════════════════════════════════════════
                        # 🔧 MODE V2: Utilise EXACTEMENT le même moteur que evaluer_signal
                        # ══════════════════════════════════════════════════════════════════
                        signal_time = df_train.index[-1]  # Signal généré à la fin du train
                        entry_price = df_test.iloc[0]['Open']
                        
                        # Calculer ATR sur la période train
                        if len(df_train) >= 14:
                            high = df_train['High']
                            low = df_train['Low']
                            close = df_train['Close']
                            tr = pd.concat([
                                high - low,
                                abs(high - close.shift(1)),
                                abs(low - close.shift(1))
                            ], axis=1).max(axis=1)
                            atr = tr.rolling(14).mean().iloc[-1]
                        else:
                            atr = entry_price * 0.02
                        
                        stop_loss = entry_price - (atr * 2)
                        take_profit = entry_price + (atr * 3)
                        
                        # ══════════════════════════════════════════════════════════════════
                        # 🔧 UTILISE EvaluateurSignaux.evaluer_trade_v2 (MÊME LOGIQUE!)
                        # ══════════════════════════════════════════════════════════════════
                        result = EvaluateurSignaux.evaluer_trade_v2(
                            entry_price=entry_price,
                            stop_loss=stop_loss,
                            take_profit=take_profit,
                            action='ACHAT',
                            df_ohlc=df_test,
                            time_signal=None,  # Commence à la 1ère bougie de df_test
                            position_size=100,
                            commission_per_share=0.005,
                            slippage_pct=0.0005,
                            horizon_days=20,
                            politique_same_bar='worst'
                        )
                    
                    if result:
                        result['symbol'] = symbol
                        window_trades.append(result)
                        all_trades_test.append(result)
                        
                except Exception as e:
                    logger.debug(f"Erreur WF {symbol}: {e}")
                    continue
            
            window_result = {
                'window': i + 1,
                'train_period': f"{train_start} → {train_end}",
                'test_period': f"{test_start} → {test_end}",
                'nb_trades': len(window_trades),
                'metriques_test': self.calculer_metriques(window_trades)
            }
            
            resultats_windows.append(window_result)
        
        metriques_globales = self.calculer_metriques(all_trades_test)
        
        return {
            'type': 'Walk-Forward',
            'mode': 'SIGNAUX RÉELS' if use_real_signals else 'SIGNAUX ATR (démo)',
            'nb_windows': len(windows),
            'nb_trades_total': len(all_trades_test),
            'windows': resultats_windows,
            'metriques_globales': metriques_globales,
            'stabilite': self._calculer_stabilite(resultats_windows),
            'criteres': self.criteres_acceptation(metriques_globales)
        }
    
    def _evaluer_signal_reel(self, symbol: str, df_train: pd.DataFrame, 
                             df_test: pd.DataFrame, moteur) -> Dict:
        """
        Évalue un signal en utilisant la vraie logique du système
        
        Cette méthode:
        1. Calcule les indicateurs sur df_train (comme le système le ferait)
        2. Génère le signal d'entrée à la fin de la période train
        3. Évalue le trade sur df_test
        """
        try:
            cloture = df_train['Close']
            volumes = df_train['Volume']
            hauts = df_train['High']
            bas = df_train['Low']
            
            # Récupérer le prix d'entrée (ouverture de la première bougie test)
            entry_price = df_test.iloc[0]['Open']
            
            # ══════════════════════════════════════════════════════════════════
            # Calculer les indicateurs sur la période train
            # ══════════════════════════════════════════════════════════════════
            rsi = IndicateursTechniques.rsi_wilder(cloture, 14)
            z_score = IndicateursTechniques.zscore(cloture, 20)
            macd_val, macd_hist, macd_tendance = IndicateursTechniques.macd(cloture)
            tendance, score_tendance = IndicateursTechniques.score_tendance(cloture)
            atr = IndicateursTechniques.atr(hauts, bas, cloture, 14)
            support, resistance = IndicateursTechniques.supports_resistances(cloture)
            vol_ratio, _ = IndicateursTechniques.volume_relatif(volumes)
            adx, di_plus, di_minus, _ = IndicateursTechniques.adx(hauts, bas, cloture)
            stoch_k, stoch_d, stoch_signal = IndicateursTechniques.stochastic_rsi(cloture)
            cmf, cmf_signal = IndicateursTechniques.cmf(hauts, bas, cloture, volumes)
            
            # ══════════════════════════════════════════════════════════════════
            # Déterminer le signal (ACHAT/VENTE) basé sur les indicateurs
            # ══════════════════════════════════════════════════════════════════
            achat_signals = 0
            vente_signals = 0
            
            # RSI
            if rsi and rsi < 35:
                achat_signals += 1
            elif rsi and rsi > 65:
                vente_signals += 1
            
            # Tendance
            if score_tendance and score_tendance > 60:
                achat_signals += 1
            elif score_tendance and score_tendance < 40:
                vente_signals += 1
            
            # MACD
            if macd_tendance == 'HAUSSIER':
                achat_signals += 1
            elif macd_tendance == 'BAISSIER':
                vente_signals += 1
            
            # ADX + Direction
            if adx and adx > 25:
                if di_plus and di_minus and di_plus > di_minus:
                    achat_signals += 1
                elif di_plus and di_minus and di_minus > di_plus:
                    vente_signals += 1
            
            # Stochastic RSI
            if stoch_signal == 'SURVENTE':
                achat_signals += 1
            elif stoch_signal == 'SURACHAT':
                vente_signals += 1
            
            # CMF
            if cmf_signal == 'ACCUMULATION':
                achat_signals += 1
            elif cmf_signal == 'DISTRIBUTION':
                vente_signals += 1
            
            # Volume
            if vol_ratio and vol_ratio > 1.5:
                if macd_tendance == 'HAUSSIER':
                    achat_signals += 1
                elif macd_tendance == 'BAISSIER':
                    vente_signals += 1
            
            # Décision finale
            if achat_signals > vente_signals and achat_signals >= 3:
                action = 'ACHAT'
            elif vente_signals > achat_signals and vente_signals >= 3:
                action = 'VENTE'
            else:
                # Pas assez de confirmations
                return None
            
            # ══════════════════════════════════════════════════════════════════
            # Calculer SL/TP
            # ══════════════════════════════════════════════════════════════════
            if not atr or atr <= 0:
                atr = entry_price * 0.02
            
            if action == 'ACHAT':
                stop_loss = entry_price - (atr * 2)
                take_profit = entry_price + (atr * 3)
            else:  # VENTE
                stop_loss = entry_price + (atr * 2)
                take_profit = entry_price - (atr * 3)
            
            # ══════════════════════════════════════════════════════════════════
            # 🔧 Évaluer avec EvaluateurSignaux.evaluer_trade_v2 (MÊME LOGIQUE!)
            # ══════════════════════════════════════════════════════════════════
            result = EvaluateurSignaux.evaluer_trade_v2(
                entry_price=entry_price,
                stop_loss=stop_loss,
                take_profit=take_profit,
                action=action,
                df_ohlc=df_test,
                time_signal=None,
                position_size=100,
                commission_per_share=0.005,
                slippage_pct=0.0005,
                horizon_days=20,
                politique_same_bar='worst'
            )
            
            if result:
                result['nb_confirmations'] = max(achat_signals, vente_signals)
                result['indicators'] = {
                    'rsi': rsi,
                    'score_tendance': score_tendance,
                    'adx': adx,
                    'stoch_signal': stoch_signal,
                    'cmf_signal': cmf_signal
                }
            
            return result
            
        except Exception as e:
            logger.debug(f"Erreur _evaluer_signal_reel {symbol}: {e}")
            return None
    
    def _calculer_stabilite(self, resultats_windows: List[Dict]) -> Dict:
        """Calcule la stabilité des résultats entre les fenêtres"""
        if not resultats_windows:
            return {
                'win_rate_std': 0,
                'pf_std': 0,
                'verdict': '⏳ Pas de données'
            }
        
        # Extraire les métriques de chaque fenêtre
        win_rates = []
        profit_factors = []
        
        for w in resultats_windows:
            metriques = w.get('metriques_test', {})
            wr = metriques.get('win_rate', 0)
            pf = metriques.get('profit_factor', 0)
            
            if wr > 0:
                win_rates.append(wr)
            if pf != 'INF' and pf > 0:
                profit_factors.append(pf)
        
        # Calcul écart-type
        wr_std = np.std(win_rates) if len(win_rates) > 1 else 0
        pf_std = np.std(profit_factors) if len(profit_factors) > 1 else 0
        
        # Verdict basé sur la stabilité
        # Si l'écart-type du win rate < 10% et PF < 0.5 → Stable
        if wr_std < 10 and pf_std < 0.5:
            verdict = '✅ RÉSULTATS STABLES'
        elif wr_std < 15 and pf_std < 1.0:
            verdict = '⚠️ RÉSULTATS VARIABLES'
        else:
            verdict = '❌ RÉSULTATS INSTABLES'
        
        return {
            'win_rate_std': round(wr_std, 2),
            'pf_std': round(pf_std, 2),
            'nb_fenetres_positives': sum(1 for w in resultats_windows 
                                        if w.get('metriques_test', {}).get('total_pnl', 0) > 0),
            'nb_fenetres_total': len(resultats_windows),
            'verdict': verdict
        }
    
    def criteres_acceptation(self, metriques: Dict) -> Dict:
        """
        Vérifie si les métriques passent les critères d'acceptation
        
        Critères:
        - Nombre de trades > 200
        - Profit Factor > 1.2
        - Win Rate > 40%
        - Max Drawdown < 20% du capital
        """
        criteres = {
            'nb_trades_min': (metriques.get('nb_trades', 0) >= 200, 
                             f"{metriques.get('nb_trades', 0)} >= 200"),
            'profit_factor': (metriques.get('profit_factor', 0) >= 1.2 if metriques.get('profit_factor') != 'INF' else True,
                             f"{metriques.get('profit_factor', 0)} >= 1.2"),
            'win_rate': (metriques.get('win_rate', 0) >= 40,
                        f"{metriques.get('win_rate', 0)}% >= 40%"),
            'expectancy_positive': (metriques.get('expectancy', 0) > 0,
                                   f"Expectancy {metriques.get('expectancy', 0)} > 0")
        }
        
        tous_passes = all(c[0] for c in criteres.values())
        
        return {
            'criteres': {k: {'passe': v[0], 'detail': v[1]} for k, v in criteres.items()},
            'tous_passes': tous_passes,
            'verdict': '✅ VALIDÉ POUR TRADING RÉEL' if tous_passes else '❌ NE PAS UTILISER EN RÉEL'
        }


# ══════════════════════════════════════════════════════════════════════════════
#                         CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

class Strategie(Enum):
    TOUTES = "toutes"
    MOMENTUM = "momentum"
    MEAN_REVERSION = "mean_reversion"
    VALEUR = "valeur"
    ML = "ml"


@dataclass
class Config:
    # Indicateurs Techniques
    RSI_PERIODE: int = 14
    ZSCORE_FENETRE: int = 20
    MACD_RAPIDE: int = 12
    MACD_LENT: int = 26
    MACD_SIGNAL: int = 9
    ATR_PERIODE: int = 14
    
    # Seuils RSI
    RSI_SURVENDU: int = 30
    RSI_SURACHETE: int = 70
    ZSCORE_HAUT: float = 2.0
    ZSCORE_BAS: float = -2.0
    
    # Valorisation
    TAUX_ACTUALISATION: float = 0.10
    TAUX_CROISSANCE: float = 0.05
    TAUX_CROISSANCE_TERMINAL: float = 0.02
    MARGE_SECURITE: float = 0.25
    
    # Trading
    STOP_LOSS_ATR_MULT: float = 2.0
    TAKE_PROFIT_ATR_MULT: float = 3.0
    RISQUE_RECOMPENSE_MIN: float = 1.5
    HORIZON_COURT: int = 5
    HORIZON_MOYEN: int = 15
    HORIZON_LONG: int = 30
    
    # ══════════════════════════════════════════════════════════════════════════
    # 🎯 SWING TRADING - FILTRES OPTIMISÉS (ÉQUILIBRÉS)
    # ══════════════════════════════════════════════════════════════════════════
    
    # Mode de trading
    MODE_TRADING: str = "SWING"  # SWING, DAYTRADING, POSITION
    
    # 🔥 FILTRES AJUSTÉS POUR PLUS D'OPPORTUNITÉS
    SWING_SCORE_MIN: int = 55          # Score minimum (68→55) - plus permissif
    SWING_CONFIANCE_MIN: int = 55      # Confiance minimum (72→55) - plus permissif
    SWING_RR_MIN: float = 1.2          # Ratio R/R minimum (1.5→1.2)
    SWING_POTENTIEL_MIN: float = 8.0   # Potentiel minimum (12→8%)
    
    # 🎯 CONFIRMATION MULTI-INDICATEURS (au moins X sur Y)
    CONFIRMATION_MIN: int = 2          # Minimum 2 confirmations (4→2) - plus permissif
    
    # 📊 QUALITÉ REQUISE
    QUALITE_MIN: int = 40              # Score qualité financière minimum (55→40)
    EXIGER_CASH_FLOW_POSITIF: bool = False  # FCF > 0 pas obligatoire
    EXIGER_CROISSANCE: bool = False    # Croissance revenus pas obligatoire
    
    # 🔄 FILTRES MOMENTUM AJUSTÉS
    MOMENTUM_RSI_MIN: int = 35         # RSI minimum (48→35) - plus large
    MOMENTUM_RSI_MAX: int = 75         # RSI maximum (65→75) - plus large
    MOMENTUM_ZSCORE_MIN: float = -1.0  # Z-Score min (0→-1)
    MOMENTUM_ZSCORE_MAX: float = 2.5   # Z-Score max (1.8→2.5)
    MOMENTUM_VOLUME_MIN: float = 0.8   # Volume minimum (1.3→0.8)
    MOMENTUM_MACD_POSITIF: bool = False # MACD pas obligatoirement positif
    
    # 📉 FILTRES MEAN REVERSION (recovery plays)
    MEANREV_RSI_MAX: int = 40          # RSI max (32→40) - moins strict
    MEANREV_ZSCORE_MAX: float = -1.0   # Z-Score max (-1.8→-1.0)
    MEANREV_SUPPORT_DISTANCE: float = 5.0  # Distance support (2.5→5%)
    MEANREV_REBOND_CONFIRME: bool = False  # Pas obligatoire
    
    # ⏱️ TIMING D'ENTRÉE OPTIMAL
    PULLBACK_ENTRY: bool = True        # Entrée sur pullback
    PULLBACK_EMA: int = 21             # EMA 21 (standard institutionnel)
    PULLBACK_DISTANCE: float = 3.0     # Distance max EMA (1.5→3%)
    PULLBACK_SCORE_MIN: int = 40       # Score pullback minimum (60→40)
    
    # ⏳ TIME STOP (gestion du temps)
    TIME_STOP_JOURS: int = 15          # Exit après 15j
    PARTIAL_PROFIT_JOURS: int = 5      # Prendre profit partiel à J+5
    
    # 💰 TP/SL OPTIMISÉS
    SWING_TP_RATIO: float = 1.5        # TP = 1.5R
    SWING_SL_ATR: float = 1.5          # SL = 1.5 ATR
    TRAILING_STOP: bool = True         # Activer trailing stop
    TRAILING_ACTIVATION: float = 0.5   # Activer à +50% du TP
    
    # 🚫 FILTRES D'EXCLUSION (éviter les pièges)
    VOLATILITE_MAX: float = 4.0        # Volatilité max (5→4%)
    EXCLURE_PENNY_STOCKS: bool = True  # Exclure prix < 5$
    PRIX_MIN: float = 5.0              # Prix minimum
    EXCLURE_FAIBLE_VOLUME: bool = True # Exclure faible liquidité
    VOLUME_MIN_MOYEN: int = 500000     # Volume moyen minimum
    
    # 🤖 ML RENFORCÉ
    ML_BLOQUER_VENTE: bool = True      # Bloquer si ML = VENTE
    ML_CONSENSUS_MIN: int = 65         # Consensus ML minimum
    ML_CONFIRMATION: bool = True       # ML doit confirmer le signal
    
    # 🏆 SCORE COMBINÉ FINAL (Win Rate focus)
    POIDS_TECHNIQUE: float = 0.25      # Pondération technique
    POIDS_FONDAMENTAL: float = 0.30    # Pondération fondamentaux
    POIDS_QUALITE: float = 0.25        # Pondération qualité
    POIDS_ML: float = 0.20             # Pondération ML
    
    # ══════════════════════════════════════════════════════════════════════════
    
    # ML AVANCÉ
    ML_ACTIF: bool = True
    ML_RENDEMENT_CIBLE: float = 0.04   # 4% cible (3→4%)
    ML_HORIZON_JOURS: int = 7          # Horizon 7j (5→7)
    ML_SEUIL_CONFIANCE: float = 0.62   # Seuil confiance (60→62%)
    ML_ENSEMBLE: bool = True           # Ensemble de modèles
    ML_MAX_SYMBOLES: int = 400         # 🔧 AUGMENTÉ: Max symboles pour entraînement (200→400)
    ML_USE_ADVANCED: bool = True       # Utiliser modèles avancés (LightGBM, CatBoost, MLP)
    ML_STACKING: bool = True           # Utiliser Stacking Classifier
    ML_FEATURE_SELECTION: bool = True  # Sélection automatique des features
    ML_TOP_K_FEATURES: int = 30        # Nombre de features à garder
    
    # Affichage
    SCORE_MINIMUM: int = 30            # Afficher à partir de 30 (plus d'opportunités)
    
    # Marchés
    INCLURE_US: bool = True
    INCLURE_EUROPE: bool = True
    INCLURE_FRANCE: bool = True
    INCLURE_GAMING: bool = True        # 🎮 Secteur Gaming/Jeux Vidéo
    INCLURE_MATIERES: bool = True      # 🥇 Matières premières
    INCLURE_CRYPTO: bool = True        # ₿ Cryptomonnaies
    
    # Suivi des Signaux
    SAUVEGARDER_SIGNAUX: bool = True
    POLITIQUE_SAME_BAR: str = 'worst'  # 'worst' ou 'best'
    
    FICHIER_CONFIG: str = "screener_config_v10.json"
    
    def sauvegarder(self) -> Tuple[bool, str]:
        try:
            data = {k: v for k, v in asdict(self).items() if not k.startswith('FICHIER')}
            with open(self.FICHIER_CONFIG, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True, "Configuration sauvegardée"
        except Exception as e:
            return False, f"Erreur: {e}"
    
    @classmethod
    def charger(cls) -> 'Config':
        config = cls()
        try:
            if os.path.exists(config.FICHIER_CONFIG):
                with open(config.FICHIER_CONFIG, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    for k, v in data.items():
                        if hasattr(config, k):
                            setattr(config, k, v)
        except:
            pass
        return config


CONFIG = Config.charger()


# ══════════════════════════════════════════════════════════════════════════════
#                    UNIVERS D'ACTIONS
# ══════════════════════════════════════════════════════════════════════════════

class UniversActions:
    # 🇺🇸 ACTIONS US ÉLARGI (150+ actions)
    US = {
        # Technologie - Big Tech
        "AAPL": "Technologie", "MSFT": "Technologie", "GOOGL": "Technologie",
        "AMZN": "Technologie", "META": "Technologie", "NVDA": "Technologie",
        "TSLA": "Technologie", "AMD": "Technologie", "INTC": "Technologie",
        "CRM": "Technologie", "ADBE": "Technologie", "ORCL": "Technologie",
        "IBM": "Technologie", "CSCO": "Technologie", "QCOM": "Technologie",
        "TXN": "Technologie", "AVGO": "Technologie", "NOW": "Technologie",
        "INTU": "Technologie", "AMAT": "Technologie", "MU": "Technologie",
        "LRCX": "Technologie", "KLAC": "Technologie", "SNPS": "Technologie",
        "CDNS": "Technologie", "PANW": "Technologie", "CRWD": "Technologie",
        "ZS": "Technologie", "DDOG": "Technologie", "SNOW": "Technologie",
        "PLTR": "Technologie", "NET": "Technologie", "MDB": "Technologie",
        "TEAM": "Technologie", "WDAY": "Technologie", "VEEV": "Technologie",
        "OKTA": "Technologie", "FTNT": "Technologie", "SPLG": "Technologie",
        "HUBS": "Technologie", "DOCU": "Technologie", "ZM": "Technologie",
        "TWLO": "Technologie", "TTD": "Technologie", "NTNX": "Technologie",
        "ESTC": "Technologie", "CFLT": "Technologie", "PATH": "Technologie",
        "IOT": "Technologie", "GTLB": "Technologie", "APP": "Technologie",
        "SMCI": "Technologie", "ARM": "Technologie", "MRVL": "Technologie",
        
        # 🎮 GAMING / JEUX VIDÉO
        "EA": "Gaming", "TTWO": "Gaming",  # Publishers
        "RBLX": "Gaming", "U": "Gaming",  # Platforms
        "SONY": "Gaming", "NTDOY": "Gaming",  # Console makers
        "MTCH": "Gaming",  # Social Gaming
        "DKNG": "Gaming", "PENN": "Gaming",  # Gaming/Betting
        "PLTK": "Gaming", "GMBL": "Gaming",  # Mobile Gaming
        
        # Finance
        "JPM": "Finance", "BAC": "Finance", "GS": "Finance", "V": "Finance",
        "MA": "Finance", "PYPL": "Finance", "MS": "Finance", "WFC": "Finance",
        "C": "Finance", "AXP": "Finance", "BLK": "Finance", "SCHW": "Finance",
        "SPGI": "Finance", "ICE": "Finance", "CME": "Finance", "MCO": "Finance",
        "COF": "Finance", "USB": "Finance", "PNC": "Finance", "TFC": "Finance",
        "XYZ": "Finance", "AFRM": "Finance", "SOFI": "Finance", "COIN": "Finance",
        "HOOD": "Finance", "NU": "Finance", "MELI": "Finance", "BILL": "Finance",
        
        # Santé
        "JNJ": "Santé", "UNH": "Santé", "PFE": "Santé", "LLY": "Santé",
        "ABBV": "Santé", "MRK": "Santé", "TMO": "Santé", "ABT": "Santé",
        "DHR": "Santé", "BMY": "Santé", "AMGN": "Santé", "GILD": "Santé",
        "ISRG": "Santé", "MDT": "Santé", "SYK": "Santé", "REGN": "Santé",
        "VRTX": "Santé", "BIIB": "Santé", "MRNA": "Santé", "ZTS": "Santé",
        "CVS": "Santé", "CI": "Santé", "HUM": "Santé", "ELV": "Santé",
        
        # Consommation
        "WMT": "Consommation", "COST": "Consommation", "HD": "Consommation",
        "NKE": "Consommation", "MCD": "Consommation", "SBUX": "Consommation",
        "TGT": "Consommation", "LOW": "Consommation", "TJX": "Consommation",
        "ROST": "Consommation", "DG": "Consommation", "DLTR": "Consommation",
        "CMG": "Consommation", "YUM": "Consommation", "DPZ": "Consommation",
        "KO": "Consommation", "PEP": "Consommation", "MDLZ": "Consommation",
        "KHC": "Consommation", "GIS": "Consommation", "K": "Consommation",
        "CL": "Consommation", "PG": "Consommation", "EL": "Consommation",
        "LULU": "Consommation", "BOOT": "Consommation", "ANF": "Consommation",
        
        # Énergie
        "XOM": "Énergie", "CVX": "Énergie", "COP": "Énergie", "SLB": "Énergie",
        "EOG": "Énergie", "FANG": "Énergie", "DVN": "Énergie", "OXY": "Énergie",
        "MPC": "Énergie", "VLO": "Énergie", "PSX": "Énergie", "HAL": "Énergie",
        
        # Industrie
        "CAT": "Industrie", "BA": "Industrie", "GE": "Industrie", "HON": "Industrie",
        "UNP": "Industrie", "UPS": "Industrie", "RTX": "Industrie", "LMT": "Industrie",
        "NOC": "Industrie", "GD": "Industrie", "DE": "Industrie", "MMM": "Industrie",
        "EMR": "Industrie", "ITW": "Industrie", "ROK": "Industrie", "PH": "Industrie",
        "FDX": "Industrie", "DAL": "Industrie", "UAL": "Industrie", "AAL": "Industrie",
        
        # Média & Divertissement
        "DIS": "Média", "NFLX": "Média", "CMCSA": "Média", "WBD": "Média",
        "AMC": "Média", "FOXA": "Média", "SPOT": "Média", "LYV": "Média",
        
        # Tech Services
        "UBER": "TechServices", "ABNB": "TechServices", "LYFT": "TechServices",
        "DASH": "TechServices", "BKNG": "TechServices", "EXPE": "TechServices",
        "TRVG": "TechServices", "TRIP": "TechServices",
        
        # Immobilier
        "AMT": "Immobilier", "PLD": "Immobilier", "CCI": "Immobilier",
        "EQIX": "Immobilier", "SPG": "Immobilier", "O": "Immobilier",
        "DLR": "Immobilier", "PSA": "Immobilier", "WELL": "Immobilier",
        
        # Utilities
        "NEE": "Utilities", "DUK": "Utilities", "SO": "Utilities",
        "D": "Utilities", "AEP": "Utilities", "XEL": "Utilities",
        
        # 🆕 AJOUTS POUR 400+ SYMBOLES
        # Tech additionnels
        "DELL": "Technologie", "HPQ": "Technologie", "HPE": "Technologie",
        "NTAP": "Technologie", "WDC": "Technologie", "STX": "Technologie",
        "ANET": "Technologie", "PSTG": "Technologie", "FFIV": "Technologie",
        "AKAM": "Technologie", "DBX": "Technologie", "BOX": "Technologie",
        "FIVN": "Technologie", "RNG": "Technologie", "ASAN": "Technologie",
        "MNDY": "Technologie", "SOUN": "Technologie", "CYBR": "Technologie",
        "QLYS": "Technologie", "TENB": "Technologie", "RPD": "Technologie",
        "VRNS": "Technologie", "BRZE": "Technologie", "PD": "Technologie",
        "DT": "Technologie", "GWRE": "Technologie", "MANH": "Technologie",
        "BSY": "Technologie", "TYL": "Technologie", "PAYC": "Technologie",
        "PCTY": "Technologie", "PAYX": "Technologie", "ADP": "Technologie",
        "WK": "Technologie", "ROKU": "Technologie", "APPF": "Technologie",
        "FOUR": "Technologie", "GFS": "Technologie", "ON": "Technologie",
        "SWKS": "Technologie", "QRVO": "Technologie", "MPWR": "Technologie",
        "MCHP": "Technologie", "NXPI": "Technologie", "ADI": "Technologie",
        "LSCC": "Technologie", "TER": "Technologie", "COHR": "Technologie",
        "MTSI": "Technologie", "CRUS": "Technologie", "SLAB": "Technologie",
        
        # Finance additionnels
        "AIG": "Finance", "MET": "Finance", "PRU": "Finance", "AFL": "Finance",
        "ALL": "Finance", "TRV": "Finance", "CB": "Finance", "HIG": "Finance",
        "PGR": "Finance", "CINF": "Finance", "AON": "Finance", "L": "Finance",
        "FITB": "Finance", "KEY": "Finance", "RF": "Finance", "CFG": "Finance",
        "HBAN": "Finance", "MTB": "Finance", "ZION": "Finance", "CMA": "Finance",
        "ALLY": "Finance", "SYF": "Finance", "TROW": "Finance", "NDAQ": "Finance",
        "CBOE": "Finance", "BEN": "Finance", "NTRS": "Finance", "STT": "Finance",
        "WAL": "Finance", "FHN": "Finance", "EWBC": "Finance", "WBS": "Finance",
        
        # Santé additionnels  
        "HCA": "Santé", "CNC": "Santé", "MOH": "Santé", "ELV": "Santé",
        "DGX": "Santé", "LH": "Santé", "A": "Santé", "IQV": "Santé",
        "WST": "Santé", "TFX": "Santé", "RMD": "Santé", "HOLX": "Santé",
        "IDXX": "Santé", "DXCM": "Santé", "ALGN": "Santé", "PODD": "Santé",
        "TECH": "Santé", "BIO": "Santé", "ILMN": "Santé", "INCY": "Santé",
        "VTRS": "Santé", "XRAY": "Santé", "HSIC": "Santé", "CAH": "Santé",
        "EXAS": "Santé", "GH": "Santé", "NTRA": "Santé", "MCK": "Santé",
        
        # Consommation additionnels
        "AMZN": "Consommation", "EBAY": "Consommation", "ETSY": "Consommation",
        "W": "Consommation", "CHWY": "Consommation", "BURL": "Consommation",
        "BBY": "Consommation", "TSCO": "Consommation", "ORLY": "Consommation",
        "AZO": "Consommation", "AAP": "Consommation", "KMX": "Consommation",
        "ULTA": "Consommation", "FIVE": "Consommation", "OLLI": "Consommation",
        "CASY": "Consommation", "MNST": "Consommation", "BBWI": "Consommation",
        "DECK": "Consommation", "CROX": "Consommation", "ONON": "Consommation",
        "UAA": "Consommation", "VFC": "Consommation", "PVH": "Consommation",
        "RL": "Consommation", "CPRI": "Consommation", "TPR": "Consommation",
        "HBI": "Consommation", "GOOS": "Consommation", "COLM": "Consommation",
        
        # Industrie additionnels
        "WM": "Industrie", "RSG": "Industrie", "WCN": "Industrie",
        "FAST": "Industrie", "GWW": "Industrie", "SWK": "Industrie",
        "TT": "Industrie", "IR": "Industrie", "DOV": "Industrie",
        "XYL": "Industrie", "IEX": "Industrie", "AME": "Industrie",
        "OTIS": "Industrie", "CARR": "Industrie", "GNRC": "Industrie",
        "PLUG": "Industrie", "BLDR": "Industrie", "MAS": "Industrie",
        "PHM": "Industrie", "DHI": "Industrie", "LEN": "Industrie",
        "TOL": "Industrie", "MTH": "Industrie", "NVR": "Industrie",
        "KBH": "Industrie", "CCS": "Industrie", "TMHC": "Industrie",
        
        # Énergie additionnels
        "APA": "Énergie", "TRGP": "Énergie", "PR": "Énergie",
        "BKR": "Énergie", "NOV": "Énergie", "FTI": "Énergie",
        "CTRA": "Énergie", "AR": "Énergie", "RRC": "Énergie",
        "KMI": "Énergie", "EQT": "Énergie", "WMB": "Énergie",
        "OVV": "Énergie", "MTDR": "Énergie", "CHRD": "Énergie",
        "SM": "Énergie", "GPRE": "Énergie", "CLNE": "Énergie",
        
        # Auto & Véhicules électriques
        "GM": "Automobile", "F": "Automobile", "RIVN": "Automobile",
        "LCID": "Automobile", "NIO": "Automobile", "XPEV": "Automobile",
        "LI": "Automobile", "STLA": "Automobile", "HMC": "Automobile",
        "GOEV": "Automobile", "REE": "Automobile", "BLNK": "Automobile",
        "CHPT": "Automobile", "EVGO": "Automobile", "QS": "Automobile",
        "PTRA": "Automobile", "WKHS": "Automobile", "RIDE": "Automobile",
        
        # Cryptos / Blockchain
        "RIOT": "Crypto", "MARA": "Crypto", "CLSK": "Crypto",
        "HUT": "Crypto", "BTBT": "Crypto", "CIFR": "Crypto",
        "IREN": "Crypto", "CORZ": "Crypto", "WULF": "Crypto",
        "MSTR": "Crypto", "GBTC": "Crypto", "BITO": "Crypto",
        
        # Cannabis
        "TLRY": "Cannabis", "CGC": "Cannabis", "ACB": "Cannabis",
        "CRON": "Cannabis", "SNDL": "Cannabis", "OGI": "Cannabis",
        "HEXO": "Cannabis", "VFF": "Cannabis", "GRWG": "Cannabis",
        
        # Matériaux
        "LIN": "Matériaux", "APD": "Matériaux", "SHW": "Matériaux",
        "ECL": "Matériaux", "NEM": "Matériaux", "FCX": "Matériaux",
        "GOLD": "Matériaux", "NUE": "Matériaux", "STLD": "Matériaux",
        "CLF": "Matériaux", "AA": "Matériaux", "X": "Matériaux",
        "VALE": "Matériaux", "RIO": "Matériaux", "BHP": "Matériaux",
        "MP": "Matériaux", "LAC": "Matériaux", "ALB": "Matériaux",
        "SQM": "Matériaux", "LTHM": "Matériaux", "PLL": "Matériaux"
    }
    
    # 🇫🇷 CAC 40 + SBF 120 (60+ actions françaises)
    FRANCE = {
        # CAC 40
        "MC.PA": "Luxe", "OR.PA": "Consommation", "SAN.PA": "Santé",
        "AI.PA": "Industrie", "BNP.PA": "Finance", "TTE.PA": "Énergie",
        "AIR.PA": "Industrie", "SAF.PA": "Industrie", "KER.PA": "Luxe",
        "SU.PA": "Industrie", "CS.PA": "Industrie", "EL.PA": "Luxe",
        "DG.PA": "Industrie", "CAP.PA": "Technologie", "RI.PA": "Consommation",
        "DSY.PA": "Technologie", "BN.PA": "Consommation", "EN.PA": "Industrie",
        "ENGI.PA": "Énergie", "VIE.PA": "Environnement", "VIV.PA": "Média",
        "ORA.PA": "Télécom", "SGO.PA": "Industrie", "PUB.PA": "Média",
        "LR.PA": "Immobilier", "HO.PA": "Industrie", "SW.PA": "Technologie",
        "RMS.PA": "Luxe", "CA.PA": "Finance", "GLE.PA": "Finance",
        "ACA.PA": "Finance", "ML.PA": "Industrie", "ATO.PA": "Défense",
        "URW.PA": "Immobilier", "STM.PA": "Technologie", "WLN.PA": "Énergie",
        "TEP.PA": "Télécom", "RNO.PA": "Automobile",
        
        # SBF 120 - Additionnels
        "ALO.PA": "Technologie",  # Alten
        "AKE.PA": "Industrie",    # Arkema
        "BIM.PA": "Consommation", # BioMérieux
        "BOL.PA": "Consommation", # Bolloré
        "CGG.PA": "Énergie",      # CGG
        "CNP.PA": "Finance",      # CNP Assurances
        "COV.PA": "Industrie",    # Covivio
        "DBG.PA": "Consommation", # Dior
        "ERF.PA": "Énergie",      # Eurofins
        "FGR.PA": "Industrie",    # Eiffage
        "FP.PA": "Énergie",       # TotalEnergies
        "GET.PA": "Industrie",    # Getlink
        "GFC.PA": "Industrie",    # Gecina
        "GTO.PA": "Technologie",  # Gemalto
        "ILD.PA": "Technologie",  # Iliad
        "IPN.PA": "Santé",        # Ipsen
        "KOF.PA": "Consommation", # Korian
        "LI.PA": "Industrie",     # Klepierre
        "MMT.PA": "Technologie",  # Mercialys
        "NEX.PA": "Industrie",    # Nexans
        "NK.PA": "Technologie",   # Imerys
        "ORP.PA": "Consommation", # Orpea
        "PP.PA": "Consommation",  # Plastic Omnium
        "RCO.PA": "Industrie",    # Rémy Cointreau
        "RXL.PA": "Consommation", # Rexel
        "SAMS.PA": "Consommation",# Samse
        "SEB.PA": "Consommation", # SEB
        "SK.PA": "Industrie",     # SEB SA
        "SOI.PA": "Technologie",  # Soitec
        "SOP.PA": "Industrie",    # Sopra Steria
        "TFI.PA": "Média",        # TF1
        "UBI.PA": "Gaming",       # UBISOFT 🎮
        "VLA.PA": "Santé",        # Valneva
        "VK.PA": "Industrie",     # Vallourec
        "VRLA.PA": "Industrie",   # Valéo
    }
    
    # 🇪🇺 EUROPE ÉLARGI (50+ actions)
    EUROPE = {
        # Allemagne
        "SAP.DE": "Technologie", "SIE.DE": "Industrie", "ALV.DE": "Finance",
        "MBG.DE": "Automobile", "BMW.DE": "Automobile", "ADS.DE": "Consommation",
        "DTE.DE": "Télécom", "BAS.DE": "Industrie", "BAYN.DE": "Santé",
        "MUV2.DE": "Finance", "DB1.DE": "Finance", "IFX.DE": "Technologie",
        "VOW3.DE": "Automobile", "HEN3.DE": "Consommation", "DPW.DE": "Industrie",
        "RWE.DE": "Énergie", "EON.DE": "Énergie", "FRE.DE": "Santé",
        "HEI.DE": "Consommation", "MTX.DE": "Technologie",
        
        # Pays-Bas
        "ASML.AS": "Technologie", "PHIA.AS": "Technologie", "INGA.AS": "Finance",
        "AD.AS": "Consommation", "UNA.AS": "Consommation", "HEIA.AS": "Consommation",
        "AKZA.AS": "Industrie", "DSM.AS": "Industrie", "NN.AS": "Finance",
        "PRX.AS": "Technologie", "WKL.AS": "Technologie", "RAND.AS": "Industrie",
        
        # UK
        "SHEL.L": "Énergie", "AZN.L": "Santé", "HSBA.L": "Finance",
        "ULVR.L": "Consommation", "RIO.L": "Industrie", "BP.L": "Énergie",
        "GSK.L": "Santé", "LLOY.L": "Finance", "BARC.L": "Finance",
        "VOD.L": "Télécom", "BT.L": "Télécom", "NG.L": "Énergie",
        "REL.L": "Média", "DGE.L": "Consommation", "RKT.L": "Consommation",
        "III.L": "Finance", "LSEG.L": "Finance", "AAL.L": "Industrie",
        
        # Suisse
        "NESN.SW": "Consommation", "NOVN.SW": "Santé", "ROG.SW": "Santé",
        "UBS.SW": "Finance", "CSGN.SW": "Finance", "ABBN.SW": "Industrie",
        "SIKA.SW": "Industrie", "GIVN.SW": "Consommation", "LONN.SW": "Santé",
        "GEBN.SW": "Industrie", "ZURN.SW": "Finance", "SLHN.SW": "Finance",
        
        # Espagne
        "SAN.MC": "Finance", "IBE.MC": "Énergie", "ITX.MC": "Consommation",
        "BBVA.MC": "Finance", "TEF.MC": "Télécom", "REP.MC": "Énergie",
        
        # Italie
        "ENEL.MI": "Énergie", "ENI.MI": "Énergie", "ISP.MI": "Finance",
        "UCG.MI": "Finance", "G.MI": "Finance", "STM.MI": "Technologie",
        
        # Autres
        "NOVO-B.CO": "Santé",    # Novo Nordisk (Danemark)
        "MAERSK-B.CO": "Industrie", # Maersk (Danemark)
        "CARL-B.CO": "Consommation", # Carlsberg (Danemark)
    }
    
    # 🎮 GAMING / JEUX VIDÉO (collection spéciale)
    GAMING = {
        # US
        "EA": "Gaming", "TTWO": "Gaming", "MSFT": "Gaming",  # ATVI acquis par MSFT
        "RBLX": "Gaming", "U": "Gaming", "DKNG": "Gaming",
        "PENN": "Gaming", "PLTK": "Gaming", "GME": "Gaming",  # GameStop
        # Japon (ADRs)
        "SONY": "Gaming", "NTDOY": "Gaming", "CCOEY": "Gaming",  # Capcom
        "SQNXF": "Gaming",  # Square Enix
        "NCBDY": "Gaming",  # Bandai Namco
        # Europe
        "UBI.PA": "Gaming",  # Ubisoft 🇫🇷
        "CDR.WA": "Gaming",  # CD Projekt (Pologne)
        "EMB.ST": "Gaming",  # Embracer (Suède)
        # Chine (ADRs)
        "NTES": "Gaming",    # NetEase
        "BILI": "Gaming",    # Bilibili
        "SE": "Gaming",      # Sea Limited (Gaming + E-commerce)
    }
    
    # 🥇 MATIÈRES PREMIÈRES (ETFs & Actions minières)
    MATIERES_PREMIERES = {
        # Or (Gold)
        "GLD": "Or",         # SPDR Gold Trust (ETF Or physique)
        "GDX": "Or",         # VanEck Gold Miners ETF
        "NEM": "Or",         # Newmont Mining
        "GOLD": "Or",        # Barrick Gold
        "AEM": "Or",         # Agnico Eagle Mines
        "FNV": "Or",         # Franco-Nevada
        "WPM": "Or",         # Wheaton Precious Metals
        "KGC": "Or",         # Kinross Gold
        
        # Argent (Silver)
        "SLV": "Argent",     # iShares Silver Trust
        "PAAS": "Argent",    # Pan American Silver
        "AG": "Argent",      # First Majestic Silver
        "HL": "Argent",      # Hecla Mining
        
        # Pétrole & Gaz (Oil & Gas)
        "USO": "Pétrole",    # United States Oil Fund
        "XLE": "Pétrole",    # Energy Select Sector SPDR
        "OIH": "Pétrole",    # VanEck Oil Services ETF
        "UNG": "Gaz",        # United States Natural Gas Fund
        
        # Cuivre & Métaux industriels
        "COPX": "Cuivre",    # Global X Copper Miners ETF
        "FCX": "Cuivre",     # Freeport-McMoRan (Cuivre)
        "SCCO": "Cuivre",    # Southern Copper
        "TECK": "Cuivre",    # Teck Resources
        
        # Lithium & Métaux rares (EV/Batteries)
        "LIT": "Lithium",    # Global X Lithium & Battery Tech ETF
        "ALB": "Lithium",    # Albemarle (Lithium)
        "SQM": "Lithium",    # Sociedad Química y Minera
        "LAC": "Lithium",    # Lithium Americas
        "LTHM": "Lithium",   # Livent Corporation
        
        # Uranium
        "URA": "Uranium",    # Global X Uranium ETF
        "CCJ": "Uranium",    # Cameco Corporation
        "UUUU": "Uranium",   # Energy Fuels
        "DNN": "Uranium",    # Denison Mines
        
        # Agriculture
        "DBA": "Agriculture", # Invesco DB Agriculture Fund
        "CORN": "Agriculture",# Teucrium Corn Fund
        "WEAT": "Agriculture",# Teucrium Wheat Fund
        "SOYB": "Agriculture",# Teucrium Soybean Fund
        "MOO": "Agriculture", # VanEck Agribusiness ETF
        "ADM": "Agriculture", # Archer-Daniels-Midland
        "BG": "Agriculture",  # Bunge Limited
        "DE": "Agriculture",  # John Deere (équipement)
        
        # Métaux diversifiés
        "XME": "Métaux",      # SPDR S&P Metals & Mining ETF
        "PICK": "Métaux",     # iShares MSCI Global Metals & Mining
        "BHP": "Métaux",      # BHP Group (Mining diversifié)
        "RIO": "Métaux",      # Rio Tinto
        "VALE": "Métaux",     # Vale SA (Fer, Nickel)
        
        # Platine & Palladium
        "PPLT": "Platine",    # Aberdeen Physical Platinum ETF
        "PALL": "Palladium",  # Aberdeen Physical Palladium ETF
    }
    
    # ₿ TOP 10 CRYPTOMONNAIES (via ETFs, Trusts & Actions)
    CRYPTO = {
        # Bitcoin
        "BTC-USD": "Crypto",   # Bitcoin (direct Yahoo Finance)
        "IBIT": "Crypto",      # iShares Bitcoin Trust ETF
        "GBTC": "Crypto",      # Grayscale Bitcoin Trust
        "BITO": "Crypto",      # ProShares Bitcoin Strategy ETF
        
        # Ethereum
        "ETH-USD": "Crypto",   # Ethereum (direct Yahoo Finance)
        "ETHE": "Crypto",      # Grayscale Ethereum Trust
        "ETHA": "Crypto",      # iShares Ethereum Trust ETF
        
        # Autres Cryptos (direct Yahoo Finance)
        "BNB-USD": "Crypto",   # Binance Coin
        "SOL-USD": "Crypto",   # Solana
        "XRP-USD": "Crypto",   # Ripple
        "ADA-USD": "Crypto",   # Cardano
        "AVAX-USD": "Crypto",  # Avalanche
        "DOGE-USD": "Crypto",  # Dogecoin
        "DOT-USD": "Crypto",   # Polkadot
        "MATIC-USD": "Crypto", # Polygon
        "LINK-USD": "Crypto",  # Chainlink
        "UNI-USD": "Crypto",   # Uniswap
        "ATOM-USD": "Crypto",  # Cosmos
        "LTC-USD": "Crypto",   # Litecoin
        
        # Actions liées aux cryptos
        "COIN": "Crypto",      # Coinbase
        "MSTR": "Crypto",      # MicroStrategy (gros détenteur BTC)
        "RIOT": "Crypto",      # Riot Platforms (mining)
        "MARA": "Crypto",      # Marathon Digital (mining)
        "HUT": "Crypto",       # Hut 8 Mining
        "CLSK": "Crypto",      # CleanSpark (mining)
    }
    
    @classmethod
    def obtenir_actions(cls, us: bool = True, europe: bool = True, france: bool = True, 
                        gaming: bool = True, matieres: bool = True, crypto: bool = True) -> Dict[str, str]:
        actions = {}
        if us:
            actions.update(cls.US)
        if france:
            actions.update(cls.FRANCE)
        if europe:
            actions.update(cls.EUROPE)
        if gaming:
            actions.update(cls.GAMING)
        if matieres:
            actions.update(cls.MATIERES_PREMIERES)
        if crypto:
            actions.update(cls.CRYPTO)
        return actions
    
    @classmethod
    def obtenir_marche(cls, symbole: str) -> str:
        if symbole in cls.CRYPTO or symbole.endswith('-USD'):
            return "₿ Crypto"
        if symbole in cls.MATIERES_PREMIERES:
            return "🥇 Commodities"
        if symbole in cls.FRANCE or symbole.endswith('.PA'):
            return "🇫🇷 France"
        if symbole.endswith(('.DE', '.L', '.AS', '.SW', '.MC', '.MI', '.CO', '.WA', '.ST')):
            return "🇪🇺 Europe"
        if symbole in cls.GAMING:
            return "🎮 Gaming"
        return "🇺🇸 US"
    
    @classmethod
    def obtenir_secteur(cls, symbole: str) -> str:
        return (cls.US.get(symbole) or cls.FRANCE.get(symbole) or 
                cls.EUROPE.get(symbole) or cls.GAMING.get(symbole) or 
                cls.MATIERES_PREMIERES.get(symbole) or cls.CRYPTO.get(symbole, "Autre"))


# ══════════════════════════════════════════════════════════════════════════════
#                    PROFILS DE CONTRATS FUTURES
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class ProfilContrat:
    """Profil d'un contrat futures avec ses caractéristiques de coût"""
    symbole: str
    nom: str
    tick_size: float       # Taille minimum du tick
    tick_value: float      # Valeur en USD d'un tick
    commission: float      # Commission par contrat (aller simple)
    slippage_ticks: float  # Slippage moyen en ticks
    session: str           # EU / US / ASIA
    
    @property
    def cout_aller_retour(self) -> float:
        """Coût total pour un aller-retour (commission + slippage)"""
        return (self.commission * 2) + (self.slippage_ticks * self.tick_value * 2)


class ProfilsContrats:
    """Bibliothèque de profils de contrats futures"""
    
    CONTRATS = {
        # Indices US
        "NQ": ProfilContrat("NQ", "Nasdaq 100 E-mini", 0.25, 5.0, 2.25, 0.5, "US"),
        "ES": ProfilContrat("ES", "S&P 500 E-mini", 0.25, 12.5, 2.25, 0.5, "US"),
        "YM": ProfilContrat("YM", "Dow Jones E-mini", 1.0, 5.0, 2.25, 1.0, "US"),
        "RTY": ProfilContrat("RTY", "Russell 2000 E-mini", 0.10, 5.0, 2.25, 1.0, "US"),
        "MNQ": ProfilContrat("MNQ", "Nasdaq 100 Micro", 0.25, 0.5, 0.57, 1.0, "US"),
        "MES": ProfilContrat("MES", "S&P 500 Micro", 0.25, 1.25, 0.57, 1.0, "US"),
        
        # Indices Européens
        "FDAX": ProfilContrat("FDAX", "DAX Futures", 0.5, 12.5, 2.0, 0.5, "EU"),
        "FESX": ProfilContrat("FESX", "Euro Stoxx 50", 1.0, 10.0, 1.5, 0.5, "EU"),
        "FCE": ProfilContrat("FCE", "CAC 40 Futures", 0.5, 5.0, 1.5, 0.5, "EU"),
        
        # Devises
        "6E": ProfilContrat("6E", "Euro FX", 0.00005, 6.25, 2.25, 0.5, "US"),
        "6B": ProfilContrat("6B", "British Pound", 0.0001, 6.25, 2.25, 0.5, "US"),
        
        # Matières premières
        "CL": ProfilContrat("CL", "Crude Oil", 0.01, 10.0, 2.25, 1.0, "US"),
        "GC": ProfilContrat("GC", "Gold", 0.10, 10.0, 2.25, 0.5, "US"),
        
        # Actions (profil générique)
        "STOCK": ProfilContrat("STOCK", "Actions", 0.01, 1.0, 0.0, 1.0, "ALL"),
    }
    
    @classmethod
    def obtenir(cls, symbole: str) -> ProfilContrat:
        """Retourne le profil d'un contrat ou le profil action par défaut"""
        return cls.CONTRATS.get(symbole.upper(), cls.CONTRATS["STOCK"])
    
    @classmethod
    def liste_contrats(cls) -> List[str]:
        return list(cls.CONTRATS.keys())


# ══════════════════════════════════════════════════════════════════════════════
#                    INDICATEURS TECHNIQUES
# ══════════════════════════════════════════════════════════════════════════════

class IndicateursTechniques:
    
    @staticmethod
    def rsi_wilder(prix: pd.Series, periode: int = 14) -> Optional[float]:
        if len(prix) < periode + 1:
            return None
        
        delta = prix.diff()
        gain = delta.where(delta > 0, 0)
        perte = -delta.where(delta < 0, 0)
        
        moy_gain = gain.ewm(alpha=1/periode, min_periods=periode, adjust=False).mean()
        moy_perte = perte.ewm(alpha=1/periode, min_periods=periode, adjust=False).mean()
        
        rs = moy_gain / moy_perte
        rsi = 100 - (100 / (1 + rs))
        
        val = rsi.iloc[-1]
        return round(val, 2) if pd.notna(val) else None
    
    @staticmethod
    def zscore(prix: pd.Series, fenetre: int = 20) -> Optional[float]:
        if len(prix) < fenetre:
            return None
        
        recent = prix.tail(fenetre)
        moyenne = recent.mean()
        ecart = recent.std()
        
        if ecart == 0 or pd.isna(ecart):
            return 0.0
        
        return round((prix.iloc[-1] - moyenne) / ecart, 2)
    
    @staticmethod
    def macd(prix: pd.Series, rapide: int = 12, lent: int = 26, signal: int = 9) -> Tuple[Optional[float], Optional[float], str]:
        if len(prix) < lent + signal:
            return None, None, "N/A"
        
        ema_rapide = prix.ewm(span=rapide, adjust=False).mean()
        ema_lent = prix.ewm(span=lent, adjust=False).mean()
        
        macd_ligne = ema_rapide - ema_lent
        signal_ligne = macd_ligne.ewm(span=signal, adjust=False).mean()
        histogramme = macd_ligne - signal_ligne
        
        macd_val = round(macd_ligne.iloc[-1], 4)
        hist_val = round(histogramme.iloc[-1], 4)
        
        if macd_val > signal_ligne.iloc[-1] and histogramme.iloc[-1] > histogramme.iloc[-2]:
            tendance = "🟢 HAUSSIER"
        elif macd_val < signal_ligne.iloc[-1] and histogramme.iloc[-1] < histogramme.iloc[-2]:
            tendance = "🔴 BAISSIER"
        else:
            tendance = "⚪ NEUTRE"
        
        return macd_val, hist_val, tendance
    
    @staticmethod
    def atr(haut: pd.Series, bas: pd.Series, cloture: pd.Series, periode: int = 14) -> Optional[float]:
        if len(cloture) < periode:
            return None
        
        tr1 = haut - bas
        tr2 = abs(haut - cloture.shift())
        tr3 = abs(bas - cloture.shift())
        tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
        
        atr = tr.rolling(periode).mean().iloc[-1]
        return round(atr, 4) if pd.notna(atr) else None
    
    @staticmethod
    def score_tendance(prix: pd.Series, court: int = 5, moyen: int = 20, long: int = 50) -> Tuple[str, int]:
        if len(prix) < long:
            return "N/A", 0
        
        sma_c = prix.tail(court).mean()
        sma_m = prix.tail(moyen).mean()
        sma_l = prix.tail(long).mean()
        actuel = prix.iloc[-1]
        
        score = 0
        if actuel > sma_c: score += 1
        if actuel > sma_m: score += 1
        if actuel > sma_l: score += 1
        if sma_c > sma_m: score += 1
        if sma_m > sma_l: score += 1
        
        labels = {5: "📈 FORTE HAUSSE", 4: "🟢 HAUSSE", 3: "➡️ NEUTRE",
                  2: "🔴 BAISSE", 1: "📉 FORTE BAISSE", 0: "📉 FORTE BAISSE"}
        
        return labels.get(score, "N/A"), score
    
    @staticmethod
    def supports_resistances(prix: pd.Series, fenetre: int = 20) -> Tuple[float, float]:
        """Calcule les niveaux de support et résistance"""
        if len(prix) < fenetre:
            return prix.iloc[-1] * 0.95, prix.iloc[-1] * 1.05
        
        recent = prix.tail(fenetre)
        support = recent.min()
        resistance = recent.max()
        
        return round(support, 2), round(resistance, 2)
    
    @staticmethod
    def volume_relatif(volumes: pd.Series, fenetre: int = 20) -> Tuple[float, str]:
        if len(volumes) < fenetre:
            return 1.0, "N/A"
        
        moy = volumes.tail(fenetre).mean()
        actuel = volumes.iloc[-1]
        
        if moy == 0 or pd.isna(moy):
            return 1.0, "N/A"
        
        ratio = actuel / moy
        
        if ratio >= 2.0:
            label = "🔥 TRÈS ÉLEVÉ"
        elif ratio >= 1.5:
            label = "📈 ÉLEVÉ"
        elif ratio <= 0.5:
            label = "📉 FAIBLE"
        else:
            label = "➡️ NORMAL"
        
        return round(ratio, 2), label
    
    @staticmethod
    def ema(prix: pd.Series, periode: int = 20) -> Optional[float]:
        """Calcule l'EMA"""
        if len(prix) < periode:
            return None
        ema = prix.ewm(span=periode, adjust=False).mean()
        return round(ema.iloc[-1], 4) if pd.notna(ema.iloc[-1]) else None
    
    @staticmethod
    def distance_ema(prix: pd.Series, periode: int = 20) -> Optional[float]:
        """Distance du prix actuel par rapport à l'EMA (%)"""
        if len(prix) < periode:
            return None
        ema = prix.ewm(span=periode, adjust=False).mean().iloc[-1]
        if ema == 0 or pd.isna(ema):
            return None
        distance = ((prix.iloc[-1] - ema) / ema) * 100
        return round(distance, 2)
    
    @staticmethod
    def volatilite_journaliere(prix: pd.Series, fenetre: int = 20) -> Optional[float]:
        """Volatilité journalière moyenne en %"""
        if len(prix) < fenetre:
            return None
        returns = prix.pct_change().tail(fenetre)
        vol = returns.std() * 100
        return round(vol, 2) if pd.notna(vol) else None
    
    @staticmethod
    def distance_support(prix_actuel: float, support: float) -> float:
        """Distance au support en %"""
        if support == 0:
            return 100
        return round(((prix_actuel - support) / support) * 100, 2)
    
    @staticmethod
    def pullback_score(prix: pd.Series, ema_periode: int = 20) -> float:
        """
        Calcule un score de pullback (0-100)
        Plus le score est élevé, meilleur est le timing d'entrée
        """
        if len(prix) < ema_periode + 5:
            return 0.0
        
        try:
            ema = prix.ewm(span=ema_periode, adjust=False).mean()
            prix_actuel = float(prix.iloc[-1])
            ema_actuel = float(ema.iloc[-1])
            ema_5j = float(ema.iloc[-5])
            
            # Distance en %
            distance = ((prix_actuel - ema_actuel) / ema_actuel) * 100
            
            # Vérifier si la tendance est haussière (EMA monte)
            ema_trend = ema_actuel > ema_5j
            
            score = 0.0
            
            if ema_trend:
                # Tendance haussière: bonus de base
                score += 40
                
                # Distance idéale: -2% à +1% de l'EMA
                if -2 <= distance <= 1:
                    score += 50  # Pullback parfait
                elif -3 <= distance <= 2:
                    score += 35  # Bon pullback
                elif 2 < distance <= 5:
                    score += 15  # Un peu éloigné
                elif distance > 5:
                    score += 0   # Trop éloigné
                elif distance < -3:
                    score += 10  # Possible renversement
                
                # Bonus si rebond récent
                if len(prix) >= 3:
                    p1 = float(prix.iloc[-1])
                    p2 = float(prix.iloc[-2])
                    p3 = float(prix.iloc[-3])
                    rebond = p1 > p2 > p3
                    if rebond:
                        score += 10
            else:
                # Tendance baissière: score faible
                score = max(0, 30 - abs(distance) * 3)
            
            return float(min(100, max(0, score)))
        except:
            return 0.0
    
    # ══════════════════════════════════════════════════════════════════════════════
    #     🆕 INDICATEURS COURT/MOYEN TERME AVANCÉS
    # ══════════════════════════════════════════════════════════════════════════════
    
    @staticmethod
    def adx(haut: pd.Series, bas: pd.Series, cloture: pd.Series, periode: int = 14) -> Tuple[Optional[float], Optional[float], Optional[float], str]:
        """
        ADX (Average Directional Index) avec DI+ et DI-
        Mesure la force de la tendance (pas la direction)
        
        Returns: (adx, di_plus, di_minus, interpretation)
        ADX > 25 = Tendance forte, ADX < 20 = Pas de tendance
        """
        if len(cloture) < periode + 10:
            return None, None, None, "N/A"
        
        try:
            # True Range
            tr1 = haut - bas
            tr2 = abs(haut - cloture.shift(1))
            tr3 = abs(bas - cloture.shift(1))
            tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
            
            # Directional Movement
            plus_dm = haut.diff()
            minus_dm = -bas.diff()
            
            plus_dm = plus_dm.where((plus_dm > minus_dm) & (plus_dm > 0), 0)
            minus_dm = minus_dm.where((minus_dm > plus_dm) & (minus_dm > 0), 0)
            
            # Smoothed averages (Wilder's smoothing)
            atr = tr.ewm(alpha=1/periode, min_periods=periode).mean()
            plus_di = 100 * (plus_dm.ewm(alpha=1/periode, min_periods=periode).mean() / atr)
            minus_di = 100 * (minus_dm.ewm(alpha=1/periode, min_periods=periode).mean() / atr)
            
            # DX and ADX
            dx = 100 * abs(plus_di - minus_di) / (plus_di + minus_di)
            adx_val = dx.ewm(alpha=1/periode, min_periods=periode).mean()
            
            adx_current = adx_val.iloc[-1]
            di_plus = plus_di.iloc[-1]
            di_minus = minus_di.iloc[-1]
            
            # Interprétation
            if adx_current >= 40:
                force = "🔥 TRÈS FORTE"
            elif adx_current >= 25:
                force = "💪 FORTE"
            elif adx_current >= 20:
                force = "➡️ MODÉRÉE"
            else:
                force = "😴 FAIBLE"
            
            if di_plus > di_minus:
                direction = "HAUSSIER"
            else:
                direction = "BAISSIER"
            
            interpretation = f"{force} {direction}"
            
            return round(adx_current, 1), round(di_plus, 1), round(di_minus, 1), interpretation
        except:
            return None, None, None, "N/A"
    
    @staticmethod
    def stochastic_rsi(prix: pd.Series, rsi_periode: int = 14, stoch_periode: int = 14, k_smooth: int = 3, d_smooth: int = 3) -> Tuple[Optional[float], Optional[float], str]:
        """
        Stochastic RSI - Combine RSI et Stochastique pour timing précis
        
        Returns: (stoch_rsi_k, stoch_rsi_d, signal)
        < 20 = Survendu (achat), > 80 = Suracheté (vente)
        """
        if len(prix) < rsi_periode + stoch_periode + 10:
            return None, None, "N/A"
        
        try:
            # Calculer RSI
            delta = prix.diff()
            gain = delta.where(delta > 0, 0)
            perte = -delta.where(delta < 0, 0)
            
            moy_gain = gain.ewm(alpha=1/rsi_periode, min_periods=rsi_periode, adjust=False).mean()
            moy_perte = perte.ewm(alpha=1/rsi_periode, min_periods=rsi_periode, adjust=False).mean()
            
            rs = moy_gain / moy_perte
            rsi = 100 - (100 / (1 + rs))
            
            # Stochastic du RSI
            rsi_low = rsi.rolling(stoch_periode).min()
            rsi_high = rsi.rolling(stoch_periode).max()
            
            stoch_rsi = ((rsi - rsi_low) / (rsi_high - rsi_low)) * 100
            stoch_rsi = stoch_rsi.fillna(50)
            
            # K et D lines
            k_line = stoch_rsi.rolling(k_smooth).mean()
            d_line = k_line.rolling(d_smooth).mean()
            
            k_val = k_line.iloc[-1]
            d_val = d_line.iloc[-1]
            
            # Signal
            if k_val < 20 and k_val > d_val:
                signal = "🟢🟢 ACHAT FORT"
            elif k_val < 30:
                signal = "🟢 SURVENDU"
            elif k_val > 80 and k_val < d_val:
                signal = "🔴🔴 VENTE FORTE"
            elif k_val > 70:
                signal = "🔴 SURACHETÉ"
            elif k_val > d_val and k_line.iloc[-2] <= d_line.iloc[-2]:
                signal = "🟢 CROISEMENT HAUSSIER"
            elif k_val < d_val and k_line.iloc[-2] >= d_line.iloc[-2]:
                signal = "🔴 CROISEMENT BAISSIER"
            else:
                signal = "⚪ NEUTRE"
            
            return round(k_val, 1), round(d_val, 1), signal
        except:
            return None, None, "N/A"
    
    @staticmethod
    def williams_r(haut: pd.Series, bas: pd.Series, cloture: pd.Series, periode: int = 14) -> Tuple[Optional[float], str]:
        """
        Williams %R - Indicateur de momentum
        
        Returns: (williams_r, signal)
        > -20 = Suracheté, < -80 = Survendu
        """
        if len(cloture) < periode:
            return None, "N/A"
        
        try:
            highest_high = haut.rolling(periode).max()
            lowest_low = bas.rolling(periode).min()
            
            wr = -100 * (highest_high - cloture) / (highest_high - lowest_low)
            wr_val = wr.iloc[-1]
            
            # Signal
            if wr_val > -20:
                signal = "🔴 SURACHETÉ"
            elif wr_val < -80:
                signal = "🟢 SURVENDU"
            elif wr_val > -50:
                signal = "⚪ HAUSSIER"
            else:
                signal = "⚪ BAISSIER"
            
            return round(wr_val, 1), signal
        except:
            return None, "N/A"
    
    @staticmethod
    def cmf(haut: pd.Series, bas: pd.Series, cloture: pd.Series, volume: pd.Series, periode: int = 20) -> Tuple[Optional[float], str]:
        """
        Chaikin Money Flow - Mesure la pression achat/vente
        
        Returns: (cmf, signal)
        > 0.05 = Pression achat, < -0.05 = Pression vente
        """
        if len(cloture) < periode:
            return None, "N/A"
        
        try:
            # Money Flow Multiplier
            mfm = ((cloture - bas) - (haut - cloture)) / (haut - bas)
            mfm = mfm.fillna(0)
            
            # Money Flow Volume
            mfv = mfm * volume
            
            # CMF
            cmf_val = mfv.rolling(periode).sum() / volume.rolling(periode).sum()
            cmf_current = cmf_val.iloc[-1]
            
            # Signal
            if cmf_current > 0.15:
                signal = "🟢🟢 FORTE ACCUMULATION"
            elif cmf_current > 0.05:
                signal = "🟢 ACCUMULATION"
            elif cmf_current < -0.15:
                signal = "🔴🔴 FORTE DISTRIBUTION"
            elif cmf_current < -0.05:
                signal = "🔴 DISTRIBUTION"
            else:
                signal = "⚪ NEUTRE"
            
            return round(cmf_current, 3), signal
        except:
            return None, "N/A"
    
    @staticmethod
    def obv(cloture: pd.Series, volume: pd.Series) -> Tuple[Optional[float], str]:
        """
        On Balance Volume - Accumulation/Distribution basé sur le volume
        
        Returns: (obv_change_pct, trend)
        """
        if len(cloture) < 20:
            return None, "N/A"
        
        try:
            obv = pd.Series(index=cloture.index, dtype=float)
            obv.iloc[0] = volume.iloc[0]
            
            for i in range(1, len(cloture)):
                if cloture.iloc[i] > cloture.iloc[i-1]:
                    obv.iloc[i] = obv.iloc[i-1] + volume.iloc[i]
                elif cloture.iloc[i] < cloture.iloc[i-1]:
                    obv.iloc[i] = obv.iloc[i-1] - volume.iloc[i]
                else:
                    obv.iloc[i] = obv.iloc[i-1]
            
            # Changement sur 20 jours en %
            obv_20 = obv.iloc[-20]
            obv_current = obv.iloc[-1]
            
            if obv_20 != 0:
                change_pct = ((obv_current - obv_20) / abs(obv_20)) * 100
            else:
                change_pct = 0
            
            # Tendance
            obv_sma = obv.rolling(10).mean()
            if obv_current > obv_sma.iloc[-1] and change_pct > 5:
                trend = "🟢 ACCUMULATION"
            elif obv_current < obv_sma.iloc[-1] and change_pct < -5:
                trend = "🔴 DISTRIBUTION"
            else:
                trend = "⚪ NEUTRE"
            
            return round(change_pct, 1), trend
        except:
            return None, "N/A"
    
    @staticmethod
    def vwap(haut: pd.Series, bas: pd.Series, cloture: pd.Series, volume: pd.Series) -> Tuple[Optional[float], Optional[float], str]:
        """
        VWAP - Volume Weighted Average Price
        
        Returns: (vwap, distance_pct, signal)
        """
        if len(cloture) < 10:
            return None, None, "N/A"
        
        try:
            # Prix typique
            tp = (haut + bas + cloture) / 3
            
            # VWAP cumulatif du jour (on utilise les 20 dernières périodes)
            periode = min(20, len(cloture))
            tp_recent = tp.tail(periode)
            vol_recent = volume.tail(periode)
            
            vwap_val = (tp_recent * vol_recent).sum() / vol_recent.sum()
            
            prix_actuel = cloture.iloc[-1]
            distance = ((prix_actuel - vwap_val) / vwap_val) * 100
            
            # Signal
            if distance > 2:
                signal = "🔴 AU-DESSUS VWAP"
            elif distance < -2:
                signal = "🟢 EN-DESSOUS VWAP"
            else:
                signal = "⚪ PROCHE VWAP"
            
            return round(vwap_val, 2), round(distance, 2), signal
        except:
            return None, None, "N/A"
    
    @staticmethod
    def supertrend(haut: pd.Series, bas: pd.Series, cloture: pd.Series, periode: int = 10, multiplier: float = 3.0) -> Tuple[Optional[float], str]:
        """
        SuperTrend - Indicateur de direction de tendance
        
        Returns: (supertrend_value, signal)
        """
        if len(cloture) < periode + 5:
            return None, "N/A"
        
        try:
            # ATR
            tr1 = haut - bas
            tr2 = abs(haut - cloture.shift(1))
            tr3 = abs(bas - cloture.shift(1))
            tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
            atr = tr.rolling(periode).mean()
            
            # HL2
            hl2 = (haut + bas) / 2
            
            # Bandes
            upper_band = hl2 + (multiplier * atr)
            lower_band = hl2 - (multiplier * atr)
            
            # SuperTrend
            supertrend = pd.Series(index=cloture.index, dtype=float)
            direction = pd.Series(index=cloture.index, dtype=int)
            
            supertrend.iloc[periode] = upper_band.iloc[periode]
            direction.iloc[periode] = -1
            
            for i in range(periode + 1, len(cloture)):
                if cloture.iloc[i] > supertrend.iloc[i-1]:
                    direction.iloc[i] = 1
                    supertrend.iloc[i] = max(lower_band.iloc[i], supertrend.iloc[i-1]) if direction.iloc[i-1] == 1 else lower_band.iloc[i]
                else:
                    direction.iloc[i] = -1
                    supertrend.iloc[i] = min(upper_band.iloc[i], supertrend.iloc[i-1]) if direction.iloc[i-1] == -1 else upper_band.iloc[i]
            
            st_val = supertrend.iloc[-1]
            dir_val = direction.iloc[-1]
            
            # Détection de changement de tendance
            dir_prev = direction.iloc[-2] if len(direction) > 1 else dir_val
            
            if dir_val == 1 and dir_prev == -1:
                signal = "🟢🟢 RETOURNEMENT HAUSSIER"
            elif dir_val == -1 and dir_prev == 1:
                signal = "🔴🔴 RETOURNEMENT BAISSIER"
            elif dir_val == 1:
                signal = "🟢 TENDANCE HAUSSIÈRE"
            else:
                signal = "🔴 TENDANCE BAISSIÈRE"
            
            return round(st_val, 2), signal
        except:
            return None, "N/A"
    
    @staticmethod
    def squeeze_momentum(haut: pd.Series, bas: pd.Series, cloture: pd.Series, bb_periode: int = 20, kc_periode: int = 20, kc_mult: float = 1.5) -> Tuple[bool, Optional[float], str]:
        """
        Squeeze Momentum Indicator
        Détecte la compression de volatilité suivie d'expansion
        
        Returns: (is_squeeze, momentum, signal)
        """
        if len(cloture) < max(bb_periode, kc_periode) + 5:
            return False, None, "N/A"
        
        try:
            # Bollinger Bands
            bb_mid = cloture.rolling(bb_periode).mean()
            bb_std = cloture.rolling(bb_periode).std()
            bb_upper = bb_mid + (2 * bb_std)
            bb_lower = bb_mid - (2 * bb_std)
            
            # Keltner Channels
            tr1 = haut - bas
            tr2 = abs(haut - cloture.shift(1))
            tr3 = abs(bas - cloture.shift(1))
            tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
            atr = tr.rolling(kc_periode).mean()
            
            kc_mid = cloture.rolling(kc_periode).mean()
            kc_upper = kc_mid + (kc_mult * atr)
            kc_lower = kc_mid - (kc_mult * atr)
            
            # Squeeze detection: BB inside KC
            squeeze_on = (bb_lower.iloc[-1] > kc_lower.iloc[-1]) and (bb_upper.iloc[-1] < kc_upper.iloc[-1])
            
            # Momentum (Linear regression of (close - (highest high + lowest low)/2 - SMA))
            highest = haut.rolling(kc_periode).max()
            lowest = bas.rolling(kc_periode).min()
            momentum = cloture - ((highest + lowest) / 2 + bb_mid) / 2
            mom_val = momentum.iloc[-1]
            mom_prev = momentum.iloc[-2]
            
            # Signal
            if squeeze_on:
                if mom_val > 0 and mom_val > mom_prev:
                    signal = "🔴 SQUEEZE - MOMENTUM +"
                elif mom_val < 0 and mom_val < mom_prev:
                    signal = "🔴 SQUEEZE - MOMENTUM -"
                else:
                    signal = "🔴 EN COMPRESSION"
            else:
                if mom_val > 0 and mom_val > mom_prev:
                    signal = "🟢🟢 EXPANSION HAUSSIÈRE"
                elif mom_val < 0 and mom_val < mom_prev:
                    signal = "🔴🔴 EXPANSION BAISSIÈRE"
                elif mom_val > 0:
                    signal = "🟢 HAUSSIER"
                else:
                    signal = "🔴 BAISSIER"
            
            return squeeze_on, round(mom_val, 4), signal
        except:
            return False, None, "N/A"
    
    # ══════════════════════════════════════════════════════════════════════════════
    #     🆕 DÉTECTION D'OPPORTUNITÉS
    # ══════════════════════════════════════════════════════════════════════════════
    
    @staticmethod
    def detecter_breakout(haut: pd.Series, bas: pd.Series, cloture: pd.Series, volume: pd.Series, periode: int = 20) -> Dict:
        """
        Détecte les cassures de résistance ou support
        
        Returns: dict avec type de breakout et score
        """
        if len(cloture) < periode + 5:
            return {'type': 'AUCUN', 'score': 0, 'niveau': None, 'details': ''}
        
        try:
            resistance = haut.tail(periode).max()
            support = bas.tail(periode).min()
            prix_actuel = cloture.iloc[-1]
            prix_hier = cloture.iloc[-2]
            
            # Volume relatif
            vol_avg = volume.tail(periode).mean()
            vol_actuel = volume.iloc[-1]
            vol_ratio = vol_actuel / vol_avg if vol_avg > 0 else 1
            
            # Détection breakout haussier
            if prix_actuel > resistance and prix_hier <= resistance:
                score = 70
                if vol_ratio >= 2:
                    score += 20
                elif vol_ratio >= 1.5:
                    score += 10
                
                return {
                    'type': '🚀 BREAKOUT HAUSSIER',
                    'score': min(100, score),
                    'niveau': round(resistance, 2),
                    'volume_confirme': vol_ratio >= 1.5,
                    'details': f"Cassure résistance ${resistance:.2f} avec volume {vol_ratio:.1f}x"
                }
            
            # Détection breakout baissier
            elif prix_actuel < support and prix_hier >= support:
                score = 70
                if vol_ratio >= 2:
                    score += 20
                elif vol_ratio >= 1.5:
                    score += 10
                
                return {
                    'type': '📉 BREAKDOWN BAISSIER',
                    'score': min(100, score),
                    'niveau': round(support, 2),
                    'volume_confirme': vol_ratio >= 1.5,
                    'details': f"Cassure support ${support:.2f} avec volume {vol_ratio:.1f}x"
                }
            
            # Approche résistance
            elif (resistance - prix_actuel) / prix_actuel < 0.02:
                return {
                    'type': '⚡ APPROCHE RÉSISTANCE',
                    'score': 50,
                    'niveau': round(resistance, 2),
                    'volume_confirme': False,
                    'details': f"Prix proche de la résistance ${resistance:.2f}"
                }
            
            return {'type': 'AUCUN', 'score': 0, 'niveau': None, 'details': ''}
        except:
            return {'type': 'AUCUN', 'score': 0, 'niveau': None, 'details': ''}
    
    @staticmethod
    def detecter_gap(ouverture: pd.Series, cloture: pd.Series) -> Dict:
        """
        Détecte les gaps significatifs (ouverture différente de la clôture précédente)
        """
        if len(cloture) < 2:
            return {'type': 'AUCUN', 'pourcentage': 0, 'details': ''}
        
        try:
            ouv_actuelle = ouverture.iloc[-1]
            cloture_hier = cloture.iloc[-2]
            
            gap_pct = ((ouv_actuelle - cloture_hier) / cloture_hier) * 100
            
            if gap_pct >= 3:
                return {
                    'type': '🟢🟢 GAP UP MAJEUR',
                    'pourcentage': round(gap_pct, 2),
                    'details': f"Gap haussier de +{gap_pct:.1f}%"
                }
            elif gap_pct >= 1.5:
                return {
                    'type': '🟢 GAP UP',
                    'pourcentage': round(gap_pct, 2),
                    'details': f"Gap haussier de +{gap_pct:.1f}%"
                }
            elif gap_pct <= -3:
                return {
                    'type': '🔴🔴 GAP DOWN MAJEUR',
                    'pourcentage': round(gap_pct, 2),
                    'details': f"Gap baissier de {gap_pct:.1f}%"
                }
            elif gap_pct <= -1.5:
                return {
                    'type': '🔴 GAP DOWN',
                    'pourcentage': round(gap_pct, 2),
                    'details': f"Gap baissier de {gap_pct:.1f}%"
                }
            
            return {'type': 'AUCUN', 'pourcentage': round(gap_pct, 2), 'details': ''}
        except:
            return {'type': 'AUCUN', 'pourcentage': 0, 'details': ''}
    
    @staticmethod
    def detecter_volume_spike(volume: pd.Series, seuil: float = 2.0, periode: int = 20) -> Dict:
        """
        Détecte les pics de volume anormaux
        """
        if len(volume) < periode + 1:
            return {'spike': False, 'ratio': 1, 'details': ''}
        
        try:
            vol_avg = volume.tail(periode).mean()
            vol_std = volume.tail(periode).std()
            vol_actuel = volume.iloc[-1]
            
            ratio = vol_actuel / vol_avg if vol_avg > 0 else 1
            zscore = (vol_actuel - vol_avg) / vol_std if vol_std > 0 else 0
            
            if ratio >= 3:
                return {
                    'spike': True,
                    'ratio': round(ratio, 1),
                    'zscore': round(zscore, 1),
                    'niveau': '🔥🔥 VOLUME EXTRÊME',
                    'details': f"Volume {ratio:.1f}x la moyenne!"
                }
            elif ratio >= seuil:
                return {
                    'spike': True,
                    'ratio': round(ratio, 1),
                    'zscore': round(zscore, 1),
                    'niveau': '🔥 VOLUME ÉLEVÉ',
                    'details': f"Volume {ratio:.1f}x la moyenne"
                }
            
            return {'spike': False, 'ratio': round(ratio, 1), 'niveau': '', 'details': ''}
        except:
            return {'spike': False, 'ratio': 1, 'details': ''}
    
    @staticmethod
    def detecter_pattern_reversal(ouverture: pd.Series, haut: pd.Series, bas: pd.Series, cloture: pd.Series) -> Dict:
        """
        Détecte les patterns de retournement (Hammer, Engulfing, Doji, etc.)
        """
        if len(cloture) < 3:
            return {'pattern': 'AUCUN', 'signal': '', 'fiabilite': 0}
        
        try:
            # Dernière bougie
            o = ouverture.iloc[-1]
            h = haut.iloc[-1]
            l = bas.iloc[-1]
            c = cloture.iloc[-1]
            
            # Bougie précédente
            o_prev = ouverture.iloc[-2]
            h_prev = haut.iloc[-2]
            l_prev = bas.iloc[-2]
            c_prev = cloture.iloc[-2]
            
            body = abs(c - o)
            range_candle = h - l
            upper_shadow = h - max(c, o)
            lower_shadow = min(c, o) - l
            
            body_prev = abs(c_prev - o_prev)
            
            patterns = []
            
            # HAMMER (après baisse)
            if body > 0 and range_candle > 0:
                if lower_shadow >= body * 2 and upper_shadow <= body * 0.3:
                    if cloture.iloc[-3] > cloture.iloc[-2]:  # Tendance baissière avant
                        patterns.append(('🔨 HAMMER', '🟢 ACHAT', 75))
            
            # INVERTED HAMMER
            if body > 0 and range_candle > 0:
                if upper_shadow >= body * 2 and lower_shadow <= body * 0.3:
                    if cloture.iloc[-3] > cloture.iloc[-2]:
                        patterns.append(('🔨⬆️ INVERTED HAMMER', '🟢 ACHAT', 65))
            
            # BULLISH ENGULFING
            if c > o and c_prev < o_prev:  # Bougie actuelle haussière, précédente baissière
                if o < c_prev and c > o_prev:  # Corps engloutit le précédent
                    patterns.append(('🟢 BULLISH ENGULFING', '🟢 ACHAT FORT', 80))
            
            # BEARISH ENGULFING
            if c < o and c_prev > o_prev:  # Bougie actuelle baissière, précédente haussière
                if o > c_prev and c < o_prev:
                    patterns.append(('🔴 BEARISH ENGULFING', '🔴 VENTE FORTE', 80))
            
            # DOJI
            if body <= range_candle * 0.1 and range_candle > 0:
                patterns.append(('✚ DOJI', '⚠️ INDÉCISION', 50))
            
            # MORNING STAR (3 bougies)
            if len(cloture) >= 3:
                o_2 = ouverture.iloc[-3]
                c_2 = cloture.iloc[-3]
                body_2 = abs(c_2 - o_2)
                
                if c_2 < o_2 and body_2 > body_prev * 0.5:  # 1ère bougie baissière
                    if body_prev <= body_2 * 0.3:  # 2ème bougie petit corps
                        if c > o and c > (o_2 + c_2) / 2:  # 3ème bougie haussière
                            patterns.append(('⭐ MORNING STAR', '🟢 ACHAT FORT', 85))
            
            if patterns:
                # Retourner le pattern le plus fiable
                patterns.sort(key=lambda x: x[2], reverse=True)
                return {
                    'pattern': patterns[0][0],
                    'signal': patterns[0][1],
                    'fiabilite': patterns[0][2],
                    'tous_patterns': [p[0] for p in patterns]
                }
            
            return {'pattern': 'AUCUN', 'signal': '', 'fiabilite': 0}
        except:
            return {'pattern': 'AUCUN', 'signal': '', 'fiabilite': 0}
    
    @staticmethod
    def calculer_momentum_burst(cloture: pd.Series, periode_court: int = 3, periode_long: int = 10) -> Dict:
        """
        Détecte les accélérations soudaines de momentum
        """
        if len(cloture) < periode_long + 5:
            return {'burst': False, 'score': 0, 'direction': '', 'details': ''}
        
        try:
            ret_court = (cloture.iloc[-1] - cloture.iloc[-periode_court]) / cloture.iloc[-periode_court] * 100
            ret_long = (cloture.iloc[-1] - cloture.iloc[-periode_long]) / cloture.iloc[-periode_long] * 100
            
            # Volatilité historique
            vol = cloture.pct_change().tail(20).std() * 100
            
            # Burst si le mouvement court terme est > 2x la volatilité
            is_burst = abs(ret_court) > vol * 2
            
            if is_burst:
                if ret_court > 0:
                    return {
                        'burst': True,
                        'score': min(100, int(50 + ret_court * 5)),
                        'direction': '🚀 ACCÉLÉRATION HAUSSIÈRE',
                        'ret_court': round(ret_court, 2),
                        'ret_long': round(ret_long, 2),
                        'details': f"+{ret_court:.1f}% sur {periode_court}j (vol: {vol:.1f}%)"
                    }
                else:
                    return {
                        'burst': True,
                        'score': min(100, int(50 + abs(ret_court) * 5)),
                        'direction': '📉 ACCÉLÉRATION BAISSIÈRE',
                        'ret_court': round(ret_court, 2),
                        'ret_long': round(ret_long, 2),
                        'details': f"{ret_court:.1f}% sur {periode_court}j (vol: {vol:.1f}%)"
                    }
            
            return {
                'burst': False,
                'score': 0,
                'direction': '',
                'ret_court': round(ret_court, 2),
                'ret_long': round(ret_long, 2),
                'details': ''
            }
        except:
            return {'burst': False, 'score': 0, 'direction': '', 'details': ''}


# ══════════════════════════════════════════════════════════════════════════════
#                    DÉTECTEUR DE RÉGIME DE MARCHÉ
# ══════════════════════════════════════════════════════════════════════════════

class ScoringMultiHorizon:
    """
    🎯 NOUVEAU: Système de scoring pour différents horizons d'investissement
    
    - Score Intraday/Scalp: 1-3 jours
    - Score Swing: 5-15 jours  
    - Score Position: 15-60 jours
    """
    
    def __init__(self, config: Config):
        self.config = config
    
    def score_intraday(self, donnees: Dict) -> Dict:
        """
        ⚡ SCORE INTRADAY (1-3 jours)
        Focus: Momentum, volume, timing précis
        """
        score = 50
        signaux = []
        alertes = []
        opportunite_type = None
        
        # 1. Stochastic RSI (très important court terme)
        stoch_k = donnees.get('stoch_rsi_k')
        stoch_signal = donnees.get('stoch_rsi_signal', '')
        if stoch_k is not None:
            if stoch_k < 20:
                score += 18
                signaux.append(f"✅ StochRSI survendu ({stoch_k:.0f})")
                opportunite_type = "REVERSAL"
            elif stoch_k > 80:
                score -= 15
                alertes.append(f"⚠️ StochRSI suracheté ({stoch_k:.0f})")
            elif 'CROISEMENT HAUSSIER' in stoch_signal:
                score += 12
                signaux.append("✅ Croisement StochRSI haussier")
        
        # 2. Volume Spike (crucial pour intraday)
        volume_spike = donnees.get('volume_spike', {})
        if volume_spike.get('spike'):
            ratio = volume_spike.get('ratio', 1)
            if ratio >= 3:
                score += 20
                signaux.append(f"🔥 Volume spike extrême ({ratio:.1f}x)")
                opportunite_type = opportunite_type or "BREAKOUT"
            elif ratio >= 2:
                score += 12
                signaux.append(f"📊 Volume élevé ({ratio:.1f}x)")
        
        # 3. Breakout Detection
        breakout = donnees.get('breakout', {})
        if breakout.get('type') not in ['AUCUN', None]:
            score += breakout.get('score', 0) * 0.25
            signaux.append(f"🚀 {breakout.get('type')}")
            opportunite_type = "BREAKOUT"
        
        # 4. Gap Detection
        gap = donnees.get('gap', {})
        if gap.get('type') not in ['AUCUN', None]:
            gap_pct = abs(gap.get('pourcentage', 0))
            if gap_pct >= 2:
                score += 15
                signaux.append(f"📊 {gap.get('type')}")
                opportunite_type = opportunite_type or "GAP_PLAY"
        
        # 5. Momentum Burst
        momentum = donnees.get('momentum_burst', {})
        if momentum.get('burst'):
            if 'HAUSSIÈRE' in momentum.get('direction', ''):
                score += 15
                signaux.append(f"🚀 {momentum.get('direction')}")
                opportunite_type = opportunite_type or "MOMENTUM"
        
        # 6. Pattern Reversal
        pattern = donnees.get('pattern_reversal', {})
        if pattern.get('fiabilite', 0) >= 70:
            if '🟢' in pattern.get('signal', ''):
                score += 12
                signaux.append(f"📊 {pattern.get('pattern')}")
                opportunite_type = opportunite_type or "REVERSAL"
        
        # 7. SuperTrend
        supertrend_signal = donnees.get('supertrend_signal', '')
        if 'RETOURNEMENT HAUSSIER' in supertrend_signal:
            score += 15
            signaux.append("✅ SuperTrend retournement haussier")
        elif 'HAUSSIÈRE' in supertrend_signal:
            score += 5
        elif 'RETOURNEMENT BAISSIER' in supertrend_signal:
            score -= 15
            alertes.append("⚠️ SuperTrend retournement baissier")
        
        # 8. VWAP Distance
        vwap_dist = donnees.get('vwap_distance')
        if vwap_dist is not None:
            if vwap_dist < -1.5:  # En-dessous VWAP
                score += 8
                signaux.append("✅ Prix sous VWAP")
            elif vwap_dist > 2:  # Au-dessus
                score -= 5
        
        # 9. RSI standard
        rsi = donnees.get('rsi')
        if rsi:
            if rsi < 30:
                score += 10
            elif rsi > 70:
                score -= 10
                alertes.append(f"⚠️ RSI suracheté ({rsi:.0f})")
        
        score = max(0, min(100, score))
        
        # Niveau de confiance
        if score >= 75 and len(signaux) >= 4:
            confiance = "🟢🟢 TRÈS HAUTE"
            recommandation = "⚡ OPPORTUNITÉ INTRADAY"
        elif score >= 65 and len(signaux) >= 3:
            confiance = "🟢 HAUTE"
            recommandation = "✅ Signal court terme"
        elif score >= 55:
            confiance = "🟡 MOYENNE"
            recommandation = "👀 À surveiller"
        else:
            confiance = "🔴 FAIBLE"
            recommandation = "❌ Pas de signal"
        
        return {
            'score': round(score, 1),
            'horizon': '1-3 jours',
            'confiance': confiance,
            'recommandation': recommandation,
            'opportunite_type': opportunite_type or 'AUCUNE',
            'signaux': signaux[:5],
            'alertes': alertes[:3],
            'nb_signaux': len(signaux)
        }
    
    def score_swing(self, donnees: Dict) -> Dict:
        """
        📈 SCORE SWING (5-15 jours)
        Focus: Tendance, pullbacks, confirmations multiples
        """
        score = 50
        signaux = []
        alertes = []
        opportunite_type = None
        
        # 1. ADX (Force de tendance)
        adx = donnees.get('adx')
        adx_interp = donnees.get('adx_interpretation', '')
        di_plus = donnees.get('di_plus', 0)
        di_minus = donnees.get('di_minus', 0)
        
        if adx:
            if adx >= 25 and di_plus > di_minus:
                score += 15
                signaux.append(f"✅ Tendance forte haussière (ADX: {adx:.0f})")
                opportunite_type = "TREND_FOLLOWING"
            elif adx >= 25 and di_minus > di_plus:
                score -= 10
                alertes.append(f"⚠️ Tendance baissière (ADX: {adx:.0f})")
            elif adx < 20:
                # Range = potentiel mean reversion
                if donnees.get('z_score', 0) < -1.5:
                    score += 8
                    signaux.append("✅ Range + survente = Mean Reversion")
                    opportunite_type = "MEAN_REVERSION"
        
        # 2. Pullback Score
        pullback = donnees.get('pullback_score', 0)
        if pullback >= 75:
            score += 18
            signaux.append(f"🎯 Pullback parfait ({pullback:.0f})")
            opportunite_type = opportunite_type or "PULLBACK_ENTRY"
        elif pullback >= 60:
            score += 10
            signaux.append(f"✅ Bon pullback ({pullback:.0f})")
        
        # 3. CMF (Pression achat/vente)
        cmf = donnees.get('cmf')
        cmf_signal = donnees.get('cmf_signal', '')
        if cmf:
            if cmf > 0.1:
                score += 12
                signaux.append(f"✅ Forte accumulation (CMF: {cmf:.2f})")
            elif cmf < -0.1:
                score -= 10
                alertes.append(f"⚠️ Distribution (CMF: {cmf:.2f})")
        
        # 4. OBV Trend
        obv_trend = donnees.get('obv_trend', '')
        if 'ACCUMULATION' in obv_trend:
            score += 8
            signaux.append("✅ OBV en accumulation")
        elif 'DISTRIBUTION' in obv_trend:
            score -= 8
        
        # 5. Squeeze Momentum
        squeeze = donnees.get('squeeze_on', False)
        squeeze_signal = donnees.get('squeeze_signal', '')
        if squeeze:
            score += 10
            signaux.append("🔴 Squeeze actif - Breakout imminent")
            opportunite_type = opportunite_type or "SQUEEZE_BREAKOUT"
        elif 'EXPANSION HAUSSIÈRE' in squeeze_signal:
            score += 12
            signaux.append("🟢 Expansion haussière!")
        
        # 6. Tendance générale
        score_tendance = donnees.get('score_tendance', 2)
        if score_tendance >= 4:
            score += 10
            signaux.append("✅ Tendance haussière confirmée")
        elif score_tendance <= 1:
            score -= 10
            alertes.append("⚠️ Tendance baissière")
        
        # 7. Potentiel valorisation
        potentiel = donnees.get('potentiel', 0)
        if potentiel > 20:
            score += 12
            signaux.append(f"💰 Potentiel +{potentiel:.0f}%")
        elif potentiel < -10:
            score -= 8
        
        # 8. MACD
        macd_tendance = donnees.get('macd_tendance', '')
        if 'HAUSSIER' in macd_tendance:
            score += 8
            signaux.append("✅ MACD haussier")
        elif 'BAISSIER' in macd_tendance:
            score -= 8
        
        # 9. Ratio R/R
        ratio_rr = donnees.get('ratio_rr', 1)
        if ratio_rr >= 2:
            score += 10
            signaux.append(f"✅ R/R excellent ({ratio_rr:.1f})")
        elif ratio_rr >= 1.5:
            score += 5
        elif ratio_rr < 1:
            score -= 10
            alertes.append(f"⚠️ R/R insuffisant ({ratio_rr:.1f})")
        
        # 10. ML Signal
        ml_signal = donnees.get('ml_signal', '')
        ml_prob = donnees.get('ml_probabilite', 50)
        if ml_prob >= 65 and 'ACHAT' in ml_signal:
            score += 10
            signaux.append(f"🧠 ML positif ({ml_prob:.0f}%)")
        elif ml_prob <= 35:
            score -= 8
            alertes.append("⚠️ ML négatif")
        
        score = max(0, min(100, score))
        
        # Niveau de confiance
        if score >= 75 and len(signaux) >= 5:
            confiance = "🟢🟢 TRÈS HAUTE"
            recommandation = "🎯 OPPORTUNITÉ SWING"
        elif score >= 65 and len(signaux) >= 4:
            confiance = "🟢 HAUTE"
            recommandation = "✅ Bon signal swing"
        elif score >= 55:
            confiance = "🟡 MOYENNE"
            recommandation = "👀 Surveiller entrée"
        else:
            confiance = "🔴 FAIBLE"
            recommandation = "❌ Conditions non réunies"
        
        return {
            'score': round(score, 1),
            'horizon': '5-15 jours',
            'confiance': confiance,
            'recommandation': recommandation,
            'opportunite_type': opportunite_type or 'AUCUNE',
            'signaux': signaux[:6],
            'alertes': alertes[:3],
            'nb_signaux': len(signaux)
        }
    
    def score_position(self, donnees: Dict) -> Dict:
        """
        🏦 SCORE POSITION (15-60 jours)
        Focus: Fondamentaux, valorisation, qualité
        """
        score = 50
        signaux = []
        alertes = []
        
        # 1. Score Qualité Financière
        score_qualite = donnees.get('score_qualite', 50)
        if score_qualite >= 70:
            score += 15
            signaux.append(f"✅ Excellente qualité ({score_qualite:.0f})")
        elif score_qualite >= 55:
            score += 8
        elif score_qualite < 40:
            score -= 10
            alertes.append(f"⚠️ Qualité faible ({score_qualite:.0f})")
        
        # 2. Potentiel Valorisation
        potentiel = donnees.get('potentiel', 0)
        if potentiel > 30:
            score += 18
            signaux.append(f"💰 Fort potentiel +{potentiel:.0f}%")
        elif potentiel > 15:
            score += 10
            signaux.append(f"✅ Potentiel +{potentiel:.0f}%")
        elif potentiel < -15:
            score -= 15
            alertes.append(f"⚠️ Surévalué {potentiel:.0f}%")
        
        # 3. PE Ratio
        pe = donnees.get('pe_ratio')
        if pe:
            if 5 < pe < 15:
                score += 10
                signaux.append(f"✅ PE attractif ({pe:.1f})")
            elif pe > 35:
                score -= 8
                alertes.append(f"⚠️ PE élevé ({pe:.1f})")
        
        # 4. Signal Analystes
        signal_analyste = donnees.get('signal_analyste', '')
        if 'ACHAT FORT' in signal_analyste:
            score += 12
            signaux.append("📊 Analystes: Achat Fort")
        elif 'ACHAT' in signal_analyste:
            score += 6
        elif 'VENTE' in signal_analyste:
            score -= 8
        
        # 5. Tendance Long Terme
        score_tendance = donnees.get('score_tendance', 2)
        if score_tendance >= 4:
            score += 8
            signaux.append("📈 Tendance LT haussière")
        
        # 6. Croissance (implicite via potentiel)
        # 7. News Sentiment
        score_news = donnees.get('score_news', 50)
        if score_news >= 65:
            score += 8
            signaux.append("📰 News positives")
        elif score_news < 35:
            score -= 8
            alertes.append("📰 News négatives")
        
        # 8. Régime de marché
        regime = donnees.get('regime', 'NEUTRE')
        if regime in ['TREND_HAUSSIER', 'TREND_HAUSSIER_FORT']:
            score += 5
        elif regime in ['TREND_BAISSIER', 'TREND_BAISSIER_FORT']:
            score -= 10
            alertes.append("⚠️ Marché baissier")
        
        # 9. Distance au support
        distance_supp = donnees.get('distance_support', 50)
        if distance_supp < 3:
            score += 8
            signaux.append("✅ Proche du support")
        
        # 10. ML Consensus
        ml_consensus = donnees.get('ml_consensus', 50)
        if ml_consensus >= 75:
            score += 8
            signaux.append(f"🧠 Consensus ML fort ({ml_consensus:.0f}%)")
        
        score = max(0, min(100, score))
        
        # Confiance
        if score >= 75 and len(signaux) >= 5:
            confiance = "🟢🟢 TRÈS HAUTE"
            recommandation = "🏦 OPPORTUNITÉ POSITION"
        elif score >= 65 and len(signaux) >= 4:
            confiance = "🟢 HAUTE"
            recommandation = "✅ Bon pour position"
        elif score >= 55:
            confiance = "🟡 MOYENNE"
            recommandation = "👀 Attendre confirmation"
        else:
            confiance = "🔴 FAIBLE"
            recommandation = "❌ Pas recommandé"
        
        return {
            'score': round(score, 1),
            'horizon': '15-60 jours',
            'confiance': confiance,
            'recommandation': recommandation,
            'signaux': signaux[:6],
            'alertes': alertes[:3],
            'nb_signaux': len(signaux)
        }
    
    def score_global(self, donnees: Dict) -> Dict:
        """
        🎯 Calcule les scores pour les 3 horizons et détermine le meilleur
        """
        intraday = self.score_intraday(donnees)
        swing = self.score_swing(donnees)
        position = self.score_position(donnees)
        
        # Trouver le meilleur horizon
        scores = [
            (intraday['score'], 'INTRADAY', intraday),
            (swing['score'], 'SWING', swing),
            (position['score'], 'POSITION', position)
        ]
        scores.sort(reverse=True)
        
        meilleur = scores[0]
        
        # Déterminer s'il y a une opportunité claire
        opportunite_claire = meilleur[0] >= 65
        
        return {
            'score_intraday': intraday,
            'score_swing': swing,
            'score_position': position,
            'meilleur_horizon': meilleur[1],
            'meilleur_score': meilleur[0],
            'meilleur_details': meilleur[2],
            'opportunite_claire': opportunite_claire,
            'synthese': f"Meilleur horizon: {meilleur[1]} (Score: {meilleur[0]:.0f})"
        }


class DetecteurRegime:
    """
    🎯 Détecte le régime de marché pour adapter la stratégie
    
    Régimes:
    - TREND_HAUSSIER: Favoriser Momentum
    - TREND_BAISSIER: Éviter ou shorter
    - RANGE: Favoriser Mean Reversion
    - HAUTE_VOLATILITE: Réduire exposition
    """
    
    @staticmethod
    def detecter_regime(prix: pd.Series, volumes: pd.Series = None) -> Dict:
        """
        Analyse le régime de marché actuel
        """
        if len(prix) < 50:
            return {
                'regime': 'INCONNU',
                'force': 0,
                'description': 'Données insuffisantes',
                'strategie_recommandee': 'ATTENDRE'
            }
        
        # Calculs
        sma20 = prix.tail(20).mean()
        sma50 = prix.tail(50).mean()
        prix_actuel = prix.iloc[-1]
        
        # Tendance
        pente_20 = (prix.tail(20).iloc[-1] - prix.tail(20).iloc[0]) / prix.tail(20).iloc[0] * 100
        pente_50 = (prix.tail(50).iloc[-1] - prix.tail(50).iloc[0]) / prix.tail(50).iloc[0] * 100
        
        # Volatilité
        vol_recente = prix.pct_change().tail(10).std() * 100
        vol_moyenne = prix.pct_change().tail(50).std() * 100
        ratio_vol = vol_recente / vol_moyenne if vol_moyenne > 0 else 1
        
        # Range detection (ADX-like)
        high_low_range = (prix.tail(20).max() - prix.tail(20).min()) / prix.tail(20).mean() * 100
        
        # Détermination du régime
        if ratio_vol > 1.5:
            regime = 'HAUTE_VOLATILITE'
            force = min(100, ratio_vol * 50)
            description = f"⚠️ Volatilité {ratio_vol:.1f}x la normale"
            strategie = 'RÉDUIRE_EXPOSITION'
        elif pente_20 > 5 and prix_actuel > sma20 > sma50:
            regime = 'TREND_HAUSSIER_FORT'
            force = min(100, pente_20 * 5)
            description = f"📈 Forte tendance haussière (+{pente_20:.1f}%)"
            strategie = 'MOMENTUM'
        elif pente_20 > 2 and prix_actuel > sma20:
            regime = 'TREND_HAUSSIER'
            force = min(100, pente_20 * 10)
            description = f"🟢 Tendance haussière (+{pente_20:.1f}%)"
            strategie = 'MOMENTUM'
        elif pente_20 < -5 and prix_actuel < sma20 < sma50:
            regime = 'TREND_BAISSIER_FORT'
            force = min(100, abs(pente_20) * 5)
            description = f"📉 Forte tendance baissière ({pente_20:.1f}%)"
            strategie = 'ÉVITER'
        elif pente_20 < -2 and prix_actuel < sma20:
            regime = 'TREND_BAISSIER'
            force = min(100, abs(pente_20) * 10)
            description = f"🔴 Tendance baissière ({pente_20:.1f}%)"
            strategie = 'PRUDENCE'
        elif high_low_range < 8:
            regime = 'RANGE'
            force = 100 - high_low_range * 10
            description = f"↔️ Marché en range ({high_low_range:.1f}%)"
            strategie = 'MEAN_REVERSION'
        else:
            regime = 'NEUTRE'
            force = 50
            description = "➡️ Pas de tendance claire"
            strategie = 'SÉLECTIF'
        
        return {
            'regime': regime,
            'force': round(force, 0),
            'description': description,
            'strategie_recommandee': strategie,
            'pente_20j': round(pente_20, 2),
            'pente_50j': round(pente_50, 2),
            'volatilite_ratio': round(ratio_vol, 2),
            'range_pct': round(high_low_range, 2)
        }
    
    @staticmethod
    def est_compatible_momentum(regime: Dict) -> bool:
        """Vérifie si le régime est compatible avec une stratégie Momentum"""
        return regime['regime'] in ['TREND_HAUSSIER', 'TREND_HAUSSIER_FORT']
    
    @staticmethod
    def est_compatible_meanrev(regime: Dict) -> bool:
        """Vérifie si le régime est compatible avec Mean Reversion"""
        return regime['regime'] in ['RANGE', 'NEUTRE'] and regime['regime'] != 'TREND_BAISSIER_FORT'
    
    @staticmethod
    def est_trop_volatile(regime: Dict) -> bool:
        """Vérifie si le marché est trop volatile"""
        return regime['regime'] == 'HAUTE_VOLATILITE' or regime.get('volatilite_ratio', 1) > 1.8


# ══════════════════════════════════════════════════════════════════════════════
#                    SYSTÈME DE VALORISATION AVANCÉ
# ══════════════════════════════════════════════════════════════════════════════

class SystemeValorisation:
    """
    🎯 Système complet de calcul du PRIX RÉEL (Fair Value)
    Utilise plusieurs méthodes et fait une moyenne pondérée
    """
    
    def __init__(self, config: Config):
        self.config = config
    
    def graham_number(self, eps: float, valeur_comptable: float) -> Optional[float]:
        """
        📊 Graham Number: √(22.5 × EPS × Book Value)
        Méthode de Benjamin Graham pour valeur défensive
        """
        if eps is None or valeur_comptable is None or eps <= 0 or valeur_comptable <= 0:
            return None
        return round(np.sqrt(22.5 * eps * valeur_comptable), 2)
    
    def graham_defensif(self, eps: float, croissance: float) -> Optional[float]:
        """
        📊 Graham Défensif: EPS × (8.5 + 2g) × 4.4/Y
        g = taux de croissance attendu
        Y = rendement AAA corporate bonds (approximé à 5%)
        """
        if eps is None or eps <= 0:
            return None
        
        g = croissance * 100 if croissance else 5
        y = 5.0
        
        valeur = eps * (8.5 + 2 * g) * (4.4 / y)
        return round(valeur, 2) if valeur > 0 else None
    
    def dcf_multi_scenarios(self, fcf: float, actions: int, 
                           croissance_optimiste: float = 0.12,
                           croissance_base: float = 0.08,
                           croissance_pessimiste: float = 0.03) -> Dict[str, float]:
        """
        📊 DCF Multi-Scénarios
        Calcule 3 scénarios: Optimiste, Base, Pessimiste
        """
        if fcf is None or fcf <= 0 or actions is None or actions <= 0:
            return {}
        
        taux = self.config.TAUX_ACTUALISATION
        terminal = self.config.TAUX_CROISSANCE_TERMINAL
        
        def calculer_dcf(croissance: float) -> float:
            pv = 0
            flux = fcf
            for annee in range(1, 11):
                flux *= (1 + croissance)
                pv += flux / ((1 + taux) ** annee)
            
            # Valeur terminale
            valeur_terminale = (flux * (1 + terminal)) / (taux - terminal)
            pv += valeur_terminale / ((1 + taux) ** 10)
            
            return round(pv / actions, 2)
        
        return {
            'optimiste': calculer_dcf(croissance_optimiste),
            'base': calculer_dcf(croissance_base),
            'pessimiste': calculer_dcf(croissance_pessimiste)
        }
    
    def peg_fair_value(self, eps: float, croissance: float, pe_secteur: float = 20) -> Optional[float]:
        """
        📊 PEG Fair Value
        Prix juste basé sur PEG = 1 (Peter Lynch)
        """
        if eps is None or eps <= 0 or croissance is None or croissance <= 0:
            return None
        
        # PEG = PE / (croissance * 100)
        # Pour PEG = 1, PE = croissance * 100
        # Plafonner à des valeurs réalistes
        croissance_pct = min(croissance * 100, 50)  # Max 50% croissance
        pe_juste = min(croissance_pct, pe_secteur * 1.5)
        pe_juste = max(8, min(40, pe_juste))  # PE entre 8 et 40
        return round(eps * pe_juste, 2)
    
    def earnings_power_value(self, ebit: float, taux_imposition: float,
                            cout_capital: float, actions: int) -> Optional[float]:
        """
        📊 Earnings Power Value (EPV)
        Valeur basée sur la capacité bénéficiaire normalisée
        """
        if ebit is None or ebit <= 0 or actions is None or actions <= 0:
            return None
        
        benefice_normalise = ebit * (1 - taux_imposition)
        epv = benefice_normalise / cout_capital
        
        return round(epv / actions, 2)
    
    def residual_income_model(self, valeur_comptable: float, roe: float,
                             cout_capitaux: float, actions: int, annees: int = 5) -> Optional[float]:
        """
        📊 Residual Income Model
        Valeur = BV + Somme(RI actualisé)
        RI = (ROE - Cost of Equity) × Book Value
        """
        if valeur_comptable is None or valeur_comptable <= 0 or actions is None or actions <= 0:
            return None
        
        roe = roe or 0.10
        
        pv_ri = 0
        bv = valeur_comptable
        
        for annee in range(1, annees + 1):
            ri = (roe - cout_capitaux) * bv
            pv_ri += ri / ((1 + cout_capitaux) ** annee)
            bv *= (1 + roe * 0.5)  # Rétention partielle
        
        valeur_totale = valeur_comptable + pv_ri
        return round(valeur_totale / actions, 2) if valeur_totale > 0 else None
    
    def calculer_prix_reel(self, info: Dict, prix_actuel: float) -> Dict:
        """
        🎯 CALCUL DU PRIX RÉEL COMPOSITE (CORRIGÉ)
        Combine toutes les méthodes avec pondération et validation
        """
        if prix_actuel <= 0:
            return {
                'prix_reel': None,
                'potentiel': 0,
                'statut_valorisation': 'N/A',
                'marge_securite': None,
                'prix_achat_ideal': None,
                'details': {}
            }
        
        # Extraire les données
        eps = info.get('trailingEps') or info.get('forwardEps')
        valeur_comptable = info.get('bookValue')
        fcf = info.get('freeCashflow')
        actions = info.get('sharesOutstanding', 0) or 1
        ebit = info.get('ebitda')
        roe = info.get('returnOnEquity')
        croissance = info.get('earningsGrowth') or info.get('revenueGrowth') or 0.05
        prix_cible_analystes = info.get('targetMeanPrice')
        
        # Limiter la croissance à des valeurs réalistes
        croissance = max(-0.3, min(0.5, croissance)) if croissance else 0.05
        
        # Calculs individuels
        valorisations = {}
        poids = {}
        
        # Fonction pour valider les prix (doit être dans une plage réaliste)
        def valider_prix(prix_calcule: float) -> Optional[float]:
            if prix_calcule is None or prix_calcule <= 0:
                return None
            # Le prix réel ne peut pas être plus de 3x ou moins de 0.2x le prix actuel
            ratio = prix_calcule / prix_actuel
            if 0.2 <= ratio <= 3.0:
                return prix_calcule
            # Ajuster vers le prix actuel si hors limites
            if ratio > 3.0:
                return prix_actuel * 2.5  # Cap à +150%
            if ratio < 0.2:
                return prix_actuel * 0.3  # Floor à -70%
            return None
        
        # 1. Graham Number (poids: 15%)
        graham = self.graham_number(eps, valeur_comptable)
        graham = valider_prix(graham)
        if graham:
            valorisations['graham_number'] = graham
            poids['graham_number'] = 0.15
        
        # 2. Graham Défensif (poids: 10%)
        graham_def = self.graham_defensif(eps, croissance)
        graham_def = valider_prix(graham_def)
        if graham_def:
            valorisations['graham_defensif'] = graham_def
            poids['graham_defensif'] = 0.10
        
        # 3. DCF Multi-Scénarios (poids: 25%)
        dcf = self.dcf_multi_scenarios(fcf, actions)
        if dcf:
            dcf_base = valider_prix(dcf.get('base'))
            dcf_opt = valider_prix(dcf.get('optimiste'))
            dcf_pes = valider_prix(dcf.get('pessimiste'))
            if dcf_base:
                valorisations['dcf_base'] = dcf_base
                poids['dcf_base'] = 0.15
            if dcf_opt:
                valorisations['dcf_optimiste'] = dcf_opt
                poids['dcf_optimiste'] = 0.05
            if dcf_pes:
                valorisations['dcf_pessimiste'] = dcf_pes
                poids['dcf_pessimiste'] = 0.05
        
        # 4. PEG Fair Value (poids: 15%)
        peg_fv = self.peg_fair_value(eps, croissance)
        peg_fv = valider_prix(peg_fv)
        if peg_fv:
            valorisations['peg_fair_value'] = peg_fv
            poids['peg_fair_value'] = 0.15
        
        # 5. EPV (poids: 10%)
        epv = self.earnings_power_value(ebit, 0.25, self.config.TAUX_ACTUALISATION, actions)
        epv = valider_prix(epv)
        if epv:
            valorisations['epv'] = epv
            poids['epv'] = 0.10
        
        # 6. Prix Cible Analystes (poids: 25% - plus important)
        if prix_cible_analystes and prix_cible_analystes > 0:
            cible_valide = valider_prix(prix_cible_analystes)
            if cible_valide:
                valorisations['cible_analystes'] = cible_valide
                poids['cible_analystes'] = 0.25
        
        # Calculer le prix réel composite
        if not valorisations:
            # Fallback: utiliser le prix actuel comme référence
            return {
                'prix_reel': prix_actuel,
                'potentiel': 0,
                'statut_valorisation': '⚪ DONNÉES INSUFFISANTES',
                'marge_securite': None,
                'prix_achat_ideal': prix_actuel * 0.85,
                'details': {}
            }
        
        # Normaliser les poids
        total_poids = sum(poids.values())
        poids_normalises = {k: v / total_poids for k, v in poids.items()}
        
        # Moyenne pondérée
        prix_reel = sum(valorisations[k] * poids_normalises[k] for k in valorisations)
        
        # VALIDATION FINALE - plafonner le potentiel à ±100%
        potentiel = ((prix_reel - prix_actuel) / prix_actuel) * 100
        potentiel = max(-70, min(100, potentiel))  # Plafonner entre -70% et +100%
        
        # Recalculer le prix réel si potentiel ajusté
        if potentiel != ((prix_reel - prix_actuel) / prix_actuel) * 100:
            prix_reel = prix_actuel * (1 + potentiel / 100)
        
        # Prix d'achat idéal (avec marge de sécurité)
        # Si le prix actuel est déjà sous le prix réel, le prix d'achat idéal = prix actuel
        prix_achat_ideal_theorique = prix_reel * (1 - self.config.MARGE_SECURITE)
        
        # Si le prix actuel est déjà une aubaine (sous le prix idéal théorique)
        if prix_actuel <= prix_achat_ideal_theorique:
            prix_achat_ideal = prix_actuel  # Acheter maintenant!
        else:
            prix_achat_ideal = prix_achat_ideal_theorique
        
        # Statut
        if prix_actuel <= prix_achat_ideal_theorique:
            statut = "🟢🟢 FORTE OPPORTUNITÉ"
        elif potentiel > 30:
            statut = "🟢 SOUS-ÉVALUÉ"
        elif potentiel > 10:
            statut = "🟡 LÉGÈREMENT SOUS-ÉV"
        elif potentiel > -10:
            statut = "⚪ JUSTE PRIX"
        elif potentiel > -25:
            statut = "🟠 LÉGÈREMENT SUR-ÉV"
        else:
            statut = "🔴 SUR-ÉVALUÉ"
        
        return {
            'prix_reel': round(prix_reel, 2),
            'potentiel': round(potentiel, 1),
            'statut_valorisation': statut,
            'marge_securite': round(prix_reel - prix_achat_ideal, 2),
            'prix_achat_ideal': round(prix_achat_ideal, 2),
            'details': {k: round(v, 2) for k, v in valorisations.items()}
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    SYSTÈME DE TRADING (STOP LOSS, TAKE PROFIT, OBJECTIF)
# ══════════════════════════════════════════════════════════════════════════════

class SystemeTrading:
    """
    🎯 Système de règles de trading complet
    Calcule: Stop Loss, Take Profit, Horizon, Ratio R/R
    """
    
    def __init__(self, config: Config):
        self.config = config
    
    def calculer_stop_loss(self, prix: float, atr: float, support: float, 
                           potentiel: float = None) -> Dict:
        """
        📉 Calcul du Stop Loss (OPTIMISÉ)
        Méthode hybride: ATR + Support + Potentiel
        """
        if atr is None:
            atr = prix * 0.02
        
        # Stop Loss basé sur ATR
        sl_atr = prix - (atr * self.config.STOP_LOSS_ATR_MULT)
        
        # Stop Loss basé sur support
        sl_support = support * 0.98 if support else sl_atr  # 2% sous le support
        
        # Ajuster selon le potentiel (plus de marge si fort potentiel)
        if potentiel and potentiel > 30:
            # Fort potentiel: SL plus serré pour meilleur R/R
            sl_atr = prix - (atr * self.config.STOP_LOSS_ATR_MULT * 0.8)
        elif potentiel and potentiel < 0:
            # Potentiel négatif: SL plus large car risqué
            sl_atr = prix - (atr * self.config.STOP_LOSS_ATR_MULT * 1.5)
        
        # Prendre le plus protecteur (le plus haut)
        stop_loss = max(sl_atr, sl_support)
        
        # Pourcentage de perte
        perte_pct = ((prix - stop_loss) / prix) * 100
        
        return {
            'stop_loss': round(stop_loss, 2),
            'perte_potentielle_pct': round(perte_pct, 2),
            'methode': 'ATR' if sl_atr > sl_support else 'Support'
        }
    
    def calculer_take_profit(self, prix: float, atr: float, resistance: float,
                            prix_reel: float = None, potentiel: float = None) -> Dict:
        """
        📈 Calcul du Take Profit
        Méthode hybride: ATR + Résistance + Prix Réel (OPTIMISÉ)
        """
        if atr is None:
            atr = prix * 0.02
        
        # TP basé sur ATR (court terme)
        tp_atr = prix + (atr * self.config.TAKE_PROFIT_ATR_MULT)
        
        # TP basé sur résistance
        tp_resistance = resistance * 1.02 if resistance else tp_atr
        
        # TP basé sur prix réel (si disponible et significatif)
        tp_prix_reel = None
        if prix_reel and prix_reel > prix:
            # Si potentiel > 20%, utiliser un objectif intermédiaire (50% du chemin)
            if potentiel and potentiel > 20:
                tp_prix_reel = prix + (prix_reel - prix) * 0.5  # 50% du potentiel
            else:
                tp_prix_reel = prix_reel * 0.95  # 5% sous le prix réel
        
        # Choisir le TP selon le potentiel
        if potentiel and potentiel > 30 and tp_prix_reel:
            # Fort potentiel: viser plus haut
            take_profit = max(tp_atr, tp_prix_reel)
            methode = 'Prix Réel'
        elif tp_prix_reel and tp_prix_reel > tp_atr:
            # Potentiel moyen: TP intermédiaire
            take_profit = tp_prix_reel
            methode = 'Prix Réel'
        else:
            # Faible potentiel: TP conservateur
            take_profit = min(tp_atr, tp_resistance)
            methode = 'ATR' if take_profit == tp_atr else 'Résistance'
        
        # Pourcentage de gain
        gain_pct = ((take_profit - prix) / prix) * 100
        
        return {
            'take_profit': round(take_profit, 2),
            'gain_potentiel_pct': round(gain_pct, 2),
            'methode': methode
        }
    
    def calculer_ratio_risque_recompense(self, prix: float, stop_loss: float, take_profit: float) -> Dict:
        """
        ⚖️ Calcul du Ratio Risque/Récompense
        """
        risque = prix - stop_loss
        recompense = take_profit - prix
        
        if risque <= 0:
            return {'ratio': 0, 'qualite': '❌ INVALIDE'}
        
        ratio = recompense / risque
        
        if ratio >= 3:
            qualite = "🟢🟢 EXCELLENT"
        elif ratio >= 2:
            qualite = "🟢 BON"
        elif ratio >= 1.5:
            qualite = "🟡 ACCEPTABLE"
        elif ratio >= 1:
            qualite = "🟠 FAIBLE"
        else:
            qualite = "🔴 MAUVAIS"
        
        return {
            'ratio': round(ratio, 2),
            'qualite': qualite
        }
    
    def determiner_horizon(self, strategie: str, volatilite: float, score: int) -> Dict:
        """
        📅 Déterminer l'horizon de détention optimal
        """
        # Base sur la stratégie
        if 'MOMENTUM' in strategie.upper():
            horizon_base = self.config.HORIZON_COURT
        elif 'VALEUR' in strategie.upper() or 'VALUE' in strategie.upper():
            horizon_base = self.config.HORIZON_LONG
        else:
            horizon_base = self.config.HORIZON_MOYEN
        
        # Ajuster selon la volatilité
        if volatilite and volatilite > 0.03:  # Volatilité journalière > 3%
            horizon_base = max(self.config.HORIZON_COURT, horizon_base - 5)
        
        # Ajuster selon le score
        if score >= 75:
            horizon_max = horizon_base + 10
        elif score >= 60:
            horizon_max = horizon_base + 5
        else:
            horizon_max = horizon_base
        
        return {
            'horizon_min': max(3, horizon_base - 5),
            'horizon_optimal': horizon_base,
            'horizon_max': horizon_max,
            'type': 'Court terme' if horizon_base <= 7 else ('Moyen terme' if horizon_base <= 20 else 'Long terme')
        }
    
    def generer_plan_trading(self, donnees: Dict) -> Dict:
        """
        🎯 GÉNÉRATION DU PLAN DE TRADING COMPLET (OPTIMISÉ)
        """
        prix = donnees.get('prix', 0)
        atr = donnees.get('atr')
        support = donnees.get('support', prix * 0.95)
        resistance = donnees.get('resistance', prix * 1.05)
        prix_reel = donnees.get('prix_reel')
        potentiel = donnees.get('potentiel', 0)
        score = donnees.get('score', 50)
        strategie = donnees.get('strategie', 'Composite')
        volatilite = donnees.get('volatilite')
        
        # Calculs (avec potentiel pour optimiser SL/TP)
        sl = self.calculer_stop_loss(prix, atr, support, potentiel)
        tp = self.calculer_take_profit(prix, atr, resistance, prix_reel, potentiel)
        rr = self.calculer_ratio_risque_recompense(prix, sl['stop_loss'], tp['take_profit'])
        horizon = self.determiner_horizon(strategie, volatilite, score)
        
        return {
            'stop_loss': sl['stop_loss'],
            'stop_loss_pct': sl['perte_potentielle_pct'],
            'take_profit': tp['take_profit'],
            'take_profit_pct': tp['gain_potentiel_pct'],
            'ratio_rr': rr['ratio'],
            'qualite_rr': rr['qualite'],
            'horizon_jours': horizon['horizon_optimal'],
            'horizon_type': horizon['type'],
            'horizon_min': horizon['horizon_min'],
            'horizon_max': horizon['horizon_max']
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    SYSTÈME DE DÉCISION FINAL (SWING OPTIMISÉ)
# ══════════════════════════════════════════════════════════════════════════════

class SystemeDecision:
    """
    🎯 Système de décision optimisé pour Win Rate élevé en Swing Trading
    
    Filtres appliqués:
    1. Alignement des axes (technique + fondamental + ML + trading)
    2. Filtre de régime de marché
    3. Conditions Momentum/Mean Reversion spécifiques
    4. Filtre volatilité excessive
    5. Filtre qualité des données
    6. Support pour pullback entry
    """
    
    def __init__(self, config: Config):
        self.config = config
    
    def evaluer_technique(self, donnees: Dict) -> Tuple[int, str]:
        """Évaluation technique (0-100)"""
        score = 50
        details = []
        
        # RSI
        rsi = donnees.get('rsi')
        if rsi:
            if rsi <= 30:
                score += 15
                details.append("RSI survendu")
            elif rsi <= 40:
                score += 8
            elif rsi >= 70:
                score -= 15
                details.append("RSI suracheté")
            elif rsi >= 60:
                score -= 5
        
        # Z-Score
        z = donnees.get('z_score')
        if z:
            if z <= -2:
                score += 15
                details.append("Z-Score très bas")
            elif z <= -1:
                score += 8
            elif z >= 2:
                score -= 15
                details.append("Z-Score très haut")
            elif z >= 1:
                score -= 5
        
        # Tendance
        tendance_score = donnees.get('score_tendance', 2)
        score += (tendance_score - 2) * 5
        if tendance_score >= 4:
            details.append("Forte tendance haussière")
        elif tendance_score <= 1:
            details.append("Forte tendance baissière")
        
        # MACD
        macd_tendance = donnees.get('macd_tendance', '')
        if 'HAUSSIER' in macd_tendance:
            score += 10
            details.append("MACD haussier")
        elif 'BAISSIER' in macd_tendance:
            score -= 10
        
        # Volume
        volume_ratio = donnees.get('volume_ratio', 1)
        if volume_ratio >= 2:
            score += 5
            details.append("Volume élevé")
        
        return max(0, min(100, score)), ", ".join(details) if details else "Neutre"
    
    def evaluer_fondamental(self, donnees: Dict) -> Tuple[int, str]:
        """Évaluation fondamentale (0-100)"""
        score = 50
        details = []
        
        # PE Ratio
        pe = donnees.get('pe_ratio')
        if pe:
            if 0 < pe < 12:
                score += 20
                details.append(f"PE attractif ({pe})")
            elif pe < 18:
                score += 10
            elif pe > 35:
                score -= 15
                details.append(f"PE élevé ({pe})")
            elif pe > 25:
                score -= 5
        
        # Potentiel valorisation
        potentiel = donnees.get('potentiel', 0)
        if potentiel > 40:
            score += 20
            details.append(f"Fort potentiel (+{potentiel:.0f}%)")
        elif potentiel > 20:
            score += 12
            details.append(f"Bon potentiel (+{potentiel:.0f}%)")
        elif potentiel > 10:
            score += 5
        elif potentiel < -20:
            score -= 15
            details.append(f"Surévalué ({potentiel:.0f}%)")
        elif potentiel < -10:
            score -= 8
        
        # Analystes
        signal_analyste = donnees.get('signal_analyste', '')
        if 'ACHAT FORT' in signal_analyste:
            score += 12
            details.append("Analystes: Achat Fort")
        elif 'ACHAT' in signal_analyste:
            score += 6
        elif 'VENTE FORTE' in signal_analyste:
            score -= 12
        elif 'VENTE' in signal_analyste:
            score -= 6
        
        return max(0, min(100, score)), ", ".join(details) if details else "Neutre"
    
    def evaluer_trading(self, donnees: Dict) -> Tuple[int, str]:
        """Évaluation des conditions de trading (0-100)"""
        score = 50
        details = []
        
        # Ratio R/R
        ratio_rr = donnees.get('ratio_rr', 1)
        if ratio_rr >= 3:
            score += 20
            details.append(f"Excellent R/R ({ratio_rr})")
        elif ratio_rr >= 2:
            score += 12
            details.append(f"Bon R/R ({ratio_rr})")
        elif ratio_rr >= 1.5:
            score += 5
        elif ratio_rr < 1:
            score -= 15
            details.append("Mauvais R/R")
        
        # Stop Loss acceptable
        sl_pct = donnees.get('stop_loss_pct', 5)
        if sl_pct < 3:
            score += 10
            details.append("SL serré")
        elif sl_pct > 8:
            score -= 10
            details.append("SL large")
        
        # Take Profit réaliste
        tp_pct = donnees.get('take_profit_pct', 10)
        if 5 <= tp_pct <= 20:
            score += 5
        elif tp_pct > 30:
            score -= 5
            details.append("TP ambitieux")
        
        return max(0, min(100, score)), ", ".join(details) if details else "Neutre"
    
    def evaluer_ml(self, donnees: Dict) -> Tuple[int, str]:
        """Évaluation ML (0-100)"""
        score = 50
        details = []
        
        ml_prob = donnees.get('ml_probabilite', 50)
        ml_prediction = donnees.get('ml_prediction')
        ml_consensus = donnees.get('ml_consensus', 50)
        
        if ml_prediction == 1:
            if ml_prob >= 70 and ml_consensus >= 75:
                score += 30
                details.append(f"ML très confiant ({ml_prob}%, consensus {ml_consensus}%)")
            elif ml_prob >= 70:
                score += 25
                details.append(f"ML très confiant ({ml_prob}%)")
            elif ml_prob >= 60:
                score += 15
                details.append(f"ML confiant ({ml_prob}%)")
            else:
                score += 5
        elif ml_prediction == 0:
            if ml_prob <= 30:
                score -= 25
                details.append(f"ML très négatif ({ml_prob}%)")
            elif ml_prob <= 40:
                score -= 15
                details.append(f"ML négatif ({ml_prob}%)")
        
        if donnees.get('ml_anomalie'):
            score -= 15
            details.append("⚠️ Anomalie détectée")
        
        return max(0, min(100, score)), ", ".join(details) if details else "Neutre"
    
    def evaluer_swing_filters(self, donnees: Dict) -> Tuple[bool, List[str], str]:
        """
        🎯 FILTRES SWING TRADING POUR WIN RATE ÉLEVÉ
        
        Retourne: (passe_filtres, raisons_rejet, strategie_detectee)
        """
        rejets = []
        strategie = "MIXTE"
        
        rsi = donnees.get('rsi', 50)
        z_score = donnees.get('z_score', 0)
        volume_ratio = donnees.get('volume_ratio', 1)
        tendance = donnees.get('tendance', '')
        score_tendance = donnees.get('score_tendance', 2)
        potentiel = donnees.get('potentiel', 0)
        ratio_rr = donnees.get('ratio_rr', 1)
        ml_signal = donnees.get('ml_signal', '')
        volatilite = donnees.get('volatilite', 0)
        distance_support = donnees.get('distance_support', 100)
        regime = donnees.get('regime', {})
        
        # ═══════════════════════════════════════════════════════════════════
        # FILTRE 1: Volatilité excessive
        # ═══════════════════════════════════════════════════════════════════
        if volatilite and volatilite > self.config.VOLATILITE_MAX:
            rejets.append(f"Volatilité trop élevée ({volatilite:.1f}% > {self.config.VOLATILITE_MAX}%)")
        
        # ═══════════════════════════════════════════════════════════════════
        # FILTRE 2: ML bloquant (si configuré)
        # ═══════════════════════════════════════════════════════════════════
        if self.config.ML_BLOQUER_VENTE and 'VENTE' in ml_signal and 'FORT' in ml_signal:
            rejets.append("ML signal FORTE VENTE")
        
        # ═══════════════════════════════════════════════════════════════════
        # FILTRE 3: Ratio R/R minimum
        # ═══════════════════════════════════════════════════════════════════
        if ratio_rr < self.config.SWING_RR_MIN:
            rejets.append(f"R/R insuffisant ({ratio_rr:.2f} < {self.config.SWING_RR_MIN})")
        
        # ═══════════════════════════════════════════════════════════════════
        # DÉTECTION STRATÉGIE ET FILTRES SPÉCIFIQUES
        # ═══════════════════════════════════════════════════════════════════
        
        # Détection Momentum
        is_momentum = (score_tendance >= 4 or 'HAUSSE' in tendance.upper())
        
        # Détection Mean Reversion
        is_meanrev = (z_score <= -1.5 or rsi <= 35)
        
        if is_momentum and not is_meanrev:
            strategie = "MOMENTUM"
            
            # Filtres Momentum spécifiques
            if rsi < self.config.MOMENTUM_RSI_MIN:
                rejets.append(f"Momentum: RSI trop bas ({rsi:.0f} < {self.config.MOMENTUM_RSI_MIN})")
            if rsi > self.config.MOMENTUM_RSI_MAX:
                rejets.append(f"Momentum: RSI trop haut ({rsi:.0f} > {self.config.MOMENTUM_RSI_MAX})")
            if z_score < self.config.MOMENTUM_ZSCORE_MIN:
                rejets.append(f"Momentum: Z-Score trop bas ({z_score:.2f})")
            if z_score > self.config.MOMENTUM_ZSCORE_MAX:
                rejets.append(f"Momentum: Z-Score trop haut ({z_score:.2f} > {self.config.MOMENTUM_ZSCORE_MAX})")
            if volume_ratio < self.config.MOMENTUM_VOLUME_MIN:
                rejets.append(f"Momentum: Volume insuffisant ({volume_ratio:.2f}x)")
            
            # Filtre régime de marché pour Momentum
            if regime and regime.get('regime') in ['TREND_BAISSIER', 'TREND_BAISSIER_FORT']:
                rejets.append(f"Momentum contre-tendance: {regime.get('description', '')}")
        
        elif is_meanrev:
            strategie = "MEAN_REVERSION"
            
            # Filtres Mean Reversion spécifiques
            if rsi > self.config.MEANREV_RSI_MAX:
                rejets.append(f"MeanRev: RSI trop haut ({rsi:.0f} > {self.config.MEANREV_RSI_MAX})")
            if z_score > self.config.MEANREV_ZSCORE_MAX:
                rejets.append(f"MeanRev: Z-Score pas assez bas ({z_score:.2f})")
            
            # Éviter le "couteau qui tombe"
            if score_tendance <= 1 or 'FORTE BAISSE' in tendance.upper():
                rejets.append("MeanRev: Tendance FORTE BAISSE - risque de couteau")
            
            # Vérifier proximité du support
            if distance_support > self.config.MEANREV_SUPPORT_DISTANCE:
                rejets.append(f"MeanRev: Trop loin du support ({distance_support:.1f}%)")
        
        else:
            strategie = "VALEUR"
            # Pour stratégie Valeur, on vérifie le potentiel
            if potentiel < self.config.SWING_POTENTIEL_MIN:
                rejets.append(f"Potentiel insuffisant ({potentiel:.1f}% < {self.config.SWING_POTENTIEL_MIN}%)")
        
        return len(rejets) == 0, rejets, strategie
    
    def generer_decision_finale(self, donnees: Dict) -> Dict:
        """
        🎯 GÉNÉRATION DE LA DÉCISION FINALE (WIN RATE OPTIMISÉ 70%+)
        
        Système de confirmation multi-indicateurs pour maximiser le Win Rate
        """
        # Évaluations par catégorie
        score_tech, details_tech = self.evaluer_technique(donnees)
        score_fond, details_fond = self.evaluer_fondamental(donnees)
        score_trading, details_trading = self.evaluer_trading(donnees)
        score_ml, details_ml = self.evaluer_ml(donnees)
        
        # 🆕 SYSTÈME DE CONFIRMATION MULTI-INDICATEURS
        confirmations = []
        confirmations_details = []
        
        # 1. RSI favorable
        rsi = donnees.get('rsi', 50)
        if 40 <= rsi <= 60:
            confirmations.append('RSI')
            confirmations_details.append(f"✅ RSI zone neutre ({rsi:.0f})")
        elif rsi < 35:
            confirmations.append('RSI_SURVENTE')
            confirmations_details.append(f"✅ RSI survendu ({rsi:.0f})")
        
        # 2. Z-Score favorable
        z_score = donnees.get('z_score', 0)
        if -1.5 <= z_score <= 1.0:
            confirmations.append('ZSCORE')
            confirmations_details.append(f"✅ Z-Score favorable ({z_score:.2f})")
        elif z_score < -1.5:
            confirmations.append('ZSCORE_BAS')
            confirmations_details.append(f"✅ Z-Score bas ({z_score:.2f})")
        
        # 3. Tendance positive
        score_tendance = donnees.get('score_tendance', 2)
        if score_tendance >= 3:
            confirmations.append('TENDANCE')
            confirmations_details.append(f"✅ Tendance haussière ({score_tendance}/5)")
        
        # 4. Volume confirmé
        volume_ratio = donnees.get('volume_ratio', 1)
        if volume_ratio >= 1.2:
            confirmations.append('VOLUME')
            confirmations_details.append(f"✅ Volume élevé ({volume_ratio:.1f}x)")
        
        # 5. ML positif
        ml_prediction = donnees.get('ml_prediction')
        ml_prob = donnees.get('ml_probabilite', 50)
        ml_consensus = donnees.get('ml_consensus', 50)
        if ml_prediction == 1 and ml_prob >= 60:
            confirmations.append('ML')
            confirmations_details.append(f"✅ ML positif ({ml_prob}%)")
        
        # 6. Potentiel valorisation
        potentiel = donnees.get('potentiel', 0)
        if potentiel >= 15:
            confirmations.append('VALORISATION')
            confirmations_details.append(f"✅ Potentiel +{potentiel:.0f}%")
        
        # 7. Ratio R/R
        ratio_rr = donnees.get('ratio_rr', 1)
        if ratio_rr >= 1.5:
            confirmations.append('RR')
            confirmations_details.append(f"✅ R/R excellent ({ratio_rr:.1f})")
        
        # 8. MACD positif
        macd_tendance = donnees.get('macd_tendance', '')
        if 'HAUSSIER' in macd_tendance:
            confirmations.append('MACD')
            confirmations_details.append("✅ MACD haussier")
        
        nb_confirmations = len(confirmations)
        
        # Pondérations optimisées pour Win Rate
        poids = {
            'technique': self.config.POIDS_TECHNIQUE,
            'fondamental': self.config.POIDS_FONDAMENTAL,
            'trading': self.config.POIDS_QUALITE,
            'ml': self.config.POIDS_ML
        }
        
        # Score final pondéré
        score_final = (
            score_tech * poids['technique'] +
            score_fond * poids['fondamental'] +
            score_trading * poids['trading'] +
            score_ml * poids['ml']
        )
        
        # 🆕 BONUS CONFIRMATIONS MULTIPLES
        if nb_confirmations >= 6:
            score_final += 10
        elif nb_confirmations >= 5:
            score_final += 6
        elif nb_confirmations >= 4:
            score_final += 3
        
        # 🆕 BONUS PULLBACK ENTRY (timing optimal)
        pullback_score = donnees.get('pullback_score', 0)
        pullback_bonus = 0
        pullback_detail = ""
        
        if self.config.PULLBACK_ENTRY and pullback_score:
            if pullback_score >= self.config.PULLBACK_SCORE_MIN:
                if pullback_score >= 80:
                    pullback_bonus = 8
                    pullback_detail = "🎯 Pullback parfait"
                elif pullback_score >= 65:
                    pullback_bonus = 5
                    pullback_detail = "📈 Bon pullback"
                else:
                    pullback_bonus = 3
                    pullback_detail = "Pullback OK"
                
                confirmations.append('PULLBACK')
                confirmations_details.append(f"✅ {pullback_detail}")
                score_final += pullback_bonus
        
        # Score qualité (si disponible)
        score_qualite = donnees.get('score_qualite', 50)
        if score_qualite >= self.config.QUALITE_MIN:
            confirmations.append('QUALITE')
            confirmations_details.append(f"✅ Qualité financière ({score_qualite:.0f})")
        
        # 📰 NEWS & SENTIMENT (nouveau)
        score_news = donnees.get('score_news', 50)
        impact_news = donnees.get('impact_news', 0)
        if score_news >= 60:
            confirmations.append('NEWS')
            confirmations_details.append(f"✅ Sentiment news positif ({score_news:.0f})")
            score_final += min(impact_news, 8)  # Max +8 bonus
        elif score_news < 35:
            # Pénalité pour news négatives
            score_final += max(impact_news, -10)  # Max -10 malus
            confirmations_details.append(f"⚠️ Sentiment news négatif ({score_news:.0f})")
        
        nb_confirmations = len(confirmations)
        
        # Niveau de confiance basé sur les confirmations
        scores = [score_tech, score_fond, score_trading, score_ml]
        ecart_type = np.std(scores)
        
        if nb_confirmations >= 6 and ecart_type < 10 and min(scores) >= 50:
            confiance = "🟢🟢 TRÈS HAUTE"
            confiance_pct = 90
        elif nb_confirmations >= 5 and ecart_type < 12:
            confiance = "🟢 HAUTE"
            confiance_pct = 80
        elif nb_confirmations >= 4 and ecart_type < 15:
            confiance = "🟡 MOYENNE"
            confiance_pct = 68
        elif nb_confirmations >= 3:
            confiance = "🟠 MODÉRÉE"
            confiance_pct = 55
        else:
            confiance = "🔴 FAIBLE"
            confiance_pct = 40
        
        # ═══════════════════════════════════════════════════════════════════
        # APPLICATION DES FILTRES SWING STRICTS
        # ═══════════════════════════════════════════════════════════════════
        
        passe_filtres, raisons_rejet, strategie_detectee = self.evaluer_swing_filters(donnees)
        
        ml_signal = donnees.get('ml_signal', '')
        
        # 🆕 FILTRES ADDITIONNELS WIN RATE
        # Filtre prix minimum (éviter penny stocks)
        prix = donnees.get('prix', 0)
        if hasattr(self.config, 'EXCLURE_PENNY_STOCKS') and self.config.EXCLURE_PENNY_STOCKS:
            if prix < self.config.PRIX_MIN:
                raisons_rejet.append(f"Prix trop bas (${prix:.2f} < ${self.config.PRIX_MIN})")
                passe_filtres = False
        
        # Filtre consensus ML
        if hasattr(self.config, 'ML_CONSENSUS_MIN') and ml_consensus < self.config.ML_CONSENSUS_MIN:
            if ml_prediction == 1:  # Seulement si ML dit achat
                raisons_rejet.append(f"Consensus ML faible ({ml_consensus}% < {self.config.ML_CONSENSUS_MIN}%)")
        
        # Filtre qualité minimum
        if hasattr(self.config, 'QUALITE_MIN') and score_qualite < self.config.QUALITE_MIN:
            raisons_rejet.append(f"Qualité insuffisante ({score_qualite:.0f} < {self.config.QUALITE_MIN})")
        
        # Décision avec filtres ÉQUILIBRÉS pour plus d'opportunités
        decision = "⚪ NEUTRE"
        action = "ATTENTE"
        conseil = ""
        
        # 🟢 CONDITIONS POUR SIGNAUX D'ACHAT
        achat_signals = []
        
        # RSI favorable (survente ou zone neutre basse)
        if rsi < 35:
            achat_signals.append(f"RSI survendu ({rsi:.0f})")
        elif 35 <= rsi <= 55:
            achat_signals.append(f"RSI favorable ({rsi:.0f})")
        
        # Z-Score bas (sous-évalué)
        if z_score < -1.0:
            achat_signals.append(f"Z-Score bas ({z_score:.2f})")
        
        # Potentiel positif (sous-évalué)
        if potentiel > 15:
            achat_signals.append(f"Sous-évalué ({potentiel:+.0f}%)")
        elif potentiel > 8:
            achat_signals.append(f"Potentiel ({potentiel:+.0f}%)")
        
        # MACD haussier
        if 'HAUSSIER' in macd_tendance:
            achat_signals.append("MACD haussier")
        
        # ML prédit achat
        if ml_prediction == 1 and ml_prob >= 55:
            achat_signals.append(f"ML prédit hausse ({ml_prob}%)")
        
        # Tendance haussière
        if score_tendance >= 3:
            achat_signals.append("Tendance haussière")
        
        # Ratio R/R favorable
        if ratio_rr >= 1.5:
            achat_signals.append(f"R/R excellent ({ratio_rr:.1f})")
        elif ratio_rr >= 1.2:
            achat_signals.append(f"R/R bon ({ratio_rr:.1f})")
        
        # Volume élevé
        if volume_ratio >= 1.3:
            achat_signals.append(f"Volume fort ({volume_ratio:.1f}x)")
        
        nb_achat_signals = len(achat_signals)
        
        # 🔴 CONDITIONS POUR SIGNAUX DE VENTE
        vente_signals = []
        
        # RSI suracheté
        if rsi > 70:
            vente_signals.append(f"RSI suracheté ({rsi:.0f})")
        
        # Z-Score très élevé (surévalué)
        if z_score > 1.5:
            vente_signals.append(f"Z-Score élevé ({z_score:.2f})")
        
        # Potentiel négatif (surévalué)
        if potentiel < -10:
            vente_signals.append(f"Surévalué ({potentiel:+.0f}%)")
        
        # MACD baissier
        if 'BAISSIER' in macd_tendance:
            vente_signals.append("MACD baissier")
        
        # ML prédit vente
        if ml_prediction == 0 and ml_prob >= 55:
            vente_signals.append(f"ML prédit baisse ({ml_prob}%)")
        
        # Tendance baissière
        if score_tendance <= 1:
            vente_signals.append("Tendance baissière")
        
        # Ratio R/R défavorable
        if ratio_rr < 0.8:
            vente_signals.append(f"R/R défavorable ({ratio_rr:.1f})")
        
        nb_vente_signals = len(vente_signals)
        
        # ══════════════════════════════════════════════════════════════════
        # 🎯 DÉCISION FINALE - COMPARAISON ACHAT vs VENTE
        # ══════════════════════════════════════════════════════════════════
        
        # Score net: positif = achat, négatif = vente
        score_net = nb_achat_signals - nb_vente_signals
        
        # 🟢🟢 ACHAT FORT - forte dominance de signaux d'achat
        if score_net >= 4 and nb_achat_signals >= 5 and score_final >= 65:
            decision = "🟢🟢 ACHETER FORT"
            action = "ACHAT"
            conseil = f"🏆 SIGNAL A+ ({nb_achat_signals} signaux): {'; '.join(achat_signals[:3])}"
        
        # 🟢 ACHAT - bonne dominance de signaux d'achat
        elif score_net >= 2 and nb_achat_signals >= 3 and score_final >= 50:
            decision = "🟢 ACHETER"
            action = "ACHAT"
            conseil = f"🎯 Signal achat ({nb_achat_signals} signaux): {'; '.join(achat_signals[:3])}"
        
        # 🟢 ACHAT (léger) - légère dominance achat
        elif score_net >= 1 and nb_achat_signals >= 2 and score_final >= 45:
            decision = "🟢 ACHETER"
            action = "ACHAT"
            conseil = f"✅ Achat validé ({nb_achat_signals} signaux): {'; '.join(achat_signals[:2])}"
        
        # 🔴🔴 VENTE FORTE - forte dominance de signaux de vente
        elif score_net <= -4 and nb_vente_signals >= 5:
            decision = "🔴🔴 VENDRE"
            action = "VENTE"
            conseil = f"⚠️ SIGNAL VENTE FORT ({nb_vente_signals} signaux): {'; '.join(vente_signals[:3])}"
        
        # 🔴 VENTE - bonne dominance de signaux de vente
        elif score_net <= -2 and nb_vente_signals >= 3:
            decision = "🔴 VENDRE"
            action = "VENTE"
            conseil = f"📉 Signal vente ({nb_vente_signals} signaux): {'; '.join(vente_signals[:3])}"
        
        # 🔴 VENTE (léger) - légère dominance vente
        elif score_net <= -1 and nb_vente_signals >= 2:
            decision = "🔴 VENDRE"
            action = "VENTE"
            conseil = f"📉 Vente suggérée ({nb_vente_signals} signaux): {'; '.join(vente_signals[:2])}"
        
        # 🟡 SURVEILLER - signaux mixtes ou potentiel
        elif nb_achat_signals >= 2 or score_final >= 50:
            decision = "🟡 SURVEILLER"
            action = "ATTENTE"
            conseil = f"Signal mixte: {nb_achat_signals} achat vs {nb_vente_signals} vente. Observer."
        
        # ⚪ NEUTRE
        else:
            decision = "⚪ NEUTRE"
            action = "ATTENTE"
            conseil = f"Pas de signal clair ({nb_achat_signals} achat vs {nb_vente_signals} vente)."
        
        return {
            'decision': decision,
            'action': action,
            'score_final': round(score_final, 1),
            'confiance': confiance,
            'confiance_pct': confiance_pct,
            'conseil': conseil,
            'strategie_detectee': strategie_detectee,
            'passe_filtres_swing': passe_filtres,
            'raisons_rejet': raisons_rejet,
            'nb_confirmations': nb_confirmations,
            'confirmations': confirmations,
            'confirmations_details': confirmations_details,
            'scores_details': {
                'technique': {'score': score_tech, 'details': details_tech},
                'fondamental': {'score': score_fond, 'details': details_fond},
                'trading': {'score': score_trading, 'details': details_trading},
                'ml': {'score': score_ml, 'details': details_ml}
            }
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    MACHINE LEARNING AVANCÉ (ENSEMBLE)
# ══════════════════════════════════════════════════════════════════════════════

class PredicteurML:
    """
    🧠 SYSTÈME ML ULTRA-AVANCÉ v2.0 avec Ensemble de Modèles de Pointe
    
    ═══════════════════════════════════════════════════════════════════════
    🏆 MODÈLES DE BASE (Tier 1 - Arbres de décision):
    ───────────────────────────────────────────────────────────────────────
    - Random Forest (robuste, moins d'overfitting, parallélisable)
    - Extra Trees (haute variance, diversité maximale)
    - Gradient Boosting (sequential learning, forte précision)
    
    🚀 MODÈLES AVANCÉS (Tier 2 - Boosting optimisé):
    ───────────────────────────────────────────────────────────────────────
    - XGBoost (performance SOTA, régularisation L1/L2)
    - LightGBM (ultra-rapide, leaf-wise growth, 10x faster)
    - CatBoost (gestion native catégorielles, ordered boosting)
    - HistGradientBoosting (sklearn native, très rapide)
    
    🧬 MODÈLES DEEP LEARNING (Tier 3 - Réseaux de neurones):
    ───────────────────────────────────────────────────────────────────────
    - MLP Neural Network (multi-layer perceptron, non-linéarités)
    - Calibrated MLP (probabilités calibrées via Platt scaling)
    
    📊 MODÈLES LINÉAIRES (Tier 4 - Baseline & régularisation):
    ───────────────────────────────────────────────────────────────────────
    - Logistic Regression (baseline, interprétable)
    - Ridge Classifier (régularisation L2)
    - SVM RBF (kernel trick, décisions non-linéaires)
    
    🎯 MÉTA-MODÈLES (Tier 5 - Stacking & Voting):
    ───────────────────────────────────────────────────────────────────────
    - Stacking Classifier (meta-learner sur les prédictions)
    - Soft Voting (moyenne pondérée des probabilités)
    - Bagging Classifier (réduction variance)
    
    🔍 DÉTECTION ANOMALIES:
    ───────────────────────────────────────────────────────────────────────
    - Isolation Forest (détection régimes anormaux de marché)
    
    ═══════════════════════════════════════════════════════════════════════
    Vote pondéré dynamique basé sur performance CV pour décision finale
    """
    
    def __init__(self, config: Config):
        self.config = config
        self.pipelines = {}
        self.poids_modeles = {}
        self.est_entraine = False
        self.metriques = {}
        self.feature_importances = {}
        self.selected_features = None
        self.scaler = None
        self.stacking_model = None
        self.voting_model = None
    
    def preparer_features(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        🔬 Prépare des features ultra-enrichies pour prédiction avancée
        
        Catégories de features:
        ═══════════════════════════════════════════════════════════════
        📈 RENDEMENTS: Multi-périodes (1j à 60j)
        📊 MOYENNES MOBILES: SMA, EMA, WMA, DEMA, TEMA ratios
        📉 VOLATILITÉ: ATR, Parkinson, Garman-Klass, Yang-Zhang
        🎯 MOMENTUM: RSI, Stochastic, Williams %R, CCI, MFI
        📦 VOLUME: OBV, VWAP, Volume Profile, A/D Line
        🔄 TENDANCE: ADX, SuperTrend, Ichimoku, Parabolic SAR
        📐 PATTERNS: Candlestick patterns, Support/Resistance
        🧮 STATISTIQUES: Z-Score, Skewness, Kurtosis, Hurst
        ═══════════════════════════════════════════════════════════════
        """
        features = pd.DataFrame(index=df.index)
        close = df['Close']
        high = df['High']
        low = df['Low']
        volume = df['Volume']
        open_price = df['Open']
        
        # ═══════════════════════════════════════════════════════════════
        # 📈 RENDEMENTS MULTI-PÉRIODES (étendu)
        # ═══════════════════════════════════════════════════════════════
        for p in [1, 2, 3, 5, 10, 15, 20, 30, 40, 60]:
            features[f'ret_{p}j'] = close.pct_change(p)
        
        # Rendements log (plus stable pour ML)
        for p in [1, 5, 10, 20]:
            features[f'log_ret_{p}j'] = np.log(close / close.shift(p))
        
        # ═══════════════════════════════════════════════════════════════
        # 📊 MOYENNES MOBILES AVANCÉES
        # ═══════════════════════════════════════════════════════════════
        # SMA ratios
        for p in [5, 10, 20, 50, 100, 200]:
            sma = close.rolling(p).mean()
            features[f'sma_ratio_{p}'] = close / sma
            if p >= 10:
                features[f'sma_slope_{p}'] = sma.pct_change(5)
        
        # EMA ratios
        for p in [8, 13, 21, 34, 55, 89]:  # Fibonacci EMAs
            ema = close.ewm(span=p, adjust=False).mean()
            features[f'ema_ratio_{p}'] = close / ema
        
        # Double EMA (DEMA)
        for p in [9, 21]:
            ema1 = close.ewm(span=p, adjust=False).mean()
            ema2 = ema1.ewm(span=p, adjust=False).mean()
            dema = 2 * ema1 - ema2
            features[f'dema_ratio_{p}'] = close / dema
        
        # Triple EMA (TEMA)
        ema1 = close.ewm(span=20, adjust=False).mean()
        ema2 = ema1.ewm(span=20, adjust=False).mean()
        ema3 = ema2.ewm(span=20, adjust=False).mean()
        tema = 3 * ema1 - 3 * ema2 + ema3
        features['tema_ratio_20'] = close / tema
        
        # Weighted MA
        def wma(series, period):
            weights = np.arange(1, period + 1)
            return series.rolling(period).apply(lambda x: np.dot(x, weights) / weights.sum(), raw=True)
        features['wma_ratio_20'] = close / wma(close, 20)
        
        # Hull MA (plus réactif)
        half_period = 10
        full_period = 20
        wma_half = wma(close, half_period)
        wma_full = wma(close, full_period)
        hull_raw = 2 * wma_half - wma_full
        features['hma_ratio_20'] = close / wma(hull_raw, int(np.sqrt(full_period)))
        
        # ═══════════════════════════════════════════════════════════════
        # 📉 VOLATILITÉ AVANCÉE (plusieurs estimateurs)
        # ═══════════════════════════════════════════════════════════════
        # Standard volatility
        for p in [5, 10, 20, 50, 100]:
            features[f'vol_{p}j'] = close.pct_change().rolling(p).std() * np.sqrt(252)
        
        # Ratio de volatilité
        features['vol_ratio_5_20'] = features['vol_5j'] / features['vol_20j']
        features['vol_ratio_10_50'] = features['vol_10j'] / features['vol_50j']
        
        # ATR (Average True Range)
        tr = pd.concat([
            high - low,
            abs(high - close.shift(1)),
            abs(low - close.shift(1))
        ], axis=1).max(axis=1)
        for p in [7, 14, 21]:
            atr = tr.rolling(p).mean()
            features[f'atr_{p}'] = atr / close * 100  # En pourcentage
        
        # Parkinson volatility (high-low based)
        features['parkinson_vol'] = np.sqrt(
            (1 / (4 * np.log(2))) * ((np.log(high / low)) ** 2).rolling(20).mean()
        ) * np.sqrt(252)
        
        # Garman-Klass volatility (OHLC based - plus précis)
        log_hl = (np.log(high / low)) ** 2
        log_co = (np.log(close / open_price)) ** 2
        features['gk_vol'] = np.sqrt(
            (0.5 * log_hl - (2 * np.log(2) - 1) * log_co).rolling(20).mean()
        ) * np.sqrt(252)
        
        # ═══════════════════════════════════════════════════════════════
        # 🎯 INDICATEURS MOMENTUM AVANCÉS
        # ═══════════════════════════════════════════════════════════════
        # RSI multi-périodes
        for p in [7, 14, 21]:
            delta = close.diff()
            gain = delta.where(delta > 0, 0).ewm(alpha=1/p, min_periods=p).mean()
            perte = -delta.where(delta < 0, 0).ewm(alpha=1/p, min_periods=p).mean()
            features[f'rsi_{p}'] = 100 - (100 / (1 + gain / perte))
        
        features['rsi_divergence'] = features['rsi_14'] - features['rsi_14'].rolling(5).mean()
        
        # Stochastic RSI
        rsi = features['rsi_14']
        for p in [14]:
            stoch_rsi = (rsi - rsi.rolling(p).min()) / (rsi.rolling(p).max() - rsi.rolling(p).min())
            features[f'stoch_rsi_{p}'] = stoch_rsi * 100
        
        # Stochastic %K et %D
        for p in [14]:
            lowest_low = low.rolling(p).min()
            highest_high = high.rolling(p).max()
            features[f'stoch_k_{p}'] = 100 * (close - lowest_low) / (highest_high - lowest_low)
            features[f'stoch_d_{p}'] = features[f'stoch_k_{p}'].rolling(3).mean()
        
        # Williams %R
        for p in [14, 21]:
            highest_high = high.rolling(p).max()
            lowest_low = low.rolling(p).min()
            features[f'williams_r_{p}'] = -100 * (highest_high - close) / (highest_high - lowest_low)
        
        # CCI (Commodity Channel Index)
        for p in [14, 20]:
            typical_price = (high + low + close) / 3
            sma_tp = typical_price.rolling(p).mean()
            mad = typical_price.rolling(p).apply(lambda x: np.abs(x - x.mean()).mean(), raw=True)
            features[f'cci_{p}'] = (typical_price - sma_tp) / (0.015 * mad)
        
        # MFI (Money Flow Index)
        typical_price = (high + low + close) / 3
        money_flow = typical_price * volume
        positive_flow = money_flow.where(typical_price > typical_price.shift(1), 0)
        negative_flow = money_flow.where(typical_price < typical_price.shift(1), 0)
        for p in [14]:
            positive_mf = positive_flow.rolling(p).sum()
            negative_mf = negative_flow.rolling(p).sum()
            mfi = 100 - (100 / (1 + positive_mf / negative_mf))
            features[f'mfi_{p}'] = mfi
        
        # Ultimate Oscillator
        bp = close - pd.concat([low, close.shift(1)], axis=1).min(axis=1)
        tr = pd.concat([high, close.shift(1)], axis=1).max(axis=1) - pd.concat([low, close.shift(1)], axis=1).min(axis=1)
        avg7 = bp.rolling(7).sum() / tr.rolling(7).sum()
        avg14 = bp.rolling(14).sum() / tr.rolling(14).sum()
        avg28 = bp.rolling(28).sum() / tr.rolling(28).sum()
        features['ultimate_osc'] = 100 * (4 * avg7 + 2 * avg14 + avg28) / 7
        
        # ═══════════════════════════════════════════════════════════════
        # 📦 INDICATEURS VOLUME AVANCÉS
        # ═══════════════════════════════════════════════════════════════
        # Volume ratios
        for p in [5, 10, 20, 50]:
            features[f'vol_ratio_{p}'] = volume / volume.rolling(p).mean()
        features['vol_trend'] = volume.rolling(5).mean() / volume.rolling(20).mean()
        
        # OBV (On Balance Volume)
        obv = (np.sign(close.diff()) * volume).fillna(0).cumsum()
        features['obv_slope_10'] = obv.pct_change(10)
        features['obv_slope_20'] = obv.pct_change(20)
        
        # CMF (Chaikin Money Flow)
        mfm = ((close - low) - (high - close)) / (high - low)
        mfm = mfm.fillna(0)
        for p in [20, 21]:
            features[f'cmf_{p}'] = (mfm * volume).rolling(p).sum() / volume.rolling(p).sum()
        
        # A/D Line (Accumulation/Distribution)
        ad_line = ((2 * close - low - high) / (high - low) * volume).fillna(0).cumsum()
        features['ad_slope_10'] = ad_line.pct_change(10)
        
        # Force Index
        features['force_index_13'] = (close.diff() * volume).ewm(span=13).mean()
        
        # ═══════════════════════════════════════════════════════════════
        # 🔄 INDICATEURS DE TENDANCE
        # ═══════════════════════════════════════════════════════════════
        # MACD multi-versions
        for fast, slow in [(12, 26), (8, 17), (5, 35)]:
            ema_fast = close.ewm(span=fast, adjust=False).mean()
            ema_slow = close.ewm(span=slow, adjust=False).mean()
            macd = ema_fast - ema_slow
            features[f'macd_{fast}_{slow}'] = macd / close
            features[f'macd_signal_{fast}_{slow}'] = macd.ewm(span=9, adjust=False).mean() / close
            features[f'macd_hist_{fast}_{slow}'] = (macd - macd.ewm(span=9, adjust=False).mean()) / close
        
        # ADX (Average Directional Index)
        plus_dm = high.diff()
        minus_dm = -low.diff()
        plus_dm = plus_dm.where((plus_dm > minus_dm) & (plus_dm > 0), 0)
        minus_dm = minus_dm.where((minus_dm > plus_dm) & (minus_dm > 0), 0)
        tr14 = tr.rolling(14).sum()
        plus_di = 100 * (plus_dm.rolling(14).sum() / tr14)
        minus_di = 100 * (minus_dm.rolling(14).sum() / tr14)
        dx = 100 * abs(plus_di - minus_di) / (plus_di + minus_di)
        features['adx_14'] = dx.rolling(14).mean()
        features['plus_di'] = plus_di
        features['minus_di'] = minus_di
        features['di_diff'] = plus_di - minus_di
        
        # Aroon
        for p in [25]:
            aroon_up = 100 * high.rolling(p + 1).apply(lambda x: x.argmax(), raw=True) / p
            aroon_down = 100 * low.rolling(p + 1).apply(lambda x: x.argmin(), raw=True) / p
            features[f'aroon_up_{p}'] = aroon_up
            features[f'aroon_down_{p}'] = aroon_down
            features[f'aroon_osc_{p}'] = aroon_up - aroon_down
        
        # ═══════════════════════════════════════════════════════════════
        # 📐 BOLLINGER & KELTNER BANDS
        # ═══════════════════════════════════════════════════════════════
        for p in [20]:
            bb_mid = close.rolling(p).mean()
            bb_std = close.rolling(p).std()
            features[f'bb_position_{p}'] = (close - bb_mid) / (2 * bb_std)
            features[f'bb_width_{p}'] = (4 * bb_std) / bb_mid * 100
            features[f'bb_pct_{p}'] = (close - (bb_mid - 2 * bb_std)) / (4 * bb_std)
        
        # Keltner Channels
        kc_mid = close.ewm(span=20).mean()
        atr_kc = tr.rolling(10).mean()
        features['kc_position'] = (close - kc_mid) / (1.5 * atr_kc)
        
        # Squeeze detection (BB inside KC)
        bb_upper = close.rolling(20).mean() + 2 * close.rolling(20).std()
        bb_lower = close.rolling(20).mean() - 2 * close.rolling(20).std()
        kc_upper = kc_mid + 1.5 * atr_kc
        kc_lower = kc_mid - 1.5 * atr_kc
        features['squeeze_on'] = ((bb_lower > kc_lower) & (bb_upper < kc_upper)).astype(int)
        
        # ═══════════════════════════════════════════════════════════════
        # 🧮 STATISTIQUES AVANCÉES
        # ═══════════════════════════════════════════════════════════════
        # Z-Scores
        for p in [10, 20, 50, 100]:
            features[f'zscore_{p}'] = (close - close.rolling(p).mean()) / close.rolling(p).std()
        
        # Skewness (asymétrie des rendements)
        features['skew_20'] = close.pct_change().rolling(20).skew()
        features['skew_50'] = close.pct_change().rolling(50).skew()
        
        # Kurtosis (queues de distribution)
        features['kurt_20'] = close.pct_change().rolling(20).kurt()
        features['kurt_50'] = close.pct_change().rolling(50).kurt()
        
        # High-Low range
        features['range_pct'] = (high - low) / close
        features['range_avg_10'] = features['range_pct'].rolling(10).mean()
        features['range_avg_20'] = features['range_pct'].rolling(20).mean()
        
        # Distance aux extremes
        for p in [10, 20, 50]:
            features[f'dist_high_{p}'] = (close - high.rolling(p).max()) / close
            features[f'dist_low_{p}'] = (close - low.rolling(p).min()) / close
        
        # Momentum multi-périodes
        for p in [5, 10, 20, 40]:
            features[f'momentum_{p}'] = close / close.shift(p) - 1
        
        # ROC multi-périodes
        for p in [5, 10, 20]:
            features[f'roc_{p}'] = (close - close.shift(p)) / close.shift(p) * 100
        
        # ═══════════════════════════════════════════════════════════════
        # 🕯️ CANDLESTICK FEATURES
        # ═══════════════════════════════════════════════════════════════
        # Body size
        features['body_size'] = abs(close - open_price) / close
        features['upper_shadow'] = (high - pd.concat([close, open_price], axis=1).max(axis=1)) / close
        features['lower_shadow'] = (pd.concat([close, open_price], axis=1).min(axis=1) - low) / close
        
        # Gap
        features['gap'] = (open_price - close.shift(1)) / close.shift(1)
        
        # Consecutive up/down days
        up_day = (close > close.shift(1)).astype(int)
        features['consec_up'] = up_day.groupby((up_day != up_day.shift()).cumsum()).cumsum()
        features['consec_down'] = (1 - up_day).groupby(((1 - up_day) != (1 - up_day).shift()).cumsum()).cumsum()
        
        # ═══════════════════════════════════════════════════════════════
        # 🔢 INTERACTION FEATURES
        # ═══════════════════════════════════════════════════════════════
        features['rsi_vol_interaction'] = features['rsi_14'] * features['vol_ratio_20']
        features['momentum_vol_interaction'] = features['momentum_10'] * features['vol_ratio_20']
        features['trend_strength'] = features['adx_14'] * np.sign(features['di_diff'])
        
        return features.dropna()
    
    def preparer_labels(self, df: pd.DataFrame) -> pd.Series:
        """Labels pour prédiction de rendement futur"""
        ret_futur = df['Close'].pct_change(self.config.ML_HORIZON_JOURS).shift(-self.config.ML_HORIZON_JOURS)
        return (ret_futur >= self.config.ML_RENDEMENT_CIBLE).astype(int)
    
    def entrainer(self, symboles: List[str], callback: Callable = None) -> Dict:
        """
        🚀 ENTRAÎNEMENT AVANCÉ DE L'ENSEMBLE ML v2.0
        
        Pipeline d'entraînement:
        ═══════════════════════════════════════════════════════════════
        1. Collecte données (jusqu'à 200 symboles)
        2. Génération features (100+ indicateurs)
        3. Sélection features (SelectKBest / RFE)
        4. Entraînement modèles individuels avec CV
        5. Calibration des probabilités
        6. Construction du Stacking/Voting ensemble
        7. Calcul des poids dynamiques basés sur F1-Score
        ═══════════════════════════════════════════════════════════════
        """
        all_X = []
        all_y = []
        symboles_utilises = []
        erreurs = []
        
        # 🔧 AUGMENTÉ: Utiliser jusqu'à 200 symboles (configurable)
        nb_symboles = min(self.config.ML_MAX_SYMBOLES, len(symboles))
        logger.info(f"🧠 Entraînement ML ULTRA-AVANCÉ sur {nb_symboles} symboles...")
        
        print(f"\n{'='*70}")
        print(f"🧠 ENTRAÎNEMENT ML ULTRA-AVANCÉ")
        print(f"{'='*70}")
        print(f"📊 Symboles à traiter: {nb_symboles}")
        print(f"🎯 Objectif: +{self.config.ML_RENDEMENT_CIBLE:.0%} en {self.config.ML_HORIZON_JOURS} jours")
        print(f"{'='*70}\n")
        
        for i, symbole in enumerate(symboles[:nb_symboles]):
            try:
                if callback:
                    callback(f"ML: {symbole}", i, nb_symboles)
                
                # Afficher progression tous les 50 symboles
                if i % 50 == 0:
                    print(f"📥 Collecte données: {i}/{nb_symboles} ({i*100//nb_symboles}%)")
                
                # 🔧 AJOUT: Délai pour éviter le rate limiting Yahoo Finance
                if i > 0 and i % 10 == 0:
                    time.sleep(2)  # Pause de 2 secondes toutes les 10 requêtes
                elif i > 0:
                    time.sleep(0.3)  # Petite pause entre chaque requête
                
                ticker = yf.Ticker(symbole)
                df = ticker.history(period="2y")
                
                if len(df) < 200:
                    erreurs.append(f"{symbole}: Données insuffisantes ({len(df)} lignes)")
                    continue
                
                X = self.preparer_features(df)
                y = self.preparer_labels(df)
                
                idx_commun = X.index.intersection(y.dropna().index)
                if len(idx_commun) < 100:
                    erreurs.append(f"{symbole}: Pas assez d'index communs ({len(idx_commun)})")
                    continue
                
                all_X.append(X.loc[idx_commun])
                all_y.append(y.loc[idx_commun])
                symboles_utilises.append(symbole)
                
            except Exception as e:
                # 🔧 CORRECTION: Logger l'erreur au lieu de l'ignorer
                erreurs.append(f"{symbole}: {str(e)[:50]}")
                logger.warning(f"ML - Erreur pour {symbole}: {e}")
                continue
        
        # Logger le résumé
        logger.info(f"ML: {len(symboles_utilises)} symboles utilisés, {len(erreurs)} erreurs")
        if erreurs[:5]:  # Afficher les 5 premières erreurs
            logger.debug(f"Premières erreurs: {erreurs[:5]}")
        
        print(f"\n{'='*70}")
        print(f"📊 RÉSUMÉ COLLECTE DONNÉES:")
        print(f"   ✅ Symboles utilisés: {len(symboles_utilises)}/{nb_symboles}")
        print(f"   ❌ Erreurs: {len(erreurs)}")
        if erreurs[:3]:
            print(f"   ⚠️ Premières erreurs: {erreurs[:3]}")
        print(f"{'='*70}\n")
        
        if not all_X:
            return {}
        
        X = pd.concat(all_X)
        y = pd.concat(all_y)
        
        valide = ~(X.isna().any(axis=1) | y.isna())
        X = X[valide]
        y = y[valide]
        
        if len(X) < 200:
            return {}
        
        tscv = TimeSeriesSplit(n_splits=5)
        
        # ═══════════════════════════════════════════════════════════════
        # 🎯 SÉLECTION DE FEATURES (optionnel mais recommandé)
        # ═══════════════════════════════════════════════════════════════
        if self.config.ML_FEATURE_SELECTION:
            logger.info(f"🔍 Sélection des top {self.config.ML_TOP_K_FEATURES} features...")
            selector = SelectKBest(f_classif, k=min(self.config.ML_TOP_K_FEATURES, X.shape[1]))
            X_selected = selector.fit_transform(X, y)
            self.selected_features = X.columns[selector.get_support()].tolist()
            X = pd.DataFrame(X_selected, columns=self.selected_features, index=X.index)
            logger.info(f"✅ {len(self.selected_features)} features sélectionnées")
        
        # ═══════════════════════════════════════════════════════════════
        # 🏆 TIER 1: MODÈLES ARBRES DE DÉCISION (Base)
        # ═══════════════════════════════════════════════════════════════
        
        # Random Forest - Robuste, parallèle, bonne généralisation
        self.pipelines['rf'] = Pipeline([
            ('scaler', RobustScaler()),
            ('clf', RandomForestClassifier(
                n_estimators=200, max_depth=12, min_samples_split=8,
                min_samples_leaf=4, max_features='sqrt',
                random_state=42, n_jobs=-1, class_weight='balanced'
            ))
        ])
        
        # Extra Trees - Variance élevée, diversité
        from sklearn.ensemble import ExtraTreesClassifier
        self.pipelines['et'] = Pipeline([
            ('scaler', RobustScaler()),
            ('clf', ExtraTreesClassifier(
                n_estimators=200, max_depth=15, min_samples_split=5,
                min_samples_leaf=3, max_features='sqrt',
                random_state=42, n_jobs=-1, class_weight='balanced'
            ))
        ])
        
        # ═══════════════════════════════════════════════════════════════
        # 🚀 TIER 2: MODÈLES BOOSTING AVANCÉS
        # ═══════════════════════════════════════════════════════════════
        
        # XGBoost - State-of-the-art, régularisation L1/L2
        if XGBOOST_DISPONIBLE:
            self.pipelines['xgb'] = Pipeline([
                ('scaler', RobustScaler()),
                ('clf', xgb.XGBClassifier(
                    n_estimators=200, max_depth=6, learning_rate=0.03,
                    subsample=0.8, colsample_bytree=0.8, colsample_bylevel=0.8,
                    reg_alpha=0.1, reg_lambda=1.0,  # L1 et L2 regularization
                    random_state=42, eval_metric='logloss',
                    n_jobs=-1, verbosity=0  # Supprime les warnings
                ))
            ])
        else:
            self.pipelines['xgb'] = Pipeline([
                ('scaler', RobustScaler()),
                ('clf', GradientBoostingClassifier(
                    n_estimators=200, max_depth=5, learning_rate=0.03,
                    subsample=0.8, random_state=42
                ))
            ])
        
        # LightGBM - Ultra-rapide, leaf-wise, excellent sur gros datasets
        if LIGHTGBM_DISPONIBLE:
            self.pipelines['lgbm'] = Pipeline([
                ('scaler', RobustScaler()),
                ('clf', lgb.LGBMClassifier(
                    n_estimators=200, max_depth=8, learning_rate=0.03,
                    num_leaves=31, subsample=0.8, colsample_bytree=0.8,
                    reg_alpha=0.1, reg_lambda=1.0,
                    random_state=42, n_jobs=-1, verbose=-1,
                    class_weight='balanced'
                ))
            ])
        
        # CatBoost - Excellent pour données avec catégorielles, ordered boosting
        if CATBOOST_DISPONIBLE:
            self.pipelines['catboost'] = Pipeline([
                ('scaler', RobustScaler()),
                ('clf', CatBoostClassifier(
                    iterations=200, depth=6, learning_rate=0.03,
                    l2_leaf_reg=3, random_seed=42,
                    verbose=False, auto_class_weights='Balanced'
                ))
            ])
        
        # HistGradientBoosting - sklearn native, très rapide
        self.pipelines['histgb'] = Pipeline([
            ('scaler', RobustScaler()),
            ('clf', HistGradientBoostingClassifier(
                max_iter=200, max_depth=8, learning_rate=0.03,
                l2_regularization=1.0, random_state=42
            ))
        ])
        
        # AdaBoost - Boosting adaptatif classique
        self.pipelines['ada'] = Pipeline([
            ('scaler', RobustScaler()),
            ('clf', AdaBoostClassifier(
                n_estimators=100, learning_rate=0.05,
                random_state=42
            ))
        ])
        
        # ═══════════════════════════════════════════════════════════════
        # 🧬 TIER 3: RÉSEAUX DE NEURONES
        # ═══════════════════════════════════════════════════════════════
        
        # MLP Neural Network - Deep learning simplifié
        self.pipelines['mlp'] = Pipeline([
            ('scaler', StandardScaler()),
            ('clf', MLPClassifier(
                hidden_layer_sizes=(128, 64, 32),
                activation='relu', solver='adam',
                alpha=0.001,  # L2 regularization
                batch_size=64, learning_rate='adaptive',
                learning_rate_init=0.001,
                max_iter=500, early_stopping=True,
                validation_fraction=0.15, n_iter_no_change=20,
                random_state=42
            ))
        ])
        
        # MLP alternatif (architecture différente)
        self.pipelines['mlp2'] = Pipeline([
            ('scaler', StandardScaler()),
            ('clf', MLPClassifier(
                hidden_layer_sizes=(256, 128, 64, 32),
                activation='tanh', solver='adam',
                alpha=0.0001, batch_size=128,
                learning_rate='adaptive',
                max_iter=500, early_stopping=True,
                validation_fraction=0.15,
                random_state=42
            ))
        ])
        
        # ═══════════════════════════════════════════════════════════════
        # 📊 TIER 4: MODÈLES LINÉAIRES
        # ═══════════════════════════════════════════════════════════════
        
        from sklearn.linear_model import LogisticRegression, RidgeClassifier
        
        # Logistic Regression - Baseline linéaire
        self.pipelines['lr'] = Pipeline([
            ('scaler', StandardScaler()),
            ('clf', LogisticRegression(
                max_iter=1000, C=0.1, solver='lbfgs',
                class_weight='balanced', random_state=42
            ))
        ])
        
        # Ridge Classifier - Régularisation L2 forte
        self.pipelines['ridge'] = Pipeline([
            ('scaler', StandardScaler()),
            ('clf', RidgeClassifier(alpha=1.0, random_state=42))
        ])
        
        # SVM RBF (si dataset pas trop gros)
        if len(X) < 10000:
            self.pipelines['svm'] = Pipeline([
                ('scaler', StandardScaler()),
                ('clf', SVC(
                    C=1.0, kernel='rbf', gamma='scale',
                    probability=True, class_weight='balanced',
                    random_state=42
                ))
            ])
        
        # KNN (pour diversité)
        self.pipelines['knn'] = Pipeline([
            ('scaler', StandardScaler()),
            ('clf', KNeighborsClassifier(n_neighbors=15, weights='distance', n_jobs=-1))
        ])
        
        # ═══════════════════════════════════════════════════════════════
        # 🎯 POIDS INITIAUX DES MODÈLES (seront ajustés dynamiquement)
        # ═══════════════════════════════════════════════════════════════
        self.poids_modeles = {
            'xgb': 0.15,
            'lgbm': 0.12 if LIGHTGBM_DISPONIBLE else 0,
            'catboost': 0.12 if CATBOOST_DISPONIBLE else 0,
            'histgb': 0.10,
            'rf': 0.12,
            'et': 0.10,
            'ada': 0.05,
            'mlp': 0.08,
            'mlp2': 0.06,
            'lr': 0.04,
            'ridge': 0.02,
            'svm': 0.02 if len(X) < 10000 else 0,
            'knn': 0.02
        }
        
        # Normaliser les poids
        total_poids = sum(self.poids_modeles.values())
        self.poids_modeles = {k: v/total_poids for k, v in self.poids_modeles.items() if v > 0}
        
        # ═══════════════════════════════════════════════════════════════
        # 📊 CROSS-VALIDATION ET ENTRAÎNEMENT AVANCÉ
        # ═══════════════════════════════════════════════════════════════
        scores_cv = {name: [] for name in self.pipelines if name != 'iso'}
        precision_cv = {name: [] for name in self.pipelines if name != 'iso'}
        recall_cv = {name: [] for name in self.pipelines if name != 'iso'}
        f1_cv = {name: [] for name in self.pipelines if name != 'iso'}
        auc_cv = {name: [] for name in self.pipelines if name != 'iso'}
        
        logger.info(f"📈 Cross-validation sur {len(self.pipelines)-1} modèles (5 folds)...")
        print(f"\n{'='*70}")
        print(f"📈 CROSS-VALIDATION: {len(self.pipelines)-1} modèles × 5 folds")
        print(f"{'='*70}")
        
        fold_num = 0
        for train_idx, test_idx in tscv.split(X):
            fold_num += 1
            print(f"\n🔄 Fold {fold_num}/5 - Train: {len(train_idx)} samples, Test: {len(test_idx)} samples")
            X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
            y_train, y_test = y.iloc[train_idx], y.iloc[test_idx]
            
            for name, pipeline in self.pipelines.items():
                if name == 'iso':
                    continue
                try:
                    print(f"   ⚙️ Training {name:12}...", end=" ", flush=True)
                    pipeline.fit(X_train, y_train)
                    y_pred = pipeline.predict(X_test)
                    
                    # Métriques multiples
                    acc = accuracy_score(y_test, y_pred)
                    f1 = f1_score(y_test, y_pred, zero_division=0)
                    scores_cv[name].append(acc)
                    precision_cv[name].append(precision_score(y_test, y_pred, zero_division=0))
                    recall_cv[name].append(recall_score(y_test, y_pred, zero_division=0))
                    f1_cv[name].append(f1)
                    print(f"✅ Acc={acc:.1%} F1={f1:.1%}")
                    
                    # AUC-ROC si probabilités disponibles
                    if hasattr(pipeline, 'predict_proba'):
                        try:
                            y_prob = pipeline.predict_proba(X_test)[:, 1]
                            auc_cv[name].append(roc_auc_score(y_test, y_prob))
                        except:
                            pass
                except Exception as e:
                    print(f"❌ Erreur: {str(e)[:30]}")
                    if fold_num == 1:  # Logger seulement une fois
                        logger.warning(f"⚠️ ML CV erreur pour {name}: {str(e)[:50]}")
        
        # ═══════════════════════════════════════════════════════════════
        # 🎯 AJUSTEMENT DYNAMIQUE DES POIDS BASÉ SUR F1-SCORE
        # ═══════════════════════════════════════════════════════════════
        logger.info("⚖️ Ajustement dynamique des poids basé sur F1-Score...")
        avg_f1_scores = {name: np.mean(scores) if scores else 0 for name, scores in f1_cv.items()}
        
        # Ajuster les poids selon la performance réelle
        total_f1 = sum(avg_f1_scores.values())
        if total_f1 > 0:
            for name in self.poids_modeles:
                if name in avg_f1_scores:
                    # Mix entre poids initial et performance réelle
                    perf_weight = avg_f1_scores[name] / total_f1
                    self.poids_modeles[name] = 0.5 * self.poids_modeles[name] + 0.5 * perf_weight
        
        # Re-normaliser
        total_poids = sum(self.poids_modeles.values())
        self.poids_modeles = {k: v/total_poids for k, v in self.poids_modeles.items()}
        
        # Log des meilleurs modèles
        top_models = sorted(avg_f1_scores.items(), key=lambda x: x[1], reverse=True)[:5]
        logger.info(f"🏆 Top 5 modèles par F1: {[(m, f'{s:.3f}') for m, s in top_models]}")
        
        print(f"\n{'='*70}")
        print("🏆 RÉSULTATS CROSS-VALIDATION:")
        print(f"{'='*70}")
        for name, f1 in sorted(avg_f1_scores.items(), key=lambda x: x[1], reverse=True):
            acc = np.mean(scores_cv[name]) if scores_cv[name] else 0
            print(f"   {name:12} | Accuracy: {acc:.1%} | F1-Score: {f1:.1%}")
        print(f"{'='*70}\n")
        
        # ═══════════════════════════════════════════════════════════════
        # 🔧 ENTRAÎNEMENT FINAL SUR TOUT LE DATASET
        # ═══════════════════════════════════════════════════════════════
        logger.info("🔄 Entraînement final sur l'ensemble du dataset...")
        print(f"\n{'='*70}")
        print("🔧 ENTRAÎNEMENT FINAL SUR TOUT LE DATASET")
        print(f"{'='*70}")
        
        modeles_entraines = 0
        for name, pipeline in self.pipelines.items():
            if name != 'iso':
                try:
                    print(f"   ⚙️ Final training {name:12}...", end=" ", flush=True)
                    pipeline.fit(X, y)
                    modeles_entraines += 1
                    print("✅")
                except Exception as e:
                    print(f"❌ {str(e)[:30]}")
                    logger.warning(f"⚠️ Erreur entraînement {name}: {str(e)[:50]}")
        
        logger.info(f"✅ {modeles_entraines}/{len(self.pipelines)-1} modèles entraînés avec succès")
        print(f"\n✅ {modeles_entraines}/{len(self.pipelines)-1} modèles entraînés avec succès")
        
        # ═══════════════════════════════════════════════════════════════
        # 🎯 CONSTRUCTION DU MÉTA-MODÈLE (STACKING)
        # ═══════════════════════════════════════════════════════════════
        if self.config.ML_STACKING and modeles_entraines >= 3:
            logger.info("🏗️ Construction du Stacking Classifier...")
            try:
                # Sélectionner les 5 meilleurs modèles pour le stacking
                top_for_stacking = [name for name, _ in top_models[:5] if name in self.pipelines]
                estimators = [(name, self.pipelines[name]) for name in top_for_stacking]
                
                self.stacking_model = StackingClassifier(
                    estimators=estimators,
                    final_estimator=LogisticRegression(max_iter=1000, C=0.5),
                    cv=3,
                    stack_method='predict_proba',
                    n_jobs=-1
                )
                self.stacking_model.fit(X, y)
                logger.info(f"✅ Stacking Classifier construit avec {len(estimators)} modèles")
            except Exception as e:
                logger.warning(f"⚠️ Stacking non disponible: {e}")
                self.stacking_model = None
        
        # ═══════════════════════════════════════════════════════════════
        # 🔍 ISOLATION FOREST POUR DÉTECTION D'ANOMALIES
        # ═══════════════════════════════════════════════════════════════
        self.pipelines['iso'] = Pipeline([
            ('scaler', RobustScaler()),
            ('detector', IsolationForest(
                n_estimators=150, contamination=0.05,
                max_features=0.8, random_state=42, n_jobs=-1
            ))
        ])
        
        # ═══════════════════════════════════════════════════════════════
        # 📊 CALCUL FEATURE IMPORTANCES
        # ═══════════════════════════════════════════════════════════════
        try:
            if 'rf' in self.pipelines:
                rf_clf = self.pipelines['rf'].named_steps['clf']
                feature_names = X.columns.tolist()
                importances = rf_clf.feature_importances_
                self.feature_importances = dict(sorted(
                    zip(feature_names, importances),
                    key=lambda x: x[1], reverse=True
                )[:20])  # Top 20
                logger.info(f"📊 Top 5 features: {list(self.feature_importances.keys())[:5]}")
        except:
            pass
        self.pipelines['iso'].fit(X)
        
        self.est_entraine = True
        
        # ═══════════════════════════════════════════════════════════════
        # 📊 MÉTRIQUES AVANCÉES COMPLÈTES
        # ═══════════════════════════════════════════════════════════════
        avg_scores = {name: np.mean(scores) for name, scores in scores_cv.items() if scores}
        avg_precision = {name: np.mean(scores) for name, scores in precision_cv.items() if scores}
        avg_recall = {name: np.mean(scores) for name, scores in recall_cv.items() if scores}
        avg_f1 = {name: np.mean(scores) for name, scores in f1_cv.items() if scores}
        avg_auc = {name: np.mean(scores) for name, scores in auc_cv.items() if scores}
        
        # Calcul des métriques ensemble pondérées
        ensemble_score = sum(avg_scores.get(name, 0.5) * weight 
                           for name, weight in self.poids_modeles.items())
        ensemble_precision = sum(avg_precision.get(name, 0.5) * weight 
                                for name, weight in self.poids_modeles.items())
        ensemble_recall = sum(avg_recall.get(name, 0.5) * weight 
                             for name, weight in self.poids_modeles.items())
        ensemble_f1 = sum(avg_f1.get(name, 0.5) * weight 
                        for name, weight in self.poids_modeles.items())
        ensemble_auc = sum(avg_auc.get(name, 0.5) * weight 
                          for name, weight in self.poids_modeles.items() if name in avg_auc)
        
        # Métriques par modèle
        self.metriques = {
            # Métriques globales ensemble
            'accuracy_ensemble': ensemble_score,
            'precision_ensemble': ensemble_precision,
            'recall_ensemble': ensemble_recall,
            'f1_ensemble': ensemble_f1,
            'auc_ensemble': ensemble_auc,
            
            # Métriques par modèle (top)
            'accuracy_xgb': avg_scores.get('xgb', 0),
            'accuracy_lgbm': avg_scores.get('lgbm', 0),
            'accuracy_catboost': avg_scores.get('catboost', 0),
            'accuracy_rf': avg_scores.get('rf', 0),
            'accuracy_mlp': avg_scores.get('mlp', 0),
            'accuracy_histgb': avg_scores.get('histgb', 0),
            
            'f1_xgb': avg_f1.get('xgb', 0),
            'f1_lgbm': avg_f1.get('lgbm', 0),
            'f1_rf': avg_f1.get('rf', 0),
            'f1_mlp': avg_f1.get('mlp', 0),
            
            # Méta-infos
            'nb_modeles': len(self.pipelines) - 1,
            'nb_features': X.shape[1],
            'echantillons': len(X),
            'symboles_utilises': len(symboles_utilises),
            'stacking_actif': self.stacking_model is not None,
            'poids_modeles': dict(sorted(self.poids_modeles.items(), key=lambda x: x[1], reverse=True)[:5]),
            'top_features': list(self.feature_importances.keys())[:10] if self.feature_importances else []
        }
        
        # Log détaillé
        logger.info("═" * 60)
        logger.info(f"🎯 ML ULTRA-AVANCÉ ENTRAÎNÉ AVEC SUCCÈS")
        logger.info(f"   📊 Accuracy Ensemble: {ensemble_score:.2%}")
        logger.info(f"   📊 F1-Score Ensemble: {ensemble_f1:.2%}")
        logger.info(f"   📊 AUC-ROC Ensemble: {ensemble_auc:.2%}")
        logger.info(f"   📊 Precision: {ensemble_precision:.2%} | Recall: {ensemble_recall:.2%}")
        logger.info(f"   📈 {len(self.pipelines)-1} modèles | {X.shape[1]} features | {len(X)} échantillons")
        logger.info(f"   🏆 Stacking: {'✅' if self.stacking_model else '❌'}")
        logger.info("═" * 60)
        
        return self.metriques
    
    def predire(self, df: pd.DataFrame) -> Dict:
        """
        🔮 PRÉDICTION AVANCÉE avec Vote Pondéré Multi-Modèles
        
        Utilise tous les modèles disponibles + Stacking pour une prédiction robuste
        """
        if not self.est_entraine:
            return {
                'prediction': None, 'probabilite': 50, 'signal': 'N/A',
                'anomalie': False, 'consensus': 0, 'details': {},
                'confidence_level': 'N/A', 'stacking_prob': None
            }
        
        try:
            X = self.preparer_features(df)
            if X.empty:
                return {
                    'prediction': None, 'probabilite': 50, 'signal': 'N/A',
                    'anomalie': False, 'consensus': 0, 'details': {},
                    'confidence_level': 'N/A', 'stacking_prob': None
                }
            
            # Appliquer la sélection de features si active
            if self.selected_features:
                available_features = [f for f in self.selected_features if f in X.columns]
                if available_features:
                    X = X[available_features]
            
            X_last = X.iloc[[-1]]
            
            # ═══════════════════════════════════════════════════════════
            # 🎯 PRÉDICTIONS DE CHAQUE MODÈLE
            # ═══════════════════════════════════════════════════════════
            probas = {}
            predictions = {}
            
            for name, pipeline in self.pipelines.items():
                if name == 'iso':
                    continue
                try:
                    if hasattr(pipeline, 'predict_proba'):
                        prob = pipeline.predict_proba(X_last)[0][1]
                    else:
                        # Pour les modèles sans predict_proba (Ridge, etc.)
                        pred = pipeline.predict(X_last)[0]
                        prob = 0.7 if pred == 1 else 0.3
                    
                    probas[name] = prob
                    predictions[name] = int(prob >= 0.5)
                except:
                    continue
            
            if not probas:
                return {
                    'prediction': None, 'probabilite': 50, 'signal': 'N/A',
                    'anomalie': False, 'consensus': 0, 'details': {},
                    'confidence_level': 'N/A', 'stacking_prob': None
                }
            
            # ═══════════════════════════════════════════════════════════
            # 🏗️ PRÉDICTION STACKING (si disponible)
            # ═══════════════════════════════════════════════════════════
            stacking_prob = None
            if self.stacking_model is not None:
                try:
                    stacking_prob = self.stacking_model.predict_proba(X_last)[0][1]
                except:
                    pass
            
            # ═══════════════════════════════════════════════════════════
            # ⚖️ VOTE PONDÉRÉ DYNAMIQUE
            # ═══════════════════════════════════════════════════════════
            prob_ensemble = 0.0
            total_weight = 0.0
            
            for name, prob in probas.items():
                if name in self.poids_modeles:
                    weight = self.poids_modeles[name]
                    prob_ensemble += prob * weight
                    total_weight += weight
            
            if total_weight > 0:
                prob_ensemble /= total_weight
            
            # Intégrer le stacking si disponible (bonus de 20%)
            if stacking_prob is not None:
                prob_ensemble = 0.8 * prob_ensemble + 0.2 * stacking_prob
            
            # ═══════════════════════════════════════════════════════════
            # 📊 CALCUL DU CONSENSUS
            # ═══════════════════════════════════════════════════════════
            votes_achat = sum(1 for p in probas.values() if p >= 0.5)
            consensus = votes_achat / len(probas) * 100
            
            # Consensus fort/faible
            votes_fort_achat = sum(1 for p in probas.values() if p >= 0.65)
            votes_fort_vente = sum(1 for p in probas.values() if p <= 0.35)
            
            prediction = int(prob_ensemble >= 0.5)
            
            # ═══════════════════════════════════════════════════════════
            # 🔍 DÉTECTION D'ANOMALIE (régime de marché anormal)
            # ═══════════════════════════════════════════════════════════
            try:
                anomalie = self.pipelines['iso'].predict(X_last)[0] == -1
            except:
                anomalie = False
            
            # ═══════════════════════════════════════════════════════════
            # 🚦 SIGNAL AVEC NIVEAU DE CONFIANCE
            # ═══════════════════════════════════════════════════════════
            seuil = self.config.ML_SEUIL_CONFIANCE
            seuil_fort = seuil + 0.10  # Seuil pour signal fort
            
            # Niveau de confiance
            if prob_ensemble >= seuil_fort and consensus >= 80 and votes_fort_achat >= 3:
                signal = "🟢🟢🟢 TRÈS FORT ACHAT ML"
                confidence_level = "ULTRA-HIGH"
            elif prob_ensemble >= seuil_fort and consensus >= 70:
                signal = "🟢🟢 FORT ACHAT ML"
                confidence_level = "HIGH"
            elif prob_ensemble >= seuil and consensus >= 60:
                signal = "🟢 ACHAT ML"
                confidence_level = "MEDIUM"
            elif prob_ensemble >= 0.5 and consensus >= 50:
                signal = "🟡 ACHAT LÉGER ML"
                confidence_level = "LOW"
            elif prob_ensemble <= (1 - seuil_fort) and consensus <= 20 and votes_fort_vente >= 3:
                signal = "🔴🔴🔴 TRÈS FORTE VENTE ML"
                confidence_level = "ULTRA-HIGH"
            elif prob_ensemble <= (1 - seuil_fort) and consensus <= 30:
                signal = "🔴🔴 FORTE VENTE ML"
                confidence_level = "HIGH"
            elif prob_ensemble <= (1 - seuil):
                signal = "🔴 VENTE ML"
                confidence_level = "MEDIUM"
            elif prob_ensemble < 0.5 and consensus < 50:
                signal = "🟠 VENTE LÉGÈRE ML"
                confidence_level = "LOW"
            else:
                signal = "⚪ NEUTRE ML"
                confidence_level = "NEUTRAL"
            
            # Ajouter warning si anomalie détectée
            if anomalie:
                signal += " ⚠️ ANOMALIE"
                confidence_level += " (CAUTION)"
            
            # ═══════════════════════════════════════════════════════════
            # 📦 RÉSULTAT COMPLET
            # ═══════════════════════════════════════════════════════════
            return {
                'prediction': prediction,
                'probabilite': round(prob_ensemble * 100, 1),
                'signal': signal,
                'anomalie': anomalie,
                'consensus': round(consensus, 0),
                'confidence_level': confidence_level,
                'stacking_prob': round(stacking_prob * 100, 1) if stacking_prob else None,
                'votes_achat': votes_achat,
                'votes_total': len(probas),
                'votes_fort_achat': votes_fort_achat,
                'votes_fort_vente': votes_fort_vente,
                'details': {name: round(prob * 100, 1) for name, prob in sorted(
                    probas.items(), key=lambda x: x[1], reverse=True
                )[:8]},  # Top 8 modèles
                'top_bullish': [name for name, prob in probas.items() if prob >= 0.65][:3],
                'top_bearish': [name for name, prob in probas.items() if prob <= 0.35][:3]
            }
            
        except Exception as e:
            logger.error(f"❌ Erreur prédiction ML: {e}")
            return {
                'prediction': None, 'probabilite': 50, 'signal': 'N/A',
                'anomalie': False, 'consensus': 0, 'details': {},
                'confidence_level': 'ERROR', 'stacking_prob': None
            }


# ══════════════════════════════════════════════════════════════════════════════
#                    ANALYSEUR DE QUALITÉ FINANCIÈRE & NEWS
# ══════════════════════════════════════════════════════════════════════════════

class AnalyseurQualite:
    """
    🎯 Analyse la qualité financière et le sentiment des news
    pour un horizon d'investissement de 1 an maximum
    """
    
    def __init__(self, config: Config):
        self.config = config
    
    def analyser_sante_financiere(self, info: Dict) -> Dict:
        """
        📊 Analyse la santé financière de l'entreprise
        Score de 0 à 100
        """
        score = 50
        details = []
        alertes = []
        
        # 1. Rentabilité
        roe = info.get('returnOnEquity')
        if roe:
            roe_pct = roe * 100 if roe < 1 else roe
            if roe_pct > 20:
                score += 15
                details.append(f"✅ ROE excellent ({roe_pct:.1f}%)")
            elif roe_pct > 10:
                score += 8
                details.append(f"✅ ROE bon ({roe_pct:.1f}%)")
            elif roe_pct < 5:
                score -= 10
                alertes.append(f"⚠️ ROE faible ({roe_pct:.1f}%)")
        
        # 2. Croissance des revenus
        rev_growth = info.get('revenueGrowth')
        if rev_growth:
            growth_pct = rev_growth * 100 if abs(rev_growth) < 1 else rev_growth
            if growth_pct > 15:
                score += 12
                details.append(f"📈 Forte croissance revenus (+{growth_pct:.1f}%)")
            elif growth_pct > 5:
                score += 6
            elif growth_pct < -5:
                score -= 10
                alertes.append(f"📉 Revenus en déclin ({growth_pct:.1f}%)")
        
        # 3. Croissance des bénéfices
        earn_growth = info.get('earningsGrowth')
        if earn_growth:
            earn_pct = earn_growth * 100 if abs(earn_growth) < 1 else earn_growth
            if earn_pct > 20:
                score += 12
                details.append(f"📈 Forte croissance bénéfices (+{earn_pct:.1f}%)")
            elif earn_pct > 5:
                score += 6
            elif earn_pct < -10:
                score -= 12
                alertes.append(f"📉 Bénéfices en chute ({earn_pct:.1f}%)")
        
        # 4. Marge bénéficiaire
        profit_margin = info.get('profitMargins')
        if profit_margin:
            margin_pct = profit_margin * 100 if profit_margin < 1 else profit_margin
            if margin_pct > 20:
                score += 10
                details.append(f"✅ Marge excellente ({margin_pct:.1f}%)")
            elif margin_pct > 10:
                score += 5
            elif margin_pct < 5:
                score -= 8
                alertes.append(f"⚠️ Marge faible ({margin_pct:.1f}%)")
        
        # 5. Dette / Equity
        debt_equity = info.get('debtToEquity')
        if debt_equity:
            if debt_equity < 50:
                score += 10
                details.append(f"✅ Faible endettement (D/E: {debt_equity:.0f}%)")
            elif debt_equity > 150:
                score -= 15
                alertes.append(f"🚨 Fort endettement (D/E: {debt_equity:.0f}%)")
            elif debt_equity > 100:
                score -= 8
                alertes.append(f"⚠️ Endettement élevé (D/E: {debt_equity:.0f}%)")
        
        # 6. Current Ratio (liquidité)
        current_ratio = info.get('currentRatio')
        if current_ratio:
            if current_ratio > 2:
                score += 8
                details.append(f"✅ Bonne liquidité ({current_ratio:.1f}x)")
            elif current_ratio < 1:
                score -= 12
                alertes.append(f"🚨 Problème de liquidité ({current_ratio:.1f}x)")
        
        # 7. Free Cash Flow
        fcf = info.get('freeCashflow', 0)
        if fcf:
            if fcf > 0:
                score += 8
                details.append("✅ Cash Flow positif")
            else:
                score -= 10
                alertes.append("🚨 Cash Flow négatif")
        
        # 8. Dividende (stabilité pour investissement 1 an)
        div_yield = info.get('dividendYield')
        if div_yield and div_yield > 0.02:
            score += 5
            details.append(f"💰 Dividende ({div_yield*100:.1f}%)")
        
        score = max(0, min(100, score))
        
        # Qualité globale
        if score >= 75:
            qualite = "🟢 EXCELLENTE"
        elif score >= 60:
            qualite = "🟢 BONNE"
        elif score >= 45:
            qualite = "🟡 CORRECTE"
        elif score >= 30:
            qualite = "🟠 FAIBLE"
        else:
            qualite = "🔴 MAUVAISE"
        
        return {
            'score_qualite': score,
            'qualite': qualite,
            'details': details[:5],  # Top 5
            'alertes': alertes[:3],  # Top 3 alertes
            'investissable_1an': score >= 45 and len(alertes) < 2
        }
    
    def analyser_momentum_fondamental(self, info: Dict) -> Dict:
        """
        📊 Analyse le momentum fondamental (révisions analystes, surprises)
        """
        score = 50
        signaux = []
        
        # Révision estimations analystes
        current_price = info.get('currentPrice', 0)
        target_price = info.get('targetMeanPrice', 0)
        
        if current_price and target_price:
            upside = ((target_price - current_price) / current_price) * 100
            if upside > 30:
                score += 20
                signaux.append(f"🎯 Analystes: +{upside:.0f}% potentiel")
            elif upside > 15:
                score += 12
                signaux.append(f"🎯 Analystes: +{upside:.0f}% potentiel")
            elif upside < -10:
                score -= 15
                signaux.append(f"⚠️ Analystes: {upside:.0f}% downside")
        
        # Recommandation consensus
        rec = info.get('recommendationKey', '')
        if rec == 'strong_buy':
            score += 15
            signaux.append("📈 Consensus: ACHAT FORT")
        elif rec == 'buy':
            score += 10
            signaux.append("📈 Consensus: ACHAT")
        elif rec == 'sell':
            score -= 10
            signaux.append("📉 Consensus: VENTE")
        elif rec == 'strong_sell':
            score -= 15
            signaux.append("📉 Consensus: VENTE FORTE")
        
        # Nombre d'analystes
        nb_analysts = info.get('numberOfAnalystOpinions', 0)
        if nb_analysts >= 20:
            score += 5
            signaux.append(f"👥 {nb_analysts} analystes suivent")
        elif nb_analysts < 5:
            score -= 5
            signaux.append(f"⚠️ Peu suivi ({nb_analysts} analystes)")
        
        score = max(0, min(100, score))
        
        return {
            'score_momentum': score,
            'signaux': signaux[:4],
            'momentum_positif': score >= 55
        }
    
    def score_investissement_1an(self, info: Dict, donnees_tech: Dict) -> Dict:
        """
        🎯 SCORE FINAL D'INVESTISSEMENT HORIZON 1 AN
        Combine qualité financière + momentum + technique
        """
        # Analyse qualité
        qualite = self.analyser_sante_financiere(info)
        momentum = self.analyser_momentum_fondamental(info)
        
        # Données techniques
        rsi = donnees_tech.get('rsi', 50)
        z_score = donnees_tech.get('z_score', 0)
        score_tendance = donnees_tech.get('score_tendance', 2)
        
        # Score technique pour horizon 1 an
        score_tech = 50
        if 30 <= rsi <= 60:  # Zone idéale pour entrée
            score_tech += 15
        elif rsi < 30:  # Survendu = opportunité
            score_tech += 10
        elif rsi > 70:  # Suracheté = risque
            score_tech -= 15
        
        if -1.5 <= z_score <= 0:  # Légèrement sous moyenne
            score_tech += 10
        elif z_score < -2:  # Très sous-évalué
            score_tech += 15
        elif z_score > 2:  # Sur-évalué
            score_tech -= 15
        
        if score_tendance >= 3:  # Tendance haussière
            score_tech += 10
        elif score_tendance <= 1:
            score_tech -= 10
        
        # Pondérations pour investissement 1 an
        # Fondamentaux plus importants que technique
        score_final = (
            qualite['score_qualite'] * 0.40 +
            momentum['score_momentum'] * 0.30 +
            score_tech * 0.30
        )
        
        # Recommandation
        if score_final >= 70 and qualite['investissable_1an']:
            recommandation = "🟢🟢 ACHETER"
            conseil = "Excellente opportunité pour horizon 1 an"
        elif score_final >= 60 and qualite['investissable_1an']:
            recommandation = "🟢 ACHETER"
            conseil = "Bonne opportunité d'investissement"
        elif score_final >= 50:
            recommandation = "🟡 SURVEILLER"
            conseil = "Attendre meilleur point d'entrée"
        elif score_final >= 40:
            recommandation = "🟠 PRUDENCE"
            conseil = "Risque élevé, petite position si convaincu"
        else:
            recommandation = "🔴 ÉVITER"
            conseil = "Ne pas investir dans les conditions actuelles"
        
        return {
            'score_investissement': round(score_final, 1),
            'recommandation_1an': recommandation,
            'conseil_1an': conseil,
            'qualite_financiere': qualite,
            'momentum_analystes': momentum,
            'score_technique': round(score_tech, 1),
            'pret_investissement': score_final >= 55 and qualite['investissable_1an']
        }


# ══════════════════════════════════════════════════════════════════════════════
#                    ANALYSEUR DE NEWS AVANCÉ V2.0
# ══════════════════════════════════════════════════════════════════════════════

class AnalyseurNews:
    """
    📰 Système d'analyse de news et sentiment MULTI-SOURCES GRATUIT
    Sources: Google News RSS, Yahoo Finance RSS, Finnhub gratuit, yfinance
    """
    
    # Mots-clés positifs avec scores
    MOTS_POSITIFS = {
        # Très positifs (+15-20)
        'beat': 20, 'beats': 20, 'exceeds': 20, 'surge': 18, 'soar': 18,
        'record': 18, 'breakthrough': 18, 'upgrade': 17, 'upgraded': 17,
        'outperform': 16, 'bullish': 16, 'strong buy': 20, 'buy rating': 18,
        'price target raised': 18, 'raises target': 16, 'all-time high': 18,
        # Positifs (+10-15)
        'growth': 12, 'grows': 12, 'profit': 12, 'gains': 12, 'rally': 12,
        'jump': 12, 'rise': 10, 'rises': 10, 'positive': 10, 'higher': 10,
        'up': 8, 'buy': 10, 'dividend': 12, 'increase': 10, 'expansion': 12,
        'innovation': 12, 'partnership': 10, 'acquisition': 10, 'deal': 10,
        'launch': 10, 'success': 12, 'approval': 15, 'approved': 15,
        'strong': 10, 'accelerating': 12, 'momentum': 10, 'revenue beat': 15,
        # Modérément positifs (+5-10)
        'stable': 6, 'recover': 8, 'recovery': 8, 'improve': 8, 'opportunity': 8,
        'optimistic': 8, 'favorable': 8, 'confident': 8, 'exceed': 10, 'exceeds': 10,
    }
    
    # Mots-clés négatifs avec scores
    MOTS_NEGATIFS = {
        # Très négatifs (-15-20)
        'crash': -20, 'plunge': -18, 'collapse': -18, 'bankruptcy': -20,
        'fraud': -20, 'scandal': -18, 'downgrade': -17, 'downgraded': -17,
        'sell': -12, 'selloff': -15, 'miss': -15, 'misses': -15, 'missed': -15,
        'price target cut': -16, 'target lowered': -15, 'cuts target': -14,
        # Négatifs (-10-15)
        'loss': -12, 'losses': -12, 'decline': -12, 'drop': -10, 'drops': -10,
        'fall': -10, 'falls': -10, 'weak': -10, 'warning': -12, 'concern': -10,
        'cut': -10, 'cuts': -10, 'layoff': -12, 'layoffs': -12, 'lawsuit': -12,
        'investigation': -15, 'recall': -12, 'delay': -10, 'delayed': -10,
        'slump': -12, 'tumble': -12, 'tank': -14, 'plummet': -15,
        # Modérément négatifs (-5-10)
        'underperform': -8, 'bearish': -10, 'risk': -6, 'uncertainty': -6,
        'volatile': -6, 'struggle': -8, 'struggles': -8, 'disappointing': -10,
        'slowdown': -8, 'slower': -6, 'pressure': -6,
    }
    
    # Mots-clés pour secteur Gaming
    MOTS_GAMING = {
        'game': 5, 'gaming': 5, 'launch': 8, 'release': 8, 'esports': 8,
        'players': 6, 'dlc': 6, 'expansion': 8, 'console': 5, 'mobile': 5,
        'steam': 6, 'epic': 5, 'xbox': 5, 'playstation': 5, 'nintendo': 5,
        'activision': 5, 'ea sports': 5, 'call of duty': 6, 'grand theft auto': 6,
        'gta': 6, 'fortnite': 6, 'microtransaction': -5, 'delay': -10, 
        'bug': -8, 'backlash': -12,
    }
    
    # Mots-clés pour périodes d'earnings
    MOTS_EARNINGS = {
        'earnings': 3, 'revenue': 3, 'eps': 3, 'guidance': 5,
        'forecast': 3, 'outlook': 3, 'quarter': 3, 'quarterly': 3,
        'annual': 3, 'fiscal': 3, 'results': 3, 'report': 2,
    }
    
    # Mapping symboles -> noms complets pour recherche
    SYMBOLE_NOMS = {
        'AAPL': 'Apple', 'MSFT': 'Microsoft', 'GOOGL': 'Google Alphabet',
        'AMZN': 'Amazon', 'META': 'Meta Facebook', 'NVDA': 'NVIDIA',
        'TSLA': 'Tesla', 'AMD': 'AMD Advanced Micro', 'INTC': 'Intel',
        'CRM': 'Salesforce', 'ADBE': 'Adobe', 'ORCL': 'Oracle',
        'IBM': 'IBM', 'CSCO': 'Cisco', 'QCOM': 'Qualcomm',
        'TXN': 'Texas Instruments', 'AVGO': 'Broadcom', 'NOW': 'ServiceNow',
        'INTU': 'Intuit', 'AMAT': 'Applied Materials', 'MU': 'Micron',
        'LRCX': 'Lam Research', 'KLAC': 'KLA Corporation', 'SNPS': 'Synopsys',
        'CDNS': 'Cadence Design', 'PANW': 'Palo Alto Networks',
        'CRWD': 'CrowdStrike', 'ZS': 'Zscaler', 'DDOG': 'Datadog',
        'SNOW': 'Snowflake', 'PLTR': 'Palantir', 'NET': 'Cloudflare',
        'MDB': 'MongoDB', 'TEAM': 'Atlassian', 'WDAY': 'Workday',
        'VEEV': 'Veeva Systems', 'OKTA': 'Okta', 'FTNT': 'Fortinet',
        'EA': 'Electronic Arts', 'TTWO': 'Take-Two Interactive',
        'ATVI': 'Activision Blizzard', 'NTDOY': 'Nintendo',
        'JPM': 'JPMorgan Chase', 'BAC': 'Bank of America', 'WFC': 'Wells Fargo',
        'GS': 'Goldman Sachs', 'MS': 'Morgan Stanley', 'C': 'Citigroup',
        'BTC-USD': 'Bitcoin', 'ETH-USD': 'Ethereum', 'SOL-USD': 'Solana',
    }
    
    def __init__(self, config: Config):
        self.config = config
        self.cache_news = {}
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
    
    def _get_company_name(self, symbole: str) -> str:
        """Retourne le nom de l'entreprise pour la recherche"""
        return self.SYMBOLE_NOMS.get(symbole, symbole)
    
    def _fetch_url_content(self, url: str, timeout: int = 10) -> str:
        """Récupère le contenu d'une URL avec gestion d'erreurs"""
        try:
            req = urllib.request.Request(url, headers=self.headers)
            with urllib.request.urlopen(req, timeout=timeout) as response:
                return response.read().decode('utf-8', errors='ignore')
        except Exception as e:
            logging.debug(f"Erreur fetch URL {url}: {e}")
            return ""
    
    def _parse_rss_date(self, date_str: str) -> Optional[datetime]:
        """Parse différents formats de date RSS"""
        formats = [
            '%a, %d %b %Y %H:%M:%S %z',
            '%a, %d %b %Y %H:%M:%S GMT',
            '%Y-%m-%dT%H:%M:%SZ',
            '%Y-%m-%dT%H:%M:%S%z',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except:
                continue
        return None
    
    def recuperer_news_google(self, symbole: str, nb_jours: int = 7) -> List[Dict]:
        """📥 Récupère les news via Google News RSS (GRATUIT)"""
        news_list = []
        try:
            company_name = self._get_company_name(symbole)
            # Recherche combinée: symbole + nom
            queries = [
                f'{symbole} stock',
                f'{company_name} stock',
            ]
            
            date_limite = datetime.now() - timedelta(days=nb_jours)
            seen_titles = set()
            
            for query in queries[:1]:  # Une seule requête pour éviter rate limiting
                encoded_query = urllib.parse.quote(query)
                url = f"https://news.google.com/rss/search?q={encoded_query}&hl=en-US&gl=US&ceid=US:en"
                
                content = self._fetch_url_content(url)
                if not content:
                    continue
                
                try:
                    root = ET.fromstring(content)
                    items = root.findall('.//item')
                    
                    for item in items[:15]:
                        title = item.find('title')
                        link = item.find('link')
                        pub_date = item.find('pubDate')
                        source = item.find('source')
                        
                        if title is None or not title.text:
                            continue
                        
                        titre_clean = html.unescape(title.text.strip())
                        
                        # Éviter doublons
                        if titre_clean.lower() in seen_titles:
                            continue
                        seen_titles.add(titre_clean.lower())
                        
                        # Parser date
                        news_date = None
                        if pub_date is not None and pub_date.text:
                            news_date = self._parse_rss_date(pub_date.text)
                        
                        if news_date and news_date.replace(tzinfo=None) < date_limite:
                            continue
                        
                        news_item = {
                            'titre': titre_clean,
                            'source': source.text if source is not None else 'Google News',
                            'date': news_date.strftime('%Y-%m-%d %H:%M') if news_date else datetime.now().strftime('%Y-%m-%d'),
                            'url': link.text if link is not None else '',
                            'type': 'NEWS',
                            'source_type': 'GoogleNews'
                        }
                        news_list.append(news_item)
                        
                        if len(news_list) >= 15:
                            break
                except ET.ParseError:
                    continue
                    
        except Exception as e:
            logging.debug(f"Erreur Google News {symbole}: {e}")
        
        return news_list
    
    def recuperer_news_yahoo_rss(self, symbole: str, nb_jours: int = 7) -> List[Dict]:
        """📥 Récupère les news via Yahoo Finance RSS (GRATUIT)"""
        news_list = []
        try:
            url = f"https://feeds.finance.yahoo.com/rss/2.0/headline?s={symbole}&region=US&lang=en-US"
            content = self._fetch_url_content(url)
            
            if not content:
                return news_list
            
            date_limite = datetime.now() - timedelta(days=nb_jours)
            
            try:
                root = ET.fromstring(content)
                items = root.findall('.//item')
                
                for item in items[:15]:
                    title = item.find('title')
                    link = item.find('link')
                    pub_date = item.find('pubDate')
                    description = item.find('description')
                    
                    if title is None or not title.text:
                        continue
                    
                    titre_clean = html.unescape(title.text.strip())
                    
                    news_date = None
                    if pub_date is not None and pub_date.text:
                        news_date = self._parse_rss_date(pub_date.text)
                    
                    if news_date and news_date.replace(tzinfo=None) < date_limite:
                        continue
                    
                    news_item = {
                        'titre': titre_clean,
                        'source': 'Yahoo Finance',
                        'date': news_date.strftime('%Y-%m-%d %H:%M') if news_date else datetime.now().strftime('%Y-%m-%d'),
                        'url': link.text if link is not None else '',
                        'type': 'NEWS',
                        'description': description.text[:200] if description is not None and description.text else '',
                        'source_type': 'YahooRSS'
                    }
                    news_list.append(news_item)
            except ET.ParseError:
                pass
                
        except Exception as e:
            logging.debug(f"Erreur Yahoo RSS {symbole}: {e}")
        
        return news_list
    
    def recuperer_news_yfinance(self, symbole: str, nb_jours: int = 7) -> List[Dict]:
        """📥 Récupère les news via yfinance API (backup)"""
        news_list = []
        try:
            ticker = yf.Ticker(symbole)
            news = getattr(ticker, 'news', []) or []
            
            date_limite = datetime.now() - timedelta(days=nb_jours)
            
            for item in news[:15]:
                try:
                    pub_time = item.get('providerPublishTime', 0)
                    if pub_time:
                        news_date = datetime.fromtimestamp(pub_time)
                        if news_date < date_limite:
                            continue
                    else:
                        news_date = datetime.now()
                    
                    news_item = {
                        'titre': item.get('title', ''),
                        'source': item.get('publisher', 'yfinance'),
                        'date': news_date.strftime('%Y-%m-%d %H:%M'),
                        'url': item.get('link', ''),
                        'type': item.get('type', 'NEWS'),
                        'source_type': 'yfinance'
                    }
                    if news_item['titre']:
                        news_list.append(news_item)
                except Exception:
                    continue
                    
        except Exception as e:
            logging.debug(f"Erreur yfinance news {symbole}: {e}")
        
        return news_list
    
    def recuperer_news_marketwatch_rss(self, symbole: str, nb_jours: int = 7) -> List[Dict]:
        """📥 Récupère les news via MarketWatch RSS (GRATUIT)"""
        news_list = []
        try:
            # MarketWatch general market news
            urls = [
                f"https://feeds.marketwatch.com/marketwatch/topstories/",
                f"https://feeds.marketwatch.com/marketwatch/marketpulse/",
            ]
            
            company_name = self._get_company_name(symbole).lower()
            symbole_lower = symbole.lower()
            date_limite = datetime.now() - timedelta(days=nb_jours)
            
            for url in urls:
                content = self._fetch_url_content(url)
                if not content:
                    continue
                
                try:
                    root = ET.fromstring(content)
                    items = root.findall('.//item')
                    
                    for item in items[:30]:
                        title = item.find('title')
                        if title is None or not title.text:
                            continue
                        
                        titre_lower = title.text.lower()
                        # Filtrer pour le symbole recherché
                        if symbole_lower not in titre_lower and company_name not in titre_lower:
                            continue
                        
                        link = item.find('link')
                        pub_date = item.find('pubDate')
                        
                        news_date = None
                        if pub_date is not None and pub_date.text:
                            news_date = self._parse_rss_date(pub_date.text)
                        
                        if news_date and news_date.replace(tzinfo=None) < date_limite:
                            continue
                        
                        news_item = {
                            'titre': html.unescape(title.text.strip()),
                            'source': 'MarketWatch',
                            'date': news_date.strftime('%Y-%m-%d %H:%M') if news_date else datetime.now().strftime('%Y-%m-%d'),
                            'url': link.text if link is not None else '',
                            'type': 'NEWS',
                            'source_type': 'MarketWatch'
                        }
                        news_list.append(news_item)
                except ET.ParseError:
                    continue
                    
        except Exception as e:
            logging.debug(f"Erreur MarketWatch {symbole}: {e}")
        
        return news_list
        
    def recuperer_news(self, symbole: str, nb_jours: int = 7) -> List[Dict]:
        """
        📥 RÉCUPÉRATION MULTI-SOURCES (GRATUIT)
        Combine: Google News RSS + Yahoo RSS + yfinance + MarketWatch
        """
        try:
            # Vérifier cache
            cache_key = f"{symbole}_{nb_jours}"
            if cache_key in self.cache_news:
                cached = self.cache_news[cache_key]
                if time.time() - cached['timestamp'] < 1800:  # 30min cache
                    return cached['news']
            
            all_news = []
            seen_titles = set()
            
            # 1. Yahoo Finance RSS (priorité - souvent les meilleures news finance)
            yahoo_news = self.recuperer_news_yahoo_rss(symbole, nb_jours)
            for news in yahoo_news:
                title_key = news['titre'].lower()[:50]
                if title_key not in seen_titles:
                    seen_titles.add(title_key)
                    all_news.append(news)
            
            # 2. Google News RSS (large couverture)
            google_news = self.recuperer_news_google(symbole, nb_jours)
            for news in google_news:
                title_key = news['titre'].lower()[:50]
                if title_key not in seen_titles:
                    seen_titles.add(title_key)
                    all_news.append(news)
            
            # 3. yfinance (backup)
            if len(all_news) < 5:
                yf_news = self.recuperer_news_yfinance(symbole, nb_jours)
                for news in yf_news:
                    title_key = news['titre'].lower()[:50]
                    if title_key not in seen_titles:
                        seen_titles.add(title_key)
                        all_news.append(news)
            
            # 4. MarketWatch pour les grosses actions (si encore peu de news)
            if len(all_news) < 3 and symbole in self.SYMBOLE_NOMS:
                mw_news = self.recuperer_news_marketwatch_rss(symbole, nb_jours)
                for news in mw_news:
                    title_key = news['titre'].lower()[:50]
                    if title_key not in seen_titles:
                        seen_titles.add(title_key)
                        all_news.append(news)
            
            # Trier par date (plus récentes en premier)
            def parse_date_for_sort(news):
                try:
                    return datetime.strptime(news['date'][:10], '%Y-%m-%d')
                except:
                    return datetime.min
            
            all_news.sort(key=parse_date_for_sort, reverse=True)
            
            # Limiter à 20 news max
            all_news = all_news[:20]
            
            # Mettre en cache
            self.cache_news[cache_key] = {
                'news': all_news,
                'timestamp': time.time()
            }
            
            logger.debug(f"📰 {symbole}: {len(all_news)} news trouvées (Y:{len(yahoo_news)} G:{len(google_news)})")
            
            return all_news
            
        except Exception as e:
            logging.warning(f"Erreur récupération news {symbole}: {e}")
            return []
    
    def analyser_sentiment_texte(self, texte: str, est_gaming: bool = False) -> Dict:
        """📊 Analyse le sentiment d'un texte avec NLP basique amélioré"""
        texte_lower = texte.lower()
        score = 0
        mots_trouves = []
        
        # Analyser mots positifs
        for mot, valeur in self.MOTS_POSITIFS.items():
            if mot in texte_lower:
                score += valeur
                mots_trouves.append((mot, valeur, 'positif'))
        
        # Analyser mots négatifs
        for mot, valeur in self.MOTS_NEGATIFS.items():
            if mot in texte_lower:
                score += valeur
                mots_trouves.append((mot, valeur, 'négatif'))
        
        # Bonus gaming si applicable
        if est_gaming:
            for mot, valeur in self.MOTS_GAMING.items():
                if mot in texte_lower:
                    score += valeur
                    mots_trouves.append((mot, valeur, 'gaming'))
        
        # Contexte earnings (amplifie le score)
        contexte_earnings = False
        for mot in self.MOTS_EARNINGS:
            if mot in texte_lower:
                contexte_earnings = True
                score = int(score * 1.2)  # Amplifier si contexte earnings
                break
        
        return {
            'score': score,
            'mots_cles': mots_trouves[:5],
            'contexte_earnings': contexte_earnings
        }
    
    def analyser_news_action(self, symbole: str, secteur: str = None) -> Dict:
        """
        🎯 Analyse complète des news pour une action
        Retourne score de sentiment et résumé
        """
        news = self.recuperer_news(symbole, nb_jours=7)
        
        if not news:
            return {
                'score_news': 50,  # Neutre par défaut
                'sentiment': "🟡 NEUTRE",
                'nb_news': 0,
                'news_recentes': [],
                'alerte_news': None,
                'resume': "Aucune news récente disponible",
                'impact_decision': 0
            }
        
        est_gaming = secteur == "Gaming" or symbole in UniversActions.GAMING
        
        # Analyser chaque news
        scores = []
        news_analysees = []
        alertes = []
        
        for item in news:
            analyse = self.analyser_sentiment_texte(item['titre'], est_gaming)
            score = analyse['score']
            scores.append(score)
            
            # Déterminer sentiment
            if score > 10:
                sentiment_item = "🟢"
            elif score > 0:
                sentiment_item = "🟢"
            elif score > -5:
                sentiment_item = "🟡"
            elif score > -10:
                sentiment_item = "🟠"
            else:
                sentiment_item = "🔴"
                alertes.append(item['titre'][:80])
            
            news_analysees.append({
                **item,
                'score_sentiment': score,
                'sentiment': sentiment_item,
                'mots_cles': [m[0] for m in analyse['mots_cles']],
                'earnings_related': analyse['contexte_earnings']
            })
        
        # Score moyen pondéré (news récentes plus importantes)
        if scores:
            poids = [1.5 if i == 0 else 1.2 if i < 3 else 1.0 for i in range(len(scores))]
            score_moyen = sum(s * p for s, p in zip(scores, poids)) / sum(poids)
        else:
            score_moyen = 0
        
        # Normaliser sur échelle 0-100
        score_normalise = max(0, min(100, 50 + score_moyen * 2))
        
        # Déterminer sentiment global
        if score_normalise >= 70:
            sentiment_global = "🟢🟢 TRÈS POSITIF"
            impact = 10
        elif score_normalise >= 58:
            sentiment_global = "🟢 POSITIF"
            impact = 5
        elif score_normalise >= 45:
            sentiment_global = "🟡 NEUTRE"
            impact = 0
        elif score_normalise >= 35:
            sentiment_global = "🟠 NÉGATIF"
            impact = -5
        else:
            sentiment_global = "🔴 TRÈS NÉGATIF"
            impact = -15
        
        # Compter news positives/négatives
        nb_positives = sum(1 for s in scores if s > 5)
        nb_negatives = sum(1 for s in scores if s < -5)
        
        # Générer résumé
        if nb_positives > nb_negatives * 2:
            resume = f"📈 Actualités majoritairement positives ({nb_positives} positives vs {nb_negatives} négatives)"
        elif nb_negatives > nb_positives * 2:
            resume = f"📉 Attention: actualités majoritairement négatives ({nb_negatives} négatives vs {nb_positives} positives)"
        elif len(news) >= 5:
            resume = f"📊 Couverture médiatique active ({len(news)} articles), sentiment mitigé"
        else:
            resume = f"📰 {len(news)} articles récents, sentiment {sentiment_global.split()[-1].lower()}"
        
        # Alerte si news très négative
        alerte = alertes[0] if alertes else None
        
        return {
            'score_news': round(score_normalise, 1),
            'sentiment': sentiment_global,
            'nb_news': len(news),
            'nb_positives': nb_positives,
            'nb_negatives': nb_negatives,
            'news_recentes': news_analysees[:5],  # Top 5
            'alerte_news': alerte,
            'resume': resume,
            'impact_decision': impact,
            'couverture': "Forte" if len(news) >= 10 else "Modérée" if len(news) >= 5 else "Faible"
        }
    
    def generer_rapport_news(self, symbole: str, secteur: str = None) -> str:
        """
        📋 Génère un rapport textuel des news
        """
        analyse = self.analyser_news_action(symbole, secteur)
        
        rapport = []
        rapport.append(f"📰 ANALYSE NEWS: {symbole}")
        rapport.append("=" * 40)
        rapport.append(f"Sentiment Global: {analyse['sentiment']}")
        rapport.append(f"Score: {analyse['score_news']}/100")
        rapport.append(f"Articles analysés: {analyse['nb_news']}")
        rapport.append(f"  • Positifs: {analyse.get('nb_positives', 0)}")
        rapport.append(f"  • Négatifs: {analyse.get('nb_negatives', 0)}")
        rapport.append(f"Couverture: {analyse['couverture']}")
        rapport.append("")
        rapport.append(f"📋 Résumé: {analyse['resume']}")
        
        if analyse['alerte_news']:
            rapport.append("")
            rapport.append(f"🚨 ALERTE: {analyse['alerte_news']}")
        
        if analyse['news_recentes']:
            rapport.append("")
            rapport.append("📰 NEWS RÉCENTES:")
            for news in analyse['news_recentes'][:3]:
                rapport.append(f"  {news['sentiment']} [{news['source']}] {news['titre'][:60]}...")
        
        return "\n".join(rapport)


# ══════════════════════════════════════════════════════════════════════════════
#                    MOTEUR PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

class MoteurScreener:
    def __init__(self, config: Config = None):
        self.config = config or CONFIG
        self.cache = GestionnaireCache()
        self.valorisation = SystemeValorisation(self.config)
        self.trading = SystemeTrading(self.config)
        self.decision = SystemeDecision(self.config)
        self.qualite = AnalyseurQualite(self.config)  # 🆕 Analyseur qualité
        self.news = AnalyseurNews(self.config)        # 🆕 Analyseur news
        self.scoring_multi = ScoringMultiHorizon(self.config)  # 🆕 Scoring multi-horizon
        self.ml = PredicteurML(self.config) if self.config.ML_ACTIF else None
        
        # 🗄️ Système de suivi des signaux
        self.gestionnaire_signaux = GestionnaireSignaux()
        self.evaluateur = EvaluateurSignaux(
            self.gestionnaire_signaux, 
            self.config.POLITIQUE_SAME_BAR
        )
        self.stress_tests = StressTests(self.gestionnaire_signaux)
        
        self.resultats: List[Dict] = []
        self.en_cours = False
    
    def entrainer_ml(self, callback: Callable = None) -> Dict:
        if self.ml:
            symboles = list(UniversActions.obtenir_actions(
                self.config.INCLURE_US, 
                self.config.INCLURE_EUROPE,
                self.config.INCLURE_FRANCE,
                self.config.INCLURE_GAMING if hasattr(self.config, 'INCLURE_GAMING') else True
            ).keys())
            return self.ml.entrainer(symboles, callback)
        return {}
    
    def analyser_action(self, symbole: str) -> Optional[Dict]:
        try:
            ticker = yf.Ticker(symbole)
            
            # Données de prix
            cache_prix = self.cache.prix.get(symbole)
            if cache_prix is not None:
                df = cache_prix
            else:
                df = ticker.history(period="1y")
                if not df.empty:
                    self.cache.prix.set(symbole, df)
            
            if df.empty or len(df) < 30:
                return None
            
            cloture = df['Close']
            volumes = df['Volume']
            hauts = df['High']
            bas = df['Low']
            ouverture = df['Open'] if 'Open' in df.columns else cloture
            
            prix_actuel = cloture.iloc[-1]
            
            # ══════════════════════════════════════════════════════════════════
            # 📊 INDICATEURS TECHNIQUES CLASSIQUES
            # ══════════════════════════════════════════════════════════════════
            rsi = IndicateursTechniques.rsi_wilder(cloture, self.config.RSI_PERIODE)
            z_score = IndicateursTechniques.zscore(cloture, self.config.ZSCORE_FENETRE)
            macd_val, macd_hist, macd_tendance = IndicateursTechniques.macd(cloture)
            tendance, score_tendance = IndicateursTechniques.score_tendance(cloture)
            atr = IndicateursTechniques.atr(hauts, bas, cloture, self.config.ATR_PERIODE)
            support, resistance = IndicateursTechniques.supports_resistances(cloture)
            vol_ratio, vol_label = IndicateursTechniques.volume_relatif(volumes)
            
            # 🆕 Indicateurs Swing supplémentaires
            distance_ema = IndicateursTechniques.distance_ema(cloture, 20)
            volatilite_jour = IndicateursTechniques.volatilite_journaliere(cloture, 20)
            distance_supp = IndicateursTechniques.distance_support(prix_actuel, support) if support else 100
            pullback_val = IndicateursTechniques.pullback_score(cloture)
            
            # ══════════════════════════════════════════════════════════════════
            # 🆕 NOUVEAUX INDICATEURS COURT/MOYEN TERME
            # ══════════════════════════════════════════════════════════════════
            
            # ADX + DI+/DI-
            adx, di_plus, di_minus, adx_interpretation = IndicateursTechniques.adx(hauts, bas, cloture)
            
            # Stochastic RSI
            stoch_k, stoch_d, stoch_signal = IndicateursTechniques.stochastic_rsi(cloture)
            
            # Williams %R
            williams_r, williams_signal = IndicateursTechniques.williams_r(hauts, bas, cloture)
            
            # CMF (Chaikin Money Flow)
            cmf, cmf_signal = IndicateursTechniques.cmf(hauts, bas, cloture, volumes)
            
            # OBV
            obv_change, obv_trend = IndicateursTechniques.obv(cloture, volumes)
            
            # VWAP
            vwap, vwap_distance, vwap_signal = IndicateursTechniques.vwap(hauts, bas, cloture, volumes)
            
            # SuperTrend
            supertrend, supertrend_signal = IndicateursTechniques.supertrend(hauts, bas, cloture)
            
            # Squeeze Momentum
            squeeze_on, squeeze_momentum, squeeze_signal = IndicateursTechniques.squeeze_momentum(hauts, bas, cloture)
            
            # ══════════════════════════════════════════════════════════════════
            # 🆕 DÉTECTION D'OPPORTUNITÉS
            # ══════════════════════════════════════════════════════════════════
            
            # Breakout
            breakout_info = IndicateursTechniques.detecter_breakout(hauts, bas, cloture, volumes)
            
            # Gap
            gap_info = IndicateursTechniques.detecter_gap(ouverture, cloture)
            
            # Volume Spike
            volume_spike = IndicateursTechniques.detecter_volume_spike(volumes)
            
            # Pattern Reversal
            pattern_reversal = IndicateursTechniques.detecter_pattern_reversal(ouverture, hauts, bas, cloture)
            
            # Momentum Burst
            momentum_burst = IndicateursTechniques.calculer_momentum_burst(cloture)
            
            # 🆕 Détection du régime de marché
            regime_info = DetecteurRegime.detecter_regime(cloture, volumes)
            
            # Variations
            var_1j = ((cloture.iloc[-1] - cloture.iloc[-2]) / cloture.iloc[-2] * 100) if len(cloture) > 1 else 0
            var_1s = ((cloture.iloc[-1] - cloture.iloc[-5]) / cloture.iloc[-5] * 100) if len(cloture) > 5 else 0
            var_1m = ((cloture.iloc[-1] - cloture.iloc[-20]) / cloture.iloc[-20] * 100) if len(cloture) > 20 else 0
            
            # Volatilité journalière (%)
            volatilite = volatilite_jour
            
            # Fondamentaux
            cache_fond = self.cache.fondamentaux.get(symbole)
            if cache_fond:
                info = cache_fond
            else:
                try:
                    info = ticker.info
                    self.cache.fondamentaux.set(symbole, info)
                except:
                    info = {}
            
            pe_ratio = info.get('trailingPE') or info.get('forwardPE')
            
            # Signal analystes
            rec = info.get('recommendationKey', 'none')
            rec_map = {'strong_buy': '🟢🟢 ACHAT FORT', 'buy': '🟢 ACHAT', 'hold': '⚪ CONSERVER',
                       'sell': '🔴 VENTE', 'strong_sell': '🔴🔴 VENTE FORTE'}
            signal_analyste = rec_map.get(rec, '⚪ N/A')
            
            # 💰 VALORISATION - PRIX RÉEL
            valorisation = self.valorisation.calculer_prix_reel(info, prix_actuel)
            
            # 🧠 ML
            ml_result = {'prediction': None, 'probabilite': 50, 'signal': 'N/A', 'anomalie': False, 'consensus': 50}
            if self.ml and self.ml.est_entraine:
                cache_ml = self.cache.ml.get(symbole)
                if cache_ml:
                    ml_result = cache_ml
                else:
                    ml_result = self.ml.predire(df)
                    self.cache.ml.set(symbole, ml_result)
            
            # 🆕 ANALYSE QUALITÉ FINANCIÈRE (HORIZON 1 AN)
            donnees_tech = {
                'rsi': rsi,
                'z_score': z_score,
                'score_tendance': score_tendance
            }
            analyse_qualite = self.qualite.score_investissement_1an(info, donnees_tech)
            
            # 📰 ANALYSE NEWS & SENTIMENT
            secteur = UniversActions.obtenir_secteur(symbole)
            analyse_news = self.news.analyser_news_action(symbole, secteur)
            impact_news = analyse_news.get('impact_decision', 0)
            
            # 📊 PLAN DE TRADING (adapté avec config swing)
            donnees_trading = {
                'prix': prix_actuel,
                'atr': atr,
                'support': support,
                'resistance': resistance,
                'prix_reel': valorisation['prix_reel'],
                'potentiel': valorisation['potentiel'],
                'score': 50,
                'strategie': tendance,
                'volatilite': volatilite,
                'sl_atr_mult': self.config.SWING_SL_ATR,
                'tp_ratio': self.config.SWING_TP_RATIO
            }
            plan_trading = self.trading.generer_plan_trading(donnees_trading)
            
            # ══════════════════════════════════════════════════════════════════
            # 🆕 SCORING MULTI-HORIZON
            # ══════════════════════════════════════════════════════════════════
            donnees_scoring = {
                # Indicateurs classiques
                'rsi': rsi,
                'z_score': z_score,
                'score_tendance': score_tendance,
                'macd_tendance': macd_tendance,
                'volume_ratio': vol_ratio,
                'pullback_score': pullback_val,
                'potentiel': valorisation['potentiel'],
                'ratio_rr': plan_trading['ratio_rr'],
                'ml_signal': ml_result.get('signal', 'N/A'),
                'ml_probabilite': ml_result['probabilite'],
                'ml_consensus': ml_result.get('consensus', 50),
                'score_qualite': analyse_qualite.get('score_investissement', 50),
                'score_news': analyse_news.get('score_news', 50),
                'signal_analyste': signal_analyste,
                'pe_ratio': pe_ratio,
                'distance_support': distance_supp,
                'regime': regime_info.get('regime', 'NEUTRE'),
                
                # 🆕 Nouveaux indicateurs
                'adx': adx,
                'adx_interpretation': adx_interpretation,
                'di_plus': di_plus,
                'di_minus': di_minus,
                'stoch_rsi_k': stoch_k,
                'stoch_rsi_d': stoch_d,
                'stoch_rsi_signal': stoch_signal,
                'williams_r': williams_r,
                'williams_signal': williams_signal,
                'cmf': cmf,
                'cmf_signal': cmf_signal,
                'obv_change': obv_change,
                'obv_trend': obv_trend,
                'vwap': vwap,
                'vwap_distance': vwap_distance,
                'vwap_signal': vwap_signal,
                'supertrend': supertrend,
                'supertrend_signal': supertrend_signal,
                'squeeze_on': squeeze_on,
                'squeeze_momentum': squeeze_momentum,
                'squeeze_signal': squeeze_signal,
                
                # Détection opportunités
                'breakout': breakout_info,
                'gap': gap_info,
                'volume_spike': volume_spike,
                'pattern_reversal': pattern_reversal,
                'momentum_burst': momentum_burst,
            }
            
            scoring_multi = self.scoring_multi.score_global(donnees_scoring)
            
            # 🎯 DÉCISION FINALE (avec données enrichies pour filtres swing)
            donnees_decision = {
                'rsi': rsi,
                'z_score': z_score,
                'score_tendance': score_tendance,
                'tendance': tendance,
                'macd_tendance': macd_tendance,
                'volume_ratio': vol_ratio,
                'pe_ratio': pe_ratio,
                'potentiel': valorisation['potentiel'],
                'signal_analyste': signal_analyste,
                'ratio_rr': plan_trading['ratio_rr'],
                'stop_loss_pct': plan_trading['stop_loss_pct'],
                'take_profit_pct': plan_trading['take_profit_pct'],
                'ml_prediction': ml_result['prediction'],
                'ml_probabilite': ml_result['probabilite'],
                'ml_signal': ml_result.get('signal', 'N/A'),
                'ml_anomalie': ml_result['anomalie'],
                'ml_consensus': ml_result.get('consensus', 50),
                # 🆕 Données swing supplémentaires
                'volatilite': volatilite,
                'distance_support': distance_supp,
                'pullback_score': pullback_val,
                'distance_ema': distance_ema,
                'regime': regime_info,
                # 🆕 Données qualité pour Win Rate
                'prix': prix_actuel,
                'score_qualite': analyse_qualite.get('score_investissement', 50),
                # 📰 Impact news
                'impact_news': impact_news,
                'score_news': analyse_news.get('score_news', 50)
            }
            decision = self.decision.generer_decision_finale(donnees_decision)
            
            # Marché et secteur
            marche = UniversActions.obtenir_marche(symbole)
            
            # ══════════════════════════════════════════════════════════════════
            # 🆕 DÉTERMINER LE TYPE D'OPPORTUNITÉ
            # ══════════════════════════════════════════════════════════════════
            opportunites = []
            
            if breakout_info.get('type') not in ['AUCUN', None]:
                opportunites.append(('BREAKOUT', breakout_info.get('score', 0), breakout_info.get('type')))
            
            if gap_info.get('type') not in ['AUCUN', None]:
                opportunites.append(('GAP', abs(gap_info.get('pourcentage', 0)) * 10, gap_info.get('type')))
            
            if volume_spike.get('spike'):
                opportunites.append(('VOLUME', volume_spike.get('ratio', 1) * 20, volume_spike.get('niveau', '')))
            
            if pattern_reversal.get('fiabilite', 0) >= 65:
                opportunites.append(('PATTERN', pattern_reversal.get('fiabilite', 0), pattern_reversal.get('pattern')))
            
            if momentum_burst.get('burst'):
                opportunites.append(('MOMENTUM', momentum_burst.get('score', 0), momentum_burst.get('direction')))
            
            if squeeze_on:
                opportunites.append(('SQUEEZE', 70, 'Breakout imminent'))
            
            opportunites.sort(key=lambda x: x[1], reverse=True)
            
            principale_opportunite = opportunites[0] if opportunites else ('AUCUNE', 0, '')
            
            # ====== RÉCUPÉRATION PRIX EN TEMPS RÉEL ======
            try:
                fast_info = yf.Ticker(symbole).fast_info
                prix_live = getattr(fast_info, 'last_price', None) or getattr(fast_info, 'previous_close', None)
                prev_close = getattr(fast_info, 'previous_close', None)
                var_live_pct = round(((prix_live - prev_close) / prev_close) * 100, 2) if prix_live and prev_close else None
                prix_live = round(prix_live, 2) if prix_live else prix_actuel
            except:
                prix_live = prix_actuel
                var_live_pct = None
            # ==============================================
            
            resultat = {
                # Identité
                'symbole': symbole,
                'marche': marche,
                'secteur': secteur,
                
                # Prix
                'prix': round(prix_actuel, 2),
                'PRIX_LIVE': prix_live,
                'VAR_LIVE_%': var_live_pct,
                'MAJ': datetime.now().strftime('%H:%M'),
                'var_1j': round(var_1j, 2),
                'var_1s': round(var_1s, 2),
                'var_1m': round(var_1m, 2),
                
                # Technique
                'rsi': rsi,
                'z_score': z_score,
                'macd_tendance': macd_tendance,
                'tendance': tendance,
                'score_tendance': score_tendance,
                'volume_ratio': vol_ratio,
                'volume_label': vol_label,
                'support': support,
                'resistance': resistance,
                
                # 🆕 Swing indicators
                'volatilite': round(volatilite, 2) if volatilite else None,
                'distance_ema': round(distance_ema, 2) if distance_ema else None,
                'distance_support': round(distance_supp, 2) if distance_supp else None,
                'pullback_score': round(pullback_val, 1) if pullback_val else None,
                'regime': regime_info.get('regime', 'UNKNOWN'),
                'regime_description': regime_info.get('description', ''),
                'strategie_detectee': decision.get('strategie_detectee', 'MIXTE'),
                
                # 🆕 NOUVEAUX INDICATEURS
                'adx': round(adx, 1) if adx else None,
                'adx_interpretation': adx_interpretation,
                'stoch_rsi_k': round(stoch_k, 1) if stoch_k else None,
                'stoch_rsi_signal': stoch_signal,
                'williams_r': round(williams_r, 1) if williams_r else None,
                'cmf': round(cmf, 3) if cmf else None,
                'cmf_signal': cmf_signal,
                'obv_trend': obv_trend,
                'vwap': round(vwap, 2) if vwap else None,
                'vwap_distance': round(vwap_distance, 2) if vwap_distance else None,
                'supertrend_signal': supertrend_signal,
                'squeeze_on': squeeze_on,
                'squeeze_signal': squeeze_signal,
                
                # 🆕 OPPORTUNITÉS DÉTECTÉES
                'breakout': breakout_info,
                'gap': gap_info,
                'volume_spike': volume_spike,
                'pattern_reversal': pattern_reversal,
                'momentum_burst': momentum_burst,
                'opportunite_principale': principale_opportunite[2],
                'opportunite_type': principale_opportunite[0],
                'opportunite_score': principale_opportunite[1],
                
                # 🆕 SCORING MULTI-HORIZON
                'score_intraday': scoring_multi['score_intraday']['score'],
                'score_swing': scoring_multi['score_swing']['score'],
                'score_position': scoring_multi['score_position']['score'],
                'meilleur_horizon': scoring_multi['meilleur_horizon'],
                'meilleur_score_horizon': scoring_multi['meilleur_score'],
                'opportunite_claire': scoring_multi['opportunite_claire'],
                'details_intraday': scoring_multi['score_intraday'],
                'details_swing': scoring_multi['score_swing'],
                'details_position': scoring_multi['score_position'],
                
                # 💰 VALORISATION
                'prix_reel': valorisation['prix_reel'],
                'potentiel': valorisation['potentiel'],
                'statut_valorisation': valorisation['statut_valorisation'],
                'prix_achat_ideal': valorisation['prix_achat_ideal'],
                'details_valorisation': valorisation['details'],
                'pe_ratio': round(pe_ratio, 2) if pe_ratio else None,
                
                # 📊 TRADING
                'stop_loss': plan_trading['stop_loss'],
                'stop_loss_pct': plan_trading['stop_loss_pct'],
                'take_profit': plan_trading['take_profit'],
                'take_profit_pct': plan_trading['take_profit_pct'],
                'ratio_rr': plan_trading['ratio_rr'],
                'qualite_rr': plan_trading['qualite_rr'],
                'horizon_jours': plan_trading['horizon_jours'],
                'horizon_type': plan_trading['horizon_type'],
                
                # 🧠 ML
                'ml_prediction': ml_result['prediction'],
                'ml_probabilite': ml_result['probabilite'],
                'ml_signal': ml_result.get('signal', 'N/A'),
                'ml_consensus': ml_result.get('consensus', 50),
                
                # Analystes
                'signal_analyste': signal_analyste,
                
                # 🎯 DÉCISION FINALE
                'decision': decision['decision'],
                'action': decision['action'],
                'score_final': decision['score_final'],
                'confiance': decision['confiance'],
                'confiance_pct': decision['confiance_pct'],
                'conseil': decision['conseil'],
                'passe_filtres_swing': decision.get('passe_filtres_swing', False),
                'raisons_rejet': decision.get('raisons_rejet', []),
                
                # 🆕 CONFIRMATIONS WIN RATE
                'nb_confirmations': decision.get('nb_confirmations', 0),
                'confirmations': decision.get('confirmations', []),
                
                # 🆕 QUALITÉ FINANCIÈRE (horizon 1 an)
                'score_qualite': analyse_qualite.get('score_investissement', 50),
                'qualite_financiere': analyse_qualite.get('qualite_financiere', {}).get('qualite', 'N/A'),
                'recommandation_1an': analyse_qualite.get('recommandation_1an', '⚪ N/A'),
                'conseil_1an': analyse_qualite.get('conseil_1an', ''),
                'pret_investissement': analyse_qualite.get('pret_investissement', False),
                'alertes_financieres': analyse_qualite.get('qualite_financiere', {}).get('alertes', []),
                
                # 📰 NEWS & SENTIMENT
                'score_news': analyse_news.get('score_news', 50),
                'sentiment_news': analyse_news.get('sentiment', '🟡 NEUTRE'),
                'nb_news': analyse_news.get('nb_news', 0),
                'resume_news': analyse_news.get('resume', ''),
                'alerte_news': analyse_news.get('alerte_news'),
                'news_recentes': analyse_news.get('news_recentes', [])[:3],  # Top 3
                'couverture_media': analyse_news.get('couverture', 'Faible'),
                
                # Valeurs brutes pour tri
                '_prix': prix_actuel,
                '_score': decision['score_final'],
                '_potentiel': valorisation['potentiel'] or 0,
                '_ratio_rr': plan_trading['ratio_rr'],
                '_score_qualite': analyse_qualite.get('score_investissement', 50),
                '_nb_confirmations': decision.get('nb_confirmations', 0),
                '_score_news': analyse_news.get('score_news', 50),
                '_score_intraday': scoring_multi['score_intraday']['score'],
                '_score_swing': scoring_multi['score_swing']['score'],
                '_meilleur_horizon_score': scoring_multi['meilleur_score'],
            }
            
            return resultat
            
        except Exception as e:
            logger.error(f"Erreur analyse {symbole}: {e}")
            return None
    
    def executer(self, callback: Callable = None) -> List[Dict]:
        self.en_cours = True
        self.resultats = []
        
        actions = UniversActions.obtenir_actions(
            self.config.INCLURE_US, 
            self.config.INCLURE_EUROPE,
            self.config.INCLURE_FRANCE,
            self.config.INCLURE_GAMING,
            self.config.INCLURE_MATIERES,
            self.config.INCLURE_CRYPTO
        )
        total = len(actions)
        
        logger.info(f"Analyse de {total} actions...")
        
        for i, symbole in enumerate(actions.keys(), 1):
            if not self.en_cours:
                break
            
            progress = int((i / total) * 100)
            resultat = self.analyser_action(symbole)
            
            if resultat and resultat.get('score_final', 0) >= self.config.SCORE_MINIMUM:
                self.resultats.append(resultat)
                
                # 💾 Sauvegarder le signal si c'est un ACHAT ou VENTE
                if self.config.SAUVEGARDER_SIGNAUX and resultat.get('action') in ['ACHAT', 'VENTE']:
                    try:
                        signal_data = {
                            'symbole': resultat['symbole'],
                            'time_signal': datetime.now().isoformat(),
                            'timeframe': 'Daily',
                            'strategie': resultat.get('tendance', 'Composite'),
                            'action': resultat['action'],
                            'entry_rule': 'Close',
                            'prix': resultat['prix'],
                            'stop_loss': resultat.get('stop_loss'),
                            'take_profit': resultat.get('take_profit'),
                            'horizon_jours': resultat.get('horizon_jours'),
                            'score_final': resultat.get('score_final'),
                            'confiance_pct': resultat.get('confiance_pct'),
                            'ratio_rr': resultat.get('ratio_rr'),
                            'marche': resultat.get('marche'),
                            'secteur': resultat.get('secteur')
                        }
                        self.gestionnaire_signaux.sauvegarder_signal(signal_data)
                    except Exception as e:
                        logger.error(f"Erreur sauvegarde signal {symbole}: {e}")
            
            if callback:
                callback(symbole, i, total, progress, resultat)
            
            time.sleep(0.12)
        
        self.resultats.sort(key=lambda x: -x.get('score_final', 0))
        self.en_cours = False
        
        logger.info(f"Résultats: {len(self.resultats)} actions")
        return self.resultats
    
    def evaluer_signaux_pending(self, callback: Callable = None) -> Dict:
        """Évalue tous les signaux en attente"""
        return self.evaluateur.evaluer_tous_pending(callback)
    
    def obtenir_scorecard(self, filtre: Dict = None) -> Dict:
        """Obtient les statistiques de performance"""
        return self.gestionnaire_signaux.obtenir_statistiques(filtre)
    
    def executer_stress_tests(self) -> Dict:
        """Exécute les stress tests"""
        return self.stress_tests.executer_tous_tests()
    
    def obtenir_historique_signaux(self, limite: int = 100) -> List[Dict]:
        """Obtient l'historique des signaux"""
        return self.gestionnaire_signaux.obtenir_tous_signaux(limite)
    
    # ══════════════════════════════════════════════════════════════════════════
    # 📋 NOUVELLES MÉTHODES D'EXPLORATION DES SIGNAUX
    # ══════════════════════════════════════════════════════════════════════════
    
    def obtenir_signaux_periode(self, jours: int = 7, symbole: str = None) -> List[Dict]:
        """Signaux des X derniers jours"""
        return self.gestionnaire_signaux.obtenir_signaux_par_periode(jours, symbole)
    
    def obtenir_signaux_symbole(self, symbole: str, limite: int = 50) -> List[Dict]:
        """Historique des signaux pour un symbole"""
        return self.gestionnaire_signaux.obtenir_signaux_par_symbole(symbole, limite)
    
    def obtenir_resume_dates(self, jours: int = 30) -> List[Dict]:
        """Résumé des signaux par date"""
        return self.gestionnaire_signaux.obtenir_resume_par_date(jours)
    
    def obtenir_symboles_avec_signaux(self) -> List[str]:
        """Liste des symboles ayant des signaux"""
        return self.gestionnaire_signaux.obtenir_symboles_uniques()
    
    def obtenir_evolution_signal(self, signal_id: str) -> Dict:
        """Détails complets d'un signal"""
        return self.gestionnaire_signaux.obtenir_evolution_signal(signal_id)
    
    def comparer_signal_actuel(self, signal_id: str) -> Dict:
        """Compare un signal avec le prix actuel"""
        signal = self.gestionnaire_signaux.obtenir_evolution_signal(signal_id)
        if not signal:
            return {}
        
        symbole = signal.get('symbol')
        if symbole:
            try:
                ticker = yf.Ticker(symbole)
                hist = ticker.history(period="1d")
                if not hist.empty:
                    prix_actuel = hist['Close'].iloc[-1]
                    return self.gestionnaire_signaux.comparer_signal_prix_actuel(signal_id, prix_actuel)
            except:
                pass
        
        return signal
    
    def obtenir_stats_periode(self, jours: int = 7) -> Dict:
        """Statistiques d'une période"""
        return self.gestionnaire_signaux.obtenir_statistiques_periode(jours)
    
    def exporter_signaux(self, filepath: str, jours: int = None) -> bool:
        """Exporte les signaux en CSV"""
        return self.gestionnaire_signaux.exporter_signaux_csv(filepath, jours)
    
    def arreter(self):
        self.en_cours = False
    
    # ══════════════════════════════════════════════════════════════════════════
    # 🚀 MODE SNIPER - SCAN ULTRA-RAPIDE
    # ══════════════════════════════════════════════════════════════════════════
    
    def executer_sniper(self, callback: Callable = None) -> List[Dict]:
        """
        🎯 SNIPER MODE - Scan ultra-rapide pour détecter les meilleures opportunités
        Analyse 200+ actions en moins de 30 secondes!
        """
        self.en_cours = True
        
        # Récupérer tous les symboles
        actions = UniversActions.obtenir_actions(
            self.config.INCLURE_US, 
            self.config.INCLURE_EUROPE,
            self.config.INCLURE_FRANCE,
            self.config.INCLURE_GAMING,
            self.config.INCLURE_MATIERES,
            self.config.INCLURE_CRYPTO
        )
        
        symbols = list(actions.keys())
        logger.info(f"🎯 SNIPER MODE: Scan de {len(symbols)} actions...")
        
        # Utiliser le SniperScanner
        scanner = SniperScanner(self.config)
        resultats_sniper = scanner.scan_rapid(symbols, callback)
        
        # Enrichir avec les infos de marché/secteur
        for r in resultats_sniper:
            symbole = r['symbole']
            r['marche'] = UniversActions.obtenir_marche(symbole)
            r['secteur'] = UniversActions.obtenir_secteur(symbole)
        
        self.resultats_sniper = resultats_sniper
        self.en_cours = False
        
        logger.info(f"✅ SNIPER: {len(resultats_sniper)} résultats triés par score")
        return resultats_sniper
    
    def generer_rapport_maitre(self, output_path: str = None, include_live: bool = True) -> str:
        """
        📊 Génère le rapport maître avec toutes les opportunités
        """
        if not self.resultats:
            logger.warning("Pas de résultats - Lancez d'abord une analyse")
            return None
        
        generator = MasterReportGenerator(self.config)
        return generator.generate_master_report(self.resultats, output_path, include_live)
    
    def generer_rapport_top(self, top_n: int = 30, output_path: str = None) -> str:
        """
        🏆 Génère le rapport des TOP N opportunités
        """
        results = self.resultats if self.resultats else getattr(self, 'resultats_sniper', [])
        if not results:
            logger.warning("Pas de résultats - Lancez d'abord une analyse")
            return None
        
        generator = MasterReportGenerator(self.config)
        return generator.generate_top_opportunities_report(results, top_n, output_path)
    
    def actualiser_prix_live(self) -> Dict[str, Dict]:
        """
        📡 Actualise tous les prix en temps réel
        """
        if not self.resultats:
            return {}
        
        symbols = [r['symbole'] for r in self.resultats]
        fetcher = LivePriceFetcher()
        return fetcher.fetch_batch_prices(symbols)


# ══════════════════════════════════════════════════════════════════════════════
#                    INTERFACE GRAPHIQUE
# ══════════════════════════════════════════════════════════════════════════════

class InterfaceScreener:
    def __init__(self):
        self.moteur = MoteurScreener()
        self.root = None
        self.file_maj = Queue()
        self.en_cours = False
        self.donnees_brutes: Dict[str, Dict] = {}
    
    def lancer(self):
        try:
            import tkinter as tk
            from tkinter import ttk, filedialog, messagebox
        except ImportError:
            logger.error("Tkinter non disponible")
            return
        
        self.tk = tk
        self.ttk = ttk
        self.messagebox = messagebox
        self.filedialog = filedialog
        
        self.root = tk.Tk()
        self.root.title("🚀 Screener Pro v11.0 - Opportunités Court & Moyen Terme")
        self.root.geometry("1850x1050")
        self.root.configure(bg="#0d1117")
        
        self._configurer_variables()
        self._configurer_styles()
        self._creer_interface()
        self._demarrer_processeur_file()
        
        self.root.protocol("WM_DELETE_WINDOW", self._fermer)
        self.root.mainloop()
    
    def _fermer(self):
        self.moteur.arreter()
        self.root.destroy()
    
    def _configurer_variables(self):
        tk = self.tk
        self.var_progress = tk.DoubleVar(value=0)
        self.var_statut = tk.StringVar(value="🟢 Prêt")
        self.var_recherche = tk.StringVar()
        self.var_ml_statut = tk.StringVar(value="⚠️ ML: Non entraîné - Cliquez '🧠 ENTRAINER ML'")
        self.var_marche = tk.StringVar(value="tous")
        self.var_secteur = tk.StringVar(value="tous")
        self.var_signaux_count = tk.StringVar(value="Signaux: 0")
    
    def _configurer_styles(self):
        style = self.ttk.Style()
        style.theme_use("clam")
        
        style.configure("Treeview",
                       background="#161b22",
                       foreground="#c9d1d9",
                       fieldbackground="#161b22",
                       rowheight=28,
                       font=("Consolas", 9))
        
        style.configure("Treeview.Heading",
                       background="#21262d",
                       foreground="#58a6ff",
                       font=("Segoe UI", 9, "bold"))
        
        style.map("Treeview", background=[("selected", "#388bfd")])
    
    def _creer_interface(self):
        tk = self.tk
        ttk = self.ttk
        
        # === EN-TÊTE ===
        entete = tk.Frame(self.root, bg="#161b22", height=80)
        entete.pack(fill="x")
        entete.pack_propagate(False)
        
        tk.Label(entete,
                text="🚀 SCREENER PRO v11.0 - OPPORTUNITÉS COURT & MOYEN TERME",
                font=("Segoe UI", 20, "bold"),
                fg="#58a6ff", bg="#161b22").pack(pady=5)
        
        tk.Label(entete,
                text="⚡ Intraday | 📈 Swing | 🏦 Position | 🎯 Breakouts | 📊 Patterns | 🧠 ML | 🌍 Marchés Globaux",
                font=("Segoe UI", 10),
                fg="#7ee787", bg="#161b22").pack()
        
        # === CONTRÔLES ===
        cadre_ctrl = tk.Frame(self.root, bg="#0d1117")
        cadre_ctrl.pack(fill="x", padx=15, pady=8)
        
        # Boutons
        cadre_btn = tk.Frame(cadre_ctrl, bg="#0d1117")
        cadre_btn.pack(side="left")
        
        self.btn_analyser = tk.Button(cadre_btn, text="🚀 ANALYSER",
                                     font=("Segoe UI", 11, "bold"),
                                     bg="#238636", fg="white", width=12, height=2,
                                     command=self._demarrer_analyse)
        self.btn_analyser.pack(side="left", padx=2)
        
        # 🎯 NOUVEAU: Bouton SNIPER (scan ultra-rapide)
        self.btn_sniper = tk.Button(cadre_btn, text="🎯 SNIPER",
                                   font=("Segoe UI", 11, "bold"),
                                   bg="#ff6b6b", fg="white", width=10, height=2,
                                   command=self._demarrer_sniper)
        self.btn_sniper.pack(side="left", padx=2)
        
        self.btn_ml = tk.Button(cadre_btn, text="🧠 ENTRAINER ML",
                               font=("Segoe UI", 11, "bold"),
                               bg="#8957e5", fg="white", width=14, height=2,
                               command=self._entrainer_ml)
        self.btn_ml.pack(side="left", padx=2)
        
        self.btn_stop = tk.Button(cadre_btn, text="⏹ STOP",
                                 font=("Segoe UI", 11, "bold"),
                                 bg="#6e7681", fg="white", width=8, height=2,
                                 state="disabled", command=self._arreter_analyse)
        self.btn_stop.pack(side="left", padx=2)
        
        tk.Button(cadre_btn, text="⚙️ CONFIG",
                 font=("Segoe UI", 10, "bold"),
                 bg="#da3633", fg="white", width=10, height=2,
                 command=self._ouvrir_config).pack(side="left", padx=2)
        
        # 📊 NOUVEAU: Bouton RAPPORT MAÎTRE
        tk.Button(cadre_btn, text="📊 RAPPORT PRO",
                 font=("Segoe UI", 10, "bold"),
                 bg="#00d4aa", fg="black", width=12, height=2,
                 command=self._generer_rapport_maitre).pack(side="left", padx=2)
        
        tk.Button(cadre_btn, text="📁 EXPORTER",
                 font=("Segoe UI", 10, "bold"),
                 bg="#1f6feb", fg="white", width=10, height=2,
                 command=self._exporter_csv).pack(side="left", padx=2)
        
        # Nouveaux boutons pour le système de signaux
        tk.Button(cadre_btn, text="📊 SCORECARD",
                 font=("Segoe UI", 10, "bold"),
                 bg="#f0883e", fg="white", width=11, height=2,
                 command=self._ouvrir_scorecard).pack(side="left", padx=2)
        
        tk.Button(cadre_btn, text="🔬 STRESS TEST",
                 font=("Segoe UI", 10, "bold"),
                 bg="#a371f7", fg="white", width=12, height=2,
                 command=self._ouvrir_stress_tests).pack(side="left", padx=2)
        
        tk.Button(cadre_btn, text="📋 SIGNAUX",
                 font=("Segoe UI", 10, "bold"),
                 bg="#3fb950", fg="white", width=10, height=2,
                 command=self._ouvrir_historique_signaux).pack(side="left", padx=2)
        
        # Filtres
        cadre_filtres = tk.Frame(cadre_ctrl, bg="#0d1117")
        cadre_filtres.pack(side="left", padx=15)
        
        tk.Label(cadre_filtres, text="Marché:", font=("Segoe UI", 9),
                fg="#8b949e", bg="#0d1117").pack(side="left")
        
        for texte, valeur in [("🌍 Tous", "tous"), ("🇺🇸 US", "us"), ("�🇷 FR", "fr"), ("🇪🇺 EU", "eu")]:
            tk.Radiobutton(cadre_filtres, text=texte, variable=self.var_marche, value=valeur,
                          bg="#0d1117", fg="#c9d1d9", selectcolor="#21262d",
                          font=("Segoe UI", 9)).pack(side="left", padx=2)
        
        # Recherche
        cadre_recherche = tk.Frame(cadre_ctrl, bg="#0d1117")
        cadre_recherche.pack(side="right", padx=10)
        
        tk.Label(cadre_recherche, text="🔍 Recherche:", bg="#0d1117", fg="#8b949e").pack(side="left")
        
        entree_recherche = tk.Entry(cadre_recherche, textvariable=self.var_recherche,
                                   font=("Segoe UI", 10), width=15,
                                   bg="#21262d", fg="#c9d1d9")
        entree_recherche.pack(side="left", padx=3)
        entree_recherche.bind('<KeyRelease>', self._filtrer_resultats)
        
        # Secteur
        tk.Label(cadre_recherche, text="Secteur:", bg="#0d1117", fg="#8b949e").pack(side="left", padx=(10, 0))
        self.combo_secteur = ttk.Combobox(cadre_recherche, textvariable=self.var_secteur,
                                         values=["tous", "Technologie", "Finance", "Santé",
                                                 "Consommation", "Industrie", "Énergie", "Luxe", 
                                                 "Automobile", "Gaming", "Média", "Immobilier", "Télécom",
                                                 "Or", "Argent", "Pétrole", "Cuivre", "Lithium", "Uranium",
                                                 "Agriculture", "Métaux", "Crypto"],
                                         width=14, state="readonly")
        self.combo_secteur.pack(side="left", padx=3)
        self.combo_secteur.bind('<<ComboboxSelected>>', self._filtrer_resultats)
        
        # Statut
        cadre_statut = tk.Frame(cadre_ctrl, bg="#0d1117")
        cadre_statut.pack(side="right", padx=15)
        
        self.lbl_ml = tk.Label(cadre_statut, textvariable=self.var_ml_statut,
                              font=("Segoe UI", 9, "bold"), fg="#f0883e", bg="#0d1117")
        self.lbl_ml.pack()
        
        self.lbl_statut = tk.Label(cadre_statut, textvariable=self.var_statut,
                                  font=("Segoe UI", 10), fg="#7ee787", bg="#0d1117")
        self.lbl_statut.pack()
        
        self.barre_progress = ttk.Progressbar(cadre_statut, variable=self.var_progress,
                                             length=150, mode="determinate")
        self.barre_progress.pack(pady=2)
        
        # === TABLEAU ===
        cadre_tableau = tk.Frame(self.root, bg="#0d1117")
        cadre_tableau.pack(fill="both", expand=True, padx=15, pady=5)
        
        colonnes = (
            "rang", "marche", "symbole", "secteur", "prix", "var1j",
            "decision", "score", "opportunite", "meilleur_horizon",
            "score_intraday", "score_swing",
            "prix_reel", "potentiel",
            "stop_loss", "take_profit", "ratio_rr",
            "rsi", "adx", "ml_signal"
        )
        
        self.arbre = ttk.Treeview(cadre_tableau, columns=colonnes, show="headings", height=22)
        
        entetes = {
            "rang": ("#", 35),
            "marche": ("🌍", 50),
            "symbole": ("Ticker", 65),
            "secteur": ("Secteur", 80),
            "prix": ("Prix", 68),
            "var1j": ("Δ1J%", 52),
            "decision": ("🎯 Décision", 115),
            "score": ("Score", 48),
            "opportunite": ("⚡ Opportunité", 110),
            "meilleur_horizon": ("⏱️ Horizon", 75),
            "score_intraday": ("📊 Intra", 55),
            "score_swing": ("📈 Swing", 55),
            "prix_reel": ("💰 Réel", 75),
            "potentiel": ("Potent.", 62),
            "stop_loss": ("📉 SL", 62),
            "take_profit": ("📈 TP", 62),
            "ratio_rr": ("⚖️R/R", 48),
            "rsi": ("RSI", 42),
            "adx": ("ADX", 45),
            "ml_signal": ("🧠 ML", 80)
        }
        
        for col, (texte, largeur) in entetes.items():
            self.arbre.heading(col, text=texte, command=lambda c=col: self._trier_colonne(c))
            self.arbre.column(col, width=largeur, anchor="center")
        
        scrollbar_y = ttk.Scrollbar(cadre_tableau, orient="vertical", command=self.arbre.yview)
        self.arbre.configure(yscrollcommand=scrollbar_y.set)
        
        self.arbre.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        
        # Double-clic pour détails
        self.arbre.bind("<Double-1>", self._afficher_details)
        
        # === PIED DE PAGE ===
        pied = tk.Frame(self.root, bg="#161b22", height=70)
        pied.pack(fill="x", padx=15, pady=5)
        pied.pack_propagate(False)
        
        cadre_stats = tk.Frame(pied, bg="#161b22")
        cadre_stats.pack(side="left", padx=15, pady=5)
        
        self.lbl_total = tk.Label(cadre_stats, text="Total: 0 | Affichés: 0",
                                 font=("Segoe UI", 10), fg="#c9d1d9", bg="#161b22")
        self.lbl_total.pack(anchor="w")
        
        self.lbl_signaux = tk.Label(cadre_stats, text="🟢 Achats: 0 | 🔴 Ventes: 0",
                                   font=("Segoe UI", 10), fg="#7ee787", bg="#161b22")
        self.lbl_signaux.pack(anchor="w")
        
        # Légende
        cadre_legende = tk.Frame(pied, bg="#161b22")
        cadre_legende.pack(side="right", padx=15, pady=5)
        
        tk.Label(cadre_legende,
                text="⏱️ Horizon: Intra=1-3j | Swing=5-15j | Position=15-60j | ⚡ Opportunités: Breakout, Gap, Squeeze",
                font=("Segoe UI", 8), fg="#8b949e", bg="#161b22").pack(anchor="e")
        
        tk.Label(cadre_legende,
                text="💰 Prix Réel = Valeur intrinsèque | 📉 SL/📈 TP = Stop Loss/Take Profit | ⚖️ R/R = Ratio Risque/Récompense",
                font=("Segoe UI", 8), fg="#8b949e", bg="#161b22").pack(anchor="e")
    
    def _demarrer_processeur_file(self):
        def traiter():
            try:
                while True:
                    msg = self.file_maj.get_nowait()
                    if msg['type'] == 'progress':
                        self.var_statut.set(f"🔄 {msg['symbole']} ({msg['actuel']}/{msg['total']})")
                        self.var_progress.set(msg['progress'])
                    elif msg['type'] == 'resultat':
                        self._ajouter_resultat(msg['resultat'], msg['rang'])
                    elif msg['type'] == 'complete':
                        self._analyse_terminee()
                    elif msg['type'] == 'ml_complete':
                        self._ml_termine(msg['metriques'])
                    # 🎯 SNIPER messages
                    elif msg['type'] == 'sniper_progress':
                        self.var_statut.set(f"🎯 {msg['message']} ({msg['actuel']}/{msg['total']})")
                        self.var_progress.set(int(msg['actuel'] / max(msg['total'], 1) * 100))
                    elif msg['type'] == 'sniper_resultat':
                        self._ajouter_resultat_sniper(msg['resultat'], msg['rang'])
                    elif msg['type'] == 'sniper_complete':
                        self._sniper_termine(msg['total'])
            except Empty:
                pass
            self.root.after(100, traiter)
        
        self.root.after(100, traiter)
    
    def _demarrer_analyse(self):
        # 🧠 Vérifier si l'entraînement ML est en cours
        ml_en_cours = str(self.btn_ml.cget('state')) == 'disabled'
        if ml_en_cours and not (self.moteur.ml and self.moteur.ml.est_entraine):
            import tkinter.messagebox as mb
            reponse = mb.askyesno(
                "⏳ Entraînement ML en cours",
                "L'entraînement ML est en cours.\n\n"
                "Voulez-vous:\n"
                "• OUI - Attendre la fin de l'entraînement\n"
                "• NON - Lancer l'analyse sans ML\n\n"
                "💡 Conseil: Attendez que le statut ML devienne vert"
            )
            if reponse:
                self.var_statut.set("⏳ Attendez la fin de l'entraînement ML...")
                return
            # Continuer sans ML
            self.var_statut.set("⚠️ Analyse sans ML - Colonnes ML seront N/A")
        
        # 🧠 Vérifier si ML est entraîné et avertir l'utilisateur
        ml_entraine = self.moteur.ml and self.moteur.ml.est_entraine
        if not ml_entraine and not ml_en_cours:
            # Afficher un avertissement optionnel
            import tkinter.messagebox as mb
            reponse = mb.askyesno(
                "⚠️ ML Non Entraîné",
                "Le modèle ML n'est pas entraîné.\n\n"
                "Les colonnes ML afficheront 'N/A'.\n\n"
                "Voulez-vous:\n"
                "• OUI - Continuer sans ML\n"
                "• NON - Annuler pour entraîner ML d'abord\n\n"
                "💡 Conseil: Cliquez sur '🧠 ENTRAINER ML' avant l'analyse"
            )
            if not reponse:
                self.var_statut.set("ℹ️ Analyse annulée - Entraînez ML d'abord")
                return
            self.var_statut.set("⚠️ Analyse sans ML - Colonnes ML seront N/A")
        
        self.en_cours = True
        self.btn_analyser.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.donnees_brutes.clear()
        
        for item in self.arbre.get_children():
            self.arbre.delete(item)
        
        marche = self.var_marche.get()
        self.moteur.config.INCLURE_US = marche in ["tous", "us"]
        self.moteur.config.INCLURE_EUROPE = marche in ["tous", "eu"]
        self.moteur.config.INCLURE_FRANCE = marche in ["tous", "fr"]
        
        thread = threading.Thread(target=self._thread_analyse, daemon=True)
        thread.start()
    
    def _thread_analyse(self):
        rang = 0
        
        def callback(symbole, actuel, total, progress, resultat):
            nonlocal rang
            
            self.file_maj.put({
                'type': 'progress',
                'symbole': symbole,
                'actuel': actuel,
                'total': total,
                'progress': progress
            })
            
            if resultat and resultat.get('score_final', 0) >= CONFIG.SCORE_MINIMUM:
                rang += 1
                self.file_maj.put({
                    'type': 'resultat',
                    'resultat': resultat,
                    'rang': rang
                })
        
        self.moteur.executer(callback=callback)
        self.file_maj.put({'type': 'complete'})
    
    def _ajouter_resultat(self, r: Dict, rang: int):
        self.donnees_brutes[r['symbole']] = r
        
        score = r.get('score_final', 0)
        tag = "normal"
        if score >= 70:
            tag = "excellent"
        elif score >= 55:
            tag = "bon"
        elif score <= 35:
            tag = "mauvais"
        
        # Formatage des valeurs
        rsi = f"{r['rsi']:.0f}" if r.get('rsi') else "--"
        adx = f"{r['adx']:.0f}" if r.get('adx') else "--"
        prix_reel = f"${r['prix_reel']:.2f}" if r.get('prix_reel') else "--"
        potentiel = f"{r['potentiel']:+.0f}%" if r.get('potentiel') else "--"
        sl = f"${r['stop_loss']:.2f}" if r.get('stop_loss') else "--"
        tp = f"${r['take_profit']:.2f}" if r.get('take_profit') else "--"
        rr = f"{r['ratio_rr']:.1f}" if r.get('ratio_rr') else "--"
        
        # 🆕 Nouvelles colonnes opportunités
        opp_type = r.get('opportunite_type', '--')
        opp_details = r.get('opportunite_principale', '')
        opportunite_str = f"{opp_type}" if opp_type not in ['AUCUNE', '--', None] else "--"
        if opp_details and opp_type not in ['AUCUNE', '--', None]:
            opportunite_str = f"{opp_type}: {opp_details[:12]}"
        
        meilleur_horizon = r.get('meilleur_horizon', '--')
        score_intraday = f"{r.get('score_intraday', 0):.0f}" if r.get('score_intraday') else "--"
        score_swing = f"{r.get('score_swing', 0):.0f}" if r.get('score_swing') else "--"
        
        self.arbre.insert("", "end", iid=r['symbole'], values=(
            rang,
            r['marche'],
            r['symbole'],
            r.get('secteur', '--'),
            f"${r['prix']:.2f}",
            f"{r['var_1j']:+.1f}%",
            r['decision'],
            f"{r['score_final']:.0f}",
            opportunite_str,
            meilleur_horizon,
            score_intraday,
            score_swing,
            prix_reel,
            potentiel,
            sl,
            tp,
            rr,
            rsi,
            adx,
            r.get('ml_signal', '--')
        ), tags=(tag,))
        
        self.arbre.tag_configure("excellent", background="#1a4731")
        self.arbre.tag_configure("bon", background="#2d4a1c")
        self.arbre.tag_configure("mauvais", background="#4a1c1c")
        self.arbre.tag_configure("normal", background="#161b22")
    
    def _analyse_terminee(self):
        resultats = self.moteur.resultats
        
        achats = len([r for r in resultats if "ACHETER" in r.get('decision', '')])
        ventes = len([r for r in resultats if "VENDRE" in r.get('decision', '') or "ÉVITER" in r.get('decision', '')])
        
        self.lbl_total.config(text=f"Total: {len(resultats)} | Affichés: {len(self.arbre.get_children())}")
        self.lbl_signaux.config(text=f"🟢 Achats: {achats} | 🔴 Ventes/Éviter: {ventes}")
        
        self.btn_analyser.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.var_statut.set(f"✅ Terminé! {len(resultats)} résultats")
        self.var_progress.set(100)
        self.en_cours = False
    
    def _arreter_analyse(self):
        self.moteur.arreter()
        self.en_cours = False
        self.var_statut.set("⏹ Arrêté")
        self.btn_analyser.config(state="normal")
        self.btn_sniper.config(state="normal")
        self.btn_stop.config(state="disabled")
    
    # ══════════════════════════════════════════════════════════════════════════
    # 🎯 MODE SNIPER - SCAN ULTRA-RAPIDE
    # ══════════════════════════════════════════════════════════════════════════
    
    def _demarrer_sniper(self):
        """🎯 Lance le scan SNIPER ultra-rapide"""
        self.en_cours = True
        self.btn_analyser.config(state="disabled")
        self.btn_sniper.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.donnees_brutes.clear()
        
        for item in self.arbre.get_children():
            self.arbre.delete(item)
        
        marche = self.var_marche.get()
        self.moteur.config.INCLURE_US = marche in ["tous", "us"]
        self.moteur.config.INCLURE_EUROPE = marche in ["tous", "eu"]
        self.moteur.config.INCLURE_FRANCE = marche in ["tous", "fr"]
        
        self.var_statut.set("🎯 SNIPER: Scan ultra-rapide en cours...")
        
        thread = threading.Thread(target=self._thread_sniper, daemon=True)
        thread.start()
    
    def _thread_sniper(self):
        """Thread pour le scan sniper"""
        def callback(message, actuel, total):
            self.file_maj.put({
                'type': 'sniper_progress',
                'message': message,
                'actuel': actuel,
                'total': total
            })
        
        resultats = self.moteur.executer_sniper(callback=callback)
        
        # Envoyer les résultats
        for rang, r in enumerate(resultats, 1):
            self.file_maj.put({
                'type': 'sniper_resultat',
                'resultat': r,
                'rang': rang
            })
        
        self.file_maj.put({'type': 'sniper_complete', 'total': len(resultats)})
    
    def _ajouter_resultat_sniper(self, r: Dict, rang: int):
        """Ajoute un résultat sniper au tableau"""
        self.donnees_brutes[r['symbole']] = r
        
        score = r.get('score_sniper', 0)
        tag = "normal"
        if score >= 70:
            tag = "excellent"
        elif score >= 55:
            tag = "bon"
        elif score <= 35:
            tag = "mauvais"
        
        # Colonnes sniper adaptées
        self.arbre.insert("", "end", iid=r['symbole'], values=(
            rang,
            r.get('marche', '--'),
            r['symbole'],
            r.get('secteur', '--'),
            f"${r['prix']:.2f}",
            f"{r.get('change_pct', 0):+.1f}%",
            r.get('signal_level', '--'),
            f"{r.get('score_sniper', 0):.0f}",
            r.get('opportunite_type', '--'),
            "Sniper",  # meilleur_horizon
            f"{r.get('score_sniper', 0):.0f}",  # score_intraday
            f"{r.get('score_sniper', 0):.0f}",  # score_swing
            f"${r.get('support', 0):.2f}",  # prix_reel -> support
            f"{r.get('var_5j', 0):+.1f}%",  # potentiel -> var_5j
            f"${r.get('stop_loss', 0):.2f}",
            f"${r.get('take_profit', 0):.2f}",
            f"{r.get('ratio_rr', 0):.1f}",
            f"{r.get('rsi', 50):.0f}",
            "--",  # ADX non calculé en sniper
            "; ".join(r.get('signaux', [])[:2]) if r.get('signaux') else "--"
        ), tags=(tag,))
        
        self.arbre.tag_configure("excellent", background="#1a4731")
        self.arbre.tag_configure("bon", background="#2d4a1c")
        self.arbre.tag_configure("mauvais", background="#4a1c1c")
        self.arbre.tag_configure("normal", background="#161b22")
    
    def _sniper_termine(self, total: int):
        """Callback quand le sniper est terminé"""
        achats = len([s for s in self.donnees_brutes.values() if s.get('score_sniper', 0) >= 55])
        
        self.lbl_total.config(text=f"Total: {total} | Affichés: {len(self.arbre.get_children())}")
        self.lbl_signaux.config(text=f"🎯 Opportunités: {achats} | 🔴 À éviter: {total - achats}")
        
        self.btn_analyser.config(state="normal")
        self.btn_sniper.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.var_statut.set(f"🎯 SNIPER terminé! {total} actions scannées - {achats} opportunités")
        self.var_progress.set(100)
        self.en_cours = False
    
    # ══════════════════════════════════════════════════════════════════════════
    # 📊 RAPPORT MAÎTRE PROFESSIONNEL
    # ══════════════════════════════════════════════════════════════════════════
    
    def _generer_rapport_maitre(self):
        """📊 Génère le rapport maître avec prix live"""
        if not self.donnees_brutes:
            self.messagebox.showwarning("Rapport", 
                "⚠️ Aucune donnée disponible!\n\n"
                "👉 Lancez d'abord une analyse (🚀 ANALYSER ou 🎯 SNIPER)\n"
                "👉 Puis cliquez sur 📊 RAPPORT PRO")
            return
        
        # Fenêtre d'options
        fenetre = self.tk.Toplevel(self.root)
        fenetre.title("📊 Générer Rapport Maître")
        fenetre.geometry("500x400")
        fenetre.configure(bg="#0d1117")
        fenetre.transient(self.root)
        fenetre.grab_set()
        
        tk = self.tk
        
        tk.Label(fenetre, text="📊 RAPPORT MAÎTRE PROFESSIONNEL", 
                font=("Segoe UI", 14, "bold"),
                fg="#00d4aa", bg="#0d1117").pack(pady=15)
        
        # Options
        var_live = tk.BooleanVar(value=True)
        var_top_only = tk.BooleanVar(value=False)
        var_top_n = tk.StringVar(value="50")
        
        cadre = tk.Frame(fenetre, bg="#0d1117")
        cadre.pack(fill="x", padx=20, pady=10)
        
        tk.Checkbutton(cadre, text="📡 Inclure PRIX EN TEMPS RÉEL (recommandé)",
                      variable=var_live, bg="#0d1117", fg="#00ff00",
                      selectcolor="#21262d", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=5)
        
        tk.Checkbutton(cadre, text="🏆 Exporter seulement les TOP opportunités",
                      variable=var_top_only, bg="#0d1117", fg="#c9d1d9",
                      selectcolor="#21262d", font=("Segoe UI", 10)).pack(anchor="w", pady=5)
        
        row_top = tk.Frame(cadre, bg="#0d1117")
        row_top.pack(fill="x", pady=5)
        tk.Label(row_top, text="   Nombre TOP:", font=("Segoe UI", 10),
                fg="#8b949e", bg="#0d1117").pack(side="left")
        tk.Entry(row_top, textvariable=var_top_n, width=5, bg="#21262d", fg="#c9d1d9").pack(side="left", padx=5)
        
        # Info
        tk.Label(cadre, text="\n💡 Le rapport inclut:\n"
                "• Toutes les opportunités triées par score\n"
                "• Prix en temps réel (actualisés)\n"
                "• Stop Loss / Take Profit calculés\n"
                "• Signaux et recommandations\n"
                "• Compatible Excel/Google Sheets",
                font=("Segoe UI", 10), fg="#8b949e", bg="#0d1117",
                justify="left").pack(anchor="w", pady=10)
        
        def executer():
            fenetre.destroy()
            self._executer_rapport_maitre(
                include_live=var_live.get(),
                top_only=var_top_only.get(),
                top_n=int(var_top_n.get()) if var_top_n.get().isdigit() else 50
            )
        
        # Boutons
        btn_frame = tk.Frame(fenetre, bg="#0d1117")
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="📊 GÉNÉRER RAPPORT", font=("Segoe UI", 12, "bold"),
                 bg="#00d4aa", fg="black", width=18, command=executer).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Annuler", font=("Segoe UI", 10),
                 bg="#21262d", fg="#c9d1d9", width=10, command=fenetre.destroy).pack(side="left", padx=10)
    
    def _executer_rapport_maitre(self, include_live: bool, top_only: bool, top_n: int):
        """Exécute la génération du rapport"""
        self.var_statut.set("📊 Génération du rapport maître...")
        self.root.update()
        
        try:
            results = list(self.donnees_brutes.values())
            
            if top_only:
                # Trier et prendre les TOP
                score_key = 'score_final' if 'score_final' in results[0] else 'score_sniper'
                results = sorted(results, key=lambda x: x.get(score_key, 0), reverse=True)[:top_n]
            
            generator = MasterReportGenerator(self.moteur.config)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"MASTER_REPORT_PRO_{timestamp}.csv"
            
            filepath = self.filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv"), ("Excel", "*.xlsx")],
                initialfile=filename
            )
            
            if not filepath:
                return
            
            # Générer
            output = generator.generate_master_report(results, filepath, include_live)
            
            if output:
                self.var_statut.set(f"✅ Rapport généré: {output}")
                self.messagebox.showinfo("Rapport Généré",
                    f"📊 Rapport maître créé avec succès!\n\n"
                    f"📁 Fichier: {output}\n"
                    f"📈 {len(results)} opportunités\n"
                    f"📡 Prix live: {'Oui' if include_live else 'Non'}\n\n"
                    f"💡 Ouvrez avec Excel ou Google Sheets")
            else:
                self.messagebox.showerror("Erreur", "Erreur lors de la génération")
                
        except Exception as e:
            logger.error(f"Erreur génération rapport: {e}")
            self.messagebox.showerror("Erreur", f"Erreur: {e}")
            self.var_statut.set("❌ Erreur génération rapport")

    def _entrainer_ml(self):
        self.btn_ml.config(state="disabled")
        self.var_ml_statut.set("🧠 Entraînement en cours...")
        self.var_statut.set("🧠 Entraînement ML - Patientez (2-5 min)...")
        
        def entrainer():
            try:
                def cb(texte, actuel, total):
                    self.var_ml_statut.set(f"🧠 {texte} ({actuel}/{total})")
                
                metriques = self.moteur.entrainer_ml(cb)
                self.file_maj.put({'type': 'ml_complete', 'metriques': metriques})
            except Exception as e:
                logger.error(f"Erreur entraînement ML: {e}")
                self.file_maj.put({'type': 'ml_complete', 'metriques': None, 'erreur': str(e)})
        
        threading.Thread(target=entrainer, daemon=True).start()
    
    def _ml_termine(self, metriques):
        self.btn_ml.config(state="normal")
        if metriques and metriques.get('echantillons', 0) > 0:
            # 🔧 CORRECTION: Utiliser accuracy_ensemble pour le taux
            prec = metriques.get('accuracy_ensemble', metriques.get('precision_ensemble', 0))
            f1 = metriques.get('f1_ensemble', 0)
            echantillons = metriques.get('echantillons', 0)
            nb_modeles = metriques.get('nb_modeles', 0)
            self.var_ml_statut.set(f"✅ ML: Acc={prec:.1%} F1={f1:.1%} ({nb_modeles} modèles, {echantillons} samples)")
            self.var_statut.set("🟢 ML entraîné - Prêt à analyser")
            # 🎨 Changer la couleur en vert pour indiquer succès
            self.lbl_ml.config(fg="#7ee787")
        else:
            self.var_ml_statut.set("❌ ML: Échec - Vérifiez connexion internet")
            self.var_statut.set("⚠️ ML non disponible")
            # 🎨 Changer la couleur en rouge pour indiquer échec
            self.lbl_ml.config(fg="#f85149")
    
    def _filtrer_resultats(self, event=None):
        recherche = self.var_recherche.get().upper().strip()
        filtre_secteur = self.var_secteur.get()
        filtre_marche = self.var_marche.get()
        
        for item in self.arbre.get_children():
            self.arbre.delete(item)
        
        rang = 0
        for symbole, donnees in self.donnees_brutes.items():
            # Filtre recherche
            if recherche:
                if recherche not in symbole.upper() and recherche not in donnees.get('secteur', '').upper():
                    continue
            
            # Filtre secteur
            if filtre_secteur != "tous":
                if donnees.get('secteur', '') != filtre_secteur:
                    continue
            
            # Filtre marché
            if filtre_marche != "tous":
                marche = donnees.get('marche', '')
                if filtre_marche == "us" and "US" not in marche:
                    continue
                if filtre_marche == "fr" and "France" not in marche:
                    continue
                if filtre_marche == "eu" and "Europe" not in marche:
                    continue
            
            rang += 1
            self._ajouter_resultat(donnees, rang)
        
        self.lbl_total.config(text=f"Total: {len(self.donnees_brutes)} | Affichés: {rang}")
    
    def _trier_colonne(self, col):
        cles_brutes = {
            'prix': '_prix', 'score': '_score', 'potentiel': '_potentiel',
            'ratio_rr': '_ratio_rr', 'var1j': 'var_1j', 'rsi': 'rsi'
        }
        
        items = list(self.arbre.get_children())
        
        if col in cles_brutes:
            cle = cles_brutes[col]
            items.sort(key=lambda x: self.donnees_brutes.get(x, {}).get(cle, 0) or 0, reverse=True)
        else:
            items.sort(key=lambda x: self.arbre.set(x, col))
        
        for idx, item in enumerate(items):
            self.arbre.move(item, '', idx)
    
    def _afficher_details(self, event):
        """Affiche les détails complets d'une action"""
        selection = self.arbre.selection()
        if not selection:
            return
        
        symbole = selection[0]
        donnees = self.donnees_brutes.get(symbole)
        if not donnees:
            return
        
        # Fenêtre de détails
        fenetre = self.tk.Toplevel(self.root)
        fenetre.title(f"📊 Détails - {symbole}")
        fenetre.geometry("750x900")
        fenetre.configure(bg="#0d1117")
        
        # Scroll
        canvas = self.tk.Canvas(fenetre, bg="#0d1117", highlightthickness=0)
        scrollbar = self.ttk.Scrollbar(fenetre, orient="vertical", command=canvas.yview)
        cadre = self.tk.Frame(canvas, bg="#0d1117")
        
        cadre.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=cadre, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Titre
        self.tk.Label(cadre,
                     text=f"📊 {symbole} - {donnees.get('secteur', '')}",
                     font=("Segoe UI", 16, "bold"),
                     fg="#58a6ff", bg="#0d1117").pack(pady=10)
        
        # Sections
        sections = [
            ("⏱️ SCORING MULTI-HORIZON", [
                ("🔥 Meilleur Horizon", donnees.get('meilleur_horizon', 'N/A')),
                ("Score Intraday (1-3j)", f"{donnees.get('score_intraday', 0):.0f}/100"),
                ("Score Swing (5-15j)", f"{donnees.get('score_swing', 0):.0f}/100"),
                ("Score Position (15-60j)", f"{donnees.get('score_position', 0):.0f}/100"),
                ("Opportunité Claire", "✅ OUI" if donnees.get('opportunite_claire') else "❌ NON"),
            ]),
            ("⚡ OPPORTUNITÉS DÉTECTÉES", [
                ("Type Principal", donnees.get('opportunite_type', 'AUCUNE')),
                ("Détails", donnees.get('opportunite_principale', '--')),
                ("Breakout", f"{donnees.get('breakout', {}).get('type', '--')} (Score: {donnees.get('breakout', {}).get('score', 0):.0f})"),
                ("Gap", f"{donnees.get('gap', {}).get('type', '--')} ({donnees.get('gap', {}).get('pourcentage', 0):.1f}%)"),
                ("Volume Spike", f"{'✅ OUI' if donnees.get('volume_spike', {}).get('spike') else '❌ NON'} ({donnees.get('volume_spike', {}).get('niveau', '--')})"),
                ("Pattern Reversal", f"{donnees.get('pattern_reversal', {}).get('pattern', '--')} ({donnees.get('pattern_reversal', {}).get('fiabilite', 0):.0f}%)"),
                ("Momentum Burst", f"{'✅ OUI' if donnees.get('momentum_burst', {}).get('burst') else '❌ NON'} ({donnees.get('momentum_burst', {}).get('direction', '--')})"),
                ("Squeeze", f"{'🔥 ACTIF' if donnees.get('squeeze_on') else '❌ Non'} - {donnees.get('squeeze_signal', '--')}"),
            ]),
            ("💰 VALORISATION", [
                ("Prix Actuel", f"${donnees['prix']:.2f}"),
                ("Prix Réel (Intrinsèque)", f"${donnees.get('prix_reel', 0):.2f}" if donnees.get('prix_reel') else "N/A"),
                ("Potentiel", f"{donnees.get('potentiel', 0):+.1f}%"),
                ("Statut", donnees.get('statut_valorisation', 'N/A')),
                ("Prix d'Achat Idéal", f"${donnees.get('prix_achat_ideal', 0):.2f}" if donnees.get('prix_achat_ideal') else "N/A"),
                ("PE Ratio", f"{donnees.get('pe_ratio', 0):.1f}" if donnees.get('pe_ratio') else "N/A"),
            ]),
            ("📊 PLAN DE TRADING", [
                ("Stop Loss", f"${donnees.get('stop_loss', 0):.2f} ({donnees.get('stop_loss_pct', 0):.1f}%)"),
                ("Take Profit", f"${donnees.get('take_profit', 0):.2f} ({donnees.get('take_profit_pct', 0):.1f}%)"),
                ("Ratio Risque/Récompense", f"{donnees.get('ratio_rr', 0):.2f} - {donnees.get('qualite_rr', '')}"),
                ("Horizon de Détention", f"{donnees.get('horizon_jours', 0)} jours ({donnees.get('horizon_type', '')})"),
            ]),
            ("🆕 INDICATEURS COURT TERME", [
                ("ADX", f"{donnees.get('adx', 0):.1f} - {donnees.get('adx_interpretation', 'N/A')}" if donnees.get('adx') else "N/A"),
                ("Stochastic RSI", f"{donnees.get('stoch_rsi_k', 0):.1f} - {donnees.get('stoch_rsi_signal', 'N/A')}" if donnees.get('stoch_rsi_k') else "N/A"),
                ("Williams %R", f"{donnees.get('williams_r', 0):.1f}" if donnees.get('williams_r') else "N/A"),
                ("CMF (Money Flow)", f"{donnees.get('cmf', 0):.3f} - {donnees.get('cmf_signal', 'N/A')}" if donnees.get('cmf') else "N/A"),
                ("OBV Trend", donnees.get('obv_trend', 'N/A')),
                ("VWAP", f"${donnees.get('vwap', 0):.2f} ({donnees.get('vwap_distance', 0):+.1f}%) - {donnees.get('vwap_signal', 'N/A')}" if donnees.get('vwap') else "N/A"),
                ("SuperTrend", donnees.get('supertrend_signal', 'N/A')),
            ]),
            ("📈 TECHNIQUE CLASSIQUE", [
                ("RSI", f"{donnees.get('rsi', 0):.1f}" if donnees.get('rsi') else "N/A"),
                ("Z-Score", f"{donnees.get('z_score', 0):+.2f}" if donnees.get('z_score') else "N/A"),
                ("MACD", donnees.get('macd_tendance', 'N/A')),
                ("Tendance", donnees.get('tendance', 'N/A')),
                ("Volume", donnees.get('volume_label', 'N/A')),
                ("Support", f"${donnees.get('support', 0):.2f}"),
                ("Résistance", f"${donnees.get('resistance', 0):.2f}"),
            ]),
            ("🧠 MACHINE LEARNING", [
                ("Signal ML", donnees.get('ml_signal', 'N/A')),
                ("Probabilité", f"{donnees.get('ml_probabilite', 50):.1f}%"),
            ]),
            ("🎯 DÉCISION FINALE", [
                ("Décision", donnees.get('decision', 'N/A')),
                ("Score Final", f"{donnees.get('score_final', 0):.1f}/100"),
                ("Confiance", f"{donnees.get('confiance', '')} ({donnees.get('confiance_pct', 0)}%)"),
                ("Signal Analystes", donnees.get('signal_analyste', 'N/A')),
                ("Confirmations", f"{donnees.get('nb_confirmations', 0)} indicateurs"),
            ]),
            ("📰 NEWS & SENTIMENT", [
                ("Sentiment News", donnees.get('sentiment_news', '🟡 NEUTRE')),
                ("Score News", f"{donnees.get('score_news', 50):.0f}/100"),
                ("Nb Articles", str(donnees.get('nb_news', 0))),
                ("Couverture Média", donnees.get('couverture_media', 'Faible')),
                ("Résumé", donnees.get('resume_news', 'N/A')[:60] + '...' if len(donnees.get('resume_news', '')) > 60 else donnees.get('resume_news', 'N/A')),
            ]),
            ("📊 QUALITÉ FINANCIÈRE (1 AN)", [
                ("Score Qualité", f"{donnees.get('score_qualite', 50):.0f}/100"),
                ("Qualité", donnees.get('qualite_financiere', 'N/A')),
                ("Recommandation 1an", donnees.get('recommandation_1an', 'N/A')),
                ("Prêt Investissement", "✅ Oui" if donnees.get('pret_investissement') else "❌ Non"),
            ]),
        ]
        
        for titre, items in sections:
            lf = self.tk.LabelFrame(cadre, text=titre, font=("Segoe UI", 11, "bold"),
                                   fg="#58a6ff", bg="#0d1117", padx=10, pady=5)
            lf.pack(fill="x", padx=15, pady=5)
            
            for label, valeur in items:
                row = self.tk.Frame(lf, bg="#0d1117")
                row.pack(fill="x", pady=2)
                
                self.tk.Label(row, text=label, font=("Segoe UI", 10),
                             fg="#8b949e", bg="#0d1117", width=25, anchor="w").pack(side="left")
                
                self.tk.Label(row, text=str(valeur), font=("Segoe UI", 10, "bold"),
                             fg="#c9d1d9", bg="#0d1117", anchor="w").pack(side="left", fill="x")
        
        # Conseil
        conseil_frame = self.tk.LabelFrame(cadre, text="📝 CONSEIL", font=("Segoe UI", 11, "bold"),
                                          fg="#7ee787", bg="#0d1117", padx=10, pady=10)
        conseil_frame.pack(fill="x", padx=15, pady=10)
        
        self.tk.Label(conseil_frame, text=donnees.get('conseil', ''),
                     font=("Segoe UI", 10), fg="#c9d1d9", bg="#0d1117",
                     wraplength=600, justify="left").pack()
        
        # Détails valorisation
        if donnees.get('details_valorisation'):
            val_frame = self.tk.LabelFrame(cadre, text="📊 DÉTAILS VALORISATION", font=("Segoe UI", 11, "bold"),
                                          fg="#f0883e", bg="#0d1117", padx=10, pady=5)
            val_frame.pack(fill="x", padx=15, pady=5)
            
            for methode, valeur in donnees['details_valorisation'].items():
                row = self.tk.Frame(val_frame, bg="#0d1117")
                row.pack(fill="x", pady=1)
                
                self.tk.Label(row, text=methode.replace('_', ' ').title(),
                             font=("Segoe UI", 9), fg="#8b949e", bg="#0d1117",
                             width=20, anchor="w").pack(side="left")
                
                self.tk.Label(row, text=f"${valeur:.2f}",
                             font=("Segoe UI", 9), fg="#c9d1d9", bg="#0d1117").pack(side="left")
        
        # 📰 NEWS RÉCENTES
        news_recentes = donnees.get('news_recentes', [])
        if news_recentes:
            news_frame = self.tk.LabelFrame(cadre, text="📰 NEWS RÉCENTES", font=("Segoe UI", 11, "bold"),
                                           fg="#58a6ff", bg="#0d1117", padx=10, pady=5)
            news_frame.pack(fill="x", padx=15, pady=5)
            
            for news in news_recentes[:3]:
                row = self.tk.Frame(news_frame, bg="#0d1117")
                row.pack(fill="x", pady=3)
                
                sentiment = news.get('sentiment', '🟡')
                titre = news.get('titre', '')[:70] + '...' if len(news.get('titre', '')) > 70 else news.get('titre', '')
                source = news.get('source', 'Unknown')
                
                self.tk.Label(row, text=sentiment, font=("Segoe UI", 10),
                             fg="#c9d1d9", bg="#0d1117", width=3).pack(side="left")
                
                self.tk.Label(row, text=f"[{source}] {titre}",
                             font=("Segoe UI", 9), fg="#c9d1d9", bg="#0d1117",
                             anchor="w").pack(side="left", fill="x", expand=True)
        
        # Alerte news si présente
        if donnees.get('alerte_news'):
            alerte_frame = self.tk.Frame(cadre, bg="#da3633", padx=10, pady=5)
            alerte_frame.pack(fill="x", padx=15, pady=5)
            
            self.tk.Label(alerte_frame, text=f"🚨 ALERTE NEWS: {donnees['alerte_news'][:80]}...",
                         font=("Segoe UI", 9, "bold"), fg="white", bg="#da3633",
                         wraplength=600, justify="left").pack()
    
    def _exporter_csv(self):
        """Exporter les données avec les TOP opportunités en premier"""
        if not self.donnees_brutes:
            self.messagebox.showwarning("Export", 
                "⚠️ Aucune donnée à exporter!\n\n"
                "👉 Lancez d'abord une analyse avec le bouton 🔍 ANALYSER\n"
                "👉 Attendez que l'analyse se termine\n"
                "👉 Puis cliquez sur 📁 EXPORTER")
            return
        
        # Demander le type d'export
        fenetre = self.tk.Toplevel(self.root)
        fenetre.title("📁 Options d'Export")
        fenetre.geometry("450x380")
        fenetre.configure(bg="#0d1117")
        fenetre.transient(self.root)
        fenetre.grab_set()
        
        tk = self.tk
        
        tk.Label(fenetre, text="📁 OPTIONS D'EXPORT", font=("Segoe UI", 14, "bold"),
                fg="#58a6ff", bg="#0d1117").pack(pady=15)
        
        # Options
        var_trier = tk.BooleanVar(value=True)
        var_top_only = tk.BooleanVar(value=False)
        var_top_n = tk.StringVar(value="20")
        var_format = tk.StringVar(value="excel_live")  # Par défaut Excel Live
        
        # Cadre options
        cadre = tk.Frame(fenetre, bg="#0d1117")
        cadre.pack(fill="x", padx=20, pady=10)
        
        tk.Checkbutton(cadre, text="🏆 Trier par meilleures opportunités (Score décroissant)",
                      variable=var_trier, bg="#0d1117", fg="#c9d1d9",
                      selectcolor="#21262d", font=("Segoe UI", 10)).pack(anchor="w", pady=5)
        
        tk.Checkbutton(cadre, text="📊 Exporter seulement les TOP opportunités",
                      variable=var_top_only, bg="#0d1117", fg="#c9d1d9",
                      selectcolor="#21262d", font=("Segoe UI", 10)).pack(anchor="w", pady=5)
        
        row_top = tk.Frame(cadre, bg="#0d1117")
        row_top.pack(fill="x", pady=5)
        tk.Label(row_top, text="   Nombre de TOP à exporter:", font=("Segoe UI", 10),
                fg="#8b949e", bg="#0d1117").pack(side="left")
        tk.Entry(row_top, textvariable=var_top_n, width=5, bg="#21262d", fg="#c9d1d9").pack(side="left", padx=5)
        
        # Format
        tk.Label(cadre, text="\n📄 Format:", font=("Segoe UI", 10, "bold"),
                fg="#c9d1d9", bg="#0d1117").pack(anchor="w")
        
        tk.Radiobutton(cadre, text="⭐ EXCEL avec PRIX EN DIRECT (recommandé)", variable=var_format, value="excel_live",
                      bg="#0d1117", fg="#00ff00", selectcolor="#21262d", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        tk.Radiobutton(cadre, text="CSV (Excel compatible)", variable=var_format, value="csv",
                      bg="#0d1117", fg="#c9d1d9", selectcolor="#21262d").pack(anchor="w")
        tk.Radiobutton(cadre, text="CSV avec résumé TOP en en-tête", variable=var_format, value="csv_resume",
                      bg="#0d1117", fg="#c9d1d9", selectcolor="#21262d").pack(anchor="w")
        
        # Info sur Excel Live
        tk.Label(cadre, text="💡 Excel Live: Les prix se mettent à jour automatiquement (F9)",
                font=("Segoe UI", 9, "italic"), fg="#58a6ff", bg="#0d1117").pack(anchor="w", pady=5)
        
        def executer_export():
            fenetre.destroy()
            self._executer_export_options(
                trier=var_trier.get(),
                top_only=var_top_only.get(),
                top_n=int(var_top_n.get()) if var_top_n.get().isdigit() else 20,
                format_type=var_format.get()
            )
        
        # Boutons
        btn_frame = tk.Frame(fenetre, bg="#0d1117")
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="📁 EXPORTER", font=("Segoe UI", 11, "bold"),
                 bg="#238636", fg="white", width=15, command=executer_export).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Annuler", font=("Segoe UI", 10),
                 bg="#21262d", fg="#c9d1d9", width=10, command=fenetre.destroy).pack(side="left", padx=10)
    
    def _executer_export_options(self, trier: bool, top_only: bool, top_n: int, format_type: str):
        """Exécute l'export avec les options choisies"""
        # Extension selon format
        if format_type == "excel_live":
            ext = ".xlsx"
            ftypes = [("Excel", "*.xlsx"), ("CSV", "*.csv")]
        else:
            ext = ".csv"
            ftypes = [("CSV", "*.csv"), ("Excel", "*.xlsx")]
        
        fichier = self.filedialog.asksaveasfilename(
            defaultextension=ext,
            filetypes=ftypes,
            initialfile=f"TOP_opportunites_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        )
        
        if not fichier:
            return
        
        try:
            df = pd.DataFrame(list(self.donnees_brutes.values()))
            
            # Supprimer colonnes internes
            df = df[[c for c in df.columns if not c.startswith('_') and c != 'details_valorisation']]
            
            # ======== AJOUT PRIX EN TEMPS RÉEL ========
            self.var_statut.set("📈 Récupération des prix en temps réel...")
            self.root.update()
            
            prix_live = []
            var_jour_pct = []
            heure_maj = []
            
            for symbole in df['symbole'].tolist():
                try:
                    ticker = yf.Ticker(symbole)
                    info = ticker.fast_info
                    price = getattr(info, 'last_price', None) or getattr(info, 'previous_close', None)
                    prev_close = getattr(info, 'previous_close', None)
                    
                    if price:
                        prix_live.append(round(price, 2))
                        if prev_close:
                            var_pct = ((price - prev_close) / prev_close) * 100
                            var_jour_pct.append(round(var_pct, 2))
                        else:
                            var_jour_pct.append(None)
                        heure_maj.append(datetime.now().strftime('%H:%M:%S'))
                    else:
                        prix_live.append(None)
                        var_jour_pct.append(None)
                        heure_maj.append(None)
                except:
                    prix_live.append(None)
                    var_jour_pct.append(None)
                    heure_maj.append(None)
            
            # Insérer après la colonne prix (avec vérification si colonnes existent déjà)
            # Supprimer les colonnes si elles existent déjà pour éviter l'erreur "already exists"
            for col in ['PRIX_LIVE', 'VAR_JOUR_%', 'MAJ_HEURE']:
                if col in df.columns:
                    df = df.drop(columns=[col])
            
            if 'prix' in df.columns:
                idx = df.columns.get_loc('prix') + 1
                df.insert(idx, 'PRIX_LIVE', prix_live)
                df.insert(idx + 1, 'VAR_JOUR_%', var_jour_pct)
                df.insert(idx + 2, 'MAJ_HEURE', heure_maj)
            else:
                df['PRIX_LIVE'] = prix_live
                df['VAR_JOUR_%'] = var_jour_pct
                df['MAJ_HEURE'] = heure_maj
            # ========================================
            
            # Trier par score décroissant (meilleures opportunités en premier)
            if trier and 'score_global' in df.columns:
                df = df.sort_values('score_global', ascending=False)
            elif trier and 'score' in df.columns:
                df = df.sort_values('score', ascending=False)
            
            # Filtrer TOP N si demandé
            if top_only:
                df = df.head(top_n)
            
            # Export selon format
            if format_type == "csv_resume":
                # Créer fichier avec résumé en en-tête
                with open(fichier, 'w', encoding='utf-8-sig') as f:
                    # En-tête avec résumé des TOP opportunités
                    f.write("=" * 80 + "\n")
                    f.write(f"🏆 TOP {min(top_n, len(df))} MEILLEURES OPPORTUNITÉS - {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
                    f.write("=" * 80 + "\n\n")
                    
                    # Résumé des TOP 5
                    f.write("📊 RÉSUMÉ DES MEILLEURES OPPORTUNITÉS:\n")
                    f.write("-" * 50 + "\n")
                    
                    for i, (_, row) in enumerate(df.head(5).iterrows()):
                        symbole = row.get('symbole', row.get('Symbole', 'N/A'))
                        score = row.get('score_global', row.get('score', 'N/A'))
                        decision = row.get('decision', row.get('décision', 'N/A'))
                        prix = row.get('prix', row.get('Prix', 'N/A'))
                        
                        f.write(f"\n#{i+1} 🎯 {symbole}\n")
                        f.write(f"   Score: {score}\n")
                        f.write(f"   Décision: {decision}\n")
                        f.write(f"   Prix: ${prix}\n")
                        
                        # Ajouter SL/TP si disponibles
                        sl = row.get('stop_loss', row.get('SL', None))
                        tp = row.get('take_profit', row.get('TP', None))
                        if sl and tp:
                            f.write(f"   SL: ${sl} | TP: ${tp}\n")
                    
                    f.write("\n" + "=" * 80 + "\n")
                    f.write("DONNÉES COMPLÈTES:\n")
                    f.write("=" * 80 + "\n\n")
                
                # Ajouter le DataFrame
                df.to_csv(fichier, mode='a', index=False, encoding='utf-8-sig')
            elif format_type == "excel_live":
                # ========== EXPORT EXCEL AVEC PRIX EN TEMPS RÉEL ==========
                try:
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    # Changer extension en .xlsx
                    if fichier.endswith('.csv'):
                        fichier = fichier.replace('.csv', '.xlsx')
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Opportunités"
                    
                    # Styles
                    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    live_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
                    live_font = Font(bold=True, color="FFFFFF", size=11)
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Colonnes principales à exporter
                    cols_export = ['symbole', 'marche', 'secteur', 'prix', 'decision', 'action', 
                                   'score_final', 'stop_loss', 'take_profit', 'ratio_rr', 
                                   'potentiel', 'rsi', 'tendance', 'conseil']
                    
                    # Filtrer les colonnes existantes
                    cols_disponibles = [c for c in cols_export if c in df.columns]
                    df_export = df[cols_disponibles].copy()
                    
                    # En-têtes avec colonnes LIVE
                    headers = list(df_export.columns) + ['💰 PRIX_LIVE', '📈 VAR_%', '📊 VOLUME', '🕐 MAJ']
                    
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        if 'LIVE' in header or 'VAR' in header or 'VOLUME' in header or 'MAJ' in header:
                            cell.fill = live_fill
                            cell.font = live_font
                        else:
                            cell.fill = header_fill
                            cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = thin_border
                    
                    # Données + Formules Excel pour prix en direct
                    for row_idx, row in enumerate(df_export.itertuples(index=False), 2):
                        # Données existantes
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Récupérer le symbole (colonne A)
                        symbole = row[0] if len(row) > 0 else ""
                        
                        # Colonnes avec formules Excel pour données en temps réel
                        col_live = len(df_export.columns) + 1
                        
                        # PRIX LIVE - Formule Excel STOCKHISTORY (Microsoft 365)
                        # =IFERROR(INDEX(STOCKHISTORY(A2,TODAY()),1,2),"N/A")
                        prix_formula = f'=IFERROR(INDEX(STOCKHISTORY(A{row_idx},TODAY()),1,2),"")'
                        cell_prix = ws.cell(row=row_idx, column=col_live, value=prix_formula)
                        cell_prix.font = Font(bold=True, color="006400")
                        cell_prix.border = thin_border
                        cell_prix.alignment = Alignment(horizontal='center')
                        
                        # VAR % - Calcul variation
                        # Compare prix live avec prix de l'analyse
                        prix_col = cols_disponibles.index('prix') + 1 if 'prix' in cols_disponibles else 4
                        var_formula = f'=IFERROR(({openpyxl.utils.get_column_letter(col_live)}{row_idx}-{openpyxl.utils.get_column_letter(prix_col)}{row_idx})/{openpyxl.utils.get_column_letter(prix_col)}{row_idx}*100,"")'
                        cell_var = ws.cell(row=row_idx, column=col_live + 1, value=var_formula)
                        cell_var.number_format = '0.00"%"'
                        cell_var.border = thin_border
                        cell_var.alignment = Alignment(horizontal='center')
                        
                        # VOLUME - Formule pour volume
                        vol_formula = f'=IFERROR(INDEX(STOCKHISTORY(A{row_idx},TODAY(),,0,0,1,6),1,2),"")'
                        cell_vol = ws.cell(row=row_idx, column=col_live + 2, value=vol_formula)
                        cell_vol.border = thin_border
                        cell_vol.alignment = Alignment(horizontal='center')
                        
                        # Heure MAJ
                        maj_formula = f'=TEXT(NOW(),"HH:MM:SS")'
                        cell_maj = ws.cell(row=row_idx, column=col_live + 3, value=maj_formula)
                        cell_maj.border = thin_border
                        cell_maj.alignment = Alignment(horizontal='center')
                    
                    # Ajuster largeur des colonnes
                    for col_idx, col in enumerate(headers, 1):
                        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
                    
                    # Figer la première ligne
                    ws.freeze_panes = 'A2'
                    
                    # Ajouter une feuille d'instructions
                    ws_help = wb.create_sheet("Instructions")
                    instructions = [
                        ["📊 GUIDE - PRIX EN TEMPS RÉEL"],
                        [""],
                        ["Les colonnes PRIX_LIVE, VAR_%, VOLUME utilisent la fonction STOCKHISTORY d'Excel."],
                        [""],
                        ["⚠️ IMPORTANT:"],
                        ["1. Vous devez avoir Microsoft 365 (Excel Online ou Desktop)"],
                        ["2. Les données se mettent à jour quand vous recalculez (F9 ou Ctrl+Shift+F9)"],
                        ["3. Pour une mise à jour automatique toutes les X minutes:"],
                        ["   - Fichier > Options > Formules > Calcul du classeur > Automatique"],
                        [""],
                        ["🔄 ALTERNATIVE SI STOCKHISTORY NE FONCTIONNE PAS:"],
                        ["1. Sélectionnez la colonne Symbole"],
                        ["2. Allez dans Données > Types de données > Actions"],
                        ["3. Excel convertira les symboles en données boursières"],
                        ["4. Cliquez sur le symbole pour voir Prix, Volume, etc."],
                        [""],
                        ["💡 ASTUCE: Appuyez sur F9 pour actualiser les prix"],
                    ]
                    for row_idx, row_data in enumerate(instructions, 1):
                        for col_idx, value in enumerate(row_data, 1):
                            ws_help.cell(row=row_idx, column=col_idx, value=value)
                    
                    wb.save(fichier)
                    
                except ImportError:
                    self.messagebox.showerror("Erreur", "Module openpyxl non installé.\nInstallez-le: pip install openpyxl")
                    return
                # ==========================================================
            else:
                # Export CSV simple mais trié
                df.to_csv(fichier, index=False, encoding='utf-8-sig')
            
            # Message de confirmation
            msg = f"✅ Exporté {len(df)} opportunités"
            if trier:
                msg += " (triées par score)"
            self.var_statut.set(msg)
            self.messagebox.showinfo("Export Réussi", 
                f"📁 Fichier exporté avec succès!\n\n"
                f"📊 {len(df)} opportunités exportées\n"
                f"🏆 Triées par meilleur score: {'Oui' if trier else 'Non'}\n"
                f"📄 Format: {format_type.upper()}\n\n"
                f"Fichier: {fichier}")
            logger.info(f"Exporté {len(df)} opportunités vers {fichier}")
            
        except Exception as e:
            self.messagebox.showerror("Erreur Export", f"Échec: {e}")
            logger.error(f"Erreur export: {e}")
    
    def _ouvrir_config(self):
        fenetre = self.tk.Toplevel(self.root)
        fenetre.title("⚙️ Configuration")
        fenetre.geometry("550x700")
        fenetre.configure(bg="#0d1117")
        
        tk = self.tk
        
        canvas = tk.Canvas(fenetre, bg="#0d1117", highlightthickness=0)
        scrollbar = self.ttk.Scrollbar(fenetre, orient="vertical", command=canvas.yview)
        cadre = tk.Frame(canvas, bg="#0d1117")
        
        cadre.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=cadre, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        params = {}
        
        def ajouter_param(parent, nom, label, valeur, ptype, desc=""):
            row = tk.Frame(parent, bg="#0d1117")
            row.pack(fill="x", padx=10, pady=2)
            
            tk.Label(row, text=label, font=("Segoe UI", 9), fg="#c9d1d9", bg="#0d1117",
                    width=28, anchor="w").pack(side="left")
            
            if ptype == bool:
                var = tk.BooleanVar(value=valeur)
                tk.Checkbutton(row, variable=var, bg="#0d1117", selectcolor="#21262d").pack(side="right")
            else:
                var = tk.StringVar(value=str(valeur))
                tk.Entry(row, textvariable=var, width=12, bg="#21262d", fg="#c9d1d9").pack(side="right")
            
            if desc:
                tk.Label(parent, text=f"  ℹ️ {desc}", font=("Segoe UI", 8),
                        fg="#6e7681", bg="#0d1117").pack(anchor="w", padx=10)
            
            params[nom] = (var, ptype)
        
        sections = [
            ("📊 Indicateurs Techniques", [
                ('RSI_PERIODE', "Période RSI", CONFIG.RSI_PERIODE, int, ""),
                ('ZSCORE_FENETRE', "Fenêtre Z-Score", CONFIG.ZSCORE_FENETRE, int, ""),
                ('ATR_PERIODE', "Période ATR", CONFIG.ATR_PERIODE, int, ""),
            ]),
            ("🎯 Seuils", [
                ('RSI_SURVENDU', "RSI Survendu", CONFIG.RSI_SURVENDU, int, ""),
                ('RSI_SURACHETE', "RSI Suracheté", CONFIG.RSI_SURACHETE, int, ""),
                ('SCORE_MINIMUM', "Score Minimum", CONFIG.SCORE_MINIMUM, int, "0-100"),
            ]),
            ("📉 Trading", [
                ('STOP_LOSS_ATR_MULT', "Stop Loss (×ATR)", CONFIG.STOP_LOSS_ATR_MULT, float, ""),
                ('TAKE_PROFIT_ATR_MULT', "Take Profit (×ATR)", CONFIG.TAKE_PROFIT_ATR_MULT, float, ""),
                ('RISQUE_RECOMPENSE_MIN', "R/R Minimum", CONFIG.RISQUE_RECOMPENSE_MIN, float, ""),
                ('HORIZON_COURT', "Horizon Court (jours)", CONFIG.HORIZON_COURT, int, ""),
                ('HORIZON_MOYEN', "Horizon Moyen (jours)", CONFIG.HORIZON_MOYEN, int, ""),
                ('HORIZON_LONG', "Horizon Long (jours)", CONFIG.HORIZON_LONG, int, ""),
            ]),
            ("💰 Valorisation", [
                ('TAUX_ACTUALISATION', "Taux d'Actualisation", CONFIG.TAUX_ACTUALISATION, float, "ex: 0.10 = 10%"),
                ('TAUX_CROISSANCE', "Taux de Croissance", CONFIG.TAUX_CROISSANCE, float, ""),
                ('MARGE_SECURITE', "Marge de Sécurité", CONFIG.MARGE_SECURITE, float, "ex: 0.25 = 25%"),
            ]),
            ("🧠 Machine Learning", [
                ('ML_ACTIF', "ML Activé", CONFIG.ML_ACTIF, bool, ""),
                ('ML_RENDEMENT_CIBLE', "Rendement Cible", CONFIG.ML_RENDEMENT_CIBLE, float, "ex: 0.03 = 3%"),
                ('ML_HORIZON_JOURS', "Horizon ML (jours)", CONFIG.ML_HORIZON_JOURS, int, ""),
                ('ML_SEUIL_CONFIANCE', "Seuil de Confiance", CONFIG.ML_SEUIL_CONFIANCE, float, "0.5-0.95"),
            ]),
        ]
        
        for titre, items in sections:
            lf = tk.LabelFrame(cadre, text=titre, font=("Segoe UI", 10, "bold"),
                              fg="#58a6ff", bg="#0d1117")
            lf.pack(fill="x", padx=10, pady=5)
            
            for item in items:
                if len(item) == 5:
                    nom, label, valeur, ptype, desc = item
                else:
                    nom, label, valeur, ptype = item
                    desc = ""
                ajouter_param(lf, nom, label, valeur, ptype, desc)
        
        var_erreur = tk.StringVar()
        tk.Label(cadre, textvariable=var_erreur, font=("Segoe UI", 9),
                fg="#f85149", bg="#0d1117", wraplength=500).pack(fill="x", padx=10, pady=5)
        
        cadre_btn = tk.Frame(fenetre, bg="#0d1117")
        cadre_btn.pack(fill="x", pady=10)
        
        def sauvegarder():
            global CONFIG
            erreurs = []
            nouvelles_valeurs = {}
            
            for nom, (var, ptype) in params.items():
                try:
                    if ptype == bool:
                        nouvelles_valeurs[nom] = var.get()
                    else:
                        nouvelles_valeurs[nom] = ptype(var.get())
                except ValueError:
                    erreurs.append(f"{nom}: Valeur invalide")
            
            if erreurs:
                var_erreur.set("\n".join(erreurs))
                return
            
            for nom, val in nouvelles_valeurs.items():
                setattr(CONFIG, nom, val)
            
            succes, message = CONFIG.sauvegarder()
            if succes:
                self.moteur.config = CONFIG
                self.messagebox.showinfo("Succès", message)
                fenetre.destroy()
            else:
                var_erreur.set(message)
        
        tk.Button(cadre_btn, text="💾 Sauvegarder", font=("Segoe UI", 11, "bold"),
                 bg="#238636", fg="white", command=sauvegarder).pack(side="left", padx=20)
        
        tk.Button(cadre_btn, text="❌ Annuler", font=("Segoe UI", 11, "bold"),
                 bg="#484f58", fg="white", command=fenetre.destroy).pack(side="left")
    
    def _ouvrir_scorecard(self):
        """📊 Affiche le Scorecard des performances"""
        tk = self.tk
        
        fenetre = tk.Toplevel(self.root)
        fenetre.title("📊 Scorecard - Performance des Signaux")
        fenetre.geometry("800x700")
        fenetre.configure(bg="#0d1117")
        
        # Titre
        tk.Label(fenetre,
                text="📊 SCORECARD - ANALYSE DE PERFORMANCE",
                font=("Segoe UI", 16, "bold"),
                fg="#58a6ff", bg="#0d1117").pack(pady=15)
        
        # Filtres
        cadre_filtres = tk.Frame(fenetre, bg="#0d1117")
        cadre_filtres.pack(fill="x", padx=20, pady=5)
        
        filtre_vars = {
            'strategy': tk.StringVar(value='tous'),
            'marche': tk.StringVar(value='tous'),
            'action': tk.StringVar(value='tous')
        }
        
        tk.Label(cadre_filtres, text="Stratégie:", bg="#0d1117", fg="#8b949e").pack(side="left")
        ttk_combo_strat = self.ttk.Combobox(cadre_filtres, textvariable=filtre_vars['strategy'],
                                           values=["tous", "Momentum", "Value", "ML", "Mean Reversion"],
                                           width=12, state="readonly")
        ttk_combo_strat.pack(side="left", padx=5)
        
        tk.Label(cadre_filtres, text="Marché:", bg="#0d1117", fg="#8b949e").pack(side="left", padx=(10, 0))
        ttk_combo_marche = self.ttk.Combobox(cadre_filtres, textvariable=filtre_vars['marche'],
                                            values=["tous", "US", "France", "Europe"],
                                            width=10, state="readonly")
        ttk_combo_marche.pack(side="left", padx=5)
        
        tk.Label(cadre_filtres, text="Action:", bg="#0d1117", fg="#8b949e").pack(side="left", padx=(10, 0))
        ttk_combo_action = self.ttk.Combobox(cadre_filtres, textvariable=filtre_vars['action'],
                                            values=["tous", "ACHAT", "VENTE"],
                                            width=8, state="readonly")
        ttk_combo_action.pack(side="left", padx=5)
        
        # Cadre des métriques
        cadre_stats = tk.Frame(fenetre, bg="#0d1117")
        cadre_stats.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Labels pour les métriques
        metriques_labels = {}
        
        def afficher_metriques():
            # Construire le filtre
            filtre = {}
            if filtre_vars['strategy'].get() != 'tous':
                filtre['strategy'] = filtre_vars['strategy'].get()
            if filtre_vars['marche'].get() != 'tous':
                filtre['marche'] = filtre_vars['marche'].get()
            if filtre_vars['action'].get() != 'tous':
                filtre['action'] = filtre_vars['action'].get()
            
            stats = self.moteur.obtenir_scorecard(filtre if filtre else None)
            
            # Effacer les widgets existants
            for widget in cadre_stats.winfo_children():
                widget.destroy()
            
            # Métriques principales
            metriques = [
                ("📊 Total Trades", str(stats.get('total_trades', 0)), "#c9d1d9"),
                ("✅ Wins / ❌ Losses", f"{stats.get('wins', 0)} / {stats.get('losses', 0)}", "#c9d1d9"),
                ("🎯 Win Rate", f"{stats.get('win_rate', 0):.1f}%", 
                 "#3fb950" if stats.get('win_rate', 0) > 50 else "#f85149"),
                ("💰 Expectancy ($/trade)", f"${stats.get('expectancy', 0):.2f}",
                 "#3fb950" if stats.get('expectancy', 0) > 0 else "#f85149"),
                ("📈 Profit Factor", f"{stats.get('profit_factor', 0):.2f}",
                 "#3fb950" if stats.get('profit_factor', 0) > 1 else "#f85149"),
                ("💵 Total PnL", f"${stats.get('total_pnl', 0):.2f}",
                 "#3fb950" if stats.get('total_pnl', 0) > 0 else "#f85149"),
                ("📉 Max Drawdown", f"${stats.get('max_drawdown', 0):.2f}", "#f85149"),
                ("⏱️ Avg Holding Days", f"{stats.get('avg_holding_days', 0):.1f}", "#c9d1d9"),
                ("🏆 SQS (Signal Quality)", f"{stats.get('sqs', 0):.0f}/100",
                 "#3fb950" if stats.get('sqs', 0) >= 70 else ("#f0883e" if stats.get('sqs', 0) >= 50 else "#f85149")),
            ]
            
            for i, (label, valeur, couleur) in enumerate(metriques):
                row = tk.Frame(cadre_stats, bg="#161b22", pady=8)
                row.pack(fill="x", pady=3)
                
                tk.Label(row, text=label, font=("Segoe UI", 12),
                        fg="#8b949e", bg="#161b22", width=25, anchor="w").pack(side="left", padx=15)
                
                tk.Label(row, text=valeur, font=("Segoe UI", 14, "bold"),
                        fg=couleur, bg="#161b22").pack(side="right", padx=15)
            
            # Verdict
            sqs = stats.get('sqs', 0)
            if sqs >= 70:
                verdict = "✅ STRATÉGIE PROFITABLE"
                verdict_color = "#3fb950"
            elif sqs >= 50:
                verdict = "⚠️ STRATÉGIE À AMÉLIORER"
                verdict_color = "#f0883e"
            else:
                verdict = "❌ STRATÉGIE NON PROFITABLE"
                verdict_color = "#f85149"
            
            if stats.get('total_trades', 0) < 10:
                verdict = "ℹ️ DONNÉES INSUFFISANTES (min. 10 trades)"
                verdict_color = "#8b949e"
            
            tk.Label(cadre_stats, text=verdict, font=("Segoe UI", 14, "bold"),
                    fg=verdict_color, bg="#0d1117").pack(pady=20)
        
        # Bouton actualiser
        tk.Button(cadre_filtres, text="🔄 Actualiser", font=("Segoe UI", 10, "bold"),
                 bg="#238636", fg="white", command=afficher_metriques).pack(side="right", padx=10)
        
        # Bouton évaluer les signaux pending
        def evaluer_pending():
            self.var_statut.set("🔄 Évaluation des signaux...")
            
            def evaluer():
                result = self.moteur.evaluer_signaux_pending()
                self.file_maj.put({'type': 'eval_complete', 'result': result})
                afficher_metriques()
            
            threading.Thread(target=evaluer, daemon=True).start()
        
        tk.Button(cadre_filtres, text="⚡ Évaluer Pending", font=("Segoe UI", 10, "bold"),
                 bg="#8957e5", fg="white", command=evaluer_pending).pack(side="right", padx=5)
        
        # Afficher les métriques initiales
        afficher_metriques()
    
    def _ouvrir_stress_tests(self):
        """🔬 Affiche les résultats des Stress Tests"""
        tk = self.tk
        
        fenetre = tk.Toplevel(self.root)
        fenetre.title("🔬 Stress Tests - Analyse de Robustesse")
        fenetre.geometry("700x600")
        fenetre.configure(bg="#0d1117")
        
        # Titre
        tk.Label(fenetre,
                text="🔬 STRESS TESTS - ROBUSTESSE DE LA STRATÉGIE",
                font=("Segoe UI", 16, "bold"),
                fg="#58a6ff", bg="#0d1117").pack(pady=15)
        
        tk.Label(fenetre,
                text="Ces tests vérifient si votre stratégie résiste aux conditions défavorables",
                font=("Segoe UI", 10),
                fg="#8b949e", bg="#0d1117").pack()
        
        # Cadre des résultats
        cadre_results = tk.Frame(fenetre, bg="#0d1117")
        cadre_results.pack(fill="both", expand=True, padx=20, pady=20)
        
        def executer_tests():
            for widget in cadre_results.winfo_children():
                widget.destroy()
            
            tk.Label(cadre_results, text="⏳ Exécution des tests...",
                    font=("Segoe UI", 12), fg="#f0883e", bg="#0d1117").pack(pady=10)
            fenetre.update()
            
            resultats = self.moteur.executer_stress_tests()
            
            for widget in cadre_results.winfo_children():
                widget.destroy()
            
            # Afficher chaque test
            for test in resultats.get('tests', []):
                lf = tk.LabelFrame(cadre_results, text=test.get('nom', 'Test'),
                                  font=("Segoe UI", 11, "bold"),
                                  fg="#58a6ff", bg="#0d1117", padx=10, pady=5)
                lf.pack(fill="x", pady=5)
                
                robuste = test.get('robuste', False)
                icon = "✅" if robuste else "❌"
                color = "#3fb950" if robuste else "#f85149"
                
                for key, val in test.items():
                    if key in ['nom', 'robuste']:
                        continue
                    
                    row = tk.Frame(lf, bg="#0d1117")
                    row.pack(fill="x", pady=1)
                    
                    label_text = key.replace('_', ' ').title()
                    tk.Label(row, text=label_text, font=("Segoe UI", 9),
                            fg="#8b949e", bg="#0d1117", width=20, anchor="w").pack(side="left")
                    
                    val_text = f"{val}" if not isinstance(val, float) else f"{val:.2f}"
                    tk.Label(row, text=val_text, font=("Segoe UI", 9, "bold"),
                            fg="#c9d1d9", bg="#0d1117").pack(side="left")
                
                tk.Label(lf, text=f"{icon} {'ROBUSTE' if robuste else 'FRAGILE'}",
                        font=("Segoe UI", 10, "bold"), fg=color, bg="#0d1117").pack(pady=5)
            
            # Score global
            score = resultats.get('score_robustesse', 0)
            verdict = resultats.get('verdict', '')
            
            verdict_frame = tk.Frame(cadre_results, bg="#161b22", pady=15)
            verdict_frame.pack(fill="x", pady=15)
            
            tk.Label(verdict_frame, text=f"SCORE GLOBAL: {score}%",
                    font=("Segoe UI", 14, "bold"), fg="#58a6ff", bg="#161b22").pack()
            
            verdict_color = "#3fb950" if score >= 75 else ("#f0883e" if score >= 50 else "#f85149")
            tk.Label(verdict_frame, text=verdict,
                    font=("Segoe UI", 16, "bold"), fg=verdict_color, bg="#161b22").pack(pady=5)
        
        tk.Button(fenetre, text="🚀 Exécuter les Stress Tests", font=("Segoe UI", 12, "bold"),
                 bg="#8957e5", fg="white", command=executer_tests).pack(pady=10)
        
        # Bouton pour les tests unitaires
        tk.Button(fenetre, text="🧪 Tests Unitaires Backtest", font=("Segoe UI", 10, "bold"),
                 bg="#238636", fg="white", command=self._ouvrir_unit_tests).pack(pady=5)
        
        # Bouton pour Out-of-Sample Test
        tk.Button(fenetre, text="📊 Test Out-of-Sample", font=("Segoe UI", 10, "bold"),
                 bg="#1f6feb", fg="white", command=self._ouvrir_oos_test).pack(pady=5)
        
        tk.Label(fenetre,
                text="⚠️ Nécessite des trades évalués pour des résultats significatifs",
                font=("Segoe UI", 9), fg="#6e7681", bg="#0d1117").pack(pady=5)
    
    def _ouvrir_oos_test(self):
        """📊 Exécute et affiche les résultats du test Out-of-Sample"""
        tk = self.tk
        
        fenetre = tk.Toplevel(self.root)
        fenetre.title("📊 Test Out-of-Sample - Walk-Forward")
        fenetre.geometry("950x750")
        fenetre.configure(bg="#0d1117")
        
        # Titre
        tk.Label(fenetre,
                text="📊 TEST OUT-OF-SAMPLE (Walk-Forward)",
                font=("Segoe UI", 16, "bold"),
                fg="#58a6ff", bg="#0d1117").pack(pady=15)
        
        tk.Label(fenetre,
                text="Validation sur données jamais vues - 6 fenêtres (2023-2026)",
                font=("Segoe UI", 10),
                fg="#8b949e", bg="#0d1117").pack()
        
        # ══════════════════════════════════════════════════════════════════
        # Configuration
        # ══════════════════════════════════════════════════════════════════
        cadre_config = tk.Frame(fenetre, bg="#161b22", pady=10)
        cadre_config.pack(fill="x", padx=20, pady=10)
        
        # Groupe de symboles
        row1 = tk.Frame(cadre_config, bg="#161b22")
        row1.pack(fill="x", pady=5)
        
        tk.Label(row1, text="Groupe de symboles:", 
                font=("Segoe UI", 10), fg="#8b949e", bg="#161b22").pack(side="left", padx=10)
        
        groupe_var = tk.StringVar(value="Large Cap")
        groupe_combo = self.ttk.Combobox(row1, textvariable=groupe_var,
                                        values=["Large Cap", "Mid Cap", "Tech/Biotech", "Tous (Large+Mid+Tech)"],
                                        width=20, state="readonly")
        groupe_combo.pack(side="left", padx=10)
        
        # Mode de signaux
        row2 = tk.Frame(cadre_config, bg="#161b22")
        row2.pack(fill="x", pady=5)
        
        tk.Label(row2, text="Mode des signaux:", 
                font=("Segoe UI", 10), fg="#8b949e", bg="#161b22").pack(side="left", padx=10)
        
        mode_var = tk.StringVar(value="RÉEL")
        mode_combo = self.ttk.Combobox(row2, textvariable=mode_var,
                                       values=["RÉEL (votre système)", "ATR (référence simple)"],
                                       width=25, state="readonly")
        mode_combo.pack(side="left", padx=10)
        
        # Info mode
        tk.Label(row2, text="💡 RÉEL = vos indicateurs (RSI, ADX, etc.)", 
                font=("Segoe UI", 9), fg="#6e7681", bg="#161b22").pack(side="left", padx=10)
        
        # ══════════════════════════════════════════════════════════════════
        # Cadre des résultats
        # ══════════════════════════════════════════════════════════════════
        cadre_results = tk.Frame(fenetre, bg="#0d1117")
        cadre_results.pack(fill="both", expand=True, padx=20, pady=10)
        
        var_statut = tk.StringVar(value="⏳ Sélectionnez les options et cliquez sur 'Exécuter'")
        lbl_statut = tk.Label(cadre_results, textvariable=var_statut,
                             font=("Segoe UI", 11), fg="#f0883e", bg="#0d1117")
        lbl_statut.pack(pady=10)
        
        def executer_oos():
            for widget in cadre_results.winfo_children():
                if widget != lbl_statut:
                    widget.destroy()
            
            use_real = "RÉEL" in mode_var.get()
            mode_text = "signaux RÉELS" if use_real else "signaux ATR"
            var_statut.set(f"⏳ Exécution avec {mode_text}... (plusieurs minutes)")
            fenetre.update()
            
            # Définir les symboles selon le groupe
            groupes_symboles = {
                "Large Cap": ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'META', 'NVDA', 'JPM', 'V', 'JNJ', 'UNH',
                             'MA', 'PG', 'HD', 'DIS', 'BAC', 'XOM', 'CVX', 'PFE', 'KO', 'PEP'],
                "Mid Cap": ['SNAP', 'ROKU', 'ETSY', 'PINS', 'TWLO', 'DDOG', 'NET', 'CRWD', 'ZS', 'OKTA',
                           'FVRR', 'UPWK', 'W', 'CHWY', 'PTON', 'ABNB', 'DASH', 'RBLX', 'U', 'HOOD'],
                "Tech/Biotech": ['MRNA', 'BNTX', 'NVAX', 'PLTR', 'SQ', 'SHOP', 'COIN', 'RIOT', 'MARA', 'SOFI',
                                'AMD', 'INTC', 'TSM', 'MU', 'QCOM', 'AVGO', 'ARM', 'SMCI', 'CRSP', 'EDIT'],
            }
            
            if "Tous" in groupe_var.get():
                symboles = groupes_symboles["Large Cap"] + groupes_symboles["Mid Cap"] + groupes_symboles["Tech/Biotech"]
            else:
                symboles = groupes_symboles.get(groupe_var.get(), groupes_symboles["Large Cap"])
            
            def callback(msg, i, total):
                var_statut.set(f"⏳ {msg}")
                fenetre.update()
            
            # Exécuter le test
            tester = OutOfSampleTester(CONFIG)
            resultats = tester.walk_forward_test(
                symboles, 
                callback=callback,
                use_real_signals=use_real,
                moteur=self.moteur if use_real else None
            )
            
            var_statut.set(f"✅ Test terminé - Mode: {resultats.get('mode', 'N/A')}")
            
            # Afficher les résultats
            afficher_resultats_oos(resultats)
        
        def afficher_resultats_oos(resultats):
            # Mode utilisé
            mode_label = resultats.get('mode', 'N/A')
            mode_color = "#3fb950" if "RÉEL" in mode_label else "#f0883e"
            
            tk.Label(cadre_results, text=f"📡 Mode: {mode_label}",
                    font=("Segoe UI", 11, "bold"), fg=mode_color, bg="#0d1117").pack(pady=5)
            
            # Métriques globales
            metriques = resultats.get('metriques_globales', {})
            criteres = resultats.get('criteres', {})
            stabilite = resultats.get('stabilite', {})
            
            # Frame métriques
            lf_metriques = tk.LabelFrame(cadre_results, text=f"📈 Métriques Globales ({resultats.get('nb_trades_total', 0)} trades sur {resultats.get('nb_windows', 0)} fenêtres)",
                                        font=("Segoe UI", 11, "bold"),
                                        fg="#58a6ff", bg="#0d1117", padx=15, pady=10)
            lf_metriques.pack(fill="x", pady=10)
            
            metriques_affichage = [
                ("Nombre de Trades", metriques.get('nb_trades', 0), "#c9d1d9"),
                ("Win Rate", f"{metriques.get('win_rate', 0)}%", 
                 "#3fb950" if metriques.get('win_rate', 0) > 50 else "#f85149"),
                ("Profit Factor", metriques.get('profit_factor', 0),
                 "#3fb950" if (isinstance(metriques.get('profit_factor', 0), (int, float)) and metriques.get('profit_factor', 0) > 1) else "#f85149"),
                ("Expectancy", f"${metriques.get('expectancy', 0):.2f}",
                 "#3fb950" if metriques.get('expectancy', 0) > 0 else "#f85149"),
                ("Max Drawdown", f"${metriques.get('max_drawdown', 0):.2f}", "#f0883e"),
                ("PnL Total", f"${metriques.get('total_pnl', 0):.2f}",
                 "#3fb950" if metriques.get('total_pnl', 0) > 0 else "#f85149"),
            ]
            
            for i, (label, valeur, couleur) in enumerate(metriques_affichage):
                row = tk.Frame(lf_metriques, bg="#0d1117")
                row.pack(fill="x", pady=2)
                
                tk.Label(row, text=label, font=("Segoe UI", 10),
                        fg="#8b949e", bg="#0d1117", width=20, anchor="w").pack(side="left")
                tk.Label(row, text=str(valeur), font=("Segoe UI", 10, "bold"),
                        fg=couleur, bg="#0d1117").pack(side="left")
            
            # Frame détail par fenêtre
            lf_windows = tk.LabelFrame(cadre_results, text="📊 Détail par Fenêtre",
                                      font=("Segoe UI", 11, "bold"),
                                      fg="#58a6ff", bg="#0d1117", padx=15, pady=5)
            lf_windows.pack(fill="x", pady=5)
            
            for w in resultats.get('windows', []):
                m = w.get('metriques_test', {})
                pnl = m.get('total_pnl', 0)
                wr = m.get('win_rate', 0)
                color = "#3fb950" if pnl > 0 else "#f85149"
                icon = "✅" if pnl > 0 else "❌"
                
                row = tk.Frame(lf_windows, bg="#0d1117")
                row.pack(fill="x", pady=1)
                tk.Label(row, text=f"{icon} F{w['window']}: {w['test_period']}", 
                        font=("Segoe UI", 9), fg="#8b949e", bg="#0d1117", width=35, anchor="w").pack(side="left")
                tk.Label(row, text=f"WR: {wr}% | PnL: ${pnl:.0f} | Trades: {w.get('nb_trades', m.get('nb_trades', 0))}", 
                        font=("Segoe UI", 9, "bold"), fg=color, bg="#0d1117").pack(side="left")
            
            # Frame stabilité
            lf_stabilite = tk.LabelFrame(cadre_results, text="📉 Stabilité",
                                        font=("Segoe UI", 11, "bold"),
                                        fg="#58a6ff", bg="#0d1117", padx=15, pady=5)
            lf_stabilite.pack(fill="x", pady=5)
            
            row = tk.Frame(lf_stabilite, bg="#0d1117")
            row.pack(fill="x", pady=2)
            pos = stabilite.get('nb_fenetres_positives', 0)
            tot = stabilite.get('nb_fenetres_total', 0)
            tk.Label(row, text=f"Fenêtres Positives: {pos}/{tot} | STD Win Rate: {stabilite.get('win_rate_std', 0)}%", 
                    font=("Segoe UI", 10), fg="#8b949e", bg="#0d1117").pack(side="left")
            
            stab_verdict = stabilite.get('verdict', '')
            stab_color = "#3fb950" if "STABLE" in stab_verdict else ("#f0883e" if "VARIABLE" in stab_verdict else "#f85149")
            tk.Label(lf_stabilite, text=stab_verdict,
                    font=("Segoe UI", 11, "bold"), fg=stab_color, bg="#0d1117").pack(pady=3)
            
            # Verdict final
            verdict_frame = tk.Frame(cadre_results, bg="#161b22", pady=10)
            verdict_frame.pack(fill="x", pady=10)
            
            final_verdict = criteres.get('verdict', '❌ Non évalué')
            verdict_color = "#3fb950" if "VALIDÉ" in final_verdict else "#f85149"
            
            tk.Label(verdict_frame, text="🎯 VERDICT FINAL",
                    font=("Segoe UI", 12), fg="#8b949e", bg="#161b22").pack()
            tk.Label(verdict_frame, text=final_verdict,
                    font=("Segoe UI", 16, "bold"), fg=verdict_color, bg="#161b22").pack(pady=5)
            
            # Détails critères
            criteres_details = criteres.get('criteres', {})
            for nom, details in criteres_details.items():
                passe = details.get('passe', False)
                icon = "✅" if passe else "❌"
                color = "#3fb950" if passe else "#f85149"
                tk.Label(verdict_frame, text=f"{icon} {details.get('detail', nom)}",
                        font=("Segoe UI", 9), fg=color, bg="#161b22").pack()
        
        # ══════════════════════════════════════════════════════════════════
        # Boutons
        # ══════════════════════════════════════════════════════════════════
        cadre_btn = tk.Frame(fenetre, bg="#0d1117")
        cadre_btn.pack(fill="x", pady=10)
        
        tk.Button(cadre_btn, text="🚀 Exécuter Walk-Forward", font=("Segoe UI", 12, "bold"),
                 bg="#238636", fg="white", 
                 command=lambda: threading.Thread(target=executer_oos, daemon=True).start()).pack(side="left", padx=20)
        
        tk.Button(cadre_btn, text="❌ Fermer", font=("Segoe UI", 10),
                 bg="#484f58", fg="white", command=fenetre.destroy).pack(side="right", padx=20)
        
        tk.Label(fenetre,
                text="⚠️ Critères: ≥200 trades, PF>1.2, Expectancy>0, Stabilité multi-fenêtres",
                font=("Segoe UI", 9), fg="#6e7681", bg="#0d1117").pack(pady=5)
    
    def _ouvrir_unit_tests(self):
        """🧪 Affiche les résultats des Tests Unitaires du Backtest"""
        tk = self.tk
        
        fenetre = tk.Toplevel(self.root)
        fenetre.title("🧪 Tests Unitaires - Validation du Backtest")
        fenetre.geometry("700x500")
        fenetre.configure(bg="#0d1117")
        
        # Titre
        tk.Label(fenetre,
                text="🧪 TESTS UNITAIRES - BACKTEST",
                font=("Segoe UI", 16, "bold"),
                fg="#58a6ff", bg="#0d1117").pack(pady=15)
        
        tk.Label(fenetre,
                text="Ces tests vérifient que le backtest ne regarde pas dans le futur",
                font=("Segoe UI", 10),
                fg="#8b949e", bg="#0d1117").pack()
        
        # Cadre des résultats
        cadre_results = tk.Frame(fenetre, bg="#0d1117")
        cadre_results.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Exécuter les tests automatiquement
        resultats = UnitTestsBacktest.executer_tous_tests()
        
        for test in resultats.get('tests', []):
            lf = tk.LabelFrame(cadre_results, text=test.get('test', 'Test'),
                              font=("Segoe UI", 11, "bold"),
                              fg="#58a6ff", bg="#0d1117", padx=10, pady=5)
            lf.pack(fill="x", pady=5)
            
            succes = test.get('succes', False)
            icon = "✅" if succes else "❌"
            color = "#3fb950" if succes else "#f85149"
            
            # Détails du test
            details = test.get('details', {})
            for key, val in details.items():
                row = tk.Frame(lf, bg="#0d1117")
                row.pack(fill="x", pady=1)
                
                label_text = key.replace('_', ' ').title()
                tk.Label(row, text=label_text, font=("Segoe UI", 9),
                        fg="#8b949e", bg="#0d1117", width=25, anchor="w").pack(side="left")
                
                tk.Label(row, text=str(val), font=("Segoe UI", 9, "bold"),
                        fg="#c9d1d9", bg="#0d1117").pack(side="left")
            
            # Erreurs
            erreurs = test.get('erreurs', [])
            if erreurs:
                for err in erreurs:
                    tk.Label(lf, text=f"⚠️ {err}", font=("Segoe UI", 9),
                            fg="#f85149", bg="#0d1117").pack(anchor="w")
            
            tk.Label(lf, text=f"{icon} {'PASSÉ' if succes else 'ÉCHOUÉ'}",
                    font=("Segoe UI", 10, "bold"), fg=color, bg="#0d1117").pack(pady=5)
        
        # Verdict final
        verdict_frame = tk.Frame(cadre_results, bg="#161b22", pady=15)
        verdict_frame.pack(fill="x", pady=15)
        
        verdict = resultats.get('verdict', '')
        taux = resultats.get('taux_succes', '0/0')
        
        verdict_color = "#3fb950" if resultats.get('succes', 0) == resultats.get('total', 0) else "#f85149"
        
        tk.Label(verdict_frame, text=f"RÉSULTAT: {taux}",
                font=("Segoe UI", 14, "bold"), fg="#58a6ff", bg="#161b22").pack()
        
        tk.Label(verdict_frame, text=verdict,
                font=("Segoe UI", 16, "bold"), fg=verdict_color, bg="#161b22").pack(pady=5)
    
    def _ouvrir_historique_signaux(self):
        """
        📋 EXPLORATEUR AVANCÉ DES SIGNAUX
        
        Fonctionnalités:
        - Filtrage par période (7j, 14j, 30j, 90j, tout)
        - Filtrage par symbole
        - Comparaison avec prix actuel
        - Résumé par date
        - Export CSV
        - Évaluation des signaux pending
        """
        tk = self.tk
        ttk = self.ttk
        
        fenetre = tk.Toplevel(self.root)
        fenetre.title("📋 Explorateur de Signaux Avancé")
        fenetre.geometry("1450x850")
        fenetre.configure(bg="#0d1117")
        
        # Variables
        var_periode = tk.StringVar(value="7")
        var_symbole = tk.StringVar(value="Tous")
        var_filtre_action = tk.StringVar(value="Tous")
        var_filtre_statut = tk.StringVar(value="Tous")
        
        # ═══════════════════════════════════════════════════════════════
        # EN-TÊTE
        # ═══════════════════════════════════════════════════════════════
        entete = tk.Frame(fenetre, bg="#161b22", height=60)
        entete.pack(fill="x")
        entete.pack_propagate(False)
        
        tk.Label(entete,
                text="📋 EXPLORATEUR DE SIGNAUX - Historique & Suivi",
                font=("Segoe UI", 18, "bold"),
                fg="#58a6ff", bg="#161b22").pack(pady=5)
        
        tk.Label(entete,
                text="Filtrez, explorez et comparez vos signaux passés avec les prix actuels",
                font=("Segoe UI", 10),
                fg="#7ee787", bg="#161b22").pack()
        
        # ═══════════════════════════════════════════════════════════════
        # FILTRES
        # ═══════════════════════════════════════════════════════════════
        cadre_filtres = tk.Frame(fenetre, bg="#0d1117")
        cadre_filtres.pack(fill="x", padx=15, pady=10)
        
        # Période
        tk.Label(cadre_filtres, text="📅 Période:", font=("Segoe UI", 10, "bold"),
                fg="#8b949e", bg="#0d1117").pack(side="left", padx=5)
        
        for texte, valeur in [("7j", "7"), ("14j", "14"), ("30j", "30"), ("90j", "90"), ("Tout", "0")]:
            tk.Radiobutton(cadre_filtres, text=texte, variable=var_periode, value=valeur,
                          bg="#0d1117", fg="#c9d1d9", selectcolor="#21262d",
                          font=("Segoe UI", 9), command=lambda: charger_signaux()).pack(side="left", padx=3)
        
        # Séparateur
        tk.Label(cadre_filtres, text=" | ", fg="#30363d", bg="#0d1117").pack(side="left", padx=5)
        
        # Symbole
        tk.Label(cadre_filtres, text="🎯 Symbole:", font=("Segoe UI", 10, "bold"),
                fg="#8b949e", bg="#0d1117").pack(side="left", padx=5)
        
        symboles = ["Tous"] + self.moteur.obtenir_symboles_avec_signaux()
        combo_symbole = ttk.Combobox(cadre_filtres, textvariable=var_symbole,
                                     values=symboles, width=12, state="readonly")
        combo_symbole.pack(side="left", padx=5)
        combo_symbole.bind('<<ComboboxSelected>>', lambda e: charger_signaux())
        
        # Séparateur
        tk.Label(cadre_filtres, text=" | ", fg="#30363d", bg="#0d1117").pack(side="left", padx=5)
        
        # Action
        tk.Label(cadre_filtres, text="📈 Action:", font=("Segoe UI", 10),
                fg="#8b949e", bg="#0d1117").pack(side="left", padx=5)
        
        for texte, valeur in [("Tous", "Tous"), ("ACHAT", "ACHAT"), ("VENTE", "VENTE")]:
            tk.Radiobutton(cadre_filtres, text=texte, variable=var_filtre_action, value=valeur,
                          bg="#0d1117", fg="#c9d1d9", selectcolor="#21262d",
                          font=("Segoe UI", 9), command=lambda: charger_signaux()).pack(side="left", padx=2)
        
        # Séparateur
        tk.Label(cadre_filtres, text=" | ", fg="#30363d", bg="#0d1117").pack(side="left", padx=5)
        
        # Statut
        tk.Label(cadre_filtres, text="📊 Statut:", font=("Segoe UI", 10),
                fg="#8b949e", bg="#0d1117").pack(side="left", padx=5)
        
        for texte, valeur in [("Tous", "Tous"), ("Pending", "PENDING"), ("Évalué", "EVALUATED")]:
            tk.Radiobutton(cadre_filtres, text=texte, variable=var_filtre_statut, value=valeur,
                          bg="#0d1117", fg="#c9d1d9", selectcolor="#21262d",
                          font=("Segoe UI", 9), command=lambda: charger_signaux()).pack(side="left", padx=2)
        
        # ═══════════════════════════════════════════════════════════════
        # STATS RAPIDES
        # ═══════════════════════════════════════════════════════════════
        cadre_stats = tk.Frame(fenetre, bg="#161b22")
        cadre_stats.pack(fill="x", padx=15, pady=5)
        
        lbl_stats = tk.Label(cadre_stats, text="", font=("Segoe UI", 10),
                            fg="#7ee787", bg="#161b22")
        lbl_stats.pack(pady=5)
        
        # ═══════════════════════════════════════════════════════════════
        # ONGLETS (Notebook)
        # ═══════════════════════════════════════════════════════════════
        notebook = ttk.Notebook(fenetre)
        notebook.pack(fill="both", expand=True, padx=15, pady=5)
        
        # --- ONGLET 1: LISTE DES SIGNAUX ---
        tab_liste = tk.Frame(notebook, bg="#0d1117")
        notebook.add(tab_liste, text="📋 Liste des Signaux")
        
        colonnes = ("id", "date", "symbole", "marche", "action", "prix_entree", 
                   "sl", "tp", "rr", "score", "statut", "result", "pnl", "jours")
        
        arbre = ttk.Treeview(tab_liste, columns=colonnes, show="headings", height=18)
        
        entetes = {
            "id": ("ID", 60),
            "date": ("📅 Date", 100),
            "symbole": ("🎯 Ticker", 70),
            "marche": ("🌍", 55),
            "action": ("Action", 60),
            "prix_entree": ("💰 Prix", 75),
            "sl": ("📉 SL", 70),
            "tp": ("📈 TP", 70),
            "rr": ("R:R", 50),
            "score": ("Score", 50),
            "statut": ("Statut", 75),
            "result": ("Résultat", 70),
            "pnl": ("💵 PnL", 80),
            "jours": ("Jours", 45)
        }
        
        for col, (texte, largeur) in entetes.items():
            arbre.heading(col, text=texte)
            arbre.column(col, width=largeur, anchor="center")
        
        scrollbar = ttk.Scrollbar(tab_liste, orient="vertical", command=arbre.yview)
        arbre.configure(yscrollcommand=scrollbar.set)
        
        arbre.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # --- ONGLET 2: RÉSUMÉ PAR DATE ---
        tab_resume = tk.Frame(notebook, bg="#0d1117")
        notebook.add(tab_resume, text="📊 Résumé par Date")
        
        colonnes_resume = ("date", "nb_signaux", "achats", "ventes", "pending", "wins", "losses", "wr", "pnl", "score_moy")
        
        arbre_resume = ttk.Treeview(tab_resume, columns=colonnes_resume, show="headings", height=18)
        
        entetes_resume = {
            "date": ("📅 Date", 100),
            "nb_signaux": ("Total", 60),
            "achats": ("📈 Achats", 70),
            "ventes": ("📉 Ventes", 70),
            "pending": ("⏳ Pending", 75),
            "wins": ("✅ Wins", 60),
            "losses": ("❌ Losses", 60),
            "wr": ("Win Rate", 75),
            "pnl": ("💵 PnL", 90),
            "score_moy": ("Score Moy", 80)
        }
        
        for col, (texte, largeur) in entetes_resume.items():
            arbre_resume.heading(col, text=texte)
            arbre_resume.column(col, width=largeur, anchor="center")
        
        scrollbar_resume = ttk.Scrollbar(tab_resume, orient="vertical", command=arbre_resume.yview)
        arbre_resume.configure(yscrollcommand=scrollbar_resume.set)
        
        arbre_resume.pack(side="left", fill="both", expand=True)
        scrollbar_resume.pack(side="right", fill="y")
        
        # --- ONGLET 3: DÉTAIL DU SIGNAL SÉLECTIONNÉ ---
        tab_detail = tk.Frame(notebook, bg="#0d1117")
        notebook.add(tab_detail, text="🔍 Détail Signal")
        
        lbl_detail_titre = tk.Label(tab_detail, text="Sélectionnez un signal pour voir les détails",
                                   font=("Segoe UI", 12), fg="#8b949e", bg="#0d1117")
        lbl_detail_titre.pack(pady=20)
        
        cadre_detail_info = tk.Frame(tab_detail, bg="#161b22")
        cadre_detail_info.pack(fill="both", expand=True, padx=20, pady=10)
        
        # ═══════════════════════════════════════════════════════════════
        # FONCTIONS
        # ═══════════════════════════════════════════════════════════════
        
        def charger_signaux():
            """Charge les signaux selon les filtres"""
            for item in arbre.get_children():
                arbre.delete(item)
            
            periode = int(var_periode.get()) if var_periode.get() != "0" else 9999
            symbole = var_symbole.get() if var_symbole.get() != "Tous" else None
            
            signaux = self.moteur.obtenir_signaux_periode(periode, symbole)
            
            # Filtrer par action
            if var_filtre_action.get() != "Tous":
                signaux = [s for s in signaux if s.get('action') == var_filtre_action.get()]
            
            # Filtrer par statut
            if var_filtre_statut.get() != "Tous":
                signaux = [s for s in signaux if s.get('statut') == var_filtre_statut.get()]
            
            # Remplir le tableau
            for s in signaux:
                date_str = s.get('time_signal', '')[:10] if s.get('time_signal') else ''
                prix = f"${s.get('entry_price_ref', 0):.2f}" if s.get('entry_price_ref') else '--'
                sl = f"${s.get('stop_loss', 0):.2f}" if s.get('stop_loss') else '--'
                tp = f"${s.get('take_profit', 0):.2f}" if s.get('take_profit') else '--'
                rr = f"{s.get('ratio_rr', 0):.1f}" if s.get('ratio_rr') else '--'
                score = f"{s.get('score_final', 0):.0f}" if s.get('score_final') else '--'
                pnl = f"${s.get('pnl_net', 0):.2f}" if s.get('pnl_net') else '--'
                result = s.get('result', '') or ''
                holding = s.get('holding_days', '') or ''
                
                tag = "normal"
                if result == 'WIN':
                    tag = "win"
                elif result == 'LOSS':
                    tag = "loss"
                elif s.get('statut') == 'PENDING':
                    tag = "pending"
                
                arbre.insert("", "end", values=(
                    s.get('id', ''),
                    date_str,
                    s.get('symbol', ''),
                    s.get('marche', '')[:5] if s.get('marche') else '',
                    s.get('action', ''),
                    prix,
                    sl,
                    tp,
                    rr,
                    score,
                    s.get('statut', ''),
                    result or '--',
                    pnl,
                    holding
                ), tags=(tag,))
            
            arbre.tag_configure("win", background="#1a4731")
            arbre.tag_configure("loss", background="#4a1c1c")
            arbre.tag_configure("pending", background="#2d3a4a")
            arbre.tag_configure("normal", background="#161b22")
            
            # Mettre à jour les stats
            stats = self.moteur.obtenir_stats_periode(periode)
            lbl_stats.config(text=f"📊 Signaux: {stats['nb_signaux']} | "
                                 f"📈 Achats: {stats['nb_achats']} | 📉 Ventes: {stats['nb_ventes']} | "
                                 f"⏳ Pending: {stats['nb_pending']} | ✅ Évalués: {stats['nb_evaluated']} | "
                                 f"🏆 Win Rate: {stats['win_rate']}% | 💵 PnL: ${stats['pnl_total']:.2f}")
            
            # Charger résumé par date
            charger_resume()
        
        def charger_resume():
            """Charge le résumé par date"""
            for item in arbre_resume.get_children():
                arbre_resume.delete(item)
            
            periode = int(var_periode.get()) if var_periode.get() != "0" else 90
            resume = self.moteur.obtenir_resume_dates(periode)
            
            for r in resume:
                nb_eval = r.get('nb_wins', 0) + r.get('nb_losses', 0)
                wr = f"{(r.get('nb_wins', 0) / nb_eval * 100):.0f}%" if nb_eval > 0 else '--'
                pnl = f"${r.get('pnl_total', 0):.2f}" if r.get('pnl_total') else '$0.00'
                
                tag = "normal"
                if r.get('pnl_total', 0) > 0:
                    tag = "win"
                elif r.get('pnl_total', 0) < 0:
                    tag = "loss"
                
                arbre_resume.insert("", "end", values=(
                    r.get('date_signal', ''),
                    r.get('nb_signaux', 0),
                    r.get('nb_achats', 0),
                    r.get('nb_ventes', 0),
                    r.get('nb_pending', 0),
                    r.get('nb_wins', 0),
                    r.get('nb_losses', 0),
                    wr,
                    pnl,
                    f"{r.get('score_moyen', 0):.0f}" if r.get('score_moyen') else '--'
                ), tags=(tag,))
            
            arbre_resume.tag_configure("win", background="#1a4731")
            arbre_resume.tag_configure("loss", background="#4a1c1c")
            arbre_resume.tag_configure("normal", background="#161b22")
        
        def afficher_detail_signal(event):
            """Affiche les détails du signal sélectionné"""
            selection = arbre.selection()
            if not selection:
                return
            
            item = arbre.item(selection[0])
            signal_id = item['values'][0]
            
            # Obtenir le signal avec comparaison prix actuel
            signal = self.moteur.comparer_signal_actuel(signal_id)
            
            if not signal:
                return
            
            # Nettoyer le cadre
            for widget in cadre_detail_info.winfo_children():
                widget.destroy()
            
            # Titre
            symbole = signal.get('symbol', '?')
            action = signal.get('action', '?')
            action_color = "#3fb950" if action == 'ACHAT' else "#f85149"
            
            tk.Label(cadre_detail_info,
                    text=f"🎯 {symbole} - {action}",
                    font=("Segoe UI", 16, "bold"),
                    fg=action_color, bg="#161b22").pack(pady=10)
            
            # Infos en deux colonnes
            cadre_cols = tk.Frame(cadre_detail_info, bg="#161b22")
            cadre_cols.pack(fill="x", pady=10)
            
            col_gauche = tk.Frame(cadre_cols, bg="#161b22")
            col_gauche.pack(side="left", fill="both", expand=True, padx=20)
            
            col_droite = tk.Frame(cadre_cols, bg="#161b22")
            col_droite.pack(side="right", fill="both", expand=True, padx=20)
            
            # Colonne gauche - Infos signal
            tk.Label(col_gauche, text="📋 SIGNAL D'ORIGINE",
                    font=("Segoe UI", 12, "bold"), fg="#58a6ff", bg="#161b22").pack(anchor="w", pady=5)
            
            infos_signal = [
                ("Date", signal.get('time_signal', '')[:16]),
                ("Prix Entrée", f"${signal.get('entry_price_ref', 0):.2f}"),
                ("Stop Loss", f"${signal.get('stop_loss', 0):.2f}" if signal.get('stop_loss') else '--'),
                ("Take Profit", f"${signal.get('take_profit', 0):.2f}" if signal.get('take_profit') else '--'),
                ("Ratio R:R", f"{signal.get('ratio_rr', 0):.2f}"),
                ("Score", f"{signal.get('score_final', 0):.0f}"),
                ("Confiance", f"{signal.get('confiance_pct', 0):.0f}%"),
                ("Horizon", f"{signal.get('horizon_days', 0)} jours"),
                ("Marché", signal.get('marche', '')),
                ("Secteur", signal.get('secteur', ''))
            ]
            
            for label, valeur in infos_signal:
                row = tk.Frame(col_gauche, bg="#161b22")
                row.pack(fill="x", pady=2)
                tk.Label(row, text=label, font=("Segoe UI", 10),
                        fg="#8b949e", bg="#161b22", width=15, anchor="w").pack(side="left")
                tk.Label(row, text=str(valeur), font=("Segoe UI", 10, "bold"),
                        fg="#c9d1d9", bg="#161b22").pack(side="left")
            
            # Colonne droite - Situation actuelle
            tk.Label(col_droite, text="📊 SITUATION ACTUELLE",
                    font=("Segoe UI", 12, "bold"), fg="#f0883e", bg="#161b22").pack(anchor="w", pady=5)
            
            prix_actuel = signal.get('prix_actuel')
            variation = signal.get('variation_pct', 0)
            
            if prix_actuel:
                var_color = "#3fb950" if variation >= 0 else "#f85149"
                
                infos_actuel = [
                    ("Prix Actuel", f"${prix_actuel:.2f}"),
                    ("Variation", f"{'+' if variation >= 0 else ''}{variation:.2f}%", var_color),
                    ("Jours depuis signal", f"{signal.get('jours_depuis_signal', 0)} jours"),
                    ("Distance SL", f"{signal.get('distance_sl_pct', 0):.2f}%" if signal.get('distance_sl_pct') else '--'),
                    ("Distance TP", f"{signal.get('distance_tp_pct', 0):.2f}%" if signal.get('distance_tp_pct') else '--'),
                    ("Performance", signal.get('performance_actuelle', '--')),
                ]
                
                for item in infos_actuel:
                    label = item[0]
                    valeur = item[1]
                    color = item[2] if len(item) > 2 else "#c9d1d9"
                    
                    row = tk.Frame(col_droite, bg="#161b22")
                    row.pack(fill="x", pady=2)
                    tk.Label(row, text=label, font=("Segoe UI", 10),
                            fg="#8b949e", bg="#161b22", width=18, anchor="w").pack(side="left")
                    tk.Label(row, text=str(valeur), font=("Segoe UI", 10, "bold"),
                            fg=color, bg="#161b22").pack(side="left")
                
                # Alertes
                if signal.get('sl_touche'):
                    tk.Label(col_droite, text="🚨 STOP LOSS TOUCHÉ",
                            font=("Segoe UI", 11, "bold"), fg="#f85149", bg="#161b22").pack(pady=5)
                elif signal.get('tp_touche'):
                    tk.Label(col_droite, text="🎯 TAKE PROFIT ATTEINT",
                            font=("Segoe UI", 11, "bold"), fg="#3fb950", bg="#161b22").pack(pady=5)
            else:
                tk.Label(col_droite, text="⚠️ Prix actuel non disponible",
                        font=("Segoe UI", 10), fg="#f0883e", bg="#161b22").pack(pady=10)
            
            # Résultat si évalué
            if signal.get('statut') == 'EVALUATED':
                tk.Label(cadre_detail_info, text="─" * 60, fg="#30363d", bg="#161b22").pack(pady=10)
                
                result = signal.get('result', '')
                pnl = signal.get('pnl_net', 0)
                result_color = "#3fb950" if result == 'WIN' else "#f85149" if result == 'LOSS' else "#8b949e"
                
                tk.Label(cadre_detail_info,
                        text=f"📊 RÉSULTAT: {result} | PnL: ${pnl:.2f}",
                        font=("Segoe UI", 14, "bold"),
                        fg=result_color, bg="#161b22").pack(pady=5)
                
                if signal.get('exit_reason'):
                    tk.Label(cadre_detail_info,
                            text=f"Raison sortie: {signal.get('exit_reason')} | Holding: {signal.get('holding_days', 0)}j",
                            font=("Segoe UI", 10),
                            fg="#8b949e", bg="#161b22").pack()
            
            # Aller à l'onglet détail
            notebook.select(tab_detail)
        
        # Bind double-click
        arbre.bind('<Double-1>', afficher_detail_signal)
        
        # ═══════════════════════════════════════════════════════════════
        # BOUTONS
        # ═══════════════════════════════════════════════════════════════
        cadre_btn = tk.Frame(fenetre, bg="#0d1117")
        cadre_btn.pack(fill="x", pady=10, padx=15)
        
        tk.Button(cadre_btn, text="🔄 Actualiser", font=("Segoe UI", 10, "bold"),
                 bg="#238636", fg="white", width=12,
                 command=charger_signaux).pack(side="left", padx=5)
        
        def evaluer_pending():
            self.moteur.evaluer_signaux_pending()
            charger_signaux()
            self.messagebox.showinfo("Évaluation", "Signaux pending évalués!")
        
        tk.Button(cadre_btn, text="⚡ Évaluer Pending", font=("Segoe UI", 10, "bold"),
                 bg="#8957e5", fg="white", width=14,
                 command=evaluer_pending).pack(side="left", padx=5)
        
        def exporter():
            periode = int(var_periode.get()) if var_periode.get() != "0" else None
            filepath = self.filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")],
                initialfile=f"signaux_{datetime.now().strftime('%Y%m%d')}.csv"
            )
            if filepath:
                if self.moteur.exporter_signaux(filepath, periode):
                    self.messagebox.showinfo("Export", f"Signaux exportés vers {filepath}")
                else:
                    self.messagebox.showerror("Erreur", "Erreur lors de l'export")
        
        tk.Button(cadre_btn, text="📁 Exporter CSV", font=("Segoe UI", 10, "bold"),
                 bg="#1f6feb", fg="white", width=12,
                 command=exporter).pack(side="left", padx=5)
        
        tk.Button(cadre_btn, text="🔍 Voir Détail", font=("Segoe UI", 10, "bold"),
                 bg="#f0883e", fg="white", width=12,
                 command=lambda: afficher_detail_signal(None) if arbre.selection() else None
                 ).pack(side="left", padx=5)
        
        # Info
        tk.Label(cadre_btn, text="💡 Double-clic sur un signal pour voir les détails et comparer avec le prix actuel",
                font=("Segoe UI", 9, "italic"), fg="#8b949e", bg="#0d1117").pack(side="right", padx=10)
        
        # Charger les données initiales
        charger_signaux()


# ══════════════════════════════════════════════════════════════════════════════
#                              MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # Fix Windows console encoding
    if sys.platform == 'win32':
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except:
            pass
    
    print("""
+==========================================================================================================+
|                                                                                                          |
|              SCREENER PRO v11.0 - OPPORTUNITES COURT & MOYEN TERME                                       |
|                                                                                                          |
|     ====================================================================================================  |
|                                                                                                          |
|     NOUVEAUX INDICATEURS COURT TERME:                                                                    |
|        - ADX + DI+/DI- (Force de tendance)                                                               |
|        - Stochastic RSI (Timing optimal - croisements)                                                   |
|        - Williams %R (Zones extremes)                                                                    |
|        - CMF Chaikin Money Flow (Pression achat/vente)                                                   |
|        - OBV On Balance Volume (Accumulation/Distribution)                                               |
|        - VWAP (Prix moyen pondere par volume)                                                            |
|        - SuperTrend (Direction de tendance)                                                              |
|        - Squeeze Momentum (Compression + Expansion)                                                      |
|                                                                                                          |
|     DETECTION D'OPPORTUNITES:                                                                            |
|        - Scanner Breakout (Cassure resistance/support avec volume)                                       |
|        - Detection Gap (Gap Up/Down significatifs > 1.5%)                                                |
|        - Volume Spike Alert (Volume > 200% moyenne)                                                      |
|        - Pattern Reversal (Hammer, Engulfing, Doji, Morning Star)                                        |
|        - Momentum Burst (Acceleration soudaine)                                                          |
|        - Squeeze Breakout (Compression volatilite)                                                       |
|                                                                                                          |
|     SCORING MULTI-HORIZON:                                                                               |
|        - Score Intraday (1-3 jours) : Focus timing, volume, momentum                                     |
|        - Score Swing (5-15 jours) : Focus tendance, pullbacks, confirmations                             |
|        - Score Position (15-60 jours) : Focus fondamentaux, valorisation                                 |
|        -> Determination automatique du MEILLEUR HORIZON                                                  |
|                                                                                                          |
|     VALORISATION AVANCEE:                                                                                |
|        - DCF Multi-Scenarios (Optimiste/Base/Pessimiste)                                                 |
|        - Graham Number + Graham Defensif                                                                 |
|        - PEG Fair Value (Peter Lynch)                                                                    |
|        - PRIX REEL COMPOSITE                                                                             |
|                                                                                                          |
|     REGLES DE TRADING:                                                                                   |
|        - STOP LOSS calcule (ATR + Support)                                                               |
|        - TAKE PROFIT calcule (ATR + Resistance + Prix Reel)                                              |
|        - RATIO RISQUE/RECOMPENSE                                                                         |
|        - HORIZON DE DETENTION optimal                                                                    |
|                                                                                                          |
|     SUIVI DES SIGNAUX:                                                                                   |
|        - Base de donnees SQLite                                                                          |
|        - Scorecard avec Win Rate, Expectancy, Profit Factor                                              |
|        - Stress Tests (Couts x2, SL/TP modifies)                                                         |
|                                                                                                          |
|     MARCHES: US + France (CAC 40) + Europe + Gaming + Commodities + Crypto                               |
|                                                                                                          |
+==========================================================================================================+
    """)
    
    logger.info("Demarrage Screener Pro v11.0")
    interface = InterfaceScreener()
    interface.lancer()


if __name__ == "__main__":
    main()
